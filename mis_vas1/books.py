from django.shortcuts import render, redirect
from django.http import HttpResponse
import requests, json
from django.contrib import messages
from MIS.functions import validate_file_format, validate_file_size
from datetime import datetime
from user_management.settings_views import *
from reportlab.lib.enums import TA_LEFT, TA_CENTER

API_STUDIO_URL = user_bundle_settings()

def book_key():
    url = "https://api.hcaschennai.edu.in/auth/token"

    payload = json.dumps({
        "secret_key": "C4ZoXbsAnHLjk1Xyz4QPT2eoiNx6K6fo"
    })
    headers = {'Content-Type': 'application/json'}

    response = requests.post(url, headers=headers, data=payload)
    if response.status_code == 200:
        res_data = response.json()
        access_token = res_data.get('access_token')
        token_type = res_data.get('token_type')
        return access_token, token_type
    return None, None

# Function to fetch staff career data
def get_book_data(access_token, token_type):
    url = "https://api.hcaschennai.edu.in/sqlviews/api/v1/auth/get_response_data"

    payload = json.dumps({
        "psk_uid": "51a531b4-bd55-491c-861d-a8d7227b325b",
        "project": "hcas",
        "data": {}
    })
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'{token_type} {access_token}'  # Fix spacing
    }

    response = requests.post(url, headers=headers, data=payload)
    if response.status_code == 200:
        return response.json()
    return []

def book_list(request):
    url = f"{API_STUDIO_URL}getapi/naac01_books_and_chapter_dc1/all"
    response = requests.get(url)
    
    if response.status_code == 200:
        books_data = response.json()
    else:
        return HttpResponse("API Call Is Not Working")

    user = get_settings(request)
    username = user.get('username')
    # username = 'CS-T155'
    filtered_books = [book for book in books_data if book.get('staff_id') == username]
    
    selected_staff_id = request.GET.get('staff_id')
    if selected_staff_id:
        filtered_books = [book for book in books_data if book.get('staff_id') == selected_staff_id]
        
    #from_dashboard = request.GET.get('from') == 'dashboard' or 'admin_hod_dash' or 'admin_dash' or 'department_dashboard'

    if not filtered_books:
        return render(request, 'Books_templates/book_list.html', {"list": []})
    
    return render(request, 'Books_templates/book_list.html', {"list": filtered_books, 
    #'from_dashboard': from_dashboard
    })

def Create(request):
    error_message = None

    # Step 1: Get API token
    access_token, token_type = book_key()
    if not access_token or not token_type:
        error_message = 'Failed to retrieve access token. Please try again later.'
        return render(request, 'Books_templates/create_book.html', {'error': error_message})

    # Step 2: Fetch book data
    book_data = get_book_data(access_token, token_type)
    if not book_data:
        error_message = 'Failed to fetch staff data. Please contact support.'
        return render(request, 'Books_templates/create_book.html', {'error': error_message})
    
    current_year = datetime.now().year
    year_of_publication = [f"{y}-{y+1}" for y in range(current_year, 2010, -1)]    
    
    user = get_settings(request)
    username = user.get('username')
    # username = 'CS-T155'
    
    
    selected_faculty = None
    
    for faculty in book_data:
        if faculty['stf_id'] == username:
            selected_faculty = faculty
            break
    stf_name = selected_faculty.get('stf_name', '') if selected_faculty else ''
    department = selected_faculty.get('department', '') if selected_faculty else ''
        
        
    # Step 3: Handle POST form submission
    if request.method == 'POST':
        try:
            # Extract form data
            staff_name = request.POST.get('staff_name')
            staff_id = username
            department_name = request.POST.get('department_name')
            publication_type = request.POST.getlist('publication_type')
            title_of_book_or_chapter = request.POST.get('title_of_book_or_chapter')
            title_of_paper = request.POST.get('title_of_paper', '')
            title_of_proceedings = request.POST.get('title_of_proceedings', '')
            name_of_conference = request.POST.get('name_of_conference', '')
            year_of_publication = request.POST.get('year_of_publication')
            national_or_international = request.POST.getlist('national_or_international')
            isbn_or_issn_number = request.POST.get('isbn_or_issn_number', '')
            affiliating_institute = request.POST.get('affiliating_institute')
            name_of_the_publisher = request.POST.get('name_of_the_publisher')

            # Convert list fields to comma-separated strings
            publication_type_str = ', '.join(publication_type)
            national_or_international_str = ', '.join(national_or_international)
            
            # selected_faculty = None
            
            for faculty in book_data:
                if faculty['stf_id'] == username:
                    selected_faculty = faculty
                    print("selected_faculty:", selected_faculty)
                    break
               
            if selected_faculty:
                department_name = selected_faculty.get('department', '')
                staff_name = selected_faculty.get('stf_name', '')
                print("selected_faculty", selected_faculty) 

            # API call to create book record
            create_url = f"{API_STUDIO_URL}postapi/create/naac01_books_and_chapter_dc1"
            payload = json.dumps({
                "data": {
                    "staff_name": staff_name,
                    "staff_id": staff_id,
                    "department_name": department_name,
                    "publication_type": publication_type_str,
                    "title_of_book_or_chapter": title_of_book_or_chapter,
                    "title_of_paper": title_of_paper,
                    "title_of_proceedings": title_of_proceedings,
                    "name_of_conference": name_of_conference,
                    "year_of_publication": year_of_publication,
                    "national_or_international": national_or_international_str,
                    "isbn_or_issn_number": isbn_or_issn_number,
                    "affiliating_institute": affiliating_institute,
                    "name_of_the_publisher": name_of_the_publisher
                }
            })
            headers = {'Content-Type': 'application/json'}
            response = requests.post(create_url, headers=headers, data=payload)

            if response.status_code == 200:
                # Successfully created book record
                response_data = response.json()
                psk_id = response_data.get('psk_id')
                messages.success(request, "Book record created successfully!")

                # Handle file uploads
                uploaded_files = request.FILES.getlist('file')
                if uploaded_files:
                    upload_url = f"{API_STUDIO_URL}crudapp/upload/media/naac01_books_and_chapter_dc1_media"
                    fields = ['CCP', 'CC', 'IP', 'POI']  # Define the fields for files (adjust if needed)
                    current_year = datetime.now().year

                    for field, uploaded_file in zip(fields, uploaded_files):
                        # Validate file size and format
                        validate_file_size(uploaded_file)
                        validate_file_format(uploaded_file)
                        file_type = uploaded_file.content_type

                        # Generate custom filename
                        custom_filename = f"{staff_id}_{field}_{current_year}_{uploaded_file.name}"
                        # print(f"Generated filename: {custom_filename}")  # Print the filename for each iteration

                        payload = {'parent_psk_id': psk_id}
                        files = {'media': (custom_filename, uploaded_file, file_type)}
                        upload_headers = {'api_name': 'naac01_books_and_chapter_dc1_media'}

                        # Upload the file
                        upload_response = requests.post(upload_url, headers=upload_headers, data=payload, files=files)

                        if upload_response.status_code != 200:
                            messages.error(
                                request,
                                f"File upload failed for {uploaded_file.name}. Error: {upload_response.text}"
                            )
                            return redirect('book_list')

                    messages.success(request, "All documents uploaded successfully!")
                else:
                    messages.warning(request, "No files were selected for upload.")

                # Redirect to the book list view
                return redirect('book_list')
            else:
                messages.error(request, "Failed to create book record. Please try again.")
                return render(request, 'Books_templates/create_book.html', {'error': response.text})

        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
            return render(request, 'Books_templates/create_book.html')

    # If the request is GET, render the form
    return render(request, 'Books_templates/create_book.html', {'book_data': book_data, 'username': username,  # Pass 'username' (which is the staff_id) to the template
        'department': department,
        'year_of_publication': year_of_publication,
        'stf_name': stf_name})

def delete(request, psk_id):
    # if request.method == 'POST':
    url = f"{API_STUDIO_URL}deleteapi/delete/naac01_books_and_chapter_dc1/{psk_id}"

    payload = {}
    headers = {}

    response = requests.request("DELETE", url, headers=headers, data=payload)

    if response.status_code == 200:
        messages.success(request, message=f"The Article was deleted successfully.")
        return redirect('book_list')
    else:
        error_msg = response.json()
        messages.error(request, message=f"{error_msg.get('book_list', 'Failed to delete participation')}")

def book_update(request, id):
    # Fetch extension details
    url = f"{API_STUDIO_URL}getapi/naac01_books_and_chapter_dc1/{id}"
    response = requests.get(url)

    if response.status_code == 200:
        books_chapter = response.json()
    else:
        return HttpResponse(f"Error fetching Books details: {response.text}", status=500)

    # Fetch media (child files) associated with this extension
    media_url = f"{API_STUDIO_URL}crudapp/get/media/naac01_books_and_chapter_dc1_media/parent/{id}"
    media_response = requests.get(media_url)

    if media_response.status_code == 200:
        child_files = media_response.json()
    else:
        return HttpResponse(f"Failed to fetch media files: {media_response.text}", status=500)
    
    current_year = datetime.now().year
    year_of_publication = [f"{y}-{y+1}" for y in range(current_year, 2010, -1)]

    # Handle form submission (POST request)
    if request.method == "POST":
        staff_name = request.POST.getlist('staff_name')
        staff_id = request.POST.getlist('staff_id')
        department_name = request.POST.getlist('department_name')

        publication_type = request.POST.getlist('publication_type')
        title_of_book_or_chapter = request.POST.get('title_of_book_or_chapter')
        title_of_paper = request.POST.get('title_of_paper')
        title_of_proceedings = request.POST.get('title_of_proceedings')
        name_of_conference = request.POST.get('name_of_conference')
        year_of_publication = request.POST.get('year_of_publication')
        national_or_international = request.POST.getlist('national_or_international')
        isbn_or_issn_number = request.POST.get('isbn_or_issn_number')
        affiliating_institute = request.POST.get('affiliating_institute')
        name_of_the_publisher = request.POST.get('name_of_the_publisher')

        publication_type_str = ', '.join(publication_type)
        national_or_international_str = ', '.join(national_or_international)

        # Prepare data for updating the extension/workshop
        update_url = f"{API_STUDIO_URL}updateapi/update/naac01_books_and_chapter_dc1/{id}"

        payload = json.dumps({
            "data": {
                "publication_type": publication_type_str,
                "staff_name": staff_name,
                "staff_id": staff_id,
                'department_name': department_name,
                "title_of_book_or_chapter": title_of_book_or_chapter,
                "title_of_paper": title_of_paper,
                "title_of_proceedings": title_of_proceedings,
                "name_of_conference": name_of_conference,
                "year_of_publication": year_of_publication,
                "national_or_international": national_or_international_str,
                "isbn_or_issn_number": isbn_or_issn_number,
                "affiliating_institute": affiliating_institute,
                "name_of_the_publisher": name_of_the_publisher
            }
        })

        headers = {'Content-Type': 'application/json'}
        update_response = requests.put(update_url, headers=headers, data=payload)

        if update_response.status_code != 200:
            return HttpResponse(f"Failed to update Books details: {update_response.text}", status=500)

        # Handle media (file) uploads
        upload_errors = []

        for child in child_files:
            upload_id = child['psk_id']
            fields = ['CCP', 'CC', 'IP', 'POI']  # Example fields to match

            for field in fields:
                # Dynamically get the list of uploaded files for this specific upload_id and field
                uploaded_files = request.FILES.getlist(f'file_{upload_id}_{field.upper()}')

                if not uploaded_files:
                    continue  # No files to upload for this field

                for uploaded_file in uploaded_files:
                    # Validate file size and format
                    try:
                        validate_file_size(uploaded_file)
                        validate_file_format(uploaded_file)
                    except Exception as e:
                        messages.error(request, str(e))
                        continue  # Skip file upload if validation fails

                    # Generate custom filename
                    current_year = datetime.now().year  # Get the current year
                    custom_filename = f"{staff_id[0]}_{field}_{current_year}_{uploaded_file.name}"

                    print(f"Generated filename for field {field}: {custom_filename}")  # Log the filename for debugging

                    upload_url = f"{API_STUDIO_URL}crudapp/upload/media/naac01_books_and_chapter_dc1_media/{upload_id}"

                    # Prepare the file payload and headers for the upload request
                    files = {'media': (custom_filename, uploaded_file, uploaded_file.content_type)}
                    payload = {'parent_psk_id': id}
                    headers = {'api_name': 'naac01_books_and_chapter_dc1_media', 'psk_id': str(upload_id)}

                    # Upload the file
                    upload_response = requests.put(upload_url, headers=headers, data=payload, files=files)

                    if upload_response.status_code != 200:
                        upload_errors.append(f"Failed to upload {uploaded_file.name}: {upload_response.text}")
                    else:
                        print(f"File uploaded successfully: {custom_filename}")  # Log successful upload

        # If there were upload errors, handle them
        if upload_errors:
            for error in upload_errors:
                messages.error(request, error)
        else:
            messages.success(request, "Files uploaded successfully.")

        # Redirect to the extension view page after successful update
        return redirect('book_list')

    # If not a POST request, render the update form
    return render(request, 'Books_templates/book_update.html', {'books': books_chapter, 'child_files': child_files, 'year_of_publication':year_of_publication})

import io
import pandas as pd
import requests
import json
from django.http import HttpResponse
from django.shortcuts import render
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from datetime import datetime
from openpyxl.styles import Font
from django.contrib import messages
from collections import Counter
from user_management.settings_views import get_settings
from MIS.functions import validate_file_format_faculty, validate_file_size
from user_management.settings_views import *

def filter_books(request):
    """
    Main filtering function for books and chapters with Excel and PDF export
    with role-based access control (Staff vs HOD)
    """
    # Step 1: Get the logged-in username
    user = get_settings(request)
    username = user.get('username', '')
    
    # Step 2: Determine user role (Staff or HOD)
    role_url = f"{API_STUDIO_URL}getapi/asa0504_01_01"
    payload = json.dumps({"queries": [{"field": "username", "value": username, "operation": "equal"}], "search_type": "first"})
    role_response = requests.post(role_url, headers={'Content-Type': 'application/json'}, data=payload)
    value_user = int(role_response.json().get("user_roles", "0").strip("{}")) if role_response.status_code == 200 else 0

    role_list = roles_tbl(request)
    user_role = next((role.get("user_role") for role in role_list if role.get("psk_id") == value_user), "Staff")
    # user_role = next((role['user_role'] for role in role_list if role['psk_id'] == value_user), 'Staff')
    
    
    # Step 3: Get research data to determine department for HOD
    access_token, token_type = book_key()
    research_data = get_book_data(access_token, token_type)
    staff_info = next((staff for staff in research_data if staff.get("stf_id") == username), {})
    department_name = staff_info.get("department", "")
    
    # Get all books data
    books_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_books_and_chapter_dc1/all"
    books_response = requests.get(books_url)
    
    if books_response.status_code != 200:
        return HttpResponse("Error fetching books data")
    
    all_books = books_response.json()
    
    
    
    # Get filter parameters from request
    stf_id = request.GET.get('stf_id', '').strip()
    from_year = request.GET.get('from_year', '').strip()
    to_year = request.GET.get('to_year', '').strip()
    publication_type = request.GET.getlist('publication_type')
    national_international = request.GET.getlist('national_international')
    export_format = request.GET.get('export', '')
    
    # ROLE-BASED FILTERING: Apply department filter for HOD
    if user_role == "Hod":
        # HOD can only see staff from their own department
        department_staff = [staff for staff in research_data if staff.get("department") == department_name]
        dept_staff_ids = [staff.get("stf_id") for staff in department_staff]
        
        # Filter books to only those in HOD's department
        all_books = [book for book in all_books if book.get('staff_id') in dept_staff_ids]
    
    # Filter books based on user selections
    filtered_books = all_books
    
    # Staff ID filter
    if user_role == "Hod" and stf_id:
        filtered_books = [book for book in filtered_books if book.get("staff_id") == stf_id]
    elif user_role == "Staff":
        filtered_books = [book for book in filtered_books if book.get("staff_id") == username]

    # Year range filter
    if from_year:
        try:
            from_year_int = int(from_year.split('-')[0])
            filtered_books = [book for book in filtered_books if book.get('year_of_publication') and extract_year_value(book.get('year_of_publication')) >= from_year_int]
        except (ValueError, AttributeError):
            pass
    
    if to_year:
        try:
            to_year_int = int(to_year.split('-')[0])
            filtered_books = [book for book in filtered_books if book.get('year_of_publication') and extract_year_value(book.get('year_of_publication')) <= to_year_int]
        except (ValueError, AttributeError):
            pass
    
    # Publication type filter
    if publication_type:
        filtered_books = [book for book in filtered_books if any(opt in (book.get('publication_type') or '') for opt in publication_type)]
    
    # National/International filter
    if national_international:
        filtered_books = [book for book in filtered_books if any(opt in (book.get('national_or_international') or '') for opt in national_international)]
    
    # Fetch media files for each book
    for book in filtered_books:
        book_id = book.get('psk_id')
        media_url = f"{API_STUDIO_URL}crudapp/get/media/naac01_books_and_chapter_dc1_media/parent/{book_id}"
        media_response = requests.get(media_url)
        
        if media_response.status_code == 200:
            media_data = media_response.json()
            # Process each media file to get individual media IDs
            processed_media_files = []
            for media in media_data:
                # Extract media ID from the media object
                media_id = media.get('id') or media.get('media_id') or media.get('psk_id') or media.get('value_id')
                
                if media_id:
                    # Create direct API URL using media ID
                    direct_api_url = f"{API_STUDIO_URL}crudapp/view/media/naac01_books_and_chapter_dc1_media/{media_id}"
                    
                    # Add the media URL and other media info
                    processed_media = {
                        'file_name': media.get('file_name', 'Unknown'),
                        'media_id': media_id,
                        'direct_api_url': direct_api_url,
                        'original_data': media
                    }
                    processed_media_files.append(processed_media)
                else:
                    # Fallback if no media ID found - use book ID as before
                    fallback_url = f"{API_STUDIO_URL}crudapp/view/media/naac01_books_and_chapter_dc1_media/{book_id}"
                    processed_media = {
                        'file_name': media.get('file_name', 'Unknown'),
                        'media_id': None,
                        'direct_api_url': fallback_url,
                        'original_data': media
                    }
                    processed_media_files.append(processed_media)
            
            book['media_files'] = processed_media_files
        else:
            book['media_files'] = []
    
    # Pre-split comma-separated fields for template display
    for book in filtered_books:
        # Split publication_type
        if book.get('publication_type'):
            book['publication_type_list'] = [part.strip() for part in book['publication_type'].split(',')]
        else:
            book['publication_type_list'] = []
        
        # Split national_or_international
        if book.get('national_or_international'):
            book['national_international_list'] = [part.strip() for part in book['national_or_international'].split(',')]
        else:
            book['national_international_list'] = []
    
    # Handle export formats
    if export_format:
        if export_format.lower() == 'excel':
            return export_books_to_excel(filtered_books, publication_type, national_international)
        elif export_format.lower() == 'pdf':
            return export_books_and_chapters_to_pdf(filtered_books, publication_type, national_international)
    
    # Get unique staff IDs for dropdown - filtered by role
    if user_role == "Hod":
        # HOD can see all staff in their department
        department_staff = [s for s in research_data if s.get("department") == department_name]
        dept_staff_ids = [s.get("stf_id") for s in department_staff]
        staff_ids = sorted(list(set(book.get('staff_id') for book in all_books if book.get('staff_id') in dept_staff_ids)))
    else:
        # Staff can only see themselves
        staff_ids = [username]
    
    # Get unique years for dropdown
    year_values = []
    for book in all_books:
        if book.get('year_of_publication'):
            try:
                year_val = extract_year_value(book.get('year_of_publication'))
                year_values.append(year_val)
            except (ValueError, AttributeError):
                continue
    
    # Convert back to "YYYY-YYYY" format for display
    years = sorted(list(set(f"{y}-{y+1}" for y in year_values)))
    
    # Define filter options
    PUBLICATION_OPTIONS = ['Book', 'Chapter', 'Conference Proceedings', 'Journal Article']
    NATIONAL_INTERNATIONAL_OPTIONS = ['National', 'International']
    
    context = {
        'books': filtered_books,
        'staff_ids': staff_ids,
        'years': years,
        'publication_options': PUBLICATION_OPTIONS,
        'national_international_options': NATIONAL_INTERNATIONAL_OPTIONS,
        'selected_stf_id': stf_id,
        'selected_from_year': from_year,
        'selected_to_year': to_year,
        'selected_publication_types': publication_type,
        'selected_national_international': national_international,
        'filter_applied': any([stf_id, from_year, to_year, publication_type, national_international]),
        'user_role': user_role,
        'username': username
    }
    
    return render(request, 'Books_templates/filter_books.html', context)

def extract_year_value(year_str):
    """Extract first year from academic year string or return None"""
    try:
        return int(str(year_str).split('-')[0]) if year_str else None
    except:
        return None

def export_books_to_excel(books, publication_types, national_international):
    """
    Export filtered books data to Excel format
    """
    # Prepare book data for export
    book_data_for_export = []
    for book in books:
        # Create the formatted book data
        book_export = {
            'Staff ID': book.get('staff_id', ''),
            'Staff Name': book.get('staff_name', ''),
            'Department': book.get('department_name', ''),
            'Publication Type': book.get('publication_type', ''),
            'Title of Book/Chapter': book.get('title_of_book_or_chapter', ''),
            'Title of Paper': book.get('title_of_paper', ''),
            'Title of Proceedings': book.get('title_of_proceedings', ''),
            'Conference Name': book.get('name_of_conference', ''),
            'Year of Publication': book.get('year_of_publication', ''),
            'National/International': book.get('national_or_international', ''),
            'ISBN/ISSN Number': book.get('isbn_or_issn_number', ''),
            'Affiliating Institute': book.get('affiliating_institute', ''),
            'Publisher Name': book.get('name_of_the_publisher', '')
        }
        
        # Handle media files
        media_files = book.get('media_files', [])
        if media_files:
            media_links = []
            for media in media_files:
                file_name = media.get('file_name', 'Unknown')
                media_url = media.get('direct_api_url', '')
                
                if media_url:
                    hyperlink_formula = f'=HYPERLINK("{media_url}", "{file_name}")'
                    media_links.append(hyperlink_formula)
                else:
                    media_links.append(file_name)
            
            book_export['Attachments'] = "\n".join(media_links)
        else:
            book_export['Attachments'] = "No media files"
        
        book_data_for_export.append(book_export)
    
    book_df = pd.DataFrame(book_data_for_export)
    
    # Create Excel writer
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write book data
        if not book_df.empty:
            book_df = book_df.fillna('')
            book_df.to_excel(writer, sheet_name='Books and Publications', index=False)
            
            # Apply hyperlink style to Attachments column
            worksheet = writer.sheets['Books and Publications']
            if 'Attachments' in book_df.columns:
                attachment_col_idx = book_df.columns.get_loc('Attachments') + 1
                hyperlink_font = Font(color="0000FF", underline="single")

                for row in range(2, len(book_df) + 2):
                    cell = worksheet.cell(row=row, column=attachment_col_idx)
                    if cell.value and cell.value.startswith("=HYPERLINK("):
                        cell.font = hyperlink_font
        
        # Write summary sheet
        summary_data = {
            'Total Books/Publications': [len(books)],
            'Selected Publication Types': [', '.join(publication_types) if publication_types else 'All'],
            'Selected Scope': [', '.join(national_international) if national_international else 'All'],
            'Export Date': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="books_and_publications_filtered.xlsx"'
    
    return response

from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from reportlab.lib.units import inch
from django.http import HttpResponse

# def export_books_and_chapters_to_pdf(parents, children=None, selected_options=None):
#     """
#     Export Books & Chapters data to PDF (landscape), grouped by staff.
#     All entries for a staff go in a single table; new page/table for a new staff.
#     Attachments are clickable links.
#     """
#     from reportlab.lib.pagesizes import A4, landscape
#     from reportlab.lib import colors
#     from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
#     from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
#     from reportlab.lib.units import inch
#     from django.http import HttpResponse

#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="books_and_chapters.pdf"'

#     doc = SimpleDocTemplate(
#         response,
#         pagesize=landscape(A4),
#         topMargin=0.5*inch,
#         bottomMargin=0.5*inch,
#         leftMargin=0.5*inch,
#         rightMargin=0.5*inch
#     )

#     elements = []
#     styles = getSampleStyleSheet()

#     info_style = ParagraphStyle('InfoStyle', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#2c3e50'))
#     table_header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontSize=9, alignment=1,
#                                         textColor=colors.white, fontName='Helvetica-Bold')
#     table_cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=8, alignment=0,
#                                       textColor=colors.HexColor('#2c3e50'))

#     # Sort by staff to group
#     parents_sorted = sorted(parents, key=lambda x: (x.get('staff_id',''), x.get('title_of_book_or_chapter','')))
#     current_staff = None
#     table_data = []

#     for parent in parents_sorted:
#         staff_id = parent.get('staff_id', 'N/A')
#         staff_name = parent.get('staff_name', 'N/A')
#         department_name = parent.get('department_name', 'N/A')

#         # New staff: build previous table + page break
#         if current_staff and current_staff != staff_id:
#             # Add previous table to elements
#             col_widths = [1.2*inch, 3*inch, 1*inch, 1*inch, 2*inch, 2*inch]
#             table = Table(table_data, colWidths=col_widths)
#             table.setStyle(TableStyle([
#                 ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
#                 ('TEXTCOLOR', (0,0), (-1,0), colors.white),
#                 ('ALIGN',(0,0),(-1,-1),'CENTER'),
#                 ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
#                 ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
#                 ('FONTSIZE', (0,0), (-1,-1), 8),
#                 ('LEFTPADDING', (0,0), (-1,-1), 4),
#                 ('RIGHTPADDING', (0,0), (-1,-1), 4),
#                 ('BOTTOMPADDING', (0,0), (-1,-1), 3),
#                 ('TOPPADDING', (0,0), (-1,-1), 3),
#             ]))
#             elements.append(table)
#             elements.append(PageBreak())
#             table_data = []

#         if current_staff != staff_id:
#             elements.append(Paragraph(f"Staff ID: {staff_id}", info_style))
#             elements.append(Paragraph(f"Name: {staff_name}", info_style))
#             elements.append(Paragraph(f"Department: {department_name}", info_style))
#             elements.append(Spacer(1, 12))

#             # Table headers
#             headers = ['Publication Type', 'Title of Book/Chapter', 'Year of Publication', 'ISBN/ISSN', 'Publisher', 'Attachments']
#             table_data.append([Paragraph(h, table_header_style) for h in headers])
#             current_staff = staff_id

#         # Media attachments as clickable links
#         media_files = parent.get('media_files', [])
#         if media_files:
#             media_text = []
#             for m in media_files:
#                 filename = m.get('file_name', 'Unknown')
#                 url = m.get('direct_api_url', '#')
#                 media_text.append(f'<link href="{url}" color="blue">{filename}</link>')
#             media_paragraph = Paragraph(', '.join(media_text), table_cell_style)
#         else:
#             media_paragraph = Paragraph("No attachments", table_cell_style)

#         # Add row for parent
#         row = [
#             Paragraph(str(parent.get('publication_type', '')), table_cell_style),
#             Paragraph(str(parent.get('title_of_book_or_chapter', '')), table_cell_style),
#             Paragraph(str(parent.get('year_of_publication', '')), table_cell_style),
#             Paragraph(str(parent.get('isbn_or_issn_number', '')), table_cell_style),
#             Paragraph(str(parent.get('name_of_the_publisher', '')), table_cell_style),
#             media_paragraph
#         ]
#         table_data.append(row)

#     # Add last staff's table
#     if table_data:
#         col_widths = [1.2*inch, 3*inch, 1*inch, 1*inch, 2*inch, 2*inch]
#         table = Table(table_data, colWidths=col_widths)
#         table.setStyle(TableStyle([
#             ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
#             ('TEXTCOLOR', (0,0), (-1,0), colors.white),
#             ('ALIGN',(0,0),(-1,-1),'CENTER'),
#             ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
#             ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
#             ('FONTSIZE', (0,0), (-1,-1), 8),
#             ('LEFTPADDING', (0,0), (-1,-1), 4),
#             ('RIGHTPADDING', (0,0), (-1,-1), 4),
#             ('BOTTOMPADDING', (0,0), (-1,-1), 3),
#             ('TOPPADDING', (0,0), (-1,-1), 3),
#         ]))
#         elements.append(table)

#     doc.build(elements)
#     return response

# def export_books_and_chapters_to_pdf(parents, children=None, selected_options=None):
#     """
#     Export Books & Chapters data to PDF (landscape), grouped by staff.
#     Title centered, Staff info aligned flexibly above table.
#     Attachments are clickable links.
#     """
#     from reportlab.lib.pagesizes import A4, landscape
#     from reportlab.lib import colors
#     from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
#     from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
#     from reportlab.lib.units import inch
#     from django.http import HttpResponse

#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="books_and_chapters.pdf"'

#     doc = SimpleDocTemplate(
#         response,
#         pagesize=landscape(A4),
#         topMargin=0.5*inch,
#         bottomMargin=0.5*inch,
#         leftMargin=0.5*inch,
#         rightMargin=0.5*inch
#     )

#     elements = []
#     styles = getSampleStyleSheet()

#     # Styles
#     title_style = ParagraphStyle('Title', parent=styles['Heading2'], alignment=1, spaceAfter=12)
#     info_style = ParagraphStyle('Info', fontSize=12, textColor=colors.HexColor('#2c3e50'))
#     table_header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontSize=9, alignment=1,
#                                         textColor=colors.white, fontName='Helvetica-Bold')
#     table_cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=8, alignment=0,
#                                       textColor=colors.HexColor('#2c3e50'))

#     # Title
#     elements.append(Paragraph("Books and Chapters", title_style))
#     elements.append(Spacer(1, 12))

#     # Sort parents by staff
#     parents_sorted = sorted(parents, key=lambda x: (x.get('staff_id',''), x.get('title_of_book_or_chapter','')))
#     current_staff = None
#     table_data = []
#     col_widths = [1.2*inch, 3*inch, 1*inch, 1*inch, 2*inch, 2*inch]

#     for parent in parents_sorted:
#         staff_id = parent.get('staff_id', 'N/A')
#         staff_name = parent.get('staff_name', 'N/A')
#         department_name = parent.get('department_name', 'N/A')

#         # New staff: flush previous table + page break
#         if current_staff and current_staff != staff_id:
#             table = Table(table_data, colWidths=col_widths)
#             table.setStyle(TableStyle([
#                 ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
#                 ('TEXTCOLOR', (0,0), (-1,0), colors.white),
#                 ('ALIGN',(0,0),(-1,-1),'CENTER'),
#                 ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
#                 ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
#                 ('FONTSIZE', (0,0), (-1,-1), 8),
#                 ('LEFTPADDING', (0,0), (-1,-1), 4),
#                 ('RIGHTPADDING', (0,0), (-1,-1), 4),
#                 ('BOTTOMPADDING', (0,0), (-1,-1), 3),
#                 ('TOPPADDING', (0,0), (-1,-1), 3),
#             ]))
#             elements.append(table)
#             elements.append(PageBreak())
#             table_data = []

#         if current_staff != staff_id:
#             # Staff info as mini-table to align left/right with table width
#             total_width = sum(col_widths)
#             info_left = Paragraph(f"Staff ID: {staff_id}<br/>Staff Name: {staff_name}", info_style)
#             info_right = Paragraph(f"Department: {department_name}", info_style)

#             info_table = Table(
#                 [[info_left, info_right]],
#                 colWidths=[total_width*0.5, total_width*0.5],
#                 hAlign='LEFT'
#             )
#             info_table.setStyle(TableStyle([
#                 ('VALIGN', (0, 0), (-1, -1), 'TOP'),
#                 ('ALIGN', (0, 0), (0, 0), 'LEFT'),
#                 ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
#                 ('LEFTPADDING', (0, 0), (-1, -1), 0),
#                 ('RIGHTPADDING', (0, 0), (-1, -1), 0),
#                 ('TOPPADDING', (0, 0), (-1, -1), 0),
#                 ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
#             ]))
#             elements.append(info_table)
#             elements.append(Spacer(1, 6))

#             # Table headers
#             headers = ['Publication Type', 'Title of Book/Chapter', 'Year of Publication', 'ISBN/ISSN', 'Publisher', 'Attachments']
#             table_data.append([Paragraph(h, table_header_style) for h in headers])
#             current_staff = staff_id

#         # Media attachments
#         media_files = parent.get('media_files', [])
#         if media_files:
#             media_text = []
#             for m in media_files:
#                 filename = m.get('file_name', 'Unknown')
#                 url = m.get('direct_api_url', '#')
#                 media_text.append(f'<link href="{url}" color="blue">{filename}</link>')
#             media_paragraph = Paragraph(', '.join(media_text), table_cell_style)
#         else:
#             media_paragraph = Paragraph("No attachments", table_cell_style)

#         row = [
#             Paragraph(str(parent.get('publication_type', '')), table_cell_style),
#             Paragraph(str(parent.get('title_of_book_or_chapter', '')), table_cell_style),
#             Paragraph(str(parent.get('year_of_publication', '')), table_cell_style),
#             Paragraph(str(parent.get('isbn_or_issn_number', '')), table_cell_style),
#             Paragraph(str(parent.get('name_of_the_publisher', '')), table_cell_style),
#             media_paragraph
#         ]
#         table_data.append(row)

#     # Add last staff table
#     if table_data:
#         table = Table(table_data, colWidths=col_widths)
#         table.setStyle(TableStyle([
#             ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
#             ('TEXTCOLOR', (0,0), (-1,0), colors.white),
#             ('ALIGN',(0,0),(-1,-1),'CENTER'),
#             ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
#             ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
#             ('FONTSIZE', (0,0), (-1,-1), 8),
#             ('LEFTPADDING', (0,0), (-1,-1), 4),
#             ('RIGHTPADDING', (0,0), (-1,-1), 4),
#             ('BOTTOMPADDING', (0,0), (-1,-1), 3),
#             ('TOPPADDING', (0,0), (-1,-1), 3),
#         ]))
#         elements.append(table)

#     doc.build(elements)
#     return response

from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
from reportlab.platypus import Paragraph, Table, TableStyle, Spacer, PageBreak

# def export_books_and_chapters_to_pdf(parents, children=None, selected_options=None):
#     """
#     Export Books & Chapters data to PDF (landscape), grouped by staff.
#     Title centered, Staff info aligned left/right above table using TA_LEFT / TA_RIGHT.
#     Attachments are clickable links.
#     """
#     from reportlab.lib.pagesizes import A4, landscape
#     from reportlab.lib import colors
#     from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
#     from reportlab.platypus import SimpleDocTemplate
#     from reportlab.lib.units import inch
#     from django.http import HttpResponse

#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="books_and_chapters.pdf"'

#     doc = SimpleDocTemplate(
#         response,
#         pagesize=landscape(A4),
#         topMargin=0.5*inch,
#         bottomMargin=0.5*inch,
#         leftMargin=0.5*inch,
#         rightMargin=0.5*inch
#     )

#     elements = []
#     styles = getSampleStyleSheet()

#     # Styles
#     title_style = ParagraphStyle('Title', parent=styles['Heading2'], alignment=TA_CENTER, spaceAfter=12)
#     left_style = ParagraphStyle('LeftStyle', parent=styles['Normal'], fontSize=9,
#                                 alignment=TA_LEFT, textColor=colors.HexColor('#2c3e50'))
#     right_style = ParagraphStyle('RightStyle', parent=styles['Normal'], fontSize=9,
#                                  alignment=TA_RIGHT, textColor=colors.HexColor('#2c3e50'))
#     table_header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontSize=9,
#                                         alignment=TA_CENTER, textColor=colors.white, fontName='Helvetica-Bold')
#     table_cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=8,
#                                       alignment=TA_CENTER, textColor=colors.HexColor('#2c3e50'))
#     attachment_link_style = ParagraphStyle('AttachmentLink', parent=styles['Normal'], fontSize=6,
#                                            alignment=TA_CENTER, textColor=colors.HexColor('#1a5276'))

#     # Title
#     elements.append(Paragraph("Books and Chapters", title_style))
#     elements.append(Spacer(1, 12))

#     # Sort parents by staff
#     parents_sorted = sorted(parents, key=lambda x: (x.get('staff_id',''), x.get('title_of_book_or_chapter','')))
#     current_staff = None
#     table_data = []
#     col_widths = [1.2*inch, 3*inch, 1*inch, 1*inch, 2*inch, 2*inch]

#     for parent in parents_sorted:
#         staff_id = parent.get('staff_id', 'N/A')
#         staff_name = parent.get('staff_name', 'N/A')
#         department_name = parent.get('department_name', 'N/A')

#         # New staff: flush previous table + page break
#         if current_staff and current_staff != staff_id:
#             table = Table(table_data, colWidths=col_widths)
#             table.setStyle(TableStyle([
#                 ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
#                 ('TEXTCOLOR', (0,0), (-1,0), colors.white),
#                 ('ALIGN',(0,0),(-1,-1),'CENTER'),
#                 ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
#                 ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
#                 ('FONTSIZE', (0,0), (-1,-1), 8),
#                 ('LEFTPADDING', (0,0), (-1,-1), 4),
#                 ('RIGHTPADDING', (0,0), (-1,-1), 4),
#                 ('BOTTOMPADDING', (0,0), (-1,-1), 3),
#                 ('TOPPADDING', (0,0), (-1,-1), 3),
#             ]))
#             elements.append(table)
#             elements.append(PageBreak())
#             table_data = []

#         if current_staff != staff_id:
#             # Staff info using TA_LEFT / TA_RIGHT
#             total_width = sum(col_widths)
#             left_paragraph = Paragraph(f"Staff ID: {staff_id}<br/>Staff Name: {staff_name}", left_style)
#             right_paragraph = Paragraph(f"Department: {department_name}", right_style)

#             info_table = Table([[left_paragraph, right_paragraph]], colWidths=[total_width*0.5]*2)
#             info_table.setStyle(TableStyle([
#                 ('VALIGN', (0, 0), (-1, -1), 'TOP'),
#                 ('ALIGN', (0,0), (0,0), 'LEFT'),
#                 ('ALIGN', (1,0), (1,0), 'RIGHT'),
#                 ('LEFTPADDING', (0,0), (-1,-1), 0),
#                 ('RIGHTPADDING', (0,0), (-1,-1), 0),
#                 ('TOPPADDING', (0,0), (-1,-1), 2),
#                 ('BOTTOMPADDING', (0,0), (-1,-1), 4),
#             ]))
#             elements.append(info_table)
#             elements.append(Spacer(1, 6))

#             # Table headers
#             headers = ['Publication Type', 'Title of Book/Chapter', 'Year of Publication', 'ISBN/ISSN', 'Publisher', 'Attachments']
#             table_data.append([Paragraph(h, table_header_style) for h in headers])
#             current_staff = staff_id

#         # Media attachments
#         media_files = parent.get('media_files', [])
#         if media_files:
#             media_text = []
#             for m in media_files:
#                 filename = m.get('file_name', 'Unknown')
#                 url = m.get('direct_api_url', '#')
#                 media_text.append(f'<link href="{url}" color="blue">{filename}</link>')
#             media_paragraph = Paragraph('<br/>'.join(media_text), attachment_link_style)
#         else:
#             media_paragraph = Paragraph("No attachments", table_cell_style)

#         row = [
#             Paragraph(str(parent.get('publication_type', '')), table_cell_style),
#             Paragraph(str(parent.get('title_of_book_or_chapter', '')), table_cell_style),
#             Paragraph(str(parent.get('year_of_publication', '')), table_cell_style),
#             Paragraph(str(parent.get('isbn_or_issn_number', '')), table_cell_style),
#             Paragraph(str(parent.get('name_of_the_publisher', '')), table_cell_style),
#             media_paragraph
#         ]
#         table_data.append(row)

#     # Add last staff table
#     if table_data:
#         table = Table(table_data, colWidths=col_widths)
#         table.setStyle(TableStyle([
#             ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
#             ('TEXTCOLOR', (0,0), (-1,0), colors.white),
#             ('ALIGN',(0,0),(-1,-1),'CENTER'),
#             ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
#             ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
#             ('FONTSIZE', (0,0), (-1,-1), 8),
#             ('LEFTPADDING', (0,0), (-1,-1), 4),
#             ('RIGHTPADDING', (0,0), (-1,-1), 4),
#             ('BOTTOMPADDING', (0,0), (-1,-1), 3),
#             ('TOPPADDING', (0,0), (-1,-1), 3),
#         ]))
#         elements.append(table)

#     doc.build(elements)
#     return response

# def export_books_and_chapters_to_pdf(parents, children=None, selected_options=None):
#     """
#     Export Books & Chapters data to PDF (landscape), grouped by staff.
#     Title centered, Staff info aligned left/right above table with bold labels.
#     Attachments are clickable links.
#     """
#     from reportlab.lib.pagesizes import A4, landscape
#     from reportlab.lib import colors
#     from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
#     from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, PageBreak
#     from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
#     from reportlab.lib.units import inch
#     from django.http import HttpResponse

#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="books_and_chapters.pdf"'

#     doc = SimpleDocTemplate(
#         response,
#         pagesize=landscape(A4),
#         topMargin=0.5*inch,
#         bottomMargin=0.5*inch,
#         leftMargin=0.5*inch,
#         rightMargin=0.5*inch
#     )

#     elements = []
#     styles = getSampleStyleSheet()

#     # Styles
#     title_style = ParagraphStyle('Title', parent=styles['Heading2'], alignment=TA_CENTER, spaceAfter=12)
#     left_style = ParagraphStyle('LeftStyle', parent=styles['Normal'], fontSize=9,
#                                 alignment=TA_LEFT, textColor=colors.HexColor('#2c3e50'))
#     right_style = ParagraphStyle('RightStyle', parent=styles['Normal'], fontSize=9,
#                                  alignment=TA_RIGHT, textColor=colors.HexColor('#2c3e50'))
#     table_header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontSize=9,
#                                         alignment=TA_CENTER, textColor=colors.white, fontName='Helvetica-Bold')
#     table_cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=8,
#                                       alignment=TA_CENTER, textColor=colors.HexColor('#2c3e50'))
#     attachment_link_style = ParagraphStyle('AttachmentLink', parent=styles['Normal'], fontSize=6,
#                                            alignment=TA_CENTER, textColor=colors.HexColor('#1a5276'))

#     # Title
#     elements.append(Paragraph("Books and Chapters", title_style))
#     elements.append(Spacer(1, 12))

#     # Sort parents by staff
#     parents_sorted = sorted(parents, key=lambda x: (x.get('staff_id',''), x.get('title_of_book_or_chapter','')))
#     current_staff = None
#     table_data = []
#     col_widths = [1.2*inch, 3*inch, 1*inch, 1*inch, 2*inch, 2*inch]

#     for parent in parents_sorted:
#         staff_id = parent.get('staff_id', 'N/A')
#         staff_name = parent.get('staff_name', 'N/A')
#         department_name = parent.get('department_name', 'N/A')

#         # New staff: flush previous table + page break
#         if current_staff and current_staff != staff_id:
#             table = Table(table_data, colWidths=col_widths)
#             table.setStyle(TableStyle([
#                 ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
#                 ('TEXTCOLOR', (0,0), (-1,0), colors.white),
#                 ('ALIGN',(0,0),(-1,-1),'CENTER'),
#                 ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
#                 ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
#                 ('FONTSIZE', (0,0), (-1,-1), 8),
#                 ('LEFTPADDING', (0,0), (-1,-1), 4),
#                 ('RIGHTPADDING', (0,0), (-1,-1), 4),
#                 ('BOTTOMPADDING', (0,0), (-1,-1), 3),
#                 ('TOPPADDING', (0,0), (-1,-1), 3),
#             ]))
#             elements.append(table)
#             elements.append(PageBreak())
#             table_data = []

#         if current_staff != staff_id:
#             # Staff heading
#             elements.append(Paragraph(f"<b>Staff: {staff_name}</b>", title_style))
#             # Staff info with bold labels
#             left_paragraph = Paragraph(
#                 f"<b>Staff ID:</b> {staff_id}<br/><b>Staff Name:</b> {staff_name}", left_style)
#             right_paragraph = Paragraph(f"<b>Department:</b> {department_name}", right_style)

#             total_width = sum(col_widths)
#             info_table = Table([[left_paragraph, right_paragraph]], colWidths=[total_width*0.5]*2)
#             info_table.setStyle(TableStyle([
#                 ('VALIGN', (0, 0), (-1, -1), 'TOP'),
#                 ('ALIGN', (0,0), (0,0), 'LEFT'),
#                 ('ALIGN', (1,0), (1,0), 'RIGHT'),
#                 ('LEFTPADDING', (0,0), (-1,-1), 0),
#                 ('RIGHTPADDING', (0,0), (-1,-1), 0),
#                 ('TOPPADDING', (0,0), (-1,-1), 2),
#                 ('BOTTOMPADDING', (0,0), (-1,-1), 4),
#             ]))
#             elements.append(info_table)
#             elements.append(Spacer(1, 6))

#             # Table headers
#             headers = ['Publication Type', 'Title of Book/Chapter', 'Year of Publication', 'ISBN/ISSN', 'Publisher', 'Attachments']
#             table_data.append([Paragraph(h, table_header_style) for h in headers])
#             current_staff = staff_id

#         # Media attachments
#         media_files = parent.get('media_files', [])
#         if media_files:
#             media_text = []
#             for m in media_files:
#                 filename = m.get('file_name', 'Unknown')
#                 url = m.get('direct_api_url', '#')
#                 media_text.append(f'<link href="{url}" color="blue">{filename}</link>')
#             media_paragraph = Paragraph('<br/>'.join(media_text), attachment_link_style)
#         else:
#             media_paragraph = Paragraph("No attachments", table_cell_style)

#         row = [
#             Paragraph(str(parent.get('publication_type', '')), table_cell_style),
#             Paragraph(str(parent.get('title_of_book_or_chapter', '')), table_cell_style),
#             Paragraph(str(parent.get('year_of_publication', '')), table_cell_style),
#             Paragraph(str(parent.get('isbn_or_issn_number', '')), table_cell_style),
#             Paragraph(str(parent.get('name_of_the_publisher', '')), table_cell_style),
#             media_paragraph
#         ]
#         table_data.append(row)

#     # Add last staff table
#     if table_data:
#         table = Table(table_data, colWidths=col_widths)
#         table.setStyle(TableStyle([
#             ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
#             ('TEXTCOLOR', (0,0), (-1,0), colors.white),
#             ('ALIGN',(0,0),(-1,-1),'CENTER'),
#             ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
#             ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
#             ('FONTSIZE', (0,0), (-1,-1), 8),
#             ('LEFTPADDING', (0,0), (-1,-1), 4),
#             ('RIGHTPADDING', (0,0), (-1,-1), 4),
#             ('BOTTOMPADDING', (0,0), (-1,-1), 3),
#             ('TOPPADDING', (0,0), (-1,-1), 3),
#         ]))
#         elements.append(table)

#     doc.build(elements)
#     return response

# def export_books_and_chapters_to_pdf(parents, children=None, selected_options=None):
#     """
#     Export Books & Chapters data to PDF (landscape), grouped by staff.
#     Each staff has a mini-title "Books and Chapters", then Staff info in bold labels, then table.
#     Attachments are clickable links.
#     """
#     from reportlab.lib.pagesizes import A4, landscape
#     from reportlab.lib import colors
#     from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
#     from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, PageBreak
#     from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
#     from reportlab.lib.units import inch
#     from django.http import HttpResponse

#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="books_and_chapters.pdf"'

#     doc = SimpleDocTemplate(
#         response,
#         pagesize=landscape(A4),
#         topMargin=0.5*inch,
#         bottomMargin=0.5*inch,
#         leftMargin=0.5*inch,
#         rightMargin=0.5*inch
#     )

#     elements = []
#     styles = getSampleStyleSheet()

#     # Styles
#     title_style = ParagraphStyle('Title', parent=styles['Heading2'], alignment=TA_CENTER, spaceAfter=12)
#     left_style = ParagraphStyle('LeftStyle', parent=styles['Normal'], fontSize=9,
#                                 alignment=TA_LEFT, textColor=colors.HexColor('#2c3e50'))
#     right_style = ParagraphStyle('RightStyle', parent=styles['Normal'], fontSize=9,
#                                  alignment=TA_RIGHT, textColor=colors.HexColor('#2c3e50'))
#     table_header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontSize=9,
#                                         alignment=TA_CENTER, textColor=colors.white, fontName='Helvetica-Bold')
#     table_cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=8,
#                                       alignment=TA_CENTER, textColor=colors.HexColor('#2c3e50'))
#     attachment_link_style = ParagraphStyle('AttachmentLink', parent=styles['Normal'], fontSize=6,
#                                            alignment=TA_CENTER, textColor=colors.HexColor('#1a5276'))

#     # Sort parents by staff
#     parents_sorted = sorted(parents, key=lambda x: (x.get('staff_id',''), x.get('title_of_book_or_chapter','')))
#     current_staff = None
#     table_data = []
#     col_widths = [1.2*inch, 3*inch, 1*inch, 1*inch, 2*inch, 2*inch]

#     for parent in parents_sorted:
#         staff_id = parent.get('staff_id', 'N/A')
#         staff_name = parent.get('staff_name', 'N/A')
#         department_name = parent.get('department_name', 'N/A')

#         # New staff: flush previous table + page break
#         if current_staff and current_staff != staff_id:
#             table = Table(table_data, colWidths=col_widths)
#             table.setStyle(TableStyle([
#                 ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
#                 ('TEXTCOLOR', (0,0), (-1,0), colors.white),
#                 ('ALIGN',(0,0),(-1,-1),'CENTER'),
#                 ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
#                 ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
#                 ('FONTSIZE', (0,0), (-1,-1), 8),
#                 ('LEFTPADDING', (0,0), (-1,-1), 4),
#                 ('RIGHTPADDING', (0,0), (-1,-1), 4),
#                 ('BOTTOMPADDING', (0,0), (-1,-1), 3),
#                 ('TOPPADDING', (0,0), (-1,-1), 3),
#             ]))
#             elements.append(table)
#             elements.append(PageBreak())
#             table_data = []

#         if current_staff != staff_id:
#             # Add mini-title for this staff
#             elements.append(Paragraph("Books and Chapters", title_style))
#             elements.append(Spacer(1, 6))

#             # Staff info with bold labels
#             left_paragraph = Paragraph(
#                 f"<b>Staff ID:</b> {staff_id}<br/><b>Staff Name:</b> {staff_name}", left_style)
#             right_paragraph = Paragraph(f"<b>Department:</b> {department_name}", right_style)

#             total_width = sum(col_widths)
#             info_table = Table([[left_paragraph, right_paragraph]], colWidths=[total_width*0.5]*2)
#             info_table.setStyle(TableStyle([
#                 ('VALIGN', (0, 0), (-1, -1), 'TOP'),
#                 ('ALIGN', (0,0), (0,0), 'LEFT'),
#                 ('ALIGN', (1,0), (1,0), 'RIGHT'),
#                 ('LEFTPADDING', (0,0), (-1,-1), 0),
#                 ('RIGHTPADDING', (0,0), (-1,-1), 0),
#                 ('TOPPADDING', (0,0), (-1,-1), 2),
#                 ('BOTTOMPADDING', (0,0), (-1,-1), 4),
#             ]))
#             elements.append(info_table)
#             elements.append(Spacer(1, 6))

#             # Table headers
#             headers = ['Publication Type', 'Title of Book/Chapter', 'Year of Publication', 'ISBN/ISSN', 'Publisher', 'Attachments']
#             table_data.append([Paragraph(h, table_header_style) for h in headers])
#             current_staff = staff_id

#         # Media attachments
#         media_files = parent.get('media_files', [])
#         if media_files:
#             media_text = []
#             for m in media_files:
#                 filename = m.get('file_name', 'Unknown')
#                 url = m.get('direct_api_url', '#')
#                 media_text.append(f'<link href="{url}" color="blue">{filename}</link>')
#             media_paragraph = Paragraph('<br/>'.join(media_text), attachment_link_style)
#         else:
#             media_paragraph = Paragraph("No attachments", table_cell_style)

#         row = [
#             Paragraph(str(parent.get('publication_type', '')), table_cell_style),
#             Paragraph(str(parent.get('title_of_book_or_chapter', '')), table_cell_style),
#             Paragraph(str(parent.get('year_of_publication', '')), table_cell_style),
#             Paragraph(str(parent.get('isbn_or_issn_number', '')), table_cell_style),
#             Paragraph(str(parent.get('name_of_the_publisher', '')), table_cell_style),
#             media_paragraph
#         ]
#         table_data.append(row)

#     # Add last staff table
#     if table_data:
#         table = Table(table_data, colWidths=col_widths)
#         table.setStyle(TableStyle([
#             ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
#             ('TEXTCOLOR', (0,0), (-1,0), colors.white),
#             ('ALIGN',(0,0),(-1,-1),'CENTER'),
#             ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
#             ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
#             ('FONTSIZE', (0,0), (-1,-1), 8),
#             ('LEFTPADDING', (0,0), (-1,-1), 4),
#             ('RIGHTPADDING', (0,0), (-1,-1), 4),
#             ('BOTTOMPADDING', (0,0), (-1,-1), 3),
#             ('TOPPADDING', (0,0), (-1,-1), 3),
#         ]))
#         elements.append(table)

#     doc.build(elements)
#     return response

def export_books_and_chapters_to_pdf(parents, children=None, selected_options=None):
    """
    Export Books & Chapters data to PDF (landscape), grouped by staff.
    - Shows only selected staff if selected_options is provided.
    - Always prints Staff ID, Name, and Department (even if no data).
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
    from django.http import HttpResponse

    # --- Create response ---
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="books_and_chapters.pdf"'

    # --- Create document ---
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch
    )

    elements = []
    styles = getSampleStyleSheet()

    # --- Styles ---
    title_style = ParagraphStyle('Title', parent=styles['Heading2'], alignment=TA_CENTER, spaceAfter=12)
    left_style = ParagraphStyle('LeftStyle', parent=styles['Normal'], fontSize=9, alignment=TA_LEFT, textColor=colors.HexColor('#2c3e50'))
    right_style = ParagraphStyle('RightStyle', parent=styles['Normal'], fontSize=9, alignment=TA_RIGHT, textColor=colors.HexColor('#2c3e50'))
    table_header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER, textColor=colors.white, fontName='Helvetica-Bold')
    table_cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.HexColor('#2c3e50'))
    attachment_link_style = ParagraphStyle('AttachmentLink', parent=styles['Normal'], fontSize=6, alignment=TA_CENTER, textColor=colors.HexColor('#1a5276'))
    no_data_style = ParagraphStyle('NoData', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, textColor=colors.HexColor('#7f8c8d'), spaceBefore=12, spaceAfter=12)

    # --- Title ---
    elements.append(Paragraph("Books and Chapters", title_style))
    elements.append(Spacer(1, 10))

    # --- Handle empty data ---
    if not parents:
        elements.append(Paragraph("No data available", no_data_style))
        doc.build(elements)
        return response

    # --- Filter selected staff ---
    if selected_options:
        selected_id = str(selected_options)
        parents_filtered = [p for p in parents if str(p.get('staff_id')) == selected_id]
    else:
        parents_filtered = parents

    # --- Handle case where no data for selected staff ---
    if not parents_filtered:
        # Try to get staff info from full data if available
        all_staff = parents
        staff_info = next((s for s in all_staff if str(s.get('staff_id')) == str(selected_options)), None)
        if staff_info:
            staff_id = staff_info.get('staff_id', 'N/A')
            staff_name = staff_info.get('staff_name', 'N/A')
            department_name = staff_info.get('department_name', 'N/A')

            # Display staff info
            total_width = 10 * inch
            left_paragraph = Paragraph(f"Staff ID: {staff_id}<br/>Staff Name: {staff_name}", left_style)
            right_paragraph = Paragraph(f"Department: {department_name}", right_style)
            info_table = Table([[left_paragraph, right_paragraph]], colWidths=[total_width * 0.5] * 2)
            info_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 10))
            elements.append(Paragraph("No data available", no_data_style))
            doc.build(elements)
            return response
        else:
            elements.append(Paragraph("No data available", no_data_style))
            doc.build(elements)
            return response

    # --- Use first staff info for header ---
    first_record = parents_filtered[0]
    staff_id = first_record.get('staff_id', 'N/A')
    staff_name = first_record.get('staff_name', 'N/A')
    department_name = first_record.get('department_name', 'N/A')

    total_width = 10 * inch
    left_paragraph = Paragraph(f"Staff ID: {staff_id}<br/>Staff Name: {staff_name}", left_style)
    right_paragraph = Paragraph(f"Department: {department_name}", right_style)
    info_table = Table([[left_paragraph, right_paragraph]], colWidths=[total_width * 0.5] * 2)
    info_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 10))

    # --- Table setup ---
    headers = ['Publication Type', 'Title of Book/Chapter', 'Year of Publication', 'ISBN/ISSN', 'Publisher', 'Attachments']
    col_widths = [1.2 * inch, 3 * inch, 1 * inch, 1 * inch, 2 * inch, 2 * inch]
    table_data = [[Paragraph(h, table_header_style) for h in headers]]

    # --- Fill rows ---
    for parent in parents_filtered:
        media_files = parent.get('media_files', [])
        if media_files:
            media_links = []
            for m in media_files:
                filename = m.get('file_name', 'Unknown')
                url = m.get('direct_api_url', '#')
                media_links.append(f'<link href="{url}" color="blue">{filename}</link>')
            media_paragraph = Paragraph('<br/>'.join(media_links), attachment_link_style)
        else:
            media_paragraph = Paragraph("No attachments", table_cell_style)

        pub_year = parent.get('year_of_publication', '')
        if isinstance(pub_year, int):
            pub_year = f"{pub_year}-{pub_year + 1}"
        elif isinstance(pub_year, str) and pub_year.isdigit():
            y = int(pub_year)
            pub_year = f"{y}-{y + 1}"
        else:
            pub_year = str(pub_year)    

        row = [
            Paragraph(str(parent.get('publication_type', '')), table_cell_style),
            Paragraph(str(parent.get('title_of_book_or_chapter', '')), table_cell_style),
            Paragraph(pub_year, table_cell_style),
            Paragraph(str(parent.get('isbn_or_issn_number', '')), table_cell_style),
            Paragraph(str(parent.get('name_of_the_publisher', '')), table_cell_style),
            media_paragraph
        ]
        table_data.append(row)

    # --- Add final table ---
    if len(table_data) > 1:
        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#02548b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d6d6d6')),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("No data available", no_data_style))

    # --- Build document ---
    doc.build(elements)
    return response

def export_books_and_chapters_to_excel(parents, children=None):
    """
    Export Books & Chapters data to Excel from filtered parents,
    including clickable media file links based on upload field codes (CCP, CC, IP, POI).
    """
    import io
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Books & Chapters"

    # Headers including media attachments
    headers = [
        "Staff ID", "Staff Name", "Department", "Publication Type",
        "Title of Book/Chapter", "Title of Paper", "Title of Proceedings",
        "Conference Name", "Year of Publication", "National/International",
        "ISBN/ISSN Number", "Affiliating Institute", "Publisher",
        "Conference Cover Page", "Copy of the Chapters", "Index Page", "Proof Of ISBN / ISSN Details"
    ]
    ws.append(headers)

    # Style headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Map upload field codes to friendly names
    field_mapping = {
        "CCP": "Conference Cover Page",
        "CC": "Copy of the Chapters",
        "IP": "Index Page",
        "POI": "Proof Of ISBN / ISSN Details"
    }

    # Populate rows
    for p in parents:
        # Initialize media categories
        media_mapping = {v: [] for v in field_mapping.values()}

        # Categorize each file based on upload code in filename
        for media in p.get("media_files", []):
            file_name = media.get("file_name", "Unknown")
            url = media.get("direct_api_url", "")
            if not url:
                continue

            lname = file_name.upper()

            # Detect the code (CCP, CC, IP, POI) from filename
            matched = False
            for code, category in field_mapping.items():
                if f"_{code}_" in lname:
                    media_mapping[category].append((file_name, url))
                    matched = True
                    break

            # Optional fallback: if none matched, put in "Copy of the Chapters"
            if not matched:
                media_mapping["Copy of the Chapters"].append((file_name, url))

        # Format publication year as [YYYY-YYYY+1]
        pub_year = p.get("year_of_publication", "")
        if isinstance(pub_year, int):
            pub_year = f"{pub_year}-{pub_year + 1}"
        elif isinstance(pub_year, str) and pub_year.isdigit():
            y = int(pub_year)
            pub_year = f"{y}-{y + 1}"
        else:
            pub_year = str(pub_year)       

        # Create the base info row (first 13 columns)
        base_row = [
            p.get("staff_id", ""),
            p.get("staff_name", ""),
            p.get("department_name", ""),
            p.get("publication_type", ""),
            p.get("title_of_book_or_chapter", ""),
            p.get("title_of_paper", ""),
            p.get("title_of_proceedings", ""),
            p.get("name_of_conference", ""),
            pub_year,
            p.get("national_or_international", ""),
            p.get("isbn_or_issn_number", ""),
            p.get("affiliating_institute", ""),
            p.get("name_of_the_publisher", "")
        ]

        # Append placeholders for media columns
        ws.append(base_row + ["", "", "", ""])
        row_idx = ws.max_row  # current row number

        # Map category to column
        media_columns = {
            "Conference Cover Page": 14,
            "Copy of the Chapters": 15,
            "Index Page": 16,
            "Proof Of ISBN / ISSN Details": 17
        }

        # Add clickable hyperlinks
        for category, files in media_mapping.items():
            cell = ws.cell(row=row_idx, column=media_columns[category])
            if files:
                # Show all filenames joined by line breaks, hyperlink first
                cell.value = "\n".join([f for f, _ in files])
                cell.hyperlink = files[0][1]  # link to first file
                cell.font = Font(color="0000EE", underline="single")
            else:
                cell.value = "-"
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-fit column widths
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

    # Adjust row heights for multiline cells
    for row_idx in range(1, ws.max_row + 1):
        max_lines = max(str(ws.cell(row=row_idx, column=col).value or "").count("\n") + 1 for col in range(1, len(headers) + 1))
        ws.row_dimensions[row_idx].height = min(max_lines * 15, 120)

    # Freeze header and enable filters
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Return as Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="books_and_chapters.xlsx"'

    with io.BytesIO() as buffer:
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())

    return response

