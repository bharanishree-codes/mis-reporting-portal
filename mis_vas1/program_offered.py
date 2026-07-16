import json
import requests
import string
import random
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from MIS.functions import validate_file_format, validate_file_size
from datetime import datetime
from user_management.settings_views import *

API_STUDIO_URL = user_bundle_settings()


def program_key():
    url = "https://api.hcaschennai.edu.in/auth/token"

    payload = json.dumps({
        "secret_key": "C4ZoXbsAnHLjk1Xyz4QPT2eoiNx6K6fo"
    })
    headers = {'Content-Type': 'application/json'}

    response = requests.post(url, headers=headers, data=payload)
    if response.status_code == 200:
        res_data = response.json()
        access_token = res_data.get('access_token')
        token_type = res_data.get('token_type')
        return access_token, token_type
    return None, None


# Function to fetch staff career data
def get_program_data(access_token, token_type):
    url = "https://api.hcaschennai.edu.in/sqlviews/api/v1/auth/get_response_data"

    payload = json.dumps({
        "psk_uid": "51a531b4-bd55-491c-861d-a8d7227b325b",
        "project": "hcas",
        "data": {}
    })
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'{token_type} {access_token}'  # Fix spacing
    }

    response = requests.post(url, headers=headers, data=payload)
    if response.status_code == 200:
        return response.json()
    return []


# Create program
def program_create(request):
    error_message = None

    # Step 1: Get API token
    access_token, token_type = program_key()
    if not access_token or not token_type:
        error_message = 'Failed to get access token from API.'
        return render(request, 'program_offered_templates/program_create.html', {'error': error_message})

    # Step 2: Fetch career data
    program_data = get_program_data(access_token, token_type)
    if not program_data:
        error_message = 'Failed to fetch staff data.'
        return render(request, 'program_offered_templates/program_create.html', {'error': error_message})

    # Function to generate random alphanumeric code for course_code
    # def generate_random_code(length=8):
    #     """Generate a random numeric string of a specified length."""
    #     characters = string.digits  # Only numeric characters (0-9)
    #     return ''.join(random.choice(characters) for _ in range(length))
    
    current_year = datetime.now().year
    year_of_offering = list(range(2000, current_year + 1))
    times_offered_in_year = list(range(2000, current_year + 1))
    
    user = get_settings(request)
    username = user['username']
    # username = 'CS-T151'
    current_year = datetime.now().year
    publication_year = [f"{y}-{y+1}" for y in range(current_year, 2010, -1)]

    
    
    selected_faculty = None
    
    for faculty in program_data:
        if faculty['stf_id'] == username:
            selected_faculty = faculty
            break
        
    staff_name = selected_faculty.get('stf_name', '') if selected_faculty else ''
    department = selected_faculty.get('department', '') if selected_faculty else ''
    

    # Step 3: Handle POST form submission
    if request.method == 'POST':
        # Extract form data
        staff_name = request.POST['staff_name']
        staff_id = username
        department = request.POST['department']
        programs_offered = request.POST['programs_offered']
        course_code = request.POST['course_code']  # This can be optional, can generate if missing
        course_duration = request.POST['course_duration']
        times_offered_in_year = request.POST['times_offered_in_year']
        students_completed = request.POST['students_completed']
        students_enrolled = request.POST['students_enrolled']
        year_of_offering = request.POST['year_of_offering']

        # If course_code is not provided, generate a random one
        # if not course_code:
        #     course_code = generate_random_code()
            
            

        selected_faculty = None
        
        for faculty in program_data:
            if faculty.get('stf_id') == staff_id:
                selected_faculty = faculty
                break

        # If a matching faculty was found, extract the data
        if selected_faculty:
            staff_name = selected_faculty.get('stf_name', '') if selected_faculty else ''
            department = selected_faculty.get('department', '') if selected_faculty else ''

            
            
        # Loop through each faculty in the research_data to find the matching stf_id
        # for faculty in program_data:
        #     if faculty['stf_id'] == staff_id:
        #         selected_faculty = faculty  # Assign the matching faculty
        #         break  # Stop the loop once the faculty is found

        # # If a matching faculty was found, extract the data
        # if selected_faculty:
        #     department = selected_faculty.get('department', '')
        #     staff_name = selected_faculty.get('stf_name', '')


        # URL to send the form data to
        url = f"{API_STUDIO_URL}postapi/create/naac01_add_on_certificate_dc1"
        payload = json.dumps({"data": {
            "staff_name": staff_name,
            "staff_id": staff_id,
            "department": department,
            "programs_offered": programs_offered,
            "course_code": course_code,
            "course_duration": course_duration,
            "times_offered_in_year": times_offered_in_year,
            "students_completed": students_completed,
            "students_enrolled": students_enrolled,
            "year_of_offering": year_of_offering
        }})
        headers = {'Content-Type': 'application/json'}

        # Make API call to create the program/workshop
        response = requests.post(url, headers=headers, data=payload)
        print(response.text)

        if response.status_code == 200:
            # Successfully created the program/workshop
            file_data = response.json()
            psk_id = file_data.get('psk_id')  # Get psk_id from the response

            # Uploading files (optional step)
            upload_url = f"{API_STUDIO_URL}crudapp/upload/media/naac01_add_on_certificate_dc1_media"

            # Get the list of uploaded files from the 'file' field in the request
            uploaded_files = request.FILES.getlist('file')

            if not uploaded_files:
                messages.error(request, message="No files selected for upload.")
                return render(request, 'program_offered_templates/extension_create.html', {'object': file_data, 'psk_id': psk_id})

            print("enter")

            # Define the fields you want to upload files for
            fields = ['Circular', 'Brochure', 'Syllabus', 'RPP', 'Attendance', 'Certificate', 'Feedback', 'QP', 'AS', 'GTP']

            # Iterate over the uploaded files
            # for field in fields:
            current_year = datetime.now().year
            staff_id = username
            for field, uploaded_file in zip(fields, uploaded_files):
                validate_file_size(uploaded_file)
                validate_file_format(uploaded_file)
                file_type = uploaded_file.content_type
                custom_filename = f"{staff_id}_{field}_{current_year}_{uploaded_file.name}"
                print(f"Generated filename: {custom_filename}")  # Print the filename for each iteration
                payload = {'parent_psk_id': psk_id}
                files = {'media': (custom_filename, uploaded_file, file_type)}
                upload_headers = {'api_name': 'naac01_add_on_certificate_dc1_media'}

                    # Make API call to upload the file
                upload_response = requests.post(upload_url, headers=upload_headers, data=payload, files=files)
                print("success")

                if upload_response.status_code != 200:
                    # File upload failed
                    messages.error(request,
                                    message=f"File upload failed for {uploaded_file.name}. Error: {upload_response.text}")
                    return redirect('program_list')

                messages.success(request, message="Documents uploaded successfully.")

            # Redirect to program view after successful creation
            return redirect('program_list')

        else:
            # API call failed for creating program
            messages.error(request, message="Failed to create program. Please try again.")
            return render(request, 'program_offered_templates/program_create.html')

    else:
        # If the request is GET, render the program creation form
        return render(request, 'program_offered_templates/program_create.html', {'program_data': program_data, 'years':year_of_offering, 'username': username,  # Pass 'username' (which is the staff_id) to the template
        'department': department,
        'publication_year': publication_year,
        'staff_name': staff_name})


# View program Details
def program_view(request, id):
    url = f"{API_STUDIO_URL}getapi/naac01_add_on_certificate_dc1/{id}"
    response = requests.get(url)

    if response.status_code == 200:
        program_data = response.json()
        return render(request, "program_offered_templates/program_view.html", {'program': program_data})

    return HttpResponse(f"Error fetching Course details: {response.text}", status=500)


# # List All programs
# def program_list(request):
    # # URL to get programs data
    # url = f"{API_STUDIO_URL}getapi/all_fields/naac01_add_on_certificate_dc1/all"
    # response = requests.get(url)

    # # Get the current username (staff_id)
    # user = get_settings(request)  # Assuming this function retrieves user settings (including staff_id)
    # username = user['username']  # Adjust if necessary to fetch staff_id or username from session
    # # username = 'AC-NT012'
    
    # # If no username (staff_id) is found, return the same page with an empty list
    # if not username:
        # return render(request, 'program_offered_templates/program_list.html', {'programs': []})

    # # If the API call was successful, filter the programs data based on staff_id (username)
    # if response.status_code == 200:
        # programs = response.json()
        
        # # Filter the programs based on the staff_id (username)
        # filtered_programs = [program for program in programs if program.get('staff_id') == username]
        
        
    # selected_staff_id = request.GET.get('staff_id')
    
    # if selected_staff_id:
        # filtered_programs = [program for program in programs if program.get('staff_id') == selected_staff_id]
        
        # # If no data is found for the username, return the same page with an empty list
        # if not filtered_programs:
            # return render(request, 'program_offered_templates/program_list.html', {'programs': programs})
        
        # # Return the filtered data to the template
        # return render(request, 'program_offered_templates/program_list.html', {'programs': filtered_programs})
    
    # # If the API call fails, return an empty list
    # return render(request, 'program_offered_templates/program_list.html', {'programs': []})
    
def program_list(request):
    url = f"{API_STUDIO_URL}getapi/naac01_add_on_certificate_dc1/all"
    response = requests.get(url)
    
    if response.status_code == 200:
        programs = response.json()
        print("programs:", programs)
    else:
        return HttpResponse("API Call Is Not Working")

    
    user = get_settings(request)
    username = user.get('username')
    # username = 'CS-T151'
    filtered_programs = [program for program in programs if program.get('staff_id') == username]
    print("filtered_programs:", filtered_programs)
    #from_dashboard = request.GET.get('from') == 'dashboard' or 'admin_hod_dash' or 'admin_dash' or 'department_dashboard'
    selected_staff_id = request.GET.get('staff_id')
    if selected_staff_id:
        filtered_programs = [program for program in programs if program.get('staff_id') == selected_staff_id]
        
    selected_department = request.GET.get('department')
    if selected_department:
        filtered_programs = [program for program in programs if program.get('department') == selected_department]


    if not filtered_programs:
        return render(request, 'program_offered_templates/program_list.html', {"programs": []})
    
    return render(request, 'program_offered_templates/program_list.html', {"programs": filtered_programs, 
    #'from_dashboard': from_dashboard
    })


# Update program
def program_update(request, id):
    # Fetch program details
    url = f"{API_STUDIO_URL}getapi/naac01_add_on_certificate_dc1/{id}"
    response = requests.get(url)

    current_year = datetime.now().year
    publication_year = [f"{y}-{y+1}" for y in range(current_year, 2010, -1)]
    
    if response.status_code == 200:
        program = response.json()
        # course_code = program.get('course_code')  # Assuming 'course_code' is the refer code
    else:
        return HttpResponse(f"Error fetching program details: {response.text}", status=500)

    # Fetch media (child files) associated with this program
    media_url = f"{API_STUDIO_URL}crudapp/get/media/naac01_add_on_certificate_dc1_media/parent/{id}"
    media_response = requests.get(media_url)

    if media_response.status_code == 200:
        child_files = media_response.json()
    else:
        return HttpResponse(f"Failed to fetch media files: {media_response.text}", status=500)
    
    current_year = datetime.now().year
    year_of_offering = list(range(2000, current_year + 1))

    # Handle form submission (POST request)
    if request.method == "POST":
        # Prepare data for updating the program/workshop
        update_url = f"{API_STUDIO_URL}updateapi/update/naac01_add_on_certificate_dc1/{id}"

        # Fetch data from POST or fallback to program object data
        staff_id = request.POST.get('staff_id', program.get('staff_id'))

        # Prepare the payload to update the program details
        payload = json.dumps({
            "data": {
                "staff_name": request.POST.get('staff_name', program.get('staff_name')),
                "staff_id": staff_id,
                "department": request.POST.get('department', program.get('department')),
                "programs_offered": request.POST.get('programs_offered', program.get('programs_offered')),
                "course_code": request.POST.get('course_code', program.get('course_code')),
                "course_duration": request.POST.get('course_duration', program.get('course_duration')),
                "times_offered_in_year": request.POST.get('times_offered_in_year', program.get('times_offered_in_year')),
                "students_completed": request.POST.get('students_completed', program.get('students_completed')),
                "students_enrolled": request.POST.get('students_enrolled', program.get('students_enrolled')),
                "year_of_offering": request.POST.get('year_of_offering', program.get('year_of_offering'))
            }
        })
        headers = {'Content-Type': 'application/json'}
        update_response = requests.put(update_url, headers=headers, data=payload)

        if update_response.status_code != 200:
            return HttpResponse(f"Failed to update program details: {update_response.text}", status=500)

        # Handle media (file) uploads
        upload_errors = []

        for child in child_files:
            upload_id = child['psk_id']
            fields = ['Circular', 'Brochure', 'Syllabus', 'RPP', 'Attendance', 'Certificate', 'Feedback', 'QP', 'AS', 'GTP']

            # Loop through the fields to handle each file
            for field in fields:
                uploaded_files = request.FILES.getlist(f'file_{upload_id}_{field.lower()}')  # Get files for the current field

                if not uploaded_files:
                    continue  # No files to upload for this field

                for uploaded_file in uploaded_files:
                    # Validate file size and format
                    try:
                        validate_file_size(uploaded_file)
                        validate_file_format(uploaded_file)
                    except Exception as e:
                        messages.error(request, str(e))
                        continue  # Skip file upload if validation fails

                    # Generate custom filename
                    current_year = datetime.now().year
                    custom_filename = f"{staff_id}_{field}_{current_year}_{uploaded_file.name}"
                    print("custom_filename: ", custom_filename)

                    # Define the upload URL for the media
                    upload_url = f"{API_STUDIO_URL}crudapp/upload/media/naac01_add_on_certificate_dc1_media/{upload_id}"

                    # Prepare the file for upload
                    files = {'media': (custom_filename, uploaded_file, uploaded_file.content_type)}

                    # Prepare additional payload and headers for the upload request
                    payload = {'parent_psk_id': id}
                    headers = {
                        'api_name': 'naac01_add_on_certificate_dc1_media',
                        'psk_id': str(upload_id)
                    }

                    # Upload the file
                    upload_response = requests.put(upload_url, headers=headers, data=payload, files=files)

                    # Check if the upload was successful
                    if upload_response.status_code != 200:
                        upload_errors.append(f"Error uploading file for {upload_id} ({field}): {upload_response.text}")
                        print(f"Error with {custom_filename}")

        # Provide feedback to the user
        if upload_errors:
            for error in upload_errors:
                messages.error(request, error)
        else:
            messages.success(request, "Files uploaded successfully.")

        # Redirect to the program view page after successful update
        return redirect('program_list')

    # If not a POST request, render the update form
    return render(request, 'program_offered_templates/program_update.html', {'program': program, 'child_files': child_files, 'years': year_of_offering, 'publication_year': publication_year})





# Delete program
def program_delete(request, id):
    delete_url = f"{API_STUDIO_URL}deleteapi/delete/naac01_add_on_certificate_dc1/{id}"
    delete_response = requests.delete(delete_url)

    if delete_response.status_code == 200:
        return redirect('program_list')
    else:
        return HttpResponse("Failed to delete participation: " + delete_response.text)


def export_programs_offered_to_excel(parents, children=None):
    """
    Export Programs Offered data to Excel from filtered parents,
    including clickable media file links based on upload field codes.
    """
    import io
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Programs Offered"

    # Headers including media attachments
    headers = [
        "Staff ID", "Staff Name", "Department",
        "Programs Offered", "Course Code", "Course Duration",
        "Times Offered in Year", "Students Enrolled", "Students Completed",
        "Year of Offering",
        "Circular", "Brochure", "Syllabus", "Resource Person Profile"
    ]
    ws.append(headers)

    # Style headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Field mapping for file categorization
    field_mapping = {
        "Circular": "Circular",
        "Brochure": "Brochure",
        "Syllabus": "Syllabus",
        "RPP": "Resource Person Profile"
    }

    # Map category to column
    media_columns = {
        "Circular": 11,
        "Brochure": 12,
        "Syllabus": 13,
        "Resource Person Profile": 14
    }

    # Populate rows
    for p in parents:
        # Initialize media categories
        media_mapping = {v: [] for v in field_mapping.values()}

        # Categorize each file based on upload code in filename
        for media in p.get("media_files", []):
            file_name = media.get("file_name", "Unknown")
            url = media.get("direct_api_url", "")
            if not url:
                continue

            # Use original filename for matching
            matched = False

            # Detect the code from filename - try multiple patterns
            for code, category in field_mapping.items():
                # Try different patterns to match the code in filename
                patterns = [
                    f"_{code}_",  # _Circular_, _Brochure_, _Syllabus_, etc.
                    f"_{code}.",  # _Circular.pdf, _Brochure.jpg, etc.
                    f"-{code}-",  # -Circular-, -Brochure-, etc.
                    f"-{code}.",  # -Circular.pdf, -Brochure.jpg, etc.
                    f"{code}_",   # Circular_, Brochure_, Syllabus_, etc.
                    f"_{code}",   # _Circular, _Brochure, _Syllabus, etc.
                ]
                
                for pattern in patterns:
                    if pattern in file_name:
                        media_mapping[category].append((file_name, url))
                        matched = True
                        print(f"Matched '{file_name}' to '{category}' using pattern '{pattern}'")
                        break
                if matched:
                    break

            # If no match found with underscore patterns, try case-insensitive matching
            if not matched:
                file_name_upper = file_name.upper()
                for code, category in field_mapping.items():
                    code_upper = code.upper()
                    # Check if code appears anywhere in filename (case insensitive)
                    if code_upper in file_name_upper:
                        # Additional check to avoid partial matches
                        patterns_upper = [
                            f"_{code_upper}_",
                            f"_{code_upper}.",
                            f"-{code_upper}-",
                            f"-{code_upper}.",
                            f"{code_upper}_",
                            f"_{code_upper}"
                        ]
                        if any(pattern in file_name_upper for pattern in patterns_upper):
                            media_mapping[category].append((file_name, url))
                            matched = True
                            print(f"Matched '{file_name}' to '{category}' using case-insensitive matching")
                            break

        # Create the base info row
        base_row = [
            p.get("staff_id", ""),
            p.get("staff_name", ""),
            p.get("department", ""),
            p.get("programs_offered", ""),
            p.get("course_code", ""),
            p.get("course_duration", ""),
            p.get("times_offered_in_year", ""),
            p.get("students_enrolled", ""),
            p.get("students_completed", ""),
            p.get("year_of_offering", ""),
        ]

        # Append the base row first
        ws.append(base_row)
        row_idx = ws.max_row  # current row number

        # Now add media files to their respective columns
        for category, files in media_mapping.items():
            if category in media_columns:
                col_idx = media_columns[category]
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if files:
                    # Show all filenames joined by line breaks
                    filenames = [f for f, _ in files]
                    cell.value = "\n".join(filenames)
                    
                    # Add hyperlink to first file if URL exists
                    if files[0][1]:
                        cell.hyperlink = files[0][1]
                        cell.font = Font(color="0000EE", underline="single")
                    else:
                        cell.font = Font()  # Regular font if no URL
                else:
                    cell.value = "-"
                    cell.font = Font()  # Regular font
                
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-fit column widths
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            if cell.value:
                # Count the maximum line length for multiline cells
                lines = str(cell.value).split('\n')
                max_line_length = max(len(line) for line in lines)
                max_length = max(max_length, max_line_length)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

    # Adjust row heights for multiline cells
    for row_idx in range(1, ws.max_row + 1):
        max_lines = 1
        for col in range(1, len(headers) + 1):
            cell_value = str(ws.cell(row=row_idx, column=col).value or "")
            line_count = cell_value.count("\n") + 1
            max_lines = max(max_lines, line_count)
        ws.row_dimensions[row_idx].height = min(max_lines * 15, 120)

    # Freeze header and enable filters
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Return as Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="programs_offered.xlsx"'

    with io.BytesIO() as buffer:
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())

    return response

def export_programs_offered_to_pdf(parents, children=None, selected_options=None):
    """
    Export Programs Offered data to PDF (portrait), grouped by staff.
    Each staff has a mini-title "Programs Offered", then Staff info in bold labels, then table.
    Attachments are clickable links.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, PageBreak
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.lib.units import inch
    from django.http import HttpResponse

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="programs_offered.pdf"'

    # Increased margins for better spacing
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,  # Changed to portrait
        topMargin=1.0*inch,
        bottomMargin=1.0*inch,
        leftMargin=0.75*inch,
        rightMargin=0.75*inch
    )

    elements = []
    styles = getSampleStyleSheet()

    # Styles
    title_style = ParagraphStyle('Title', parent=styles['Heading2'], alignment=TA_CENTER, spaceAfter=12)
    left_style = ParagraphStyle('LeftStyle', parent=styles['Normal'], fontSize=9,
                                alignment=TA_LEFT, textColor=colors.HexColor('#2c3e50'))
    right_style = ParagraphStyle('RightStyle', parent=styles['Normal'], fontSize=9,
                                 alignment=TA_RIGHT, textColor=colors.HexColor('#2c3e50'))
    table_header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontSize=8,
                                        alignment=TA_CENTER, textColor=colors.white, fontName='Helvetica-Bold')
    table_cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=7,
                                      alignment=TA_LEFT, textColor=colors.HexColor('#2c3e50'), leading=9)
    table_cell_center_style = ParagraphStyle('TableCellCenter', parent=styles['Normal'], fontSize=7,
                                            alignment=TA_CENTER, textColor=colors.HexColor('#2c3e50'), leading=9)
    attachment_link_style = ParagraphStyle('AttachmentLink', parent=styles['Normal'], fontSize=6,
                                           alignment=TA_LEFT, textColor=colors.HexColor('#1a5276'), leading=8)
    no_data_style = ParagraphStyle('NoData', parent=styles['Normal'], fontSize=10,
                                   alignment=TA_CENTER, textColor=colors.HexColor('#7f8c8d'),
                                   fontStyle='italic', spaceBefore=12, spaceAfter=12)

    # Field mapping for file categorization
    field_mapping = {
        "Circular": "Circular",
        "Brochure": "Brochure",
        "Syllabus": "Syllabus",
        "RPP": "Resource Person Profile"
    }

    # Check if there are any parents
    if not parents:
        elements.append(Paragraph("Programs Offered", title_style))
        elements.append(Spacer(1, 0.5*inch))
        elements.append(Paragraph("No data available", no_data_style))
        doc.build(elements)
        return response

    # Sort parents by staff
    parents_sorted = sorted(parents, key=lambda x: (x.get('staff_id',''), x.get('programs_offered','')))
    current_staff = None
    table_data = []
    
    # Adjusted column widths for portrait orientation
    col_widths = [
        1.5*inch,  # Programs Offered
        1.0*inch,  # Course Code
        0.8*inch,  # Course Duration
        0.8*inch,  # Times Offered
        0.8*inch,  # Students Enrolled
        0.8*inch,  # Students Completed
        0.7*inch,  # Year of Offering
        1.6*inch   # Attachments
    ]

    for parent in parents_sorted:
        staff_id = parent.get('staff_id', 'N/A')
        staff_name = parent.get('staff_name', 'N/A')
        department = parent.get('department', 'N/A')

        # New staff: flush previous table + page break
        if current_staff and current_staff != staff_id:
            # Only add table if there's data (more than just headers)
            if len(table_data) > 1:
                table = Table(table_data, colWidths=col_widths)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                    ('ALIGN', (0,0), (-1,0), 'CENTER'),  # Headers centered
                    ('ALIGN', (0,1), (0,-1), 'LEFT'),    # Programs Offered left aligned
                    ('ALIGN', (1,1), (1,-1), 'LEFT'),    # Course Code left aligned
                    ('ALIGN', (2,1), (2,-1), 'CENTER'),  # Course Duration centered
                    ('ALIGN', (3,1), (3,-1), 'CENTER'),  # Times Offered centered
                    ('ALIGN', (4,1), (4,-1), 'CENTER'),  # Students Enrolled centered
                    ('ALIGN', (5,1), (5,-1), 'CENTER'),  # Students Completed centered
                    ('ALIGN', (6,1), (6,-1), 'CENTER'),  # Year of Offering centered
                    ('ALIGN', (7,1), (7,-1), 'LEFT'),    # Attachments left aligned
                    ('VALIGN', (0,0), (-1,-1), 'TOP'),   # Top alignment for all cells
                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
                    ('FONTSIZE', (0,0), (-1,-1), 7),
                    ('LEFTPADDING', (0,0), (-1,-1), 4),
                    ('RIGHTPADDING', (0,0), (-1,-1), 4),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 3),
                    ('TOPPADDING', (0,0), (-1,-1), 3),
                ]))
                elements.append(table)
            else:
                # Only headers exist, meaning no data for this staff
                elements.append(Spacer(1, 0.3*inch))
                elements.append(Paragraph("No data available for this staff", no_data_style))
            
            elements.append(PageBreak())
            table_data = []

        if current_staff != staff_id:
            # Add mini-title for this staff
            elements.append(Paragraph("Programs Offered", title_style))
            elements.append(Spacer(1, 0.3*inch))

            # Staff info with bold labels
            left_paragraph = Paragraph(
                f"<b>Staff ID:</b> {staff_id}<br/><b>Staff Name:</b> {staff_name}", left_style)
            right_paragraph = Paragraph(f"<b>Department:</b> {department}", right_style)

            total_width = sum(col_widths)
            info_table = Table([[left_paragraph, right_paragraph]], colWidths=[total_width*0.5]*2)
            info_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0,0), (0,0), 'LEFT'),
                ('ALIGN', (1,0), (1,0), 'RIGHT'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('TOPPADDING', (0,0), (-1,-1), 2),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 0.3*inch))

            # Table headers
            headers = [
                'Programs Offered', 'Course Code', 'Duration', 'Times Offered', 
                'Enrolled', 'Completed', 'Year', 'Attachments'
            ]
            table_data.append([Paragraph(h, table_header_style) for h in headers])
            current_staff = staff_id

        # Process media attachments with categorization
        media_files = parent.get('media_files', [])
        categorized_files = {v: [] for v in field_mapping.values()}
        
        # Categorize media files
        for media in media_files:
            file_name = media.get('file_name', 'Unknown')
            url = media.get('direct_api_url', '#')
            
            matched = False
            for code, category in field_mapping.items():
                patterns = [
                    f"_{code}_", f"_{code}.", f"-{code}-", f"-{code}.",
                    f"{code}_", f"_{code}"
                ]
                for pattern in patterns:
                    if pattern in file_name:
                        categorized_files[category].append((file_name, url))
                        matched = True
                        break
                if matched:
                    break
                    
            if not matched:
                file_name_upper = file_name.upper()
                for code, category in field_mapping.items():
                    code_upper = code.upper()
                    patterns_upper = [
                        f"_{code_upper}_", f"_{code_upper}.",
                        f"-{code_upper}-", f"-{code_upper}.",
                        f"{code_upper}_", f"_{code_upper}"
                    ]
                    if any(pattern in file_name_upper for pattern in patterns_upper):
                        categorized_files[category].append((file_name, url))
                        break

        # Build attachments text with categorized files
        if any(categorized_files.values()):
            media_text = []
            for category, files in categorized_files.items():
                if files:
                    for file_name, url in files:
                        # Truncate long filenames for better display
                        display_name = file_name if len(file_name) <= 15 else file_name[:25] + "..."
                        media_text.append(f'<link href="{url}" color="blue">{display_name}</link>')
            
            if media_text:
                media_paragraph = Paragraph('<br/>'.join(media_text), attachment_link_style)
            else:
                media_paragraph = Paragraph("No attachments", table_cell_style)
        else:
            media_paragraph = Paragraph("No attachments", table_cell_style)

        # Create rows with appropriate alignment styles
        row = [
            Paragraph(str(parent.get('programs_offered', '')), table_cell_style),
            Paragraph(str(parent.get('course_code', '')), table_cell_style),
            Paragraph(str(parent.get('course_duration', '')), table_cell_center_style),
            Paragraph(str(parent.get('times_offered_in_year', '')), table_cell_center_style),
            Paragraph(str(parent.get('students_enrolled', '')), table_cell_center_style),
            Paragraph(str(parent.get('students_completed', '')), table_cell_center_style),
            Paragraph(str(parent.get('year_of_offering', '')), table_cell_center_style),
            media_paragraph
        ]
        table_data.append(row)

    # Add last staff table
    if table_data:
        # Check if there's actual data (more than just headers)
        if len(table_data) > 1:
            table = Table(table_data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('ALIGN', (0,0), (-1,0), 'CENTER'),  # Headers centered
                ('ALIGN', (0,1), (0,-1), 'LEFT'),    # Programs Offered left aligned
                ('ALIGN', (1,1), (1,-1), 'LEFT'),    # Course Code left aligned
                ('ALIGN', (2,1), (2,-1), 'CENTER'),  # Course Duration centered
                ('ALIGN', (3,1), (3,-1), 'CENTER'),  # Times Offered centered
                ('ALIGN', (4,1), (4,-1), 'CENTER'),  # Students Enrolled centered
                ('ALIGN', (5,1), (5,-1), 'CENTER'),  # Students Completed centered
                ('ALIGN', (6,1), (6,-1), 'CENTER'),  # Year of Offering centered
                ('ALIGN', (7,1), (7,-1), 'LEFT'),    # Attachments left aligned
                ('VALIGN', (0,0), (-1,-1), 'TOP'),   # Top alignment for all cells
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
                ('FONTSIZE', (0,0), (-1,-1), 7),
                ('LEFTPADDING', (0,0), (-1,-1), 4),
                ('RIGHTPADDING', (0,0), (-1,-1), 4),
                ('BOTTOMPADDING', (0,0), (-1,-1), 3),
                ('TOPPADDING', (0,0), (-1,-1), 3),
            ]))
            elements.append(table)
        else:
            # Only headers exist, meaning no data for this staff
            elements.append(Spacer(1, 0.3*inch))
            elements.append(Paragraph("No data available for this staff", no_data_style))

    doc.build(elements)
    return response