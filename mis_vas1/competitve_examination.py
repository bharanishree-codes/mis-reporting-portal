import json
import requests
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from MIS.functions import validate_file_format, validate_file_size
from datetime import date, datetime
from user_management.settings_views import *

API_STUDIO_URL = user_bundle_settings()

def competive_key():
    url = "https://api.hcaschennai.edu.in/auth/token"

    payload = json.dumps({
        "secret_key": "C4ZoXbsAnHLjk1Xyz4QPT2eoiNx6K6fo"
    })
    headers = {'Content-Type': 'application/json'}

    response = requests.post(url, headers=headers, data=payload)
    if response.status_code == 200:
        res_data = response.json()
        access_token = res_data.get('access_token')
        token_type = res_data.get('token_type')
        return access_token, token_type
    return None, None


# Function to fetch staff career data
def get_competive_data(access_token, token_type):
    url = "https://api.hcaschennai.edu.in/sqlviews/api/v1/auth/get_response_data"

    payload = json.dumps({
        "psk_uid": "51a531b4-bd55-491c-861d-a8d7227b325b",
        "project": "hcas",
        "data": {}
    })
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'{token_type} {access_token}'  # Fix spacing
    }

    response = requests.post(url, headers=headers, data=payload)
    if response.status_code == 200:
        return response.json()
    return []


# Create competitve_examination
def competitve_examination_create(request):
    error_message = None

    # Step 1: Get API token
    access_token, token_type = competive_key()
    if not access_token or not token_type:
        error_message = 'Failed to get access token from API.'
        return render(request, 'competitve_examination_templates/competitve_examination_create.html', {'error': error_message})

    # Step 2: Fetch career data
    competive_data = get_competive_data(access_token, token_type)
    if not competive_data:
        error_message = 'Failed to fetch staff data.'
        return render(request, 'competitve_examination_templates/competitve_examination_create.html', {'error': error_message})
    
    user = get_settings(request)
    username = user["username"]
    # username = 'AC-NT012'
    
    selected_faculty = None
    
    for faculty in competive_data:
        if faculty['stf_id'] == username:
            selected_faculty = faculty
            break
        
    staff_name = selected_faculty.get('stf_name') if selected_faculty else ''
    department_name = selected_faculty.get('department') if selected_faculty else ""
    
    # Generate academic years for dropdown
    current_year = datetime.now().year
    publication_year = [f"{y}-{y+1}" for y in range(current_year, 2010, -1)]

    if request.method == 'POST':
        # Get form data from the POST request
        staff_name = request.POST.get('staff_name')
        staff_id = username
        department_name = request.POST.get('department_name')
        year_range = request.POST.get('year_of_activity')  # This will be in "YYYY-YYYY" format
        name_of_the_activity = request.POST.get('name_of_the_activity')
        number_of_students_attended = request.POST.get('number_of_students_attended')
        number_of_students_placed = request.POST.get('number_of_students_placed')
        
        # Convert year range to date format for storage
        if year_range and '-' in year_range:
            try:
                start_year = int(year_range.split('-')[0])
                # Create a date object (using June 1st as default)
                year_of_activity = date(start_year, 6, 1).isoformat()
            except (ValueError, IndexError):
                year_of_activity = None
        else:
            year_of_activity = None
        
        for faculty in competive_data:
            if faculty['stf_id'] == username:
                selected_faculty = faculty
                break
        
        staff_name = selected_faculty.get('stf_name') if selected_faculty else ''
        department_name = selected_faculty.get('department') if selected_faculty else ""

        
        # URL to send the form data to
        url = f"{API_STUDIO_URL}postapi/create/naac01_students_benefited_exam_guidance_dc1"
        payload = json.dumps({"data":
            {
                "staff_name": staff_name,
                "staff_id": staff_id,
                "department_name": department_name,
                "name_of_the_activity": name_of_the_activity,
                "year_of_activity": year_of_activity,  # Use the converted date
                "number_of_students_attended": number_of_students_attended,
                "number_of_students_placed": number_of_students_placed
            }})
        headers = {'Content-Type': 'application/json'}

        # Make API call to create the competitve_examination/workshop
        try:
            response = requests.post(url, headers=headers, data=payload)
            response.raise_for_status()  # This will raise an exception if the status code is not 2xx
        except requests.exceptions.HTTPError as http_err:
            # HTTP error
            messages.error(request, message=f"HTTP error occurred: {http_err}")
            return render(request, 'competitve_examination_templates/competitve_examination_create.html')
        except requests.exceptions.RequestException as req_err:
            # Other types of requests errors
            messages.error(request, message=f"Request error occurred: {req_err}")
            return render(request, 'competitve_examination_templates/competitve_examination_create.html')

        if response.status_code == 200:
            # Successfully created the competitve_examination/workshop
            file_data = response.json()
            psk_id = file_data.get('psk_id')  # Get psk_id from the response

            if not psk_id:
                messages.error(request, message="Failed to create the competitve examination. No psk_id returned.")
                return render(request, 'competitve_examination_templates/competitve_examination_create.html')

            # Uploading files (optional step)
            upload_url = f"{API_STUDIO_URL}crudapp/upload/media/naac01_students_benefited_exam_guidance_dc1_media"
            uploaded_files = request.FILES.getlist('file')

            if not uploaded_files:
                messages.error(request, message="No files selected for upload.")
                return render(request, 'competitve_examination_templates/competitve_examination_create.html')

            fields = ['PL', 'Circular', 'Brochure', 'RPP', 'DR', 'GTP', 'SA']
            for field, uploaded_file in zip(fields, uploaded_files):
                try:
                    validate_file_size(uploaded_file)
                    validate_file_format(uploaded_file)
                    file_type = uploaded_file.content_type
                    current_year = datetime.now().year
                    custom_filename = f"{staff_id}_{field}_{current_year}_{uploaded_file.name}"
                    print(f"Generated filename: {custom_filename}")  # Print the filename for each iteration
                    
                    payload = {'parent_psk_id': psk_id}
                    files = {'media': (custom_filename, uploaded_file, file_type)}
                    upload_headers = {'api_name': 'naac01_students_benefited_exam_guidance_dc1_media'}

                    # Make API call to upload the file
                    upload_response = requests.post(upload_url, headers=upload_headers, data=payload, files=files)

                    if upload_response.status_code != 200:
                        # File upload failed
                        messages.error(request, message=f"File upload failed for {uploaded_file.name}. Error: {upload_response.text}")
                        return redirect('competitve_examination_view', id=psk_id)

                except Exception as e:
                    messages.error(request, message=f"Error during file upload: {str(e)}")
                    return redirect('competitve_examination_view', id=psk_id)

            messages.success(request, message="Documents uploaded successfully.")
            return redirect('competitve_examination_list')

        else:
            # API call failed for creating competitve_examination
            messages.error(request, message="Failed to create competitve_examination. Please try again.")
            return render(request, 'competitve_examination_templates/competitve_examination_create.html')
    else:
        # If the request is GET, render the competitve_examination creation form
        return render(request, 'competitve_examination_templates/competitve_examination_create.html', {
            'competive_data': competive_data, 
            "username": username, 
            "staff_name": staff_name, 
            "department": department_name,
            'publication_year': publication_year  # Pass years to template
        })



# View competitve_examination Details
def competitve_examination_view(request, id):
    url = f"{API_STUDIO_URL}getapi/naac01_students_benefited_exam_guidance_dc1/{id}"
    response = requests.get(url)

    if response.status_code == 200:
        competitve_examination_data = response.json()
        return render(request, "competitve_examination_templates/competitve_examination_view.html",
                      {'competitve_examination': competitve_examination_data})

    return HttpResponse(f"Error fetching Course details: {response.text}", status=500)


# List All competitve_examinations
def competitve_examination_list(request):
    # URL to get competitive examination data
    url = f"{API_STUDIO_URL}getapi/all_fields/naac01_students_benefited_exam_guidance_dc1/all"
    response = requests.get(url)

    if response.status_code == 200:
        competitve_examinations = response.json()
        print("competitve_examinations:", competitve_examinations)
        
        # Format the year for each entry to show as range
        for exam in competitve_examinations:
            year_str = exam.get('year_of_activity', '')
            if year_str and isinstance(year_str, str) and len(year_str) >= 4:
                try:
                    # Extract year from date string (e.g., "2005-06-01" -> 2005)
                    year = year_str[:4]
                    exam['year_range'] = f"{year}-{int(year) + 1}"
                except (ValueError, TypeError):
                    exam['year_range'] = year_str
            else:
                exam['year_range'] = year_str
                
    else:
        competitve_examinations = []

    # Get the current username (staff_id)
    user = get_settings(request)
    username = user["username"]
    # username = 'AC-NT012'
    
    filtered_competitve_examinations = [exam for exam in competitve_examinations if exam.get('staff_id') == username]
    
    selected_staff_id = request.GET.get('staff_id')
    
    if selected_staff_id:
        filtered_competitve_examinations = [exam for exam in competitve_examinations if exam.get('staff_id') == selected_staff_id]

    selected_department = request.GET.get('department')
    if selected_department:
        filtered_competitve_examinations = [exam for exam in competitve_examinations if exam.get('department_name') == selected_department]
        print("filtered_competitve_examinations:", filtered_competitve_examinations)

    if not filtered_competitve_examinations:
        return render(request, 'competitve_examination_templates/competitve_examination_list.html', {'competitve_examinations': []})

    return render(request, 'competitve_examination_templates/competitve_examination_list.html', {
        'competitve_examinations': filtered_competitve_examinations
    })

# Update competitve_examination
def competitve_examination_update(request, id):
    # Generate academic years for dropdown
    current_year = datetime.now().year
    publication_year = [f"{y}-{y+1}" for y in range(current_year, 2010, -1)]

    # Fetch competitve_examination details
    url = f"{API_STUDIO_URL}getapi/naac01_students_benefited_exam_guidance_dc1/{id}"
    response = requests.get(url)

    if response.status_code == 200:
        competitve_examination = response.json()
        
        # Convert date format "2005-06-01" to "2005-2006" for display
        year_str = competitve_examination.get('year_of_activity', '')
        if year_str and isinstance(year_str, str) and len(year_str) >= 4:
            try:
                # Extract the year part and create range format
                start_year = year_str[:4]
                competitve_examination['year_range'] = f"{start_year}-{int(start_year) + 1}"
            except (ValueError, TypeError):
                competitve_examination['year_range'] = year_str
        else:
            competitve_examination['year_range'] = year_str
            
    else:
        return HttpResponse(f"Error fetching competitve_examination details: {response.text}", status=500)

    # Fetch media (child files) associated with this competitve_examination
    media_url = f"{API_STUDIO_URL}crudapp/get/media/naac01_students_benefited_exam_guidance_dc1_media/parent/{id}"
    media_response = requests.get(media_url)

    if media_response.status_code == 200:
        child_files = media_response.json()
    else:
        return HttpResponse(f"Failed to fetch media files: {media_response.text}", status=500)

    # Handle form submission (POST request)
    if request.method == "POST":
        # Prepare data for updating the competitve_examination/workshop
        update_url = f"{API_STUDIO_URL}updateapi/update/naac01_students_benefited_exam_guidance_dc1/{id}"

        # Convert selected year range back to date format for storage
        year_range = request.POST.get('year_of_activity')
        year_of_activity_date = None
        
        if year_range and '-' in year_range:
            try:
                start_year = int(year_range.split('-')[0])
                # Create a date object (using June 1st as default)
                year_of_activity_date = date(start_year, 6, 1).isoformat()
            except (ValueError, IndexError):
                year_of_activity_date = competitve_examination.get('year_of_activity')
        else:
            year_of_activity_date = competitve_examination.get('year_of_activity')

        payload = json.dumps({"data": {
            "name_of_the_activity": request.POST.get('name_of_the_activity', competitve_examination.get('name_of_the_activity')),
            "staff_name": request.POST.get('staff_name', competitve_examination.get('staff_name')),
            "staff_id": request.POST.get('staff_id', competitve_examination.get('staff_id')),
            "department_name": request.POST.get('department_name', competitve_examination.get('department_name')),
            "year_of_activity": year_of_activity_date,  # Use the converted date
            "number_of_students_attended": request.POST.get('number_of_students_attended', competitve_examination.get('number_of_students_attended')),
            "number_of_students_placed": request.POST.get('number_of_students_placed', competitve_examination.get('number_of_students_placed'))
        }})

        headers = {'Content-Type': 'application/json'}
        update_response = requests.put(update_url, headers=headers, data=payload)

        if update_response.status_code != 200:
            return HttpResponse(f"Failed to update competitve_examination details: {update_response.text}", status=500)

        # Handle media (file) uploads
        upload_errors = []

        for child in child_files:
            upload_id = child['psk_id']
            fields = ['PL', 'Circular', 'Brochure', 'RPP', 'DR', 'GTP', 'SA']

            for field in fields:
                uploaded_files = request.FILES.getlist(f'file_{upload_id}_{field}')

                if not uploaded_files:
                    continue  # Skip if no files are uploaded for this field

                for uploaded_file in uploaded_files:
                    # Validate file size and format
                    validate_file_size(uploaded_file)
                    validate_file_format(uploaded_file)

                    # Generate custom filename
                    staff_id = request.POST.get('staff_id')
                    current_year = datetime.now().year
                    print(staff_id)
                    custom_filename = f"{staff_id}_{field}_{current_year}_{uploaded_file.name}"

                    print(f"Generated filename for field {field}: {custom_filename}")  # Debug log

                    # Construct the upload URL and payload
                    upload_url = f"{API_STUDIO_URL}crudapp/upload/media/naac01_students_benefited_exam_guidance_dc1_media/{upload_id}"
                    files = {'media': (custom_filename, uploaded_file, uploaded_file.content_type)}
                    headers = {
                        'api_name': 'naac01_students_benefited_exam_guidance_dc1_media',
                        'psk_id': str(upload_id)
                    }
                    payload = {'parent_psk_id': id}

                    # Upload the file
                    upload_response = requests.put(upload_url, headers=headers, data=payload, files=files)

                if upload_response.status_code != 200:
                    upload_errors.append(f"Error uploading file for {upload_id}: {upload_response.text}")

        # Provide feedback to the user
        if upload_errors:
            for error in upload_errors:
                messages.error(request, error)
        else:
            messages.success(request, "Files uploaded successfully.")

        # Redirect to the competitve_examination view page after successful update
        return redirect('competitve_examination_list')

    # If not a POST request, render the update form
    return render(request, 'competitve_examination_templates/competitve_examination_update.html', {
        'competitve_examination': competitve_examination, 
        'child_files': child_files,
        'publication_year': publication_year  # Pass years to template
    })

# Delete competitve_examination
def competitve_examination_delete(request, id):
    # url = f"{API_STUDIO_URL}getapi/naac01_students_benefited_exam_guidance_dc1/{id}"
    # response = requests.get(url)
    # competitve_examination = response.json()

    # if request.method == 'POST':
    delete_url = f"{API_STUDIO_URL}deleteapi/delete/naac01_students_benefited_exam_guidance_dc1/{id}"
    payload = ""
    headers = {}
    delete_response = requests.request("DELETE", delete_url, headers=headers, data=payload)

    if delete_response.status_code == 200:
        return redirect('competitve_examination_list')
    else:
        return HttpResponse("Failed to delete participation: " + delete_response.text)

    # return render(request, 'competitve_examination_templates/competitve_examination_delete.html', {'competitve_examination': competitve_examination})


# def export_competitive_examination_to_pdf(parents, children=None, selected_options=None):
#     """
#     Export Competitive Examination data to PDF (landscape), grouped by staff.
#     Media files are listed only once per examination without category labels.
#     Fields: Activity Name, Year, Students Attended, Students Placed, Attachments
#     """
#     from reportlab.lib.pagesizes import A4, landscape
#     from reportlab.lib import colors
#     from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
#     from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, PageBreak
#     from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
#     from reportlab.lib.units import inch
#     from django.http import HttpResponse

#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="competitive_examination.pdf"'

#     doc = SimpleDocTemplate(
#         response,
#         pagesize=landscape(A4),
#         topMargin=0.5 * inch,
#         bottomMargin=0.5 * inch,
#         leftMargin=0.5 * inch,
#         rightMargin=0.5 * inch
#     )

#     elements = []
#     styles = getSampleStyleSheet()

#     # ====== Custom Styles ======
#     title_style = ParagraphStyle('Title', parent=styles['Heading2'], alignment=TA_CENTER, spaceAfter=12)
#     table_header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontSize=9,
#                                         alignment=TA_CENTER, textColor=colors.white, fontName='Helvetica-Bold')
#     table_cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=8,
#                                       alignment=TA_LEFT, wordWrap='CJK')
#     attachment_link_style = ParagraphStyle('AttachmentLink', parent=styles['Normal'], fontSize=7,
#                                            alignment=TA_LEFT, textColor=colors.HexColor('#1a5276'), wordWrap='CJK')
#     no_data_style = ParagraphStyle('NoData', parent=styles['Normal'], fontSize=10,
#                                    alignment=TA_CENTER, textColor=colors.HexColor('#7f8c8d'),
#                                    fontStyle='italic', spaceBefore=12, spaceAfter=12)

#     # ====== Check if there are any parents ======
#     if not parents:
#         elements.append(Paragraph("Competitive Examination", title_style))
#         elements.append(Paragraph("No data available", no_data_style))
#         doc.build(elements)
#         return response

#     # ====== Media Lookup ======
#     media_lookup = {}
#     if children:
#         for child in children:
#             psk_id = child.get('psk_id')
#             if psk_id:
#                 media_lookup[psk_id] = child.get('media_files', [])
#     else:
#         for p in parents:
#             media_lookup[p.get('psk_id')] = p.get('media_files', [])

#     parents_sorted = sorted(parents, key=lambda x: (x.get('staff_id', ''), x.get('name_of_the_activity', '')))
#     current_staff = None
#     table_data = []
#     col_widths = [3 * inch, 1 * inch, 1.2 * inch, 1.2 * inch, 2 * inch]

#     for parent in parents_sorted:
#         staff_id = parent.get('staff_id', 'N/A')
#         staff_name = parent.get('staff_name', 'N/A')
#         department_name = parent.get('department_name', 'N/A')
#         psk_id = parent.get('psk_id')

#         # ====== Flush Previous Staff Table ======
#         if current_staff and current_staff != staff_id:
#             # Only add table if there's data (more than just headers)
#             if len(table_data) > 1:
#                 table = Table(table_data, colWidths=col_widths, repeatRows=1, splitByRow=True)
#                 table.setStyle(TableStyle([
#                     ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#02548b')),
#                     ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
#                     ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#                     ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
#                     ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d6d6d6')),
#                     ('LEFTPADDING', (0, 0), (-1, -1), 4),
#                     ('RIGHTPADDING', (0, 0), (-1, -1), 4),
#                     ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
#                     ('TOPPADDING', (0, 0), (-1, -1), 4),
#                 ]))
#                 elements.append(table)
#             else:
#                 # Only headers exist, meaning no data for this staff
#                 elements.append(Paragraph("No data available for this staff", no_data_style))
            
#             elements.append(PageBreak())
#             table_data = []

#         # ====== Add New Staff Section ======
#         if current_staff != staff_id:
#             elements.append(Paragraph("Competitive Examination", title_style))
#             elements.append(Spacer(1, 6))

#             # Staff info section
#             left_style = ParagraphStyle('LeftAlign', parent=table_cell_style, alignment=TA_LEFT)
#             right_style = ParagraphStyle('RightAlign', parent=table_cell_style, alignment=TA_RIGHT)
#             center_style = ParagraphStyle('CenterAlign', parent=table_cell_style, alignment=TA_CENTER)

#             # Line 1: Staff ID (left) | Department (right)
#             staff_id_para = Paragraph(f"<b>Staff ID:</b> {staff_id}", left_style)
#             dept_para = Paragraph(f"<b>Department:</b> {department_name}", right_style)
#             line1 = Table([[staff_id_para, dept_para]], colWidths=[5.5 * inch, 4.5 * inch])
#             line1.setStyle(TableStyle([
#                 ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
#                 ('ALIGN', (0, 0), (0, -1), 'LEFT'),
#                 ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
#                 ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
#                 ('TOPPADDING', (0, 0), (-1, -1), 3),
#             ]))
#             elements.append(line1)

#             # Line 2: Staff Name (left only)
#             staff_name_para = Paragraph(f"<b>Staff Name:</b> {staff_name}", left_style)
#             line2 = Table([[staff_name_para]], colWidths=[10 * inch])
#             line2.setStyle(TableStyle([
#                 ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
#                 ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
#                 ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
#                 ('TOPPADDING', (0, 0), (-1, -1), 3),
#             ]))
#             elements.append(line2)
#             elements.append(Spacer(1, 10))

#             # Table headers
#             headers = ['Activity Name', 'Year', 'Students Attended', 'Students Placed', 'Attachments']
#             table_data.append([Paragraph(h, table_header_style) for h in headers])
#             current_staff = staff_id

#         # ====== Media Files ======
#         media_files = media_lookup.get(psk_id, [])
#         media_text = []
#         seen_files = set()
#         for media in media_files:
#             fname = media.get("file_name", "Unknown")
#             url = media.get("direct_api_url", "")
#             if not url or fname in seen_files:
#                 continue
#             seen_files.add(fname)
#             display_name = fname if len(fname) <= 50 else fname[:50] + "..."
#             media_text.append(f'<link href="{url}" color="blue">{display_name}</link>')

#         media_paragraph = Paragraph('<br/>'.join(media_text) if media_text else "No attachments", attachment_link_style)

#         # Data row
#         row = [
#             Paragraph(str(parent.get('name_of_the_activity', '')), table_cell_style),
#             Paragraph(str(parent.get('year_range', parent.get('year_of_activity', ''))), table_cell_style),
#             Paragraph(str(parent.get('number_of_students_attended', '')), table_cell_style),
#             Paragraph(str(parent.get('number_of_students_placed', '')), table_cell_style),
#             media_paragraph
#         ]
#         table_data.append(row)

#     # ====== Final Table Flush ======
#     if table_data:
#         # Check if there's actual data (more than just headers)
#         if len(table_data) > 1:
#             table = Table(table_data, colWidths=col_widths, repeatRows=1, splitByRow=True)
#             table.setStyle(TableStyle([
#                 ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#02548b')),
#                 ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
#                 ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#                 ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
#                 ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d6d6d6')),
#                 ('LEFTPADDING', (0, 0), (-1, -1), 4),
#                 ('RIGHTPADDING', (0, 0), (-1, -1), 4),
#                 ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
#                 ('TOPPADDING', (0, 0), (-1, -1), 4),
#             ]))
#             elements.append(table)
#         else:
#             # Only headers exist, meaning no data for this staff
#             elements.append(Paragraph("No data available for this staff", no_data_style))

#     doc.build(elements)
#     return response

def export_competitive_examination_to_pdf(parents, children=None, selected_options=None):
    """
    Export Competitive Examination data to PDF (portrait), grouped by staff.
    Each staff has a mini-title "Competitive Examination", then Staff info in bold labels, then table.
    Attachments are clickable links.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, PageBreak
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.lib.units import inch
    from django.http import HttpResponse

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="collaborative_students.pdf"'

    # Increased margins for better spacing
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,  # Changed to portrait
        topMargin=1.0*inch,
        bottomMargin=1.0*inch,
        leftMargin=0.75*inch,
        rightMargin=0.75*inch
    )

    elements = []
    styles = getSampleStyleSheet()

    # Styles
    title_style = ParagraphStyle('Title', parent=styles['Heading2'], alignment=TA_CENTER, spaceAfter=12)
    left_style = ParagraphStyle('LeftStyle', parent=styles['Normal'], fontSize=9,
                                alignment=TA_LEFT, textColor=colors.HexColor('#2c3e50'))
    right_style = ParagraphStyle('RightStyle', parent=styles['Normal'], fontSize=9,
                                 alignment=TA_RIGHT, textColor=colors.HexColor('#2c3e50'))
    table_header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontSize=8,
                                        alignment=TA_CENTER, textColor=colors.white, fontName='Helvetica-Bold')
    table_cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=7,
                                      alignment=TA_LEFT, textColor=colors.HexColor('#2c3e50'), leading=9)
    table_cell_center_style = ParagraphStyle('TableCellCenter', parent=styles['Normal'], fontSize=7,
                                            alignment=TA_CENTER, textColor=colors.HexColor('#2c3e50'), leading=9)
    attachment_link_style = ParagraphStyle('AttachmentLink', parent=styles['Normal'], fontSize=6,
                                           alignment=TA_LEFT, textColor=colors.HexColor('#1a5276'), leading=8)
    no_data_style = ParagraphStyle('NoData', parent=styles['Normal'], fontSize=10,
                                   alignment=TA_CENTER, textColor=colors.HexColor('#7f8c8d'),
                                   fontStyle='italic', spaceBefore=12, spaceAfter=12)

    # Check if there are any parents
    if not parents:
        elements.append(Paragraph("Competitive Examination", title_style))
        elements.append(Spacer(1, 0.5*inch))
        elements.append(Paragraph("No data available", no_data_style))
        doc.build(elements)
        return response

    # Sort parents by staff
    parents_sorted = sorted(parents, key=lambda x: (x.get('staff_id',''), x.get('activity_title','')))
    current_staff = None
    table_data = []
    
    # Adjusted column widths for portrait orientation
    col_widths = [
        2.0*inch,  # Activity Title
        1.3*inch,  # Agency Name
        1.0*inch,  # Participant Name
        0.7*inch,  # Year
        2.0*inch   # Attachments
    ]

    for parent in parents_sorted:
        staff_id = parent.get('staff_id', 'N/A')
        staff_name = parent.get('staff_name', 'N/A')
        department_name = parent.get('department_name', 'N/A')

        # New staff: flush previous table + page break
        if current_staff and current_staff != staff_id:
            # Only add table if there's data (more than just headers)
            if len(table_data) > 1:
                table = Table(table_data, colWidths=col_widths)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                    ('ALIGN', (0,0), (-1,0), 'CENTER'),  # Headers centered
                    ('ALIGN', (0,1), (0,-1), 'LEFT'),    # Activity Title left aligned
                    ('ALIGN', (1,1), (1,-1), 'LEFT'),    # Agency Name left aligned
                    ('ALIGN', (2,1), (2,-1), 'LEFT'),    # Participant Name left aligned
                    ('ALIGN', (3,1), (3,-1), 'CENTER'),  # Year centered
                    ('ALIGN', (4,1), (4,-1), 'LEFT'),    # Activity Name left aligned
                    ('ALIGN', (5,1), (5,-1), 'LEFT'),    # Attachments left aligned
                    ('VALIGN', (0,0), (-1,-1), 'TOP'),   # Top alignment for all cells
                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
                    ('FONTSIZE', (0,0), (-1,-1), 7),
                    ('LEFTPADDING', (0,0), (-1,-1), 4),
                    ('RIGHTPADDING', (0,0), (-1,-1), 4),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 3),
                    ('TOPPADDING', (0,0), (-1,-1), 3),
                ]))
                elements.append(table)
            else:
                # Only headers exist, meaning no data for this staff
                elements.append(Spacer(1, 0.3*inch))
                elements.append(Paragraph("No data available for this staff", no_data_style))
            
            elements.append(PageBreak())
            table_data = []

        if current_staff != staff_id:
            # Add mini-title for this staff
            elements.append(Paragraph("Competitive Examination", title_style))
            elements.append(Spacer(1, 0.3*inch))

            # Staff info with bold labels
            left_paragraph = Paragraph(
                f"<b>Staff ID:</b> {staff_id}<br/><b>Staff Name:</b> {staff_name}", left_style)
            right_paragraph = Paragraph(f"<b>Department:</b> {department_name}", right_style)

            total_width = sum(col_widths)
            info_table = Table([[left_paragraph, right_paragraph]], colWidths=[total_width*0.5]*2)
            info_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0,0), (0,0), 'LEFT'),
                ('ALIGN', (1,0), (1,0), 'RIGHT'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('TOPPADDING', (0,0), (-1,-1), 2),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 0.3*inch))

            # Table headers
            headers = [
                'Name of the activity', 'Number of students Attended', 'Number of students placed', 'Year', 'Attachments'
            ]
            table_data.append([Paragraph(h, table_header_style) for h in headers])
            current_staff = staff_id

        # Media attachments
        media_files = parent.get('media_files', [])
        if media_files:
            media_text = []
            for m in media_files:
                filename = m.get('file_name', 'Unknown')
                url = m.get('direct_api_url', '#')
                # Truncate long filenames for better display in portrait mode
                display_name = filename if len(filename) <= 25 else filename[:22] + "..."
                media_text.append(f'<link href="{url}" color="blue">{display_name}</link>')
            media_paragraph = Paragraph('<br/>'.join(media_text), attachment_link_style)
        else:
            media_paragraph = Paragraph("No attachments", table_cell_style)

        # Create rows with appropriate alignment styles
        row = [
            Paragraph(str(parent.get('name_of_the_activity', '')), table_cell_style),
            Paragraph(str(parent.get('number_of_students_attended', '')), table_cell_style),
            Paragraph(str(parent.get('number_of_students_placed', '')), table_cell_style),
            Paragraph(str(parent.get('year_of_activity', '')), table_cell_center_style),
            media_paragraph
        ]
        table_data.append(row)

    # Add last staff table
    if table_data:
        # Check if there's actual data (more than just headers)
        if len(table_data) > 1:
            table = Table(table_data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('ALIGN', (0,0), (-1,0), 'CENTER'),  # Headers centered
                ('ALIGN', (0,1), (0,-1), 'LEFT'),    # Activity Title left aligned
                ('ALIGN', (1,1), (1,-1), 'LEFT'),    # Agency Name left aligned
                ('ALIGN', (2,1), (2,-1), 'LEFT'),    # Participant Name left aligned
                ('ALIGN', (3,1), (3,-1), 'CENTER'),  # Year centered
                ('ALIGN', (4,1), (4,-1), 'LEFT'),    # Activity Name left aligned
                ('ALIGN', (5,1), (5,-1), 'LEFT'),    # Attachments left aligned
                ('VALIGN', (0,0), (-1,-1), 'TOP'),   # Top alignment for all cells
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
                ('FONTSIZE', (0,0), (-1,-1), 7),
                ('LEFTPADDING', (0,0), (-1,-1), 4),
                ('RIGHTPADDING', (0,0), (-1,-1), 4),
                ('BOTTOMPADDING', (0,0), (-1,-1), 3),
                ('TOPPADDING', (0,0), (-1,-1), 3),
            ]))
            elements.append(table)
        else:
            # Only headers exist, meaning no data for this staff
            elements.append(Spacer(1, 0.3*inch))
            elements.append(Paragraph("No data available for this staff", no_data_style))

    doc.build(elements)
    return response

# def export_competitive_examination_to_excel(parents, children=None, selected_options=None):
#     """
#     Export Competitive Examination data to Excel, including clickable media file links.
#     Media categories: Permission Letter, Circular, Brochure, Resource Person Profile, 
#     Detailed Report, Geo Tagged Photos, Student Attendance.
#     Files are mapped based on naming convention: staffid_field_year_filename.pdf
#     """
#     import io
#     from django.http import HttpResponse
#     from openpyxl import Workbook
#     from openpyxl.styles import Font, Alignment

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Competitive Examination"

#     headers = [
#         "Staff ID", "Staff Name", "Department", "Activity Name",
#         "Year", "Students Attended", "Students Placed",
#         "Permission Letter", "Circular", "Brochure",
#         "Resource Person Profile", "Detailed Report", "Geo Tagged Photos", "Student Attendance"
#     ]
#     ws.append(headers)

#     # Style headers
#     for col in range(1, len(headers)+1):
#         cell = ws.cell(row=1, column=col)
#         cell.font = Font(bold=True)
#         cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

#     # Map upload field codes to friendly names
#     field_mapping = {
#         "PL": "Permission Letter",
#         "Circular": "Circular",
#         "Brochure": "Brochure",
#         "RPP": "Resource Person Profile",
#         "DR": "Detailed Report",
#         "GTP": "Geo Tagged Photos",
#         "SA": "Student Attendance"
#     }

#     # Map category to column
#     media_columns = {
#         "Permission Letter": 8,
#         "Circular": 9,
#         "Brochure": 10,
#         "Resource Person Profile": 11,
#         "Detailed Report": 12,
#         "Geo Tagged Photos": 13,
#         "Student Attendance": 14
#     }

#     for p in parents:
#         # Initialize media categories
#         media_mapping = {v: [] for v in field_mapping.values()}

#         # Categorize each file based on upload code in filename
#         for media in p.get("media_files", []):
#             fname = media.get("file_name", "Unknown")
#             url = media.get("direct_api_url", "")
#             if not url:
#                 continue

#             # Determine category based on filename field
#             matched = False
#             for field, category in field_mapping.items():
#                 if f"_{field}_" in fname.upper():  # Matches "_PL_", "_Circular_", etc.
#                     media_mapping[category].append((fname, url))
#                     matched = True
#                     break
#             if not matched:
#                 # Default to Permission Letter if no match
#                 media_mapping["Permission Letter"].append((fname, url))

#         base_row = [
#             p.get("staff_id", ""), p.get("staff_name", ""), p.get("department_name", ""),
#             p.get("name_of_the_activity", ""), 
#             p.get("year_range", p.get("year_of_activity", "")),
#             p.get("number_of_students_attended", ""), p.get("number_of_students_placed", "")
#         ]
#         ws.append(base_row + ["", "", "", "", "", "", ""])
#         row_idx = ws.max_row

#         # Add media hyperlinks
#         for category, files in media_mapping.items():
#             if category in media_columns:
#                 col_idx = media_columns[category]
#                 cell = ws.cell(row=row_idx, column=col_idx)
#                 if files:
#                     # Show all filenames joined by line breaks
#                     filenames = [f for f, _ in files]
#                     cell.value = "\n".join(filenames)
#                     if files[0][1]:
#                         cell.hyperlink = files[0][1]
#                         cell.font = Font(color="0000EE", underline="single")
#                 else:
#                     cell.value = "-"
#                 cell.alignment = Alignment(wrap_text=True, vertical="top")

#     # Auto-fit columns
#     for col_idx, column_cells in enumerate(ws.columns, 1):
#         max_length = 0
#         for cell in column_cells:
#             if cell.value:
#                 lines = str(cell.value).split('\n')
#                 max_line_length = max(len(line) for line in lines)
#                 max_length = max(max_length, max_line_length)
#         ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

#     # Adjust row heights
#     for row_idx in range(1, ws.max_row+1):
#         max_lines = 1
#         for col in range(1, len(headers)+1):
#             cell_value = str(ws.cell(row=row_idx, column=col).value or "")
#             line_count = cell_value.count("\n")+1
#             max_lines = max(max_lines, line_count)
#         ws.row_dimensions[row_idx].height = min(max_lines*15, 120)

#     ws.freeze_panes = "A2"
#     ws.auto_filter.ref = ws.dimensions

#     response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#     response["Content-Disposition"] = 'attachment; filename="competitive_examination.xlsx"'
#     with io.BytesIO() as buffer:
#         wb.save(buffer)
#         buffer.seek(0)
#         response.write(buffer.getvalue())
#     return response

def export_competitive_examination_to_excel(parents, children=None):
    """
    Export Competitive Examination data to Excel from filtered parents,
    including clickable media file links based on upload field codes.
    """
    import io
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Competitive Examination"

    # Headers including media attachments
    headers = [
        "Staff ID", "Staff Name", "Department Code", "Department Name",
        "Name of the Activity", "Number of Students Attended", 
        "Number of Students Placed", "Year of Activity",
        "Permission Letter", "Circular", "Brochure", "Resource person profile",
        "Detailed Report-Objective & Outcome", "Geo Tagged Photos", 
        "Students Attendance with Authorised signature"
    ]
    ws.append(headers)

    # Style headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Updated field mapping based on the actual categories
    field_mapping = {
        "PL": "Permission Letter",
        "Circular": "Circular",
        "Brochure": "Brochure",
        "RPP": "Resource person profile",
        "DR": "Detailed Report-Objective & Outcome",
        "GTP": "Geo Tagged Photos",
        "SA": "Students Attendance with Authorised signature"
    }

    # Map category to column
    media_columns = {
        "Permission Letter": 9,
        "Circular": 10,
        "Brochure": 11,
        "Resource person profile": 12,
        "Detailed Report-Objective & Outcome": 13,
        "Geo Tagged Photos": 14,
        "Students Attendance with Authorised signature": 15
    }

    # Populate rows
    for p in parents:
        # Initialize media categories - only use categories that exist in field_mapping
        media_mapping = {v: [] for v in field_mapping.values()}

        # Categorize each file based on upload code in filename
        for media in p.get("media_files", []):
            file_name = media.get("file_name", "Unknown")
            url = media.get("direct_api_url", "")
            if not url:
                continue

            # Use original filename for matching
            matched = False

            # Detect the code from filename - try multiple patterns
            for code, category in field_mapping.items():
                # Try different patterns to match the code in filename
                patterns = [
                    f"_{code}_",  # _PL_, _RPP_, _GTP_, etc.
                    f"_{code}.",  # _PL.pdf, _RPP.docx, etc.
                    f"-{code}-",  # -PL-, -RPP-, etc.
                    f"-{code}.",  # -PL.pdf, -RPP.docx, etc.
                    f"{code}_",   # PL_, RPP_, etc. (at start after prefix)
                    f"_{code}",   # _PL, _RPP, etc. (at end before extension)
                ]
                
                for pattern in patterns:
                    if pattern in file_name:
                        media_mapping[category].append((file_name, url))
                        matched = True
                        print(f"Matched '{file_name}' to '{category}' using pattern '{pattern}'")
                        break
                if matched:
                    break

            # If no match found with underscore patterns, try case-insensitive matching
            if not matched:
                file_name_upper = file_name.upper()
                for code, category in field_mapping.items():
                    code_upper = code.upper()
                    # Check if code appears anywhere in filename (case insensitive)
                    if code_upper in file_name_upper:
                        # Additional check to avoid partial matches
                        patterns_upper = [
                            f"_{code_upper}_",
                            f"_{code_upper}.",
                            f"-{code_upper}-",
                            f"-{code_upper}.",
                            f"{code_upper}_",
                            f"_{code_upper}"
                        ]
                        if any(pattern in file_name_upper for pattern in patterns_upper):
                            media_mapping[category].append((file_name, url))
                            matched = True
                            print(f"Matched '{file_name}' to '{category}' using case-insensitive matching")
                            break

        # Create the base info row
        base_row = [
            p.get("staff_id", ""),
            p.get("staff_name", ""),
            p.get("department_code", ""),
            p.get("department_name", ""),
            p.get("name_of_the_activity", ""),
            p.get("number_of_students_attended", ""),
            p.get("number_of_students_placed", ""),
            p.get("year_of_activity", ""),
        ]

        # Append the base row first
        ws.append(base_row)
        row_idx = ws.max_row  # current row number

        # Now add media files to their respective columns
        for category, files in media_mapping.items():
            if category in media_columns:  # Double check the category exists
                col_idx = media_columns[category]
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if files:
                    # Show all filenames joined by line breaks
                    filenames = [f for f, _ in files]
                    cell.value = "\n".join(filenames)
                    
                    # Add hyperlink to first file if URL exists
                    if files[0][1]:
                        cell.hyperlink = files[0][1]
                        cell.font = Font(color="0000EE", underline="single")
                    else:
                        cell.font = Font()  # Regular font if no URL
                else:
                    cell.value = "-"
                    cell.font = Font()  # Regular font
                
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-fit column widths
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            if cell.value:
                # Count the maximum line length for multiline cells
                lines = str(cell.value).split('\n')
                max_line_length = max(len(line) for line in lines)
                max_length = max(max_length, max_line_length)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

    # Adjust row heights for multiline cells
    for row_idx in range(1, ws.max_row + 1):
        max_lines = 1
        for col in range(1, len(headers) + 1):
            cell_value = str(ws.cell(row=row_idx, column=col).value or "")
            line_count = cell_value.count("\n") + 1
            max_lines = max(max_lines, line_count)
        ws.row_dimensions[row_idx].height = min(max_lines * 15, 120)

    # Freeze header and enable filters
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Return as Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="collaborative_students.xlsx"'

    with io.BytesIO() as buffer:
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())

    return response