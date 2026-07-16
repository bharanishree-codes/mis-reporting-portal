import json
import requests
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from MIS.functions import validate_file_format, validate_file_size
from datetime import datetime
from user_management.settings_views import *

API_STUDIO_URL = user_bundle_settings()


def extension_key():
    url = "https://api.hcaschennai.edu.in/auth/token"

    payload = json.dumps({"secret_key": "C4ZoXbsAnHLjk1Xyz4QPT2eoiNx6K6fo"})
    headers = {'Content-Type': 'application/json'}

    response = requests.post(url, headers=headers, data=payload)
    if response.status_code == 200:
        res_data = response.json()
        access_token = res_data.get('access_token')
        token_type = res_data.get('token_type')
        return access_token, token_type
    return None, None


# Function to fetch staff career data
def get_extension_data(access_token, token_type):
    url = "https://api.hcaschennai.edu.in/sqlviews/api/v1/auth/get_response_data"

    payload = json.dumps({"psk_uid": "51a531b4-bd55-491c-861d-a8d7227b325b","project": "hcas","data": {}})
    headers = {'Content-Type': 'application/json','Authorization': f'{token_type} {access_token}'}

    response = requests.post(url, headers=headers, data=payload)
    if response.status_code == 200:
        return response.json()
    return []


# Create extension
def extension_create(request):
    error_message = None

    # Step 1: Get API token
    access_token, token_type = extension_key()
    if not access_token or not token_type:
        error_message = 'Failed to get access token from API.'
        return render(request, 'Extentions_templates/extension_create.html', {'error': error_message})

    # Step 2: Fetch career data
    extension_data = get_extension_data(access_token, token_type)
    if not extension_data:
        error_message = 'Failed to fetch staff data.'
        return render(request, 'Extentions_templates/extension_create.html', {'error': error_message})
    
    
    user = get_settings(request)
    username = user['username']
    # username = 'AC-NT012'
    
    current_year = datetime.now().year    
    publication_year = [f"{y}-{y+1}" for y in range(current_year, 2010, -1)]
    
    selected_faculty = None
    for faculty in extension_data:
        if faculty['stf_id'] == username:
            selected_faculty = faculty
            break
        
    staff_name = selected_faculty.get('stf_name') if selected_faculty else ''
    department_name = selected_faculty.get('department') if selected_faculty else ''


    # Step 3: Handle POST form submission
    if request.method == 'POST':
        # Extract form data
        staff_name = request.POST.get('staff_name')
        staff_id = username
        department_name = request.POST.get('department_name')
        name_of_the_activity = request.POST.get('name_of_the_activity')
        organizing_agency_collaborating = request.POST.get('organizing_agency_collaborating')
        name_of_the_scheme = request.POST.get('name_of_the_scheme')
        year_of_activity = request.POST.get('year_of_activity')
        number_of_students = request.POST.get('number_of_students')
        number_of_beneficiaries = request.POST.get('number_of_beneficiaries')
        # URL to send the form data to
        
        
        selected_faculty = None
        for faculty in extension_data:
            if faculty['stf_id'] == username:
                selected_faculty = faculty
                break
            
        staff_name = selected_faculty.get('stf_name') if selected_faculty else ''
        department_name = selected_faculty.get('department') if selected_faculty else ''
        
        url = f"{API_STUDIO_URL}postapi/create/naac01_no_of_extension_and_outreach_programs_dc1"
        payload = json.dumps({"data": {
            "staff_name": staff_name,
            "staff_id": staff_id,
            "department_name": department_name,
            "name_of_the_activity": name_of_the_activity,
            "organizing_agency_collaborating": organizing_agency_collaborating,
            "name_of_the_scheme": name_of_the_scheme, "year_of_activity": year_of_activity,
            "number_of_students": number_of_students,
            "number_of_beneficiaries": number_of_beneficiaries}})
        headers = {'Content-Type': 'application/json'}

        # Make API call to create the extension/workshop
        response = requests.post(url, headers=headers, data=payload)

        if response.status_code == 200:
            # Successfully created the extension/workshop
            file_data = response.json()
            psk_id = file_data.get('psk_id')  # Get psk_id from the response

            # Uploading files (optional step)
        upload_url = f"{API_STUDIO_URL}crudapp/upload/media/naac01_no_of_extension_and_outreach_programs_dc1_media"
        uploaded_files = request.FILES.getlist('file')

        if not uploaded_files:
            messages.error(request, message="No files selected for upload.")
            return render(request, 'Extentions_templates/extension_create.html')
        fields = ['PL', 'Circular', 'Brochure', 'RPP', 'DR', 'GTP', 'SA']
        for field in fields:
            current_year = datetime.now().year
            staff_id = selected_faculty.get('stf_id')
            print("staff_id: ", staff_id)
            for field, uploaded_file in zip(fields, uploaded_files):
                validate_file_size(uploaded_file)
                validate_file_format(uploaded_file)
                file_type = uploaded_file.content_type
                custom_filename = f"{staff_id}_{field}_{current_year}_{uploaded_file.name}"
                print(f"Generated filename: {custom_filename}")  # Print the filename for each iteration
                payload = {'parent_psk_id': psk_id}
                files = {'media': (custom_filename, uploaded_file, file_type)}
                upload_headers = {'api_name': 'naac01_no_of_extension_and_outreach_programs_dc1_media'}

                # Make API call to upload the file
                upload_response = requests.post(upload_url, headers=upload_headers, data=payload, files=files)
                print("success")

                if upload_response.status_code != 200:
                    # File upload failed
                    messages.error(request,
                                   message=f"File upload failed for {uploaded_file.name}. Error: {upload_response.text}")
                    return redirect('extension_list')

                messages.success(request, message="Documents uploaded successfully.")

            # Redirect to extension view after successful creation
            return redirect('extension_list')

        else:
            # API call failed for creating extension
            messages.error(request, message="Failed to create extension. Please try again.")
            return render(request, 'Extentions_templates/extension_create.html')

    else:
        # If the request is GET, render the extension creation form
        return render(request, 'Extentions_templates/extension_create.html', {'extension_data': extension_data, "username": username, "staff_name": staff_name, "department": department_name, 'publication_year': publication_year})


# View extension Details
def extension_view(request, id):
    url = f"{API_STUDIO_URL}getapi/naac01_no_of_extension_and_outreach_programs_dc1/{id}"
    response = requests.get(url)

    if response.status_code == 200:
        extension_data = response.json()
        return render(request, "Extentions_templates/extension_view.html", {'extension': extension_data})

    return HttpResponse(f"Error fetching Course details: {response.text}", status=500)


# List All extensions
# def extension_list(request):
    # # URL to get extension and outreach programs data
    # url = f"{API_STUDIO_URL}getapi/all_fields/naac01_no_of_extension_and_outreach_programs_dc1/all"
    # response = requests.get(url)

    # # Get the current username (staff_id)
    # user = get_settings(request)  # Assuming this function retrieves user settings (including staff_id)
    # username = user.get('username', None)  # Adjust if necessary to fetch staff_id or username from session
    # # username = 'AC-NT012'

    # # If no username (staff_id) is found, return the same page with an empty list
    # if not username:
        # return render(request, 'Extentions_templates/extension_list.html', {'extensions': []})

    # # If the API call was successful, filter the extension data based on staff_id (username)
    # if response.status_code == 200:
        # extensions = response.json()
        
        # # Filter the extensions based on the staff_id (username)
        # filtered_extensions = [extension for extension in extensions if extension.get('staff_id') == username]
        
    # selected_staff_id = request.GET.get('staff_id')
    
    # if selected_staff_id:
        # filtered_extensions = [extension for extension in extensions if extension.get('staff_id') == selected_staff_id]

        
        # # If no data is found for the username, return the same page with an empty list
        # if not filtered_extensions:
            # return render(request, 'Extentions_templates/extension_list.html', {'extensions': extensions})
        
        # # Return the filtered data to the template
        # return render(request, 'Extentions_templates/extension_list.html', {'extensions': filtered_extensions})
    
    # # If the API call fails, return an empty list
    # return render(request, 'Extentions_templates/extension_list.html', {'extensions': []})
    
# def extension_list(request):

#     url = f"{API_STUDIO_URL}getapi/naac01_no_of_extension_and_outreach_programs_dc1/all"
#     response = requests.get(url)
    
#     if response.status_code == 200:
#         extensions = response.json()
#         print("extensions:", extensions)
#     else:
#         return HttpResponse("API Call Is Not Working")

#     user = get_settings(request)
#     username = user.get('username')
#     # username = 'AC-NT012'
#     filtered_extensions = [program for program in extensions if program.get('staff_id') == username]
    
#     print("filtered_extensions:", filtered_extensions)
#     #from_dashboard = request.GET.get('from') == 'dashboard' or 'admin_hod_dash' or 'admin_dash' or 'department_dashboard'
#     selected_staff_id = request.GET.get('staff_id')
#     if selected_staff_id:
#         filtered_extensions = [extension for extension in extensions if extension.get('staff_id') == selected_staff_id]
        
#     selected_department = request.GET.get('department')
#     if selected_department:
#         filtered_extensions = [extension for extension in extensions if extension.get('department_name') == selected_department]


#     if not filtered_extensions:
#         return render(request, 'Extentions_templates/extension_list.html', {"extensions": extensions})
    
#     return render(request, 'Extentions_templates/extension_list.html', {"extensions": filtered_extensions, 
#     #'from_dashboard': from_dashboard
#     })


def extension_list(request):
    # Step 1: Get all extensions from API
    url = f"{API_STUDIO_URL}getapi/naac01_no_of_extension_and_outreach_programs_dc1/all"
    response = requests.get(url)
    if response.status_code != 200:
        return HttpResponse("API Call Is Not Working")

    extensions = response.json()

    # Step 2: Get logged-in user
    user = get_settings(request)
    username = user.get('username')

    # Step 3: Start filtering with logged-in staff
    filtered_extensions = [program for program in extensions if program.get('staff_id') == username]

    # Step 4: Further filter by department (based on user’s department)
    user_department = user.get('department_name')  # assuming your `get_settings` returns department_name
    if user_department:
        filtered_extensions = [ext for ext in filtered_extensions if ext.get('department_name') == user_department]

    # Step 5: Optionally allow query params (if admin/HOD needs to check)
    selected_staff_id = request.GET.get('staff_id')
    if selected_staff_id:
        filtered_extensions = [ext for ext in extensions if ext.get('staff_id') == selected_staff_id]

    selected_department = request.GET.get('department')
    if selected_department:
        filtered_extensions = [ext for ext in filtered_extensions if ext.get('department_name') == selected_department]

    # Step 6: Always show only filtered, no fallback to all
    return render(
        request,
        'Extentions_templates/extension_list.html',
        {"extensions": filtered_extensions}
    )



# Update extension
def extension_update(request, id):
    current_year = datetime.now().year
    publication_year = [f"{y}-{y+1}" for y in range(current_year, 2010, -1)]

    
    # Fetch extension details
    url = f"{API_STUDIO_URL}getapi/naac01_no_of_extension_and_outreach_programs_dc1/{id}"
    response = requests.get(url)

    if response.status_code == 200:
        extension = response.json()
    else:
        return HttpResponse(f"Error fetching extension details: {response.text}", status=500)

    # Fetch media (child files) associated with this extension
    media_url = f"{API_STUDIO_URL}crudapp/get/media/naac01_no_of_extension_and_outreach_programs_dc1_media/parent/{id}"
    media_response = requests.get(media_url)

    if media_response.status_code == 200:
        child_files = media_response.json()
    else:
        return HttpResponse(f"Failed to fetch media files: {media_response.text}", status=500)

    # Handle form submission (POST request)
    if request.method == "POST":
        # Prepare data for updating the extension/workshop
        update_url = f"{API_STUDIO_URL}updateapi/update/naac01_no_of_extension_and_outreach_programs_dc1/{id}"

        payload = json.dumps({"data": {
            "staff_id": request.POST.get('staff_id', extension.get('staff_id')),
            "staff_name": request.POST.get('staff_name', extension.get('staff_name')),
            "department_name": request.POST.get('department_name', extension.get('department_name')),
            "department_code": request.POST.get('department_code', extension.get('department_code')),
            "name_of_the_activity": request.POST.get('name_of_the_activity', extension.get('name_of_the_activity')),
            "organizing_agency_collaborating": request.POST.get('organizing_agency_collaborating',
                                                                extension.get('organizing_agency_collaborating')),
            "name_of_the_scheme": request.POST.get('name_of_the_scheme', extension.get('name_of_the_scheme')),
            "year_of_activity": request.POST.get('year_of_activity', extension.get('year_of_activity')),
            "number_of_students": request.POST.get('number_of_students', extension.get('number_of_students')),
            "number_of_beneficiaries": request.POST.get('number_of_beneficiaries',
                                                        extension.get('number_of_beneficiaries'))}})

        headers = {'Content-Type': 'application/json'}
        update_response = requests.put(update_url, headers=headers, data=payload)

        if update_response.status_code != 200:
            return HttpResponse(f"Failed to update extension details: {update_response.text}", status=500)

        # Handle media (file) uploads
        upload_errors = []

        for child in child_files:
            upload_id = child['psk_id']
            fields = ['PL', 'Circular', 'Brochure', 'RPP', 'DR', 'GTP', 'SA']

            for field in fields:
                uploaded_files = request.FILES.getlist(f'file_{upload_id}_{field}')

                if not uploaded_files:
                    continue  # Skip if no files are uploaded for this field

                for uploaded_file in uploaded_files:
                    try:
                        # Validate file size and format
                        validate_file_size(uploaded_file)
                        validate_file_format(uploaded_file)

                        # Generate custom filename
                        current_year = datetime.now().year
                        staff_id = request.POST.get('staff_id', 'unknown')
                        custom_filename = f"{staff_id}_{field}_{current_year}_{uploaded_file.name}"

                        print(f"Generated filename for field {field}: {custom_filename}")  # Debug log

                        # Construct the upload URL and payload
                        upload_url = f"{API_STUDIO_URL}crudapp/upload/media/naac01_no_of_extension_and_outreach_programs_dc1_media/{upload_id}"
                        files = {'media': (custom_filename, uploaded_file, uploaded_file.content_type)}
                        headers = {
                            'api_name': 'naac01_no_of_extension_and_outreach_programs_dc1_media',
                            'psk_id': str(upload_id)
                        }
                        payload = {'parent_psk_id': id}

                        # Upload the file
                        upload_response = requests.put(upload_url, headers=headers, data=payload, files=files)

                        # Check the response
                        if upload_response.status_code != 200:
                            upload_errors.append(
                                f"Error uploading file '{custom_filename}' for {upload_id}: {upload_response.text}"
                            )
                    except Exception as e:
                        upload_errors.append(
                            f"Error processing file '{uploaded_file.name}' for {upload_id}: {str(e)}"
                        )

        # Provide feedback to the user
        if upload_errors:
            for error in upload_errors:
                messages.error(request, error)
        else:
            messages.success(request, "Files uploaded successfully.")

        # Redirect to the extension view page after a successful update
        return redirect('extension_list')

    # If not a POST request, render the update form
    return render(request, 'Extentions_templates/extension_update.html', {'extension': extension, 'child_files': child_files, 'publication_year': publication_year})



# Delete extension
def extension_delete(request, id):
    # url = f"{API_STUDIO_URL}getapi/naac01_no_of_extension_and_outreach_programs_dc1/{id}"
    # response = requests.get(url)
    # extension = response.json()

    # if request.method == 'POST':
    delete_url = f"{API_STUDIO_URL}deleteapi/delete/naac01_no_of_extension_and_outreach_programs_dc1/{id}"
    delete_response = requests.delete(delete_url)

    if delete_response.status_code == 200:
        return redirect('extension_list')
    else:
        return HttpResponse("Failed to delete participation: " + delete_response.text)

# return render(request, 'Extentions_templates/extension_delete.html', {'extension': extension})

def export_Extentions_to_excel(parents, children=None):
    """
    Export Extentions data to Excel from filtered parents,
    including clickable media file links based on upload field codes.
    """
    import io
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Extentions"

    # Headers including media attachments
    headers = [
        "Staff ID", "Staff Name", "Department Name",
        "Name of the Activity", "Organizing Agency/Collaborating", 
        "Name of the Scheme", "Year of Activity", "Number of Students",
        "Number of Beneficiaries",
        "Permission Letter", "Circular", "Brochure", "Resource Person Profile",
        "Detailed Report", "Geo Tagged Photos", "Student Attendance"
    ]
    ws.append(headers)

    # Style headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Field mapping for file categorization
    field_mapping = {
        "PL": "Permission Letter",
        "Circular": "Circular",
        "Brochure": "Brochure",
        "RPP": "Resource Person Profile",
        "DR": "Detailed Report",
        "GTP": "Geo Tagged Photos",
        "SA": "Student Attendance"
    }

    # Map category to column
    media_columns = {
        "Permission Letter": 10,
        "Circular": 11,
        "Brochure": 12,
        "Resource Person Profile": 13,
        "Detailed Report": 14,
        "Geo Tagged Photos": 15,
        "Student Attendance": 16
    }

    # Populate rows
    for p in parents:
        # Initialize media categories
        media_mapping = {v: [] for v in field_mapping.values()}

        # Categorize each file based on upload code in filename
        for media in p.get("media_files", []):
            file_name = media.get("file_name", "Unknown")
            url = media.get("direct_api_url", "")
            if not url:
                continue

            # Use original filename for matching
            matched = False

            # Detect the code from filename - try multiple patterns
            for code, category in field_mapping.items():
                # Try different patterns to match the code in filename
                patterns = [
                    f"_{code}_",  # _PL_, _RPP_, _GTP_, etc.
                    f"_{code}.",  # _PL.pdf, _RPP.docx, etc.
                    f"-{code}-",  # -PL-, -RPP-, etc.
                    f"-{code}.",  # -PL.pdf, -RPP.docx, etc.
                    f"{code}_",   # PL_, RPP_, etc. (at start after prefix)
                    f"_{code}",   # _PL, _RPP, etc. (at end before extension)
                ]
                
                for pattern in patterns:
                    if pattern in file_name:
                        media_mapping[category].append((file_name, url))
                        matched = True
                        print(f"Matched '{file_name}' to '{category}' using pattern '{pattern}'")
                        break
                if matched:
                    break

            # If no match found with underscore patterns, try case-insensitive matching
            if not matched:
                file_name_upper = file_name.upper()
                for code, category in field_mapping.items():
                    code_upper = code.upper()
                    # Check if code appears anywhere in filename (case insensitive)
                    if code_upper in file_name_upper:
                        # Additional check to avoid partial matches
                        patterns_upper = [
                            f"_{code_upper}_",
                            f"_{code_upper}.",
                            f"-{code_upper}-",
                            f"-{code_upper}.",
                            f"{code_upper}_",
                            f"_{code_upper}"
                        ]
                        if any(pattern in file_name_upper for pattern in patterns_upper):
                            media_mapping[category].append((file_name, url))
                            matched = True
                            print(f"Matched '{file_name}' to '{category}' using case-insensitive matching")
                            break

        # Create the base info row using the exact field names from your form
        base_row = [
            p.get("staff_id", ""),
            p.get("staff_name", ""),
            p.get("department_name", ""),
            p.get("name_of_the_activity", ""),
            p.get("organizing_agency_collaborating", ""),
            p.get("name_of_the_scheme", ""),
            p.get("year_of_activity", ""),
            p.get("number_of_students", ""),
            p.get("number_of_beneficiaries", ""),
        ]

        # Append the base row first
        ws.append(base_row)
        row_idx = ws.max_row  # current row number

        # Now add media files to their respective columns
        for category, files in media_mapping.items():
            if category in media_columns:
                col_idx = media_columns[category]
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if files:
                    # Show all filenames joined by line breaks
                    filenames = [f for f, _ in files]
                    cell.value = "\n".join(filenames)
                    
                    # Add hyperlink to first file if URL exists
                    if files[0][1]:
                        cell.hyperlink = files[0][1]
                        cell.font = Font(color="0000EE", underline="single")
                    else:
                        cell.font = Font()  # Regular font if no URL
                else:
                    cell.value = "-"
                    cell.font = Font()  # Regular font
                
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-fit column widths
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            if cell.value:
                # Count the maximum line length for multiline cells
                lines = str(cell.value).split('\n')
                max_line_length = max(len(line) for line in lines)
                max_length = max(max_length, max_line_length)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

    # Adjust row heights for multiline cells
    for row_idx in range(1, ws.max_row + 1):
        max_lines = 1
        for col in range(1, len(headers) + 1):
            cell_value = str(ws.cell(row=row_idx, column=col).value or "")
            line_count = cell_value.count("\n") + 1
            max_lines = max(max_lines, line_count)
        ws.row_dimensions[row_idx].height = min(max_lines * 15, 120)

    # Freeze header and enable filters
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Return as Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Extensions_outreach.xlsx"'

    with io.BytesIO() as buffer:
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())

    return response

def export_Extentions_to_pdf(parents, children=None, selected_options=None):
    """
    Export Extentions data to PDF (portrait), grouped by staff.
    Each staff has a mini-title "Extentions", then Staff info in bold labels, then table.
    Attachments are clickable links.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, PageBreak
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.lib.units import inch
    from django.http import HttpResponse

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="collaborative_students.pdf"'

    # Increased margins for better spacing
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,  # Changed to portrait
        topMargin=1.0*inch,
        bottomMargin=1.0*inch,
        leftMargin=0.75*inch,
        rightMargin=0.75*inch
    )

    elements = []
    styles = getSampleStyleSheet()

    # Styles
    title_style = ParagraphStyle('Title', parent=styles['Heading2'], alignment=TA_CENTER, spaceAfter=12)
    left_style = ParagraphStyle('LeftStyle', parent=styles['Normal'], fontSize=9,
                                alignment=TA_LEFT, textColor=colors.HexColor('#2c3e50'))
    right_style = ParagraphStyle('RightStyle', parent=styles['Normal'], fontSize=9,
                                 alignment=TA_RIGHT, textColor=colors.HexColor('#2c3e50'))
    table_header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontSize=8,
                                        alignment=TA_CENTER, textColor=colors.white, fontName='Helvetica-Bold')
    table_cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=7,
                                      alignment=TA_LEFT, textColor=colors.HexColor('#2c3e50'), leading=9)
    table_cell_center_style = ParagraphStyle('TableCellCenter', parent=styles['Normal'], fontSize=7,
                                            alignment=TA_CENTER, textColor=colors.HexColor('#2c3e50'), leading=9)
    attachment_link_style = ParagraphStyle('AttachmentLink', parent=styles['Normal'], fontSize=6,
                                           alignment=TA_LEFT, textColor=colors.HexColor('#1a5276'), leading=8)
    no_data_style = ParagraphStyle('NoData', parent=styles['Normal'], fontSize=10,
                                   alignment=TA_CENTER, textColor=colors.HexColor('#7f8c8d'),
                                   fontStyle='italic', spaceBefore=12, spaceAfter=12)

    # Check if there are any parents
    if not parents:
        elements.append(Paragraph("Extentions", title_style))
        elements.append(Spacer(1, 0.5*inch))
        elements.append(Paragraph("No data available", no_data_style))
        doc.build(elements)
        return response

    # Sort parents by staff
    parents_sorted = sorted(parents, key=lambda x: (x.get('staff_id',''), x.get('activity_title','')))
    current_staff = None
    table_data = []
    
    # Adjusted column widths for portrait orientation
    col_widths = [
        2.0*inch,  # Activity Title
        1.3*inch,  # Agency Name
        1.0*inch,  # Participant Name
        0.7*inch,  # Year
        2.0*inch   # Attachments
    ]

    for parent in parents_sorted:
        staff_id = parent.get('staff_id', 'N/A')
        staff_name = parent.get('staff_name', 'N/A')
        department_name = parent.get('department_name', 'N/A')

        # New staff: flush previous table + page break
        if current_staff and current_staff != staff_id:
            # Only add table if there's data (more than just headers)
            if len(table_data) > 1:
                table = Table(table_data, colWidths=col_widths)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                    ('ALIGN', (0,0), (-1,0), 'CENTER'),  # Headers centered
                    ('ALIGN', (0,1), (0,-1), 'LEFT'),    # Activity Title left aligned
                    ('ALIGN', (1,1), (1,-1), 'LEFT'),    # Agency Name left aligned
                    ('ALIGN', (2,1), (2,-1), 'LEFT'),    # Participant Name left aligned
                    ('ALIGN', (3,1), (3,-1), 'CENTER'),  # Year centered
                    ('ALIGN', (4,1), (4,-1), 'LEFT'),    # Activity Name left aligned
                    ('ALIGN', (5,1), (5,-1), 'LEFT'),    # Attachments left aligned
                    ('VALIGN', (0,0), (-1,-1), 'TOP'),   # Top alignment for all cells
                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
                    ('FONTSIZE', (0,0), (-1,-1), 7),
                    ('LEFTPADDING', (0,0), (-1,-1), 4),
                    ('RIGHTPADDING', (0,0), (-1,-1), 4),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 3),
                    ('TOPPADDING', (0,0), (-1,-1), 3),
                ]))
                elements.append(table)
            else:
                # Only headers exist, meaning no data for this staff
                elements.append(Spacer(1, 0.3*inch))
                elements.append(Paragraph("No data available for this staff", no_data_style))
            
            elements.append(PageBreak())
            table_data = []

        if current_staff != staff_id:
            # Add mini-title for this staff
            elements.append(Paragraph("Extentions", title_style))
            elements.append(Spacer(1, 0.3*inch))

            # Staff info with bold labels
            left_paragraph = Paragraph(
                f"<b>Staff ID:</b> {staff_id}<br/><b>Staff Name:</b> {staff_name}", left_style)
            right_paragraph = Paragraph(f"<b>Department:</b> {department_name}", right_style)

            total_width = sum(col_widths)
            info_table = Table([[left_paragraph, right_paragraph]], colWidths=[total_width*0.5]*2)
            info_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0,0), (0,0), 'LEFT'),
                ('ALIGN', (1,0), (1,0), 'RIGHT'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('TOPPADDING', (0,0), (-1,-1), 2),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 0.3*inch))

            # Table headers
            headers = [
                'Name of the activity', 'Number of students Attended', 'Number of students placed', 'Year', 'Attachments'
            ]
            table_data.append([Paragraph(h, table_header_style) for h in headers])
            current_staff = staff_id

        # Media attachments
        media_files = parent.get('media_files', [])
        if media_files:
            media_text = []
            for m in media_files:
                filename = m.get('file_name', 'Unknown')
                url = m.get('direct_api_url', '#')
                # Truncate long filenames for better display in portrait mode
                display_name = filename if len(filename) <= 25 else filename[:22] + "..."
                media_text.append(f'<link href="{url}" color="blue">{display_name}</link>')
            media_paragraph = Paragraph('<br/>'.join(media_text), attachment_link_style)
        else:
            media_paragraph = Paragraph("No attachments", table_cell_style)

        # Create rows with appropriate alignment styles
        row = [
            Paragraph(str(parent.get('name_of_the_activity', '')), table_cell_style),
            Paragraph(str(parent.get('number_of_students_attended', '')), table_cell_style),
            Paragraph(str(parent.get('number_of_students_placed', '')), table_cell_style),
            Paragraph(str(parent.get('year_of_activity', '')), table_cell_center_style),
            media_paragraph
        ]
        table_data.append(row)

    # Add last staff table
    if table_data:
        # Check if there's actual data (more than just headers)
        if len(table_data) > 1:
            table = Table(table_data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('ALIGN', (0,0), (-1,0), 'CENTER'),  # Headers centered
                ('ALIGN', (0,1), (0,-1), 'LEFT'),    # Activity Title left aligned
                ('ALIGN', (1,1), (1,-1), 'LEFT'),    # Agency Name left aligned
                ('ALIGN', (2,1), (2,-1), 'LEFT'),    # Participant Name left aligned
                ('ALIGN', (3,1), (3,-1), 'CENTER'),  # Year centered
                ('ALIGN', (4,1), (4,-1), 'LEFT'),    # Activity Name left aligned
                ('ALIGN', (5,1), (5,-1), 'LEFT'),    # Attachments left aligned
                ('VALIGN', (0,0), (-1,-1), 'TOP'),   # Top alignment for all cells
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
                ('FONTSIZE', (0,0), (-1,-1), 7),
                ('LEFTPADDING', (0,0), (-1,-1), 4),
                ('RIGHTPADDING', (0,0), (-1,-1), 4),
                ('BOTTOMPADDING', (0,0), (-1,-1), 3),
                ('TOPPADDING', (0,0), (-1,-1), 3),
            ]))
            elements.append(table)
        else:
            # Only headers exist, meaning no data for this staff
            elements.append(Spacer(1, 0.3*inch))
            elements.append(Paragraph("No data available for this staff", no_data_style))

    doc.build(elements)
    return response
