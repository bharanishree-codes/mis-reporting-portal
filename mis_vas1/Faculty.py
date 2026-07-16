from pandas.api.types import is_scalar
from openpyxl.styles import Font
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
import requests
import json
from datetime import datetime
from collections import Counter
from user_management.settings_views import get_settings, roles_tbl
from MIS.functions import validate_file_format_faculty, validate_file_size

API_STUDIO_URL = "https://api.hcaschennai.edu.in/"

PARTICIPATION_OPTIONS = [
    'Board of Studies', 'Question Paper Setting', 'Evaluation', 
    'Add-On (Other University & College)', 'Certificate Courses',
    'External Examiner', 'Conference', 'Seminar', 'Workshop', 'Faculty Development Program'
]

def faculty_auth():
    url = "https://api.hcaschennai.edu.in/auth/token"
    payload = json.dumps({"secret_key": "C4ZoXbsAnHLjk1Xyz4QPT2eoiNx6K6fo"})
    headers = {'Content-Type': 'application/json'}
    response = requests.post(url, headers=headers, data=payload)
    if response.status_code == 200:
        response_data = response.json()
        return response_data.get('access_token'), response_data.get('token_type')
    return None, None

def faculty_token(access_token, token_type):
    url = "https://api.hcaschennai.edu.in/sqlviews/api/v1/auth/get_response_data"
    payload = json.dumps({
        "psk_uid": "51a531b4-bd55-491c-861d-a8d7227b325b",
        "project": "hcas",
        "data": {}
    })
    headers = {
        'Content-Type': 'application/json', 
        'Authorization': f'{token_type} {access_token}'
    }
    response = requests.post(url, headers=headers, data=payload)
    return response.json() if response.status_code == 200 else None

def create_participation(request):
    access_token, token_type = faculty_auth()
    if not access_token or not token_type:
        return render(request, 'ParticipationFaculty_templates/create_participation.html', 
                     {'error': 'Failed to get access token from API.'})

    faculty_data = faculty_token(access_token, token_type)
    if not faculty_data:
        return render(request, 'ParticipationFaculty_templates/create_participation.html', 
                     {'error': 'Failed to fetch staff data.'})

    current_year = datetime.now().year
    years = [f"{y}-{y+1}" for y in range(current_year, 2010, -1)]
    
    user = get_settings(request)
    username = user.get('username', '')
    # username = 'CS-T151'
    
    selected_faculty = None
    for faculty in faculty_data:
        if faculty.get('stf_id') == username:
            selected_faculty = faculty
            break
    
    stf_name = selected_faculty.get('stf_name', '') if selected_faculty else ''
    department = selected_faculty.get('department', '') if selected_faculty else ''

    if request.method == 'POST':
        selected_options = request.POST.getlist('items')
        year = request.POST.get('year')
        name = request.POST.get('name')
        stf_id = username

        selected_faculty = None
        for faculty in faculty_data:
            if faculty.get('stf_id') == stf_id:
                selected_faculty = faculty
                break

        if selected_faculty:
            depcode = selected_faculty.get('depcode', '')
            department = selected_faculty.get('department', '')
            stf_name = selected_faculty.get('stf_name', '')
            
            url = f"{API_STUDIO_URL}postapi/create/naac01_faculty_participation_dc1"
            create_data = {
                "data": {
                    "name": name, 
                    "year": year, 
                    "participation": ', '.join(selected_options), 
                    "stf_id": stf_id, 
                    "depcode": depcode, 
                    "department": department, 
                    "stf_name": stf_name
                }
            }
            
            headers = {'Content-Type': 'application/json', 'Authorization': f'{token_type} {access_token}'}
            response = requests.post(url, headers=headers, data=json.dumps(create_data))
            
            if response.status_code == 200:
                psk_id = response.json().get('psk_id')
                return redirect('detail_view', id=psk_id)
            else:
                return HttpResponse("Failure: " + response.text)
        else:
            return render(request, 'ParticipationFaculty_templates/create_participation.html', 
                         {'error': "Selected faculty not found.", 'options': PARTICIPATION_OPTIONS, 
                          'faculty_data': faculty_data, 'years': years, 'staff_id': username, 
                          'stf_name': stf_name, 'department': department})
    
    return render(request, 'ParticipationFaculty_templates/create_participation.html', 
                 {'options': PARTICIPATION_OPTIONS, 'faculty_data': faculty_data, 'years': years, 
                  'staff_id': username, 'stf_name': stf_name, 'department': department})

def detail_view(request, id):
    current_year = datetime.now().year
    years = [f"{y}-{y+1}" for y in range(current_year, 2010, -1)]
    
    parent_url = f"{API_STUDIO_URL}getapi/naac01_faculty_participation_dc1/{id}"
    parent_res = requests.get(parent_url)

    if parent_res.status_code != 200:
        return HttpResponse(f"Error fetching participation details: {parent_res.text}")

    participation = parent_res.json()
    parent_id = participation.get('psk_id')
    selected_options = participation.get('participation', '').split(', ')
    year = request.POST.get('year', participation.get('year', ''))

    payload = json.dumps({
        "queries": [{"field": "transaction_id", "value": parent_id, "operation": "equal"}],
        "search_type": "all"
    })
    
    child_res = requests.get(
        url=f"{API_STUDIO_URL}getapi/naac01_faculty_participation_dc2", 
        headers={'Content-Type': 'application/json'}, 
        data=payload
    )

    children = child_res.json() if child_res.status_code == 200 else []

    for child in children:
        child_id = child.get('psk_id')
        media_url = f"{API_STUDIO_URL}crudapp/get/media/naac01_faculty_participation_dc2_media/parent/{child_id}"
        media_res = requests.get(media_url)
        child['media'] = [item.get('file_name') for item in media_res.json()] if media_res.status_code == 200 else []

        # Convert date_from and date_to to DD-MM-YYYY format
        for date_field in ['date_from', 'date_to']:
            date_val = child.get(date_field)
            if date_val:
                parts = date_val.split('-')
                # handle YYYY-MM-DD → DD-MM-YYYY
                if len(parts[0]) == 4:
                    child[date_field] = f"{parts[2]}-{parts[1]}-{parts[0]}"
                # already DD-MM-YYYY
                else:
                    child[date_field] = date_val
            else:
                child[date_field] = ""


    field_map = {
        'board of studies': ['board_university_college_name', 'board_designation'],
        'question paper setting': ['qp_university_college_name', 'qp_subject'],
        'evaluation': ['eval_university_college_name', 'eval_subject'],
        'Add-On (Other University & College)': ['design_development_university_college_name', 'design_development_addon'],
        'certificate courses': ['certificate_university_college_name', 'certificate_course'],
        'external examiner': ['external_examiner_university_college_name'],
        'conference': ['conference_university_college_name', 'conference_name'],
        'seminar': ['seminar_university_college_name', 'seminar_name'],
        'workshop': ['workshop_university_college_name', 'workshop_name'],
        'faculty development program': ['name_of_the_resource_person', 'name_of_the_program', 'date_from', 'date_to']
    }

    missing_children = []
    if request.method == 'POST':
        for option in selected_options:
            key = option.lower()
            fields = field_map.get(key, [])
            has_valid_child = any(all(child.get(f) for f in fields) for child in children)
            if not has_valid_child:
                missing_children.append(option)

        if missing_children:
            messages.error(request, f"The following selections are missing required data: {', '.join(missing_children)}")
            return render(request, "ParticipationFaculty_templates/detail_view.html", 
                         {'participation': participation, 'selected_options': selected_options, 
                          'children': children, 'year': year, 'missing_children': missing_children})

        messages.success(request, "Participation data processed successfully!")
        return redirect('list_participations')

    return render(request, "ParticipationFaculty_templates/detail_view.html", 
                 {'participation': participation, 'selected_options': selected_options, 
                  'children': children, 'year': year, 'missing_children': [], 'years': years})

def list_participations(request):
    cleanup_orphan_participations()
    
    url = f"{API_STUDIO_URL}getapi/all_fields/naac01_faculty_participation_dc1/all"
    response = requests.get(url)
    
    if response.status_code != 200:
        return HttpResponse("API Call Is Not Working")

    participations = response.json()
    user = get_settings(request)
    username = user.get('username', '')
    # username = 'CS-T151'
    
    filtered_participations = [
        participation for participation in participations 
        if participation.get('stf_id') == username
    ]
    
    selected_staff_id = request.GET.get('staff_id')
    if selected_staff_id:
        filtered_participations = [
            participation for participation in participations 
            if participation.get('stf_id') == selected_staff_id
        ]
    
    if not filtered_participations:
        return render(request, 'ParticipationFaculty_templates/list_participations.html', 
                     {"participations": []})
    
    return render(request, 'ParticipationFaculty_templates/list_participations.html', 
                 {"participations": filtered_participations})

def cleanup_orphan_participations():
    parent_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_faculty_participation_dc1/all"
    parent_res = requests.get(parent_url)
    
    if parent_res.status_code != 200:
        return
    
    all_parents = parent_res.json()
    
    for parent in all_parents:
        parent_id = parent.get("psk_id")
        if not parent_id:
            continue
        
        child_url = f"{API_STUDIO_URL}getapi/naac01_faculty_participation_dc2"
        payload = json.dumps({
            "queries": [{"field": "transaction_id", "value": parent_id, "operation": "equal"}],
            "search_type": "all"
        })
        headers = {'Content-Type': 'application/json'}
        children_response = requests.get(child_url, headers=headers, data=payload)
        
        children = children_response.json() if children_response.status_code == 200 else []
        
        if not children:
            delete_url = f"{API_STUDIO_URL}deleteapi/delete/naac01_faculty_participation_dc1/{parent_id}"
            requests.delete(delete_url)

def update_participation(request, id):
    try:
        url = f"{API_STUDIO_URL}getapi/naac01_faculty_participation_dc1/{id}"
        response = requests.get(url)
        
        if response.status_code != 200:
            return HttpResponse(f"Error fetching participation data: {response.text}", status=500)
            
        participation = response.json()
        
        current_year = datetime.now().year
        years = [f"{y}-{y+1}" for y in range(current_year, 2010, -1)]
        selected_options = participation.get('participation', '').split(', ') if 'participation' in participation else []
        
        my_child = []
        try:
            children_url = f"{API_STUDIO_URL}getapi/naac01_faculty_participation_dc2"
            payload = json.dumps({
                "queries": [{"field": "transaction_id", "value": id, "operation": "equal"}], 
                "search_type": "all"
            })
            headers = {'Content-Type': 'application/json'}
            
            children_response = requests.post(children_url, headers=headers, data=payload)
            
            if children_response.status_code == 200:
                my_child = children_response.json()
        except Exception:
            pass
        
        if request.method == "POST":
            selected_options_from_form = request.POST.getlist('items')
            
            options_with_children = []
            
            if my_child:
                field_map = {
                    'board of studies': ['board_university_college_name', 'board_designation'],
                    'question paper setting': ['qp_university_college_name', 'qp_subject'],
                    'evaluation': ['eval_university_college_name', 'eval_subject'],
                    'Add-On (Other University & College)': ['design_development_university_college_name', 'design_development_addon'],
                    'certificate courses': ['certificate_university_college_name', 'certificate_course'],
                    'external examiner': ['external_examiner_university_college_name'],
                    'conference': ['conference_university_college_name', 'conference_name'],
                    'seminar': ['seminar_university_college_name', 'seminar_name'],
                    'workshop': ['workshop_university_college_name', 'workshop_name'],
                    'faculty development program': ['name_of_the_resource_person', 'name_of_the_program', 'date_from', 'date_to']
                }
                
                for option in selected_options:
                    option_key = option.lower()
                    has_children = False
                    
                    required_fields = field_map.get(option_key, [])
                    if required_fields:
                        for child in my_child:
                            if all(child.get(field) for field in required_fields):
                                has_children = True
                                break
                                
                        if has_children and option not in selected_options_from_form:
                            options_with_children.append(option)
            
            if options_with_children:
                error_message = f"Cannot deselect the following options because they have associated children: {', '.join(options_with_children)}"
                messages.error(request, error_message)
                
                return render(request, 'ParticipationFaculty_templates/update_participation.html', {
                    'participation': participation,
                    'selected_options': selected_options,
                    'years': years,
                    'options': PARTICIPATION_OPTIONS
                })
            
            name = request.POST.get('name', participation.get('name', ''))
            year = request.POST.get('year', participation.get('year', ''))
            participation_type = ', '.join(request.POST.getlist('items'))
            
            update_url = f"{API_STUDIO_URL}updateapi/update/naac01_faculty_participation_dc1/{id}"
            payload = json.dumps({
                "data": {
                    "name": name, 
                    "year": year, 
                    "participation": participation_type
                }
            })
            headers = {'Content-Type': 'application/json'}
            update_response = requests.put(update_url, headers=headers, data=payload)
            
            if update_response.status_code == 200:
                return redirect('detail_view', id=participation.get('psk_id'))
            else:
                return HttpResponse("Failed to update participation: " + update_response.text)
        
        return render(request, 'ParticipationFaculty_templates/update_participation.html', {
            'participation': participation,
            'selected_options': selected_options,
            'years': years,
            'options': PARTICIPATION_OPTIONS
        })
        
    except Exception as e:
        return HttpResponse(f"Unexpected error: {str(e)}", status=500)

# def delete_participation(request, id):
#     delete_url = f"{API_STUDIO_URL}deleteapi/delete/naac01_faculty_participation_dc1/{id}"
    
#     delete_response = requests.delete(delete_url)

#     if delete_response.status_code == 200:
#         messages.success(request, "The Article was deleted successfully.")
#         return redirect('list_participations')
#     else:
#         error_msg = delete_response.json()
#         messages.error(request, f"{error_msg.get('detail', 'Failed to delete participation')}")


def delete_participation(request, id):
    # id = parent psk_id (example: 48)

    child_get_url = (
        "https://api.hcaschennai.edu.in/"
        "getapi/all_fields/naac01_faculty_participation_dc2/all"
    )

    child_response = requests.get(child_get_url)

    if child_response.status_code == 200:
        for child in child_response.json():
            # Match parent-child relation
            if child.get("transaction_id") == id:
                child_psk_id = child.get("psk_id")

                if child_psk_id:
                    child_delete_url = (
                        f"{API_STUDIO_URL}deleteapi/delete/"
                        f"naac01_faculty_participation_dc2/{child_psk_id}"
                    )
                    requests.delete(child_delete_url)

    # Delete parent after children
    parent_delete_url = (
        f"{API_STUDIO_URL}deleteapi/delete/"
        f"naac01_faculty_participation_dc1/{id}"
    )

    parent_response = requests.delete(parent_delete_url)

    if parent_response.status_code == 200:
        messages.success(
            request,
            "Parent and all related child records deleted successfully."
        )
    else:
        messages.error(request, "Failed to delete parent record")

    return redirect("list_participations")


def list_all_participation_children(request):
    participation_api = f"{API_STUDIO_URL}getapi/naac01_faculty_participation_dc1/all"
    response = requests.get(participation_api)
    if response.status_code != 200:
        return HttpResponse("Could not load participation data")
    participation_data = response.json()

    user = get_settings(request)
    username = user.get('username', '')
    # username = 'CS-T151'
    final_staff_id = request.GET.get("staff_id") or username

    selected_category = request.GET.get("category")
    categories = [
        "Board of Studies", "Question Paper Setting", "Evaluation", "Design and Development", 
        "Certificate Courses", "External Examiner", "Conference", "Seminar", "Workshop", 
        'Faculty Development Program'
    ]

    filtered_data = [
        participation for participation in participation_data 
        if participation.get("stf_id") == final_staff_id and 
        (not selected_category or selected_category in [
            category.strip() for category in participation.get("participation", "").split(",")
        ])
    ]

    selected_types = {selected_category} if selected_category in categories else {
        c.strip() for p in filtered_data for c in p.get("participation", "").split(",") if c.strip() in categories
    }

    counts = Counter(
        t.strip().replace(" ", "_") for p in filtered_data 
        for t in p.get("participation", "").split(",") if t.strip()
    )

    child_api = f"{API_STUDIO_URL}getapi/all_fields/naac01_faculty_participation_dc2/all"
    child_response = requests.get(child_api)
    if child_response.status_code != 200:
        return HttpResponse("Could not load child records")
    child_data = child_response.json()

    staff_psk_ids = {p.get("psk_id") for p in filtered_data}
    final_children = []

    for cat in selected_types:
        key_map = {
            "Board of Studies": "board_university_college_name",
            "Question Paper Setting": "qp_university_college_name",
            "Evaluation": "eval_university_college_name",
            "Design and Development": "design_development_university_college_name",
            "Certificate Courses": "certificate_university_college_name",
            "External Examiner": "external_examiner_university_college_name",
            "Conference": "conference_university_college_name",
            "Seminar": "seminar_university_college_name",
            "Workshop": "workshop_university_college_name",
            "Faculty Development Program": 'name_of_the_resource_person',
        }
        field = key_map.get(cat)
        matching_children = [
            child for child in child_data 
            if child.get("transaction_id") in staff_psk_ids and child.get(field)
        ]

        for child in matching_children:
            child["media"] = []
            psk_id = child.get("psk_id")
            if psk_id:
                media_url = f"{API_STUDIO_URL}crudapp/get/media/naac01_faculty_participation_dc2_media/parent/{psk_id}"
                media_response = requests.get(media_url)
                if media_response.status_code == 200:
                    child["media"] = [m.get("file_name") for m in media_response.json()]

        final_children.append((cat, matching_children))

    return render(request, "ParticipationFaculty_templates/list_participation_children.html", {
        "participation": filtered_data,
        "selected_options": selected_types,
        "children_by_option": final_children,
        "participation_counts": counts,
        "username": final_staff_id
    })
   
def create_participation_child(request, participation_id, val):
    # --- Fetch parent participation ---
    url = f"{API_STUDIO_URL}getapi/naac01_faculty_participation_dc1/{participation_id}"
    participation_parent_response = requests.get(url)

    if participation_parent_response.status_code != 200:
        return HttpResponse("Error fetching participation details: " + participation_parent_response.text)

    participation_parent = participation_parent_response.json()
    transaction_id = participation_parent.get("psk_id")
    selected_options = participation_parent.get("participation", "").split(", ") if participation_parent.get("participation") else []

    if request.method == "POST":
        child_url = f"{API_STUDIO_URL}postapi/create/naac01_faculty_participation_dc2"

        payload = {
            "board_university_college_name": request.POST.get("board_university"),
            "board_designation": request.POST.get("board_designation"),
            "qp_university_college_name": request.POST.get("qp_university"),
            "qp_subject": request.POST.get("qp_subject"),
            "eval_university_college_name": request.POST.get("eval_institute"),
            "eval_subject": request.POST.get("eval_role"),
            "certificate_university_college_name": request.POST.get("cert_institute"),
            "certificate_course": request.POST.get("cert_course_name"),
            "external_examiner_university_college_name": request.POST.get("external_institute"),
            "conference_university_college_name": request.POST.get("conf_name"),
            "conference_name": request.POST.get("conf_role"),
            "seminar_university_college_name": request.POST.get("seminar_title"),
            "seminar_name": request.POST.get("seminar_role"),
            "workshop_university_college_name": request.POST.get("workshop_title"),
            "workshop_name": request.POST.get("workshop_role"),
            "design_development_university_college_name": request.POST.get("curriculum_institute"),
            "design_development_addon": request.POST.get("curriculum_role"),
            "name_of_the_resource_person": request.POST.get("name_of_the_resource_person"),
            "name_of_the_program": request.POST.get("name_of_the_program"),
            "date_from": request.POST.get("date_from"),
            "date_to": request.POST.get("date_to"),
            # REMOVED: date_from and date_to fields
            "transaction_id": transaction_id,
        }

        # Remove None values from payload to avoid sending empty fields
        payload = {k: v for k, v in payload.items() if v is not None}

        headers = {"Content-Type": "application/json"}
        response = requests.post(child_url, headers=headers, data=json.dumps({"data": payload}))

        if response.status_code != 200:
            return HttpResponse("Failed to create participation child: " + response.text)

        # --- Get child ID ---
        child_data = response.json()
        child_id = child_data.get("psk_id")

        # --- Handle media upload ---
        uploaded_file = request.FILES.get("file")
        if uploaded_file:
            validate_file_size(uploaded_file)
            validate_file_format_faculty(uploaded_file)

            media_url = f"{API_STUDIO_URL}crudapp/upload/media/naac01_faculty_participation_dc2_media/"
            file_type = uploaded_file.content_type
            payload = {"parent_psk_id": child_id}
            files = {"media": (uploaded_file.name, uploaded_file, file_type)}
            headers = {"api_name": "naac01_faculty_participation_dc2_media"}

            media_response = requests.post(media_url, headers=headers, data=payload, files=files)
            if media_response.status_code != 200:
                return HttpResponse("Failed to upload media files: " + media_response.text)

        return redirect("detail_view", id=participation_id)

    # --- Render form ---
    return render(
        request,
        "ParticipationFaculty_templates/create_participation_child.html",
        {"val": val, "participation": participation_parent, "selected_options": selected_options},
    )    

# def update_participation_child(request, participation_id, child_id, val):
#     url = f"{API_STUDIO_URL}getapi/naac01_faculty_participation_dc1/{participation_id}"
#     participation_parent_response = requests.get(url)
    
#     if participation_parent_response.status_code != 200:
#         return HttpResponse(f"Error fetching participation details: {participation_parent_response.text}")

#     participation_parent = participation_parent_response.json()
#     name = participation_parent.get('name')
#     year = participation_parent.get('year')
#     selected_options = participation_parent.get('participation', '').split(', ') if participation_parent.get('participation') else []

#     child_url = f"{API_STUDIO_URL}getapi/naac01_faculty_participation_dc2/{child_id}"
#     headers = {'Content-Type': 'application/json'}
#     response = requests.get(child_url, headers=headers)

#     if response.status_code != 200:
#         return HttpResponse(f"Failed to fetch child data: {response.text}")

#     child_data = response.json()

#     media_url = f"{API_STUDIO_URL}crudapp/get/media/naac01_faculty_participation_dc2_media/parent/{child_id}"
#     media_response = requests.get(media_url, headers=headers)

#     value_id = None
#     file_name = None
#     if media_response.status_code == 200:
#         media_data = media_response.json()
#         if media_data:
#             value_id = media_data[0]['psk_id']
#             file_name = media_data[0]['file_name']

#     if request.method == 'POST':
#         update_url = f"{API_STUDIO_URL}updateapi/update/naac01_faculty_participation_dc2/{child_id}"

#         payload = json.dumps({"data": {
#             "board_university_college_name": request.POST.get('board_university', ''),
#             "board_designation": request.POST.get('board_designation', ''),
#             "qp_university_college_name": request.POST.get('qp_university', ''),
#             "qp_subject": request.POST.get('qp_subject', ''),
#             "eval_university_college_name": request.POST.get('eval_institute', ''),
#             "eval_subject": request.POST.get('eval_role', ''),
#             "certificate_university_college_name": request.POST.get('cert_institute', ''),
#             "certificate_course": request.POST.get('cert_course_name', ''),
#             "external_examiner_university_college_name": request.POST.get('external_institute', ''),
#             "conference_university_college_name": request.POST.get('conf_name', ''),
#             "conference_name": request.POST.get('conf_role', ''),
#             "seminar_university_college_name": request.POST.get('seminar_title', ''),
#             "seminar_name": request.POST.get('seminar_role', ''),
#             "workshop_university_college_name": request.POST.get('workshop_title', ''),
#             "workshop_name": request.POST.get('workshop_role', ''),
#             "design_development_university_college_name": request.POST.get('curriculum_institute', ''),
#             "design_development_addon": request.POST.get('curriculum_role', ''),
#             "name_of_the_resource_person": request.POST.get('name_of_the_resource_person', ''),
#             "name_of_the_program": request.POST.get('name_of_the_program', ''),
#             "date_from": request.POST.get('date_from', ''),
#             "date_to": request.POST.get('date_to', ''),
#         }})
        
#         headers = {'Content-Type': 'application/json'}
#         update_response = requests.put(update_url, headers=headers, data=payload)

#         if update_response.status_code != 200:
#             return HttpResponse(f"Failed to update participation child: {update_response.text}")

#         uploaded_file = request.FILES.get('file')
#         if uploaded_file and value_id:
#             validate_file_size(uploaded_file)
#             validate_file_format_faculty(uploaded_file)
            
#             url = f"{API_STUDIO_URL}crudapp/upload/media/naac01_faculty_participation_dc2_media/{value_id}"
#             file_type = uploaded_file.content_type
#             payload = {'parent_psk_id': child_id}
#             files = {'media': (uploaded_file.name, uploaded_file, file_type)}
#             headers = {'api_name': 'naac01_faculty_participation_dc2_media', 'psk_id': str(value_id)}

#             upload_response = requests.put(url, headers=headers, data=payload, files=files)

#             if upload_response.status_code != 200:
#                 return HttpResponse(f"Failed to upload media files: {upload_response.text}")

#         return redirect('detail_view', id=participation_id)

#     return render(request, 'ParticipationFaculty_templates/update_participation_child.html', {
#         'val': val, 'participation': participation_id, 'child': child_data, 'name': name, 
#         'year': year, 'selected_options': selected_options, 'value_id': value_id, 
#         'id': participation_id, 'file_name': file_name
#     })

def update_participation_child(request, participation_id, child_id, val):
    url = f"{API_STUDIO_URL}getapi/naac01_faculty_participation_dc1/{participation_id}"
    participation_parent_response = requests.get(url)
    
    if participation_parent_response.status_code != 200:
        return HttpResponse(f"Error fetching participation details: {participation_parent_response.text}")

    participation_parent = participation_parent_response.json()
    name = participation_parent.get('name')
    year = participation_parent.get('year')
    selected_options = participation_parent.get('participation', '').split(', ') if participation_parent.get('participation') else []

    child_url = f"{API_STUDIO_URL}getapi/naac01_faculty_participation_dc2/{child_id}"
    headers = {'Content-Type': 'application/json'}
    response = requests.get(child_url, headers=headers)

    if response.status_code != 200:
        return HttpResponse(f"Failed to fetch child data: {response.text}")

    child_data = response.json()

    media_url = f"{API_STUDIO_URL}crudapp/get/media/naac01_faculty_participation_dc2_media/parent/{child_id}"
    media_response = requests.get(media_url, headers=headers)

    value_id = None
    file_name = None
    if media_response.status_code == 200:
        media_data = media_response.json()
        if media_data:
            value_id = media_data[0]['psk_id']
            file_name = media_data[0]['file_name']

    if request.method == 'POST':
        update_url = f"{API_STUDIO_URL}updateapi/update/naac01_faculty_participation_dc2/{child_id}"

        date_from = request.POST.get('date_from')
        date_to = request.POST.get('date_to')


        payload = {
            "board_university_college_name": request.POST.get('board_university'),
            "board_designation": request.POST.get('board_designation'),
            "qp_university_college_name": request.POST.get('qp_university'),
            "qp_subject": request.POST.get('qp_subject'),
            "eval_university_college_name": request.POST.get('eval_institute'),
            "eval_subject": request.POST.get('eval_role'),
            "certificate_university_college_name": request.POST.get('cert_institute'),
            "certificate_course": request.POST.get('cert_course_name'),
            "external_examiner_university_college_name": request.POST.get('external_institute'),
            "conference_university_college_name": request.POST.get('conf_name'),
            "conference_name": request.POST.get('conf_role'),
            "seminar_university_college_name": request.POST.get('seminar_title'),
            "seminar_name": request.POST.get('seminar_role'),
            "workshop_university_college_name": request.POST.get('workshop_title'),
            "workshop_name": request.POST.get('workshop_role'),
            "design_development_university_college_name": request.POST.get('curriculum_institute'),
            "design_development_addon": request.POST.get('curriculum_role'),
            "name_of_the_resource_person": request.POST.get('name_of_the_resource_person'),
            "name_of_the_program": request.POST.get('name_of_the_program'),
        }

        # ✅ Only add date fields if value exists
        if date_from:
            payload["date_from"] = date_from
        else:
            payload["date_from"] = None

        if date_to:
            payload["date_to"] = date_to
        else:
            payload["date_to"] = None

        payload = {k: v for k, v in payload.items() if v not in [None, ""]}    

        
        headers = {'Content-Type': 'application/json'}
        update_response = requests.put(
            update_url,
            headers={'Content-Type': 'application/json'},
            data=json.dumps({"data": payload})
        )

        if update_response.status_code != 200:
            return HttpResponse(f"Failed to update participation child: {update_response.text}")

        uploaded_file = request.FILES.get('file')
        if uploaded_file and value_id:
            validate_file_size(uploaded_file)
            validate_file_format_faculty(uploaded_file)
            
            url = f"{API_STUDIO_URL}crudapp/upload/media/naac01_faculty_participation_dc2_media/{value_id}"
            file_type = uploaded_file.content_type
            payload = {'parent_psk_id': child_id}
            files = {'media': (uploaded_file.name, uploaded_file, file_type)}
            headers = {'api_name': 'naac01_faculty_participation_dc2_media', 'psk_id': str(value_id)}

            upload_response = requests.put(url, headers=headers, data=payload, files=files)

            if upload_response.status_code != 200:
                return HttpResponse(f"Failed to upload media files: {upload_response.text}")

        messages.success(request, "Record updated successfully.")   

        return redirect('detail_view', id=participation_id)

    return render(request, 'ParticipationFaculty_templates/update_participation_child.html', {
        'val': val, 'participation': participation_id, 'child': child_data, 'name': name, 
        'year': year, 'selected_options': selected_options, 'value_id': value_id, 
        'id': participation_id, 'file_name': file_name
    })


def delete_participation_child(request, child_id, participation_id):
    delete_url = f"{API_STUDIO_URL}deleteapi/delete/naac01_faculty_participation_dc2/{child_id}"

    delete_response = requests.delete(delete_url)

    if delete_response.status_code == 200:
        messages.success(request, "The child participation was deleted successfully.")
    else:
        error_msg = delete_response.json() if delete_response.status_code == 400 else {}
        messages.error(request, f"Failed to delete child participation: {error_msg.get('detail', 'Unknown error')}")
    
    return redirect('detail_view', id=participation_id)

def list_options_participations(request):
    url = f"{API_STUDIO_URL}getapi/all_fields/naac01_faculty_participation_dc1/all"
    response = requests.get(url)
    
    if response.status_code != 200:
        return HttpResponse("API Call Is Not Working")

    participations = response.json()

    option_counter = Counter()
    for participation in participations:
        options = participation.get('participation', '')
        if options:
            option_list = [opt.strip() for opt in options.split(',')]
            option_counter.update(option_list)

    participation_option_counts = {option: option_counter.get(option, 0) for option in PARTICIPATION_OPTIONS}

    return render(request, 'ParticipationFaculty_templates/list_options_participations.html', {
        'participations': participations,
        'participation_option_counts': participation_option_counts
    })


    


"""
path('participations/filter/', filter_participations, name='filter_participations'),
"""

import io
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import json
from datetime import datetime

# views.py
import pandas as pd
import io
import requests
import json
from django.http import HttpResponse
from django.shortcuts import render
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from datetime import datetime

def extract_year_value(year_str):
    """Extract first year from academic year string or return None"""
    try:
        return int(str(year_str).split('-')[0]) if year_str else None
    except:
        return None

def filter_participations(request):
    """
    Main filtering function for participations with Excel and PDF export
    with role-based access control (Staff vs HOD)
    """
    # Step 1: Get the logged-in username
    user = get_settings(request)
    username = user.get('username', '')
    
    # Step 2: Determine user role (Staff or HOD)
    role_url = f"{API_STUDIO_URL}getapi/asa0504_01_01"
    payload = json.dumps({"queries": [{"field": "username", "value": username, "operation": "equal"}], "search_type": "first"})
    role_response = requests.post(role_url, headers={'Content-Type': 'application/json'}, data=payload)
    value_user = int(role_response.json().get("user_roles", "0").strip("{}")) if role_response.status_code == 200 else 0

    role_list = roles_tbl(request)
    user_role = next((role.get("user_role") for role in role_list if role.get("psk_id") == value_user), "Staff")
    
    # Step 3: Get research data to determine department for HOD
    access_token, token_type = faculty_auth()
    research_data = faculty_token(access_token, token_type)
    staff_info = next((s for s in research_data if s.get("stf_id") == username), {})
    department_name = staff_info.get("department", "")
    
    # Get all parent participations
    parent_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_faculty_participation_dc1/all"
    parent_response = requests.get(parent_url)
    
    if parent_response.status_code != 200:
        return HttpResponse("Error fetching participation data")
    
    all_parents = parent_response.json()
    
    # Get all child participations
    child_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_faculty_participation_dc2/all"
    child_response = requests.get(child_url)
    
    all_children = child_response.json() if child_response.status_code == 200 else []
    
    # Get filter parameters from request
    stf_id = request.GET.get('stf_id', '').strip()
    from_year = request.GET.get('from_year', '').strip()
    to_year = request.GET.get('to_year', '').strip()
    selected_options = request.GET.getlist('options')
    export_format = request.GET.get('export', '')
    
    # ROLE-BASED FILTERING: Apply department filter for HOD
    if user_role == "Hod":
        # HOD can only see staff from their own department
        department_staff = [s for s in research_data if s.get("department") == department_name]
        dept_staff_ids = [s.get("stf_id") for s in department_staff]
        
        # Filter parents to only those in HOD's department
        all_parents = [parent for parent in all_parents if parent.get('stf_id') in dept_staff_ids]
    
    # Filter parents based on user selections
    filtered_parents = all_parents
    
    if user_role == "Hod":
        if stf_id:  # Only filter by staff ID if one is selected
            filtered_parents = [p for p in filtered_parents if p.get("stf_id") == stf_id]
    else:
        # Staff can only see themselves
        filtered_parents = [p for p in filtered_parents if p.get("stf_id") == username]

    # Convert year values to integers for comparison
    if from_year:
        try:
            from_year_int = int(from_year.split('-')[0])
            filtered_parents = [parent for parent in filtered_parents if parent.get('year') and extract_year_value(parent.get('year')) >= from_year_int]
        except (ValueError, AttributeError):
            pass
    
    if to_year:
        try:
            to_year_int = int(to_year.split('-')[0])
            filtered_parents = [parent for parent in filtered_parents if parent.get('year') and extract_year_value(parent.get('year')) <= to_year_int]
        except (ValueError, AttributeError):
            pass
    
    if selected_options:
        filtered_parents = [parent for parent in filtered_parents if any(opt in (parent.get('participation') or '') for opt in selected_options)]
    
    # Get parent IDs for child filtering
    parent_ids = [parent.get('psk_id') for parent in filtered_parents]
    
    # Filter children based on parent IDs
    filtered_children = [child for child in all_children if child.get('transaction_id') in parent_ids]
    
    # Fetch media files for each child - MODIFIED TO USE MEDIA ID INSTEAD OF CHILD ID
    for child in filtered_children:
        child_id = child.get('psk_id')
        media_url = f"{API_STUDIO_URL}crudapp/get/media/naac01_faculty_participation_dc2_media/parent/{child_id}"
        media_response = requests.get(media_url)
        
        if media_response.status_code == 200:
            media_data = media_response.json()
            # Process each media file to get individual media IDs
            processed_media_files = []
            for media in media_data:
                # Extract media ID from the media object
                media_id = media.get('id') or media.get('media_id') or media.get('psk_id') or media.get('value_id')
                
                if media_id:
                    # Create direct API URL using media ID
                    direct_api_url = f"{API_STUDIO_URL}crudapp/view/media/naac01_faculty_participation_dc2_media/{media_id}"
                    
                    # Add the media URL and other media info
                    processed_media = {
                        'file_name': media.get('file_name', 'Unknown'),
                        'media_id': media_id,
                        'direct_api_url': direct_api_url,
                        'original_data': media  # Keep original data for reference
                    }
                    processed_media_files.append(processed_media)
                else:
                    # Fallback if no media ID found - use child ID as before
                    fallback_url = f"{API_STUDIO_URL}crudapp/view/media/naac01_faculty_participation_dc2_media/{child_id}"
                    processed_media = {
                        'file_name': media.get('file_name', 'Unknown'),
                        'media_id': None,
                        'direct_api_url': fallback_url,
                        'original_data': media
                    }
                    processed_media_files.append(processed_media)
            
            child['media_files'] = processed_media_files
        else:
            child['media_files'] = []
    
    # Create mapping of parent ID to parent data for easy access
    parent_map = {parent['psk_id']: parent for parent in filtered_parents}
    
    # Add parent information to each child
    for child in filtered_children:
        parent_id = child.get('transaction_id')
        if parent_id in parent_map:
            child['parent_data'] = parent_map[parent_id]
    
    # ADD THIS: Pre-split participation strings for template (using comma separator)
    for parent in filtered_parents:
        if parent.get('participation'):
            # Split by comma and strip whitespace from each part
            parent['participation_list'] = [part.strip() for part in parent['participation'].split(',')]
        else:
            parent['participation_list'] = []
    
    # Handle export formats - PASS THE FILTERED DATA CORRECTLY
    if export_format:
        if export_format.lower() == 'excel':
            return export_to_excel(filtered_parents, filtered_children, selected_options)
        elif export_format.lower() == 'pdf':
            return export_to_pdf(filtered_parents, filtered_children, selected_options)
    
    # Get unique staff IDs for dropdown - filtered by role
    if user_role == "Hod":
        # HOD can see all staff in their department
        department_staff = [s for s in research_data if s.get("department") == department_name]
        dept_staff_ids = [s.get("stf_id") for s in department_staff]
        staff_ids = sorted(list(set(parent.get('stf_id') for parent in all_parents if parent.get('stf_id') in dept_staff_ids)))
    else:
        # Staff can only see themselves
        staff_ids = [username]
    
    # Get unique years for dropdown
    year_values = []
    for parent in all_parents:
        if parent.get('year'):
            try:
                year_val = extract_year_value(parent.get('year'))
                year_values.append(year_val)
            except (ValueError, AttributeError):
                continue
    
    # Convert back to "YYYY-YYYY" format for display
    years = sorted(list(set(f"{y}-{y+1}" for y in year_values)))
    
    context = {
        'parents': filtered_parents,
        'children': filtered_children,
        'staff_ids': staff_ids,
        'years': years,
        'options': PARTICIPATION_OPTIONS,
        'selected_stf_id': stf_id,
        'selected_from_year': from_year,
        'selected_to_year': to_year,
        'selected_options': selected_options,
        'filter_applied': any([stf_id, from_year, to_year, selected_options]),
        'user_role': user_role,
        'username': username
    }
    
    return render(request, 'ParticipationFaculty_templates/filter_participations.html', context)


def export_to_excel(parents, children, selected_forms=None):
    """
    Export filtered participations into Excel.
    - If no options are selected, shows only "No data available for your request."
    - Otherwise exports filtered participation data.
    """
    import io
    import re
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    try:
        # ---------- Setup ----------
        if not selected_forms:  # if None or empty list
            wb = Workbook()
            ws = wb.active
            ws.title = "No Data"
            ws.append(["No data available for your request."])

            # Center align and style it a bit
            cell = ws["A1"]
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.column_dimensions["A"].width = 40
            ws.row_dimensions[1].height = 25

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="participations.xlsx"'
            return response

        # ---------- Continue normal export if options selected ----------
        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet

        hyperlink_pattern = re.compile(r'=HYPERLINK\("([^"]+)",\s*"([^"]+)"\)')

        form_configs = {
            'Board of Studies': {'fields': ['board_university_college_name'], 'detail': 'board_designation'},
            'Question Paper Setting': {'fields': ['qp_university_college_name'], 'detail': 'qp_subject'},
            'Evaluation': {'fields': ['eval_university_college_name'], 'detail': 'eval_subject'},
            'Add-On Program': {'fields': ['design_development_university_college_name'], 'detail': 'design_development_addon'},
            'Certificate Courses': {'fields': ['certificate_university_college_name'], 'detail': 'certificate_course'},
            'External Examiner': {'fields': ['external_examiner_university_college_name'], 'detail': ''},
            'Conference': {'fields': ['conference_university_college_name'], 'detail': 'conference_name'},
            'Seminar': {'fields': ['seminar_university_college_name'], 'detail': 'seminar_name'},
            'Workshop': {'fields': ['workshop_university_college_name'], 'detail': 'workshop_name'},
            'FDP': {'fields': ['name_of_the_resource_person'], 'detail': 'name_of_the_program'}
        }

        # Filter form configs
        form_configs = {k: v for k, v in form_configs.items() if k in selected_forms}

        headers = ["Staff ID", "Staff Name", "Academic Year", "Activity Type", "Institution", "Details", "Attachment"]

        ws = wb.create_sheet("Faculty Participation")
        ws.append(headers)

        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_font = Font(bold=True, color="333333")
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Organize data
        staff_children = {}
        for child in children:
            parent = child.get('parent_data', {})
            staff_id = parent.get('stf_id', '')
            staff_children.setdefault(staff_id, []).append(child)

        unique_parents = {p.get('stf_id'): p for p in parents if p.get('stf_id')}

        for parent in unique_parents.values():
            staff_id = parent.get('stf_id', '')
            staff_name = parent.get('stf_name', '')
            raw_year = parent.get('year', '')
            academic_year = f"{int(raw_year)}-{int(raw_year) + 1}" if str(raw_year).isdigit() else str(raw_year) or "N/A"

            child_list = staff_children.get(staff_id, [])
            if not child_list:
                continue

            for child in child_list:
                form_detected = None
                institution = ''
                details = ''

                for form_name, config in form_configs.items():
                    for field in config['fields']:
                        if child.get(field):
                            form_detected = form_name
                            institution = child.get(field, '')
                            details = child.get(config['detail'], '') if config['detail'] else ''
                            break
                    if form_detected:
                        break

                if form_detected not in selected_forms:
                    continue

                attachments = []
                for media in child.get('media_files', []):
                    file_name = media.get('file_name', 'File')
                    media_url = media.get('direct_api_url', '')
                    if media_url:
                        attachments.append(f'=HYPERLINK("{media_url}", "{file_name}")')
                attachment_cell = "\n".join(attachments) if attachments else "-"

                ws.append([
                    staff_id,
                    staff_name,
                    academic_year,
                    form_detected,
                    institution or "-",
                    details or "-",
                    attachment_cell
                ])

        for row in ws.iter_rows(min_row=2, max_col=len(headers), max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if isinstance(cell.value, str) and cell.value.startswith("=HYPERLINK("):
                    cell.font = Font(color="0000FF", underline="single")

        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = 0
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            for cell in col_cells:
                val = str(cell.value) if cell.value else ""
                for line in val.split("\n"):
                    max_len = max(max_len, len(line))
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 40)

        ws.freeze_panes = ws['A2']
        ws.auto_filter.ref = ws.dimensions

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="participations.xlsx"'
        return response

    except Exception as e:
        return HttpResponse(f"Error generating Excel file: {str(e)}")

# def export_to_pdf(parents, children, selected_options=None):
#     """
#     Export Faculty Participation data to PDF with separate pages per staff.
#     - If `selected_options` is None or empty → show only staff info (no participation data)
#     - If no participation data → show an empty row labeled 'No data available for this staff.'
#     """
#     from reportlab.lib import colors
#     from reportlab.lib.pagesizes import A4
#     from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
#     from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, PageBreak
#     from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
#     from reportlab.lib.units import inch
#     from django.http import HttpResponse
#     from collections import defaultdict

#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="participations_filtered.pdf"'

#     doc = SimpleDocTemplate(
#         response,
#         pagesize=A4,
#         topMargin=0.5 * inch,
#         bottomMargin=0.5 * inch,
#         leftMargin=0.5 * inch,
#         rightMargin=0.5 * inch
#     )
#     elements = []
#     styles = getSampleStyleSheet()

#     # ---------------- Custom Styles ----------------
#     title_style = ParagraphStyle(
#         'CustomTitle', parent=styles['Heading1'], fontSize=14, spaceAfter=15,
#         alignment=TA_CENTER, textColor=colors.HexColor('#2c3e50'), fontName='Helvetica-Bold'
#     )
#     info_style = ParagraphStyle(
#         'InfoStyle', parent=styles['Normal'], fontSize=9,
#         textColor=colors.HexColor('#2c3e50'), leading=12
#     )
#     right_aligned_info_style = ParagraphStyle('RightInfoStyle', parent=info_style, alignment=TA_RIGHT)
#     table_header_style = ParagraphStyle(
#         'TableHeader', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER,
#         textColor=colors.white, fontName='Helvetica-Bold'
#     )
#     table_cell_style = ParagraphStyle(
#         'TableCell', parent=styles['Normal'], fontSize=7, alignment=TA_LEFT,
#         textColor=colors.HexColor('#2c3e50'), leading=10, wordWrap='CJK'
#     )
#     attachment_link_style = ParagraphStyle(
#         'AttachmentLink', parent=styles['Normal'], fontSize=6, alignment=TA_LEFT,
#         textColor=colors.HexColor('#1a5276'), leading=8, wordWrap='CJK'
#     )

#     # ---------------- Group Children by Staff ----------------
#     staff_children = defaultdict(list)
#     if selected_options:  # Only group children if filters are applied
#         for child in children:
#             # Detect activity type
#             activity_type = 'Unknown'
#             if child.get('board_university_college_name'):
#                 activity_type = 'Board of Studies'
#             elif child.get('qp_university_college_name'):
#                 activity_type = 'Question Paper Setting'
#             elif child.get('eval_university_college_name'):
#                 activity_type = 'Evaluation'
#             elif child.get('design_development_university_college_name'):
#                 activity_type = 'Add-On Program'
#             elif child.get('certificate_university_college_name'):
#                 activity_type = 'Certificate Courses'
#             elif child.get('external_examiner_university_college_name'):
#                 activity_type = 'External Examiner'
#             elif child.get('conference_university_college_name'):
#                 activity_type = 'Conference'
#             elif child.get('seminar_university_college_name'):
#                 activity_type = 'Seminar'
#             elif child.get('workshop_university_college_name'):
#                 activity_type = 'Workshop'
#             elif child.get('name_of_the_resource_person'):
#                 activity_type = 'FDP'

#             # Skip if not part of selected filter
#             if activity_type not in selected_options:
#                 continue

#             staff_id = child.get('parent_data', {}).get('stf_id', 'Unknown')
#             staff_children[staff_id].append(child)

#     # ---------------- Deduplicate parent staff list ----------------
#     unique_parents = {}
#     for p in parents:
#         sid = p.get('stf_id')
#         if sid and sid not in unique_parents:
#             unique_parents[sid] = p

#     # ---------------- Generate Each Staff Page ----------------
#     for idx, (staff_id, parent) in enumerate(unique_parents.items()):
#         if idx > 0:
#             elements.append(PageBreak())

#         # If selected_options is None, skip all children (show empty)
#         child_list = staff_children.get(staff_id, []) if selected_options else []

#         # Title
#         elements.append(Paragraph("FACULTY PARTICIPATION REPORT", title_style))

#         # Staff info
#         staff_info_data = [
#             [Paragraph(f"<b>Staff ID:</b> {parent.get('stf_id', 'N/A')}", info_style),
#              Paragraph(f"<b>Department:</b> {parent.get('department', 'N/A')}", right_aligned_info_style)],
#             [Paragraph(f"<b>Name:</b> {parent.get('stf_name', 'N/A')}", info_style),
#              Paragraph("", info_style)]
#         ]
#         staff_info_table = Table(staff_info_data, colWidths=[3.3 * inch, 3.3 * inch])
#         staff_info_table.setStyle(TableStyle([
#             ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
#             ('LEFTPADDING', (0, 0), (-1, -1), 6),
#             ('RIGHTPADDING', (0, 0), (-1, -1), 6),
#             ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
#             ('TOPPADDING', (0, 0), (-1, -1), 2),
#         ]))
#         elements.append(staff_info_table)
#         elements.append(Spacer(1, 10))

#         # Table headers
#         headers = ['Activity Type', 'Institution', 'Details', 'Academic Year', 'Attachments']
#         child_data = [[Paragraph(f"<b>{h}</b>", table_header_style) for h in headers]]

#         # --- CASE: No participation data ---
#         if not child_list:
#             child_data.append([
#                 Paragraph('<i>No data available for this staff.</i>', table_cell_style),
#                 Paragraph('', table_cell_style),
#                 Paragraph('', table_cell_style),
#                 Paragraph('', table_cell_style),
#                 Paragraph('', table_cell_style),
#             ])
#         else:
#             # --- CASE: Has participation data ---
#             for child in child_list:
#                 activity_type = 'Unknown'
#                 institution, details = '', ''
#                 if child.get('board_university_college_name'):
#                     activity_type = 'Board of Studies'
#                     institution = child.get('board_university_college_name', '')
#                     details = child.get('board_designation', '')
#                 elif child.get('qp_university_college_name'):
#                     activity_type = 'Question Paper Setting'
#                     institution = child.get('qp_university_college_name', '')
#                     details = child.get('qp_subject', '')
#                 elif child.get('eval_university_college_name'):
#                     activity_type = 'Evaluation'
#                     institution = child.get('eval_university_college_name', '')
#                     details = child.get('eval_subject', '')
#                 elif child.get('design_development_university_college_name'):
#                     activity_type = 'Add-On Program'
#                     institution = child.get('design_development_university_college_name', '')
#                     details = child.get('design_development_addon', '')
#                 elif child.get('certificate_university_college_name'):
#                     activity_type = 'Certificate Courses'
#                     institution = child.get('certificate_university_college_name', '')
#                     details = child.get('certificate_course', '')
#                 elif child.get('external_examiner_university_college_name'):
#                     activity_type = 'External Examiner'
#                     institution = child.get('external_examiner_university_college_name', '')
#                     details = 'External Examiner'
#                 elif child.get('conference_university_college_name'):
#                     activity_type = 'Conference'
#                     institution = child.get('conference_university_college_name', '')
#                     details = child.get('conference_name', '')
#                 elif child.get('seminar_university_college_name'):
#                     activity_type = 'Seminar'
#                     institution = child.get('seminar_university_college_name', '')
#                     details = child.get('seminar_name', '')
#                 elif child.get('workshop_university_college_name'):
#                     activity_type = 'Workshop'
#                     institution = child.get('workshop_university_college_name', '')
#                     details = child.get('workshop_name', '')
#                 elif child.get('name_of_the_resource_person'):
#                     activity_type = 'FDP'
#                     institution = child.get('name_of_the_resource_person', '')
#                     details = child.get('name_of_the_program', '')

#                 # Media links
#                 media_files = child.get('media_files', [])
#                 media_links = []
#                 for media in media_files:
#                     fname = media.get('file_name', 'Unknown')
#                     url = media.get('direct_api_url', '')
#                     if url:
#                         media_links.append(f'<a href="{url}" color="blue">{fname}</a>')
#                     else:
#                         media_links.append(fname)
#                 media_info = "<br/>".join(media_links) if media_links else "No attachments"

#                 raw_year = child.get('parent_data', {}).get('year', 'N/A')
#                 academic_year = f"{int(raw_year)}-{int(raw_year)+1}" if str(raw_year).isdigit() else raw_year

#                 child_data.append([
#                     Paragraph(str(activity_type), table_cell_style),
#                     Paragraph(str(institution), table_cell_style),
#                     Paragraph(str(details), table_cell_style),
#                     Paragraph(str(academic_year), table_cell_style),
#                     Paragraph(media_info, attachment_link_style)
#                 ])

#         # Child table styling
#         col_widths = [1.2 * inch, 1.6 * inch, 1.6 * inch, 0.8 * inch, 1.4 * inch]
#         child_table = Table(child_data, colWidths=col_widths, repeatRows=1)
#         child_table.setStyle(TableStyle([
#             ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#02548b")),
#             ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
#             ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#             ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
#             ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d6d6d6')),
#             ('FONTSIZE', (0, 0), (-1, -1), 7),
#             ('LEFTPADDING', (0, 0), (-1, -1), 4),
#             ('RIGHTPADDING', (0, 0), (-1, -1), 4),
#             ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
#             ('TOPPADDING', (0, 0), (-1, -1), 3),
#         ]))

#         elements.append(child_table)
#         elements.append(Spacer(1, 10))
#         elements.append(Paragraph(
#             f"<i>Total participations for {parent.get('stf_name', 'this staff')}: {len(child_list)}</i>",
#             table_cell_style
#         ))

#     doc.build(elements)
#     return response

def export_to_pdf(parents, children, selected_options=None):
    """
    Export Faculty Participation data to PDF with separate pages per staff.
    - If `selected_options` is None or empty → show only staff info (no participation data)
    - If no participation data → show an empty row labeled 'No data available for this staff.'
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, PageBreak
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.lib.units import inch
    from django.http import HttpResponse
    from collections import defaultdict

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="participations_filtered.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch
    )
    elements = []
    styles = getSampleStyleSheet()

    # ---------------- Custom Styles ----------------
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'], fontSize=14, spaceAfter=15,
        alignment=TA_CENTER, textColor=colors.HexColor('#2c3e50'), fontName='Helvetica-Bold'
    )
    info_style = ParagraphStyle(
        'InfoStyle', parent=styles['Normal'], fontSize=9,
        textColor=colors.HexColor('#2c3e50'), leading=12
    )
    right_aligned_info_style = ParagraphStyle('RightInfoStyle', parent=info_style, alignment=TA_RIGHT)
    table_header_style = ParagraphStyle(
        'TableHeader', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER,
        textColor=colors.white, fontName='Helvetica-Bold'
    )
    table_cell_style = ParagraphStyle(
        'TableCell', parent=styles['Normal'], fontSize=7, alignment=TA_LEFT,
        textColor=colors.HexColor('#2c3e50'), leading=10, wordWrap='CJK'
    )
    attachment_link_style = ParagraphStyle(
        'AttachmentLink', parent=styles['Normal'], fontSize=6, alignment=TA_LEFT,
        textColor=colors.HexColor('#1a5276'), leading=8, wordWrap='CJK'
    )
    no_data_style = ParagraphStyle(
        'NoData', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER,
        textColor=colors.HexColor('#7f8c8d'), fontStyle='italic'
    )

    # ---------------- Display Selected Staff Information ----------------
    if selected_options and 'staff_names' in selected_options and selected_options['staff_names']:
        staff_names = selected_options['staff_names']
        if isinstance(staff_names, list):
            staff_info = ", ".join(staff_names)
        else:
            staff_info = str(staff_names)
        elements.append(Paragraph(f"<b>Selected Staff:</b> {staff_info}", ParagraphStyle(
            'SelectedStaff', parent=styles['Normal'], fontSize=10, alignment=TA_LEFT, 
            textColor=colors.HexColor('#2c3e50'), spaceAfter=6
        )))
        elements.append(Spacer(1, 6))

    # ---------------- Group Children by Staff ----------------
    staff_children = defaultdict(list)
    if selected_options and selected_options != {}:  # Only group children if filters are applied and not empty
        for child in children:
            # Detect activity type
            activity_type = 'Unknown'
            if child.get('board_university_college_name'):
                activity_type = 'Board of Studies'
            elif child.get('qp_university_college_name'):
                activity_type = 'Question Paper Setting'
            elif child.get('eval_university_college_name'):
                activity_type = 'Evaluation'
            elif child.get('design_development_university_college_name'):
                activity_type = 'Add-On Program'
            elif child.get('certificate_university_college_name'):
                activity_type = 'Certificate Courses'
            elif child.get('external_examiner_university_college_name'):
                activity_type = 'External Examiner'
            elif child.get('conference_university_college_name'):
                activity_type = 'Conference'
            elif child.get('seminar_university_college_name'):
                activity_type = 'Seminar'
            elif child.get('workshop_university_college_name'):
                activity_type = 'Workshop'
            elif child.get('name_of_the_resource_person'):
                activity_type = 'FDP'

            # Skip if not part of selected filter
            if activity_type not in selected_options:
                continue

            staff_id = child.get('parent_data', {}).get('stf_id', 'Unknown')
            staff_children[staff_id].append(child)

    # ---------------- Deduplicate parent staff list ----------------
    unique_parents = {}
    for p in parents:
        sid = p.get('stf_id')
        if sid and sid not in unique_parents:
            unique_parents[sid] = p

    # ---------------- Check if there are any parents ----------------
    if not unique_parents:
        elements.append(Paragraph("No staff data available", no_data_style))
        doc.build(elements)
        return response

    # ---------------- Generate Each Staff Page ----------------
    for idx, (staff_id, parent) in enumerate(unique_parents.items()):
        if idx > 0:
            elements.append(PageBreak())

        # If selected_options is None or empty, skip all children (show empty)
        child_list = staff_children.get(staff_id, []) if selected_options and selected_options != {} else []

        # Title
        elements.append(Paragraph("FACULTY PARTICIPATION REPORT", title_style))

        # Staff info
        staff_info_data = [
            [Paragraph(f"<b>Staff ID:</b> {parent.get('stf_id', 'N/A')}", info_style),
             Paragraph(f"<b>Department:</b> {parent.get('department', 'N/A')}", right_aligned_info_style)],
            [Paragraph(f"<b>Name:</b> {parent.get('stf_name', 'N/A')}", info_style),
             Paragraph("", info_style)]
        ]
        staff_info_table = Table(staff_info_data, colWidths=[3.3 * inch, 3.3 * inch])
        staff_info_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
        ]))
        elements.append(staff_info_table)
        elements.append(Spacer(1, 10))

        # Table headers
        headers = ['Activity Type', 'Institution', 'Details', 'Academic Year', 'Attachments']
        child_data = [[Paragraph(f"<b>{h}</b>", table_header_style) for h in headers]]

        # --- CASE: No participation data ---
        if not child_list:
            # Create a separate table for "No data available" message that spans all columns
            no_data_table_data = [[Paragraph('<i>No data available for this staff.</i>', no_data_style)]]
            no_data_table = Table(no_data_table_data, colWidths=[sum([1.2 * inch, 1.6 * inch, 1.6 * inch, 0.8 * inch, 1.4 * inch])])
            no_data_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d6d6d6')),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
            ]))
            
            # Add headers table
            headers_table = Table([headers], colWidths=[1.2 * inch, 1.6 * inch, 1.6 * inch, 0.8 * inch, 1.4 * inch])
            headers_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#02548b")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d6d6d6')),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
            ]))
            
            elements.append(headers_table)
            elements.append(no_data_table)
            
        else:
            # --- CASE: Has participation data ---
            for child in child_list:
                activity_type = 'Unknown'
                institution, details = '', ''
                if child.get('board_university_college_name'):
                    activity_type = 'Board of Studies'
                    institution = child.get('board_university_college_name', '')
                    details = child.get('board_designation', '')
                elif child.get('qp_university_college_name'):
                    activity_type = 'Question Paper Setting'
                    institution = child.get('qp_university_college_name', '')
                    details = child.get('qp_subject', '')
                elif child.get('eval_university_college_name'):
                    activity_type = 'Evaluation'
                    institution = child.get('eval_university_college_name', '')
                    details = child.get('eval_subject', '')
                elif child.get('design_development_university_college_name'):
                    activity_type = 'Add-On Program'
                    institution = child.get('design_development_university_college_name', '')
                    details = child.get('design_development_addon', '')
                elif child.get('certificate_university_college_name'):
                    activity_type = 'Certificate Courses'
                    institution = child.get('certificate_university_college_name', '')
                    details = child.get('certificate_course', '')
                elif child.get('external_examiner_university_college_name'):
                    activity_type = 'External Examiner'
                    institution = child.get('external_examiner_university_college_name', '')
                    details = 'External Examiner'
                elif child.get('conference_university_college_name'):
                    activity_type = 'Conference'
                    institution = child.get('conference_university_college_name', '')
                    details = child.get('conference_name', '')
                elif child.get('seminar_university_college_name'):
                    activity_type = 'Seminar'
                    institution = child.get('seminar_university_college_name', '')
                    details = child.get('seminar_name', '')
                elif child.get('workshop_university_college_name'):
                    activity_type = 'Workshop'
                    institution = child.get('workshop_university_college_name', '')
                    details = child.get('workshop_name', '')
                elif child.get('name_of_the_resource_person'):
                    activity_type = 'FDP'
                    institution = child.get('name_of_the_resource_person', '')
                    details = child.get('name_of_the_program', '')

                # Media links
                media_files = child.get('media_files', [])
                media_links = []
                for media in media_files:
                    fname = media.get('file_name', 'Unknown')
                    url = media.get('direct_api_url', '')
                    if url:
                        media_links.append(f'<a href="{url}" color="blue">{fname}</a>')
                    else:
                        media_links.append(fname)
                media_info = "<br/>".join(media_links) if media_links else "No attachments"

                raw_year = child.get('parent_data', {}).get('year', 'N/A')
                academic_year = f"{int(raw_year)}-{int(raw_year)+1}" if str(raw_year).isdigit() else raw_year

                child_data.append([
                    Paragraph(str(activity_type), table_cell_style),
                    Paragraph(str(institution), table_cell_style),
                    Paragraph(str(details), table_cell_style),
                    Paragraph(str(academic_year), table_cell_style),
                    Paragraph(media_info, attachment_link_style)
                ])

            # Create table for data case
            col_widths = [1.2 * inch, 1.6 * inch, 1.6 * inch, 0.8 * inch, 1.4 * inch]
            child_table = Table(child_data, colWidths=col_widths, repeatRows=1)
            child_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#02548b")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d6d6d6')),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
            ]))
            elements.append(child_table)

        # Only show total participations count if there is data
        if child_list:
            elements.append(Spacer(1, 10))
            elements.append(Paragraph(
                f"<i>Total participations for {parent.get('stf_name', 'this staff')}: {len(child_list)}</i>",
                table_cell_style
            ))

    doc.build(elements)
    return response