import json
import requests
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from MIS.functions import validate_file_format, validate_file_size
from datetime import date, datetime
from user_management.settings_views import *

API_STUDIO_URL = user_bundle_settings()

def collaborative_key():
    url = "https://api.hcaschennai.edu.in/auth/token"

    payload = json.dumps({
        "secret_key": "C4ZoXbsAnHLjk1Xyz4QPT2eoiNx6K6fo"
    })
    headers = {'Content-Type': 'application/json'}

    response = requests.post(url, headers=headers, data=payload)
    if response.status_code == 200:
        res_data = response.json()
        access_token = res_data.get('access_token')
        token_type = res_data.get('token_type')
        return access_token, token_type
    return None, None


def get_collaborative_data(access_token, token_type):
    url = "https://api.hcaschennai.edu.in/sqlviews/api/v1/auth/get_response_data"

    payload = json.dumps({
        "psk_uid": "51a531b4-bd55-491c-861d-a8d7227b325b",
        "project": "hcas",
        "data": {}
    })
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'{token_type} {access_token}'  # Fix spacing
    }

    response = requests.post(url, headers=headers, data=payload)
    if response.status_code == 200:
        return response.json()
    return []


def collaborative_students_create(request):
    error_message = None

    # Step 1: Get API token
    access_token, token_type = collaborative_key()
    if not access_token or not token_type:
        error_message = 'Failed to get access token from API.'
        return render(request, 'collaborative_students_templates/collaborative_students_create.html', {'error': error_message})

    # Step 2: Fetch career data
    student_data = get_collaborative_data(access_token, token_type)
    if not student_data:
        error_message = 'Failed to fetch staff data.'
        return render(request, 'collaborative_students_templates/collaborative_students_create.html', {'error': error_message})
    
    user = get_settings(request)
    username = user['username']
    # username = 'CS-T151'
    
    selected_faculty = None
    
    for faculty in student_data:
        if faculty['stf_id'] == username:
            selected_faculty = faculty
            break
        
    staff_name = selected_faculty.get('stf_name') if selected_faculty else ''
    department_name = selected_faculty.get('department') if selected_faculty else ''
    
    # Generate academic years for dropdown
    current_year = datetime.now().year
    publication_year = [f"{y}-{y+1}" for y in range(current_year, 2010, -1)]

    if request.method == 'POST':
        # Get form data from the POST request
        staff_id = username
        staff_name = request.POST.get('staff_name')
        department_name = request.POST.get('department_name')
        activity_title = request.POST.get('activity_title')
        agency_name = request.POST.get('agency_name')
        phone_number = request.POST.get('phone_number')
        participant_name = request.POST.get('participant_name')
        year_range = request.POST.get('year')  # This will be in "YYYY-YYYY" format
        date_from = request.POST.get('date_from')
        date_to = request.POST.get('date_to')
        activity_name = request.POST.get('activity_name')
        
        # Convert year range to date format for storage
        if year_range and '-' in year_range:
            try:
                start_year = int(year_range.split('-')[0])
                # Create a date object (using June 1st as default)
                year_date = date(start_year, 6, 1).isoformat()
            except (ValueError, IndexError):
                year_date = None
        else:
            year_date = None
        
        for faculty in student_data:
            if faculty['stf_id'] == staff_id:
                selected_faculty = faculty
                break
            
        staff_name = selected_faculty.get('stf_name') if selected_faculty else ''
        department_name = selected_faculty.get('department') if selected_faculty else ''

        # API URL to send the form data to
        url = f"{API_STUDIO_URL}postapi/create/naac01_students_benefited_dc1"
        payload = json.dumps({
            "data": {
                "staff_name": staff_name,
                "staff_id": staff_id,
                "department_name": department_name,
                "activity_title": activity_title,
                "agency_name": agency_name,
                "phone_number": phone_number,
                "participant_name": participant_name,
                "year": year_date,  # Use the converted date
                "date_from": date_from,
                "date_to": date_to,
                "activity_name": activity_name
            }
        })

        headers = {'Content-Type': 'application/json'}

        # Make API call to create the collaborative students/workshop
        response = requests.post(url, headers=headers, data=payload)

        if response.status_code == 200:
            # Successfully created the collaborative students/workshop
            file_data = response.json()
            psk_id = file_data.get('psk_id')  # Get psk_id from the response

            # Handle file uploads
            upload_url = f"{API_STUDIO_URL}crudapp/upload/media/naac01_students_benefited_dc1_media"
            uploaded_files = request.FILES.getlist('file')

            if not uploaded_files:
                messages.warning(request, "No files selected for upload.")
            else:
                upload_errors = []
                current_year = datetime.now().year
                fields = ['PL', 'Certificate', 'RPP', 'DR', 'GTP', 'SA']

                for field, uploaded_file in zip(fields, uploaded_files):
                    try:
                        # Validate file size and format
                        validate_file_size(uploaded_file)
                        validate_file_format(uploaded_file)

                        # Generate custom filename
                        custom_filename = f"{staff_id}_{field}_{current_year}_{uploaded_file.name}"
                        print(f"Generated filename: {custom_filename}")  # Debug log

                        # Prepare upload payload and headers
                        files = {'media': (custom_filename, uploaded_file, uploaded_file.content_type)}
                        upload_payload = {'parent_psk_id': psk_id}
                        upload_headers = {'api_name': 'naac01_students_benefited_dc1_media'}

                        # Upload the file
                        upload_response = requests.post(upload_url, headers=upload_headers, data=upload_payload, files=files)

                        if upload_response.status_code != 200:
                            upload_errors.append(
                                f"File upload failed for {uploaded_file.name}. Error: {upload_response.text}"
                            )
                    except Exception as e:
                        upload_errors.append(f"Error processing file '{uploaded_file.name}': {str(e)}")

                if upload_errors:
                    for error in upload_errors:
                        messages.error(request, error)
                else:
                    messages.success(request, "Documents uploaded successfully.")

            # Redirect to collaborative students view after successful creation
            return redirect('collaborative_students_list')

        else:
            # API call failed for creating collaborative students
            messages.error(request, "Failed to create collaborative students. Please try again.")
            return render(request, 'collaborative_students_templates/collaborative_students_create.html')

    else:
        # If the request is GET, render the collaborative students creation form
        return render(request, 'collaborative_students_templates/collaborative_students_create.html', {
            'student_data': student_data, 
            "username": username, 
            "staff_name": staff_name, 
            "department": department_name,
            'publication_year': publication_year  # Pass years to template
        })



# View collaborative_students Details
def collaborative_students_view(request, id):
    url = f"{API_STUDIO_URL}getapi/naac01_students_benefited_dc1/{id}"
    response = requests.get(url)

    if response.status_code == 200:
        collaborative_students_data = response.json()
        return render(request, "collaborative_students_templates/collaborative_students_view.html",
                      {'collaborative_students': collaborative_students_data})

    return HttpResponse(f"Error fetching Course details: {response.text}", status=500)


# List All collaborative_studentss
def collaborative_students_list(request):
    # URL to get collaborative students data
    url = f"{API_STUDIO_URL}getapi/all_fields/naac01_students_benefited_dc1/all"
    response = requests.get(url)

    if response.status_code == 200:
        collaborative_students = response.json()
        print("collaborative_students:", collaborative_students)
    else:
        collaborative_students = []

    # Get the current username (staff_id)
    user = get_settings(request)  # Assuming this function retrieves user settings (including staff_id)
    username = user.get('username')  # Adjust if necessary to fetch staff_id or username from session
    # username = 'CS-T151'
    filtered_collaborative_students = [student for student in collaborative_students if student.get('staff_id') == username]
    
    #from_dashboard = request.GET.get('from') == 'dashboard' or 'admin_hod_dash' or 'admin_dash' or 'department_dashboard'
    # If no username (staff_id) is found, return the same page with an empty list
    # if not username:
    #     return render(request, 'collaborative_students_templates/collaborative_students_list.html', {'collaborative_studentss': []})

    # Filter the collaborative students data based on the staff_id
    
    
    selected_staff_id = request.GET.get('staff_id')
    if selected_staff_id:
        filtered_collaborative_students = [student for student in collaborative_students if student.get('staff_id') == selected_staff_id]


    selected_department = request.GET.get('department')
    print("selected_department:", selected_department)
    if selected_department:
        filtered_collaborative_students = [student for student in collaborative_students if student.get('department_name') == selected_department]

    
    # If no data is found for the username, return the same page with an empty list
    if not filtered_collaborative_students:
        return render(request, 'collaborative_students_templates/collaborative_students_list.html', {'collaborative_studentss': []})

    # Return the filtered data to the template
    return render(request, 'collaborative_students_templates/collaborative_students_list.html', {'collaborative_studentss': filtered_collaborative_students, 
    #'from_dashboard': from_dashboard
    })


# Update collaborative_students
def collaborative_students_update(request, id):
    # Generate academic years for dropdown
    current_year = datetime.now().year
    publication_year = [f"{y}-{y+1}" for y in range(current_year, 2010, -1)]

    # Fetch collaborative_students details
    url = f"{API_STUDIO_URL}getapi/naac01_students_benefited_dc1/{id}"
    response = requests.get(url)

    if response.status_code == 200:
        collaborative_students = response.json()
        
        # Convert date format "2005-06-01" to "2005-2006" for display
        year_str = collaborative_students.get('year', '')
        if year_str and isinstance(year_str, str) and len(year_str) >= 4:
            try:
                # Extract the year part and create range format
                start_year = year_str[:4]
                collaborative_students['year_range'] = f"{start_year}-{int(start_year) + 1}"
            except (ValueError, TypeError):
                collaborative_students['year_range'] = year_str
        else:
            collaborative_students['year_range'] = year_str
            
    else:
        return HttpResponse(f"Error fetching collaborative_students details: {response.text}", status=500)

    # Fetch media (child files) associated with this collaborative_students
    media_url = f"{API_STUDIO_URL}crudapp/get/media/naac01_students_benefited_dc1_media/parent/{id}"
    media_response = requests.get(media_url)

    if media_response.status_code == 200:
        child_files = media_response.json()
    else:
        return HttpResponse(f"Failed to fetch media files: {media_response.text}", status=500)

    # Handle form submission (POST request)
    if request.method == "POST":
        # Prepare data for updating the collaborative_students/workshop
        update_url = f"{API_STUDIO_URL}updateapi/update/naac01_students_benefited_dc1/{id}"

        # Convert selected year range back to date format for storage
        year_range = request.POST.get('year')
        year_date = None
        
        if year_range and '-' in year_range:
            try:
                start_year = int(year_range.split('-')[0])
                # Create a date object (using June 1st as default)
                year_date = date(start_year, 6, 1).isoformat()
            except (ValueError, IndexError):
                year_date = collaborative_students.get('year')
        else:
            year_date = collaborative_students.get('year')

        payload = json.dumps({"data": {
            "department_name": request.POST.get('department_name', collaborative_students.get('department_name')),
            "staff_id": request.POST.get('staff_id', collaborative_students.get('staff_id')),
            "staff_name": request.POST.get('staff_name', collaborative_students.get('staff_name')),
            "activity_title": request.POST.get('activity_title', collaborative_students.get('activity_title')),
            "agency_name": request.POST.get('agency_name', collaborative_students.get('agency_name')),
            "phone_number": request.POST.get('phone_number', collaborative_students.get('phone_number')),
            "participant_name": request.POST.get('participant_name', collaborative_students.get('participant_name')),
            "year": year_date,  # Use the converted date
            "date_from": request.POST.get('date_from', collaborative_students.get('date_from')),
            "date_to": request.POST.get('date_to', collaborative_students.get('date_to')),
            "activity_name": request.POST.get('activity_name', collaborative_students.get('activity_name'))}})
        
        headers = {'Content-Type': 'application/json'}
        update_response = requests.put(update_url, headers=headers, data=payload)

        if update_response.status_code != 200:
            return HttpResponse(f"Failed to update collaborative_students details: {update_response.text}", status=500)

        # Handle media (file) uploads
        upload_errors = []

        for child in child_files:
            upload_id = child['psk_id']
            fields = ['PL', 'Certificate', 'RPP', 'DR', 'GTP', 'SA']

            for field in fields:
                uploaded_files = request.FILES.getlist(f'file_{upload_id}_{field}')

                if not uploaded_files:
                    continue  # Skip if no files are uploaded for this field

                for uploaded_file in uploaded_files:
                    # Validate file size and format
                    validate_file_size(uploaded_file)
                    validate_file_format(uploaded_file)

                    # Generate custom filename
                    current_year = datetime.now().year
                    staff_id = request.POST.get('staff_id')
                    custom_filename = f"{staff_id}_{field}_{current_year}_{uploaded_file.name}"

                    print(f"Generated filename for field {field}: {custom_filename}")  # Debug log

                    # Construct the upload URL and payload
                    upload_url = f"{API_STUDIO_URL}crudapp/upload/media/naac01_students_benefited_dc1_media/{upload_id}"
                    files = {'media': (custom_filename, uploaded_file, uploaded_file.content_type)}
                    headers = {
                        'api_name': 'naac01_students_benefited_dc1_media',
                        'psk_id': str(upload_id)
                    }
                    payload = {'parent_psk_id': id}

                    # Upload the file
                    upload_response = requests.put(upload_url, headers=headers, data=payload, files=files)

                if upload_response.status_code != 200:
                    upload_errors.append(f"Error uploading file for {upload_id}: {upload_response.text}")

        # Provide feedback to the user
        if upload_errors:
            for error in upload_errors:
                messages.error(request, error)
        else:
            messages.success(request, "Files uploaded successfully.")

        # Redirect to the collaborative_students view page after successful update
        return redirect('collaborative_students_list')

    # If not a POST request, render the update form
    return render(request, 'collaborative_students_templates/collaborative_students_update.html', {
        'collaborative_students': collaborative_students, 
        'child_files': child_files,
        'publication_year': publication_year  # Pass years to template
    })

# Delete collaborative_students
def collaborative_students_delete(request, id):
    # url = f"{API_STUDIO_URL}getapi/naac01_students_benefited_dc1/{id}"
    # response = requests.get(url)
    # collaborative_students = response.json()

    # if request.method == 'POST':
    delete_url = f"{API_STUDIO_URL}deleteapi/delete/naac01_students_benefited_dc1/{id}"
    payload = ""
    headers = {}
    delete_response = requests.request("DELETE", delete_url, headers=headers, data=payload)

    if delete_response.status_code == 200:
        return redirect('collaborative_students_list')
    else:
        return HttpResponse("Failed to delete participation: " + delete_response.text)

# return render(request, 'collaborative_students_delete.html', {'collaborative_students': collaborative_students})

# def export_collaborative_students_to_pdf(parents, children=None, selected_options=None):
#     """
#     Export Collaborative Students data to PDF (landscape), grouped by staff.
#     Each staff has a mini-title "Collaborative Students", then Staff info in bold labels, then table.
#     Attachments are clickable links.
#     """
#     from reportlab.lib.pagesizes import A4, landscape
#     from reportlab.lib import colors
#     from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
#     from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, PageBreak
#     from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
#     from reportlab.lib.units import inch
#     from django.http import HttpResponse

#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="collaborative_students.pdf"'

#     doc = SimpleDocTemplate(
#         response,
#         pagesize=landscape(A4),
#         topMargin=0.5*inch,
#         bottomMargin=0.5*inch,
#         leftMargin=0.5*inch,
#         rightMargin=0.5*inch
#     )

#     elements = []
#     styles = getSampleStyleSheet()

#     # Styles
#     title_style = ParagraphStyle('Title', parent=styles['Heading2'], alignment=TA_CENTER, spaceAfter=12)
#     left_style = ParagraphStyle('LeftStyle', parent=styles['Normal'], fontSize=9,
#                                 alignment=TA_LEFT, textColor=colors.HexColor('#2c3e50'))
#     right_style = ParagraphStyle('RightStyle', parent=styles['Normal'], fontSize=9,
#                                  alignment=TA_RIGHT, textColor=colors.HexColor('#2c3e50'))
#     table_header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontSize=8,
#                                         alignment=TA_CENTER, textColor=colors.white, fontName='Helvetica-Bold')
#     table_cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=7,
#                                       alignment=TA_LEFT, textColor=colors.HexColor('#2c3e50'), leading=9)
#     attachment_link_style = ParagraphStyle('AttachmentLink', parent=styles['Normal'], fontSize=6,
#                                            alignment=TA_LEFT, textColor=colors.HexColor('#1a5276'), leading=8)
#     no_data_style = ParagraphStyle('NoData', parent=styles['Normal'], fontSize=10,
#                                    alignment=TA_CENTER, textColor=colors.HexColor('#7f8c8d'),
#                                    fontStyle='italic', spaceBefore=12, spaceAfter=12)

#     # Check if there are any parents
#     if not parents:
#         elements.append(Paragraph("Collaborative Students", title_style))
#         elements.append(Paragraph("No data available", no_data_style))
#         doc.build(elements)
#         return response

#     # Sort parents by staff
#     parents_sorted = sorted(parents, key=lambda x: (x.get('staff_id',''), x.get('activity_title','')))
#     current_staff = None
#     table_data = []
#     col_widths = [1.0*inch, 1.2*inch, 1.2*inch, 1.5*inch, 1.2*inch, 1.0*inch, 1.2*inch, 0.7*inch, 0.8*inch, 0.8*inch, 1.2*inch, 1.5*inch]

#     for parent in parents_sorted:
#         staff_id = parent.get('staff_id', 'N/A')
#         staff_name = parent.get('staff_name', 'N/A')
#         department_name = parent.get('department_name', 'N/A')

#         # New staff: flush previous table + page break
#         if current_staff and current_staff != staff_id:
#             # Only add table if there's data (more than just headers)
#             if len(table_data) > 1:
#                 table = Table(table_data, colWidths=col_widths)
#                 table.setStyle(TableStyle([
#                     ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
#                     ('TEXTCOLOR', (0,0), (-1,0), colors.white),
#                     ('ALIGN',(0,0),(-1,-1),'CENTER'),
#                     ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
#                     ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
#                     ('FONTSIZE', (0,0), (-1,-1), 7),
#                     ('LEFTPADDING', (0,0), (-1,-1), 4),
#                     ('RIGHTPADDING', (0,0), (-1,-1), 4),
#                     ('BOTTOMPADDING', (0,0), (-1,-1), 3),
#                     ('TOPPADDING', (0,0), (-1,-1), 3),
#                 ]))
#                 elements.append(table)
#             else:
#                 # Only headers exist, meaning no data for this staff
#                 elements.append(Paragraph("No data available for this staff", no_data_style))
            
#             elements.append(PageBreak())
#             table_data = []

#         if current_staff != staff_id:
#             # Add mini-title for this staff
#             elements.append(Paragraph("Collaborative Students", title_style))
#             elements.append(Spacer(1, 6))

#             # Staff info with bold labels
#             left_paragraph = Paragraph(
#                 f"<b>Staff ID:</b> {staff_id}<br/><b>Staff Name:</b> {staff_name}", left_style)
#             right_paragraph = Paragraph(f"<b>Department:</b> {department_name}", right_style)

#             total_width = sum(col_widths)
#             info_table = Table([[left_paragraph, right_paragraph]], colWidths=[total_width*0.5]*2)
#             info_table.setStyle(TableStyle([
#                 ('VALIGN', (0, 0), (-1, -1), 'TOP'),
#                 ('ALIGN', (0,0), (0,0), 'LEFT'),
#                 ('ALIGN', (1,0), (1,0), 'RIGHT'),
#                 ('LEFTPADDING', (0,0), (-1,-1), 0),
#                 ('RIGHTPADDING', (0,0), (-1,-1), 0),
#                 ('TOPPADDING', (0,0), (-1,-1), 2),
#                 ('BOTTOMPADDING', (0,0), (-1,-1), 4),
#             ]))
#             elements.append(info_table)
#             elements.append(Spacer(1, 6))

#             # Table headers
#             headers = [
#                 'Activity Title', 'Agency Name', 'Phone Number', 'Participant Name', 'Year',
#                 'Date From', 'Date To', 'Activity Name', 'Attachments'
#             ]
#             table_data.append([Paragraph(h, table_header_style) for h in headers])
#             current_staff = staff_id

#         # Media attachments
#         media_files = parent.get('media_files', [])
#         if media_files:
#             media_text = []
#             for m in media_files:
#                 filename = m.get('file_name', 'Unknown')
#                 url = m.get('direct_api_url', '#')
#                 media_text.append(f'<link href="{url}" color="blue">{filename}</link>')
#             media_paragraph = Paragraph('<br/>'.join(media_text), attachment_link_style)
#         else:
#             media_paragraph = Paragraph("No attachments", table_cell_style)

#         row = [
#             Paragraph(str(parent.get('activity_title', '')), table_cell_style),
#             Paragraph(str(parent.get('agency_name', '')), table_cell_style),
#             Paragraph(str(parent.get('phone_number', '')), table_cell_style),
#             Paragraph(str(parent.get('participant_name', '')), table_cell_style),
#             Paragraph(str(parent.get('year', '')), table_cell_style),
#             Paragraph(str(parent.get('date_from', '')), table_cell_style),
#             Paragraph(str(parent.get('date_to', '')), table_cell_style),
#             Paragraph(str(parent.get('activity_name', '')), table_cell_style),
#             media_paragraph
#         ]
#         table_data.append(row)

#     # Add last staff table
#     if table_data:
#         # Check if there's actual data (more than just headers)
#         if len(table_data) > 1:
#             table = Table(table_data, colWidths=col_widths)
#             table.setStyle(TableStyle([
#                 ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
#                 ('TEXTCOLOR', (0,0), (-1,0), colors.white),
#                 ('ALIGN',(0,0),(-1,-1),'CENTER'),
#                 ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
#                 ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
#                 ('FONTSIZE', (0,0), (-1,-1), 7),
#                 ('LEFTPADDING', (0,0), (-1,-1), 4),
#                 ('RIGHTPADDING', (0,0), (-1,-1), 4),
#                 ('BOTTOMPADDING', (0,0), (-1,-1), 3),
#                 ('TOPPADDING', (0,0), (-1,-1), 3),
#             ]))
#             elements.append(table)
#         else:
#             # Only headers exist, meaning no data for this staff
#             elements.append(Paragraph("No data available for this staff", no_data_style))

#     doc.build(elements)
#     return response


def export_collaborative_students_to_pdf(parents, children=None, selected_options=None):
    """
    Export Collaborative Students data to PDF (portrait), grouped by staff.
    Each staff has a mini-title "Collaborative Students", then Staff info in bold labels, then table.
    Attachments are clickable links.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, PageBreak
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.lib.units import inch
    from django.http import HttpResponse

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="collaborative_students.pdf"'

    # Increased margins for better spacing
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,  # Changed to portrait
        topMargin=1.0*inch,
        bottomMargin=1.0*inch,
        leftMargin=0.75*inch,
        rightMargin=0.75*inch
    )

    elements = []
    styles = getSampleStyleSheet()

    # Styles
    title_style = ParagraphStyle('Title', parent=styles['Heading2'], alignment=TA_CENTER, spaceAfter=12)
    left_style = ParagraphStyle('LeftStyle', parent=styles['Normal'], fontSize=9,
                                alignment=TA_LEFT, textColor=colors.HexColor('#2c3e50'))
    right_style = ParagraphStyle('RightStyle', parent=styles['Normal'], fontSize=9,
                                 alignment=TA_RIGHT, textColor=colors.HexColor('#2c3e50'))
    table_header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontSize=8,
                                        alignment=TA_CENTER, textColor=colors.white, fontName='Helvetica-Bold')
    table_cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=7,
                                      alignment=TA_LEFT, textColor=colors.HexColor('#2c3e50'), leading=9)
    table_cell_center_style = ParagraphStyle('TableCellCenter', parent=styles['Normal'], fontSize=7,
                                            alignment=TA_CENTER, textColor=colors.HexColor('#2c3e50'), leading=9)
    attachment_link_style = ParagraphStyle('AttachmentLink', parent=styles['Normal'], fontSize=6,
                                           alignment=TA_LEFT, textColor=colors.HexColor('#1a5276'), leading=8)
    no_data_style = ParagraphStyle('NoData', parent=styles['Normal'], fontSize=10,
                                   alignment=TA_CENTER, textColor=colors.HexColor('#7f8c8d'),
                                   fontStyle='italic', spaceBefore=12, spaceAfter=12)

    # Check if there are any parents
    if not parents:
        elements.append(Paragraph("Collaborative Students", title_style))
        elements.append(Spacer(1, 0.5*inch))
        elements.append(Paragraph("No data available", no_data_style))
        doc.build(elements)
        return response

    # Sort parents by staff
    parents_sorted = sorted(parents, key=lambda x: (x.get('staff_id',''), x.get('activity_title','')))
    current_staff = None
    table_data = []
    
    # Adjusted column widths for portrait orientation
    col_widths = [
        1.0*inch,  # Activity Title
        1.3*inch,  # Agency Name
        1.0*inch,  # Participant Name
        0.7*inch,  # Year
        1.0*inch,  # Activity Name
        1.5*inch   # Attachments
    ]

    for parent in parents_sorted:
        staff_id = parent.get('staff_id', 'N/A')
        staff_name = parent.get('staff_name', 'N/A')
        department_name = parent.get('department_name', 'N/A')

        # New staff: flush previous table + page break
        if current_staff and current_staff != staff_id:
            # Only add table if there's data (more than just headers)
            if len(table_data) > 1:
                table = Table(table_data, colWidths=col_widths)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                    ('ALIGN', (0,0), (-1,0), 'CENTER'),  # Headers centered
                    ('ALIGN', (0,1), (0,-1), 'LEFT'),    # Activity Title left aligned
                    ('ALIGN', (1,1), (1,-1), 'LEFT'),    # Agency Name left aligned
                    ('ALIGN', (2,1), (2,-1), 'LEFT'),    # Participant Name left aligned
                    ('ALIGN', (3,1), (3,-1), 'CENTER'),  # Year centered
                    ('ALIGN', (4,1), (4,-1), 'LEFT'),    # Activity Name left aligned
                    ('ALIGN', (5,1), (5,-1), 'LEFT'),    # Attachments left aligned
                    ('VALIGN', (0,0), (-1,-1), 'TOP'),   # Top alignment for all cells
                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
                    ('FONTSIZE', (0,0), (-1,-1), 7),
                    ('LEFTPADDING', (0,0), (-1,-1), 4),
                    ('RIGHTPADDING', (0,0), (-1,-1), 4),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 3),
                    ('TOPPADDING', (0,0), (-1,-1), 3),
                ]))
                elements.append(table)
            else:
                # Only headers exist, meaning no data for this staff
                elements.append(Spacer(1, 0.3*inch))
                elements.append(Paragraph("No data available for this staff", no_data_style))
            
            elements.append(PageBreak())
            table_data = []

        if current_staff != staff_id:
            # Add mini-title for this staff
            elements.append(Paragraph("Collaborative Students", title_style))
            elements.append(Spacer(1, 0.3*inch))

            # Staff info with bold labels
            left_paragraph = Paragraph(
                f"<b>Staff ID:</b> {staff_id}<br/><b>Staff Name:</b> {staff_name}", left_style)
            right_paragraph = Paragraph(f"<b>Department:</b> {department_name}", right_style)

            total_width = sum(col_widths)
            info_table = Table([[left_paragraph, right_paragraph]], colWidths=[total_width*0.5]*2)
            info_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0,0), (0,0), 'LEFT'),
                ('ALIGN', (1,0), (1,0), 'RIGHT'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('TOPPADDING', (0,0), (-1,-1), 2),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 0.3*inch))

            # Table headers
            headers = [
                'Activity Title', 'Agency Name', 'Participant Name', 'Year', 'Activity Name', 'Attachments'
            ]
            table_data.append([Paragraph(h, table_header_style) for h in headers])
            current_staff = staff_id

        # Media attachments
        media_files = parent.get('media_files', [])
        if media_files:
            media_text = []
            for m in media_files:
                filename = m.get('file_name', 'Unknown')
                url = m.get('direct_api_url', '#')
                # Truncate long filenames for better display in portrait mode
                display_name = filename if len(filename) <= 25 else filename[:22] + "..."
                media_text.append(f'<link href="{url}" color="blue">{display_name}</link>')
            media_paragraph = Paragraph('<br/>'.join(media_text), attachment_link_style)
        else:
            media_paragraph = Paragraph("No attachments", table_cell_style)

        # Create rows with appropriate alignment styles
        row = [
            Paragraph(str(parent.get('activity_title', '')), table_cell_style),
            Paragraph(str(parent.get('agency_name', '')), table_cell_style),
            Paragraph(str(parent.get('participant_name', '')), table_cell_style),
            Paragraph(str(parent.get('year', '')), table_cell_center_style),
            Paragraph(str(parent.get('activity_name', '')), table_cell_style),
            media_paragraph
        ]
        table_data.append(row)

    # Add last staff table
    if table_data:
        # Check if there's actual data (more than just headers)
        if len(table_data) > 1:
            table = Table(table_data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('ALIGN', (0,0), (-1,0), 'CENTER'),  # Headers centered
                ('ALIGN', (0,1), (0,-1), 'LEFT'),    # Activity Title left aligned
                ('ALIGN', (1,1), (1,-1), 'LEFT'),    # Agency Name left aligned
                ('ALIGN', (2,1), (2,-1), 'LEFT'),    # Participant Name left aligned
                ('ALIGN', (3,1), (3,-1), 'CENTER'),  # Year centered
                ('ALIGN', (4,1), (4,-1), 'LEFT'),    # Activity Name left aligned
                ('ALIGN', (5,1), (5,-1), 'LEFT'),    # Attachments left aligned
                ('VALIGN', (0,0), (-1,-1), 'TOP'),   # Top alignment for all cells
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
                ('FONTSIZE', (0,0), (-1,-1), 7),
                ('LEFTPADDING', (0,0), (-1,-1), 4),
                ('RIGHTPADDING', (0,0), (-1,-1), 4),
                ('BOTTOMPADDING', (0,0), (-1,-1), 3),
                ('TOPPADDING', (0,0), (-1,-1), 3),
            ]))
            elements.append(table)
        else:
            # Only headers exist, meaning no data for this staff
            elements.append(Spacer(1, 0.3*inch))
            elements.append(Paragraph("No data available for this staff", no_data_style))

    doc.build(elements)
    return response

def export_collaborative_students_to_excel(parents, children=None):
    """
    Export Collaborative Students data to Excel from filtered parents,
    including clickable media file links based on upload field codes.
    """
    import io
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Collaborative Students"

    # Headers including media attachments
    headers = [
        "Staff ID", "Staff Name", "Department", "Activity Title",
        "Agency Name", "Phone Number", "Participant Name", "Year",
        "Date From", "Date To", "Activity Name",
        "Permission Letter", "Internship Certificate", "Objective & Outcome",
        "Detailed Report", "Geo Tagged Photos", "Faculty Attendance with signature"
    ]
    ws.append(headers)

    # Style headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Map upload field codes to friendly names
    field_mapping = {
        "PL": "Permission Letter",
        "Certificate": "Internship Certificate",
        "RPP": "Objective & Outcome",
        "DR": "Detailed Report",
        "GTP": "Geo Tagged Photos",
        "SA": "Faculty Attendance with signature"
    }

    # Map category to column
    media_columns = {
        "Permission Letter": 12,
        "Internship Certificate": 13,
        "Objective & Outcome": 14,
        "Detailed Report": 15,
        "Geo Tagged Photos": 16,
        "Faculty Attendance with signature": 17
    }

    # Populate rows
    for p in parents:
        # Initialize media categories - only use categories that exist in field_mapping
        media_mapping = {v: [] for v in field_mapping.values()}

        # Categorize each file based on upload code in filename
        for media in p.get("media_files", []):
            file_name = media.get("file_name", "Unknown")
            url = media.get("direct_api_url", "")
            if not url:
                continue

            # Use original filename for matching (case sensitive)
            matched = False

            # Detect the code from filename - try multiple patterns
            for code, category in field_mapping.items():
                # Try different patterns to match the code in filename (case sensitive)
                patterns = [
                    f"_{code}_",  # _PL_, _Certificate_, _RPP_
                    f"_{code}.",  # _PL.pdf, _Certificate.jpg, etc.
                    f"-{code}-",  # -PL-, -Certificate-, etc.
                    f"-{code}.",  # -PL.pdf, -Certificate.jpg, etc.
                    f"{code}_",   # PL_, Certificate_, RPP_ (at start)
                    f"_{code}",   # _PL, _Certificate, _RPP (at end)
                ]
                
                for pattern in patterns:
                    if pattern in file_name:
                        media_mapping[category].append((file_name, url))
                        matched = True
                        print(f"Matched '{file_name}' to '{category}' using pattern '{pattern}'")
                        break
                if matched:
                    break

            # If no match found with patterns, try case-insensitive partial matching
            if not matched:
                file_name_upper = file_name.upper()
                for code, category in field_mapping.items():
                    code_upper = code.upper()
                    # Check if code appears anywhere in filename (case insensitive)
                    if code_upper in file_name_upper:
                        media_mapping[category].append((file_name, url))
                        matched = True
                        print(f"Matched '{file_name}' to '{category}' using case-insensitive matching")
                        break

        # Create the base info row
        base_row = [
            p.get("staff_id", ""),
            p.get("staff_name", ""),
            p.get("department_name", ""),
            p.get("activity_title", ""),
            p.get("agency_name", ""),
            p.get("phone_number", ""),
            p.get("participant_name", ""),
            p.get("year", ""),
            p.get("date_from", ""),
            p.get("date_to", ""),
            p.get("activity_name", ""),
        ]

        # Append the base row first
        ws.append(base_row)
        row_idx = ws.max_row  # current row number

        # Now add media files to their respective columns
        for category, files in media_mapping.items():
            if category in media_columns:  # Double check the category exists
                col_idx = media_columns[category]
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if files:
                    # Show all filenames joined by line breaks
                    filenames = [f for f, _ in files]
                    cell.value = "\n".join(filenames)
                    
                    # Add hyperlink to first file if URL exists
                    if files[0][1]:
                        cell.hyperlink = files[0][1]
                        cell.font = Font(color="0000EE", underline="single")
                    else:
                        cell.font = Font()  # Regular font if no URL
                else:
                    cell.value = "-"
                    cell.font = Font()  # Regular font
                
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-fit column widths
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            if cell.value:
                # Count the maximum line length for multiline cells
                lines = str(cell.value).split('\n')
                max_line_length = max(len(line) for line in lines)
                max_length = max(max_length, max_line_length)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

    # Adjust row heights for multiline cells
    for row_idx in range(1, ws.max_row + 1):
        max_lines = 1
        for col in range(1, len(headers) + 1):
            cell_value = str(ws.cell(row=row_idx, column=col).value or "")
            line_count = cell_value.count("\n") + 1
            max_lines = max(max_lines, line_count)
        ws.row_dimensions[row_idx].height = min(max_lines * 15, 120)

    # Freeze header and enable filters
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Return as Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="collaborative_students.xlsx"'

    with io.BytesIO() as buffer:
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())

    return response
