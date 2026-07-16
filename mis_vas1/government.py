import json
import requests
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from MIS.functions import validate_file_format, validate_file_size
from datetime import datetime
from user_management.settings_views import *

API_STUDIO_URL = user_bundle_settings()


def get_government():
    url = "https://api.hcaschennai.edu.in/auth/token"

    payload = json.dumps({
        "secret_key": "C4ZoXbsAnHLjk1Xyz4QPT2eoiNx6K6fo"
    })
    headers = {
        'Content-Type': 'application/json'
    }

    response = requests.request("POST", url, headers=headers, data=payload)

    if response.status_code == 200:
        res_data = response.json()
        access_token = res_data.get('access_token')
        token_type = res_data.get('token_type')

        return access_token, token_type
    return None, None


def get_governmentdata(access_token, token_type):
    url = "https://api.hcaschennai.edu.in/sqlviews/api/v1/auth/get_response_data"

    payload = json.dumps({
        "psk_uid": "51a531b4-bd55-491c-861d-a8d7227b325b",
        "project": "hcas",
        "data": {}
    })
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'{token_type} {access_token}'
    }

    response = requests.request("POST", url, headers=headers, data=payload)

    if response.status_code == 200:
        return response.json()
    return []


def government_grants_create(request):
    error_message = None

    # Step 1: Get API token
    access_token, token_type = get_government()
    if not access_token or not token_type:
        error_message = 'Failed to get access token from API.'
        return render(request, 'government_templates/government_grants_create.html', {'error': error_message})

    # Step 2: Fetch career data
    government_data = get_governmentdata(access_token, token_type)
    if not government_data:
        error_message = 'Failed to fetch staff data.'
        return render(request, 'government_templates/government_grants_create.html', {'error': error_message})
    
    current_year = datetime.now().year
    publication_year = [f"{y}-{y+1}" for y in range(current_year, 2010, -1)]

    user = get_settings(request)
    username = user['username']
    # username = 'CS-T151'
    
    
    selected_faculty =None
    for faculty in government_data:
        if faculty['stf_id'] == username:
            selected_faculty = faculty
            break
        
    staff_name = selected_faculty.get('stf_name', '') if selected_faculty else ''
    department = selected_faculty.get('department', '') if selected_faculty else ''


    if request.method == 'POST':
        # Get form data from the POST request
        department_name = request.POST.get('department_name')
        staff_id = username
        staff_name = request.POST.get('staff_name')
        amount_sanctioned = request.POST.get('amount_sanctioned')
        department_of_principal_investigator = request.POST.get('department_of_principal_investigator')
        duration_end_date = request.POST.get('duration_end_date')
        name_of_funding_agency = request.POST.get('name_of_funding_agency')
        name_of_principal_investigator = request.POST.get('name_of_principal_investigator')
        project_name = request.POST.get('project_name')
        duration_start_date = request.POST.get('duration_start_date')
        government_type = request.POST.get('government_type')
        year_of_award = request.POST.get('year_of_award')
        
        
        for faculty in government_data:
            if faculty['stf_id'] == staff_id:
                selected_faculty = faculty
                break
            
        if selected_faculty:
            department_name = selected_faculty.get('department', '')
            staff_name = selected_faculty.get('stf_name', '')


        # API URL to send the form data to
        url = f"{API_STUDIO_URL}postapi/create/naac01_government_grants_dc1"
        payload = json.dumps({"data": {
            "staff_id": staff_id,
            "staff_name": staff_name,
            "amount_sanctioned": amount_sanctioned,
            "department_of_principal_investigator": department_of_principal_investigator,
            "duration_end_date": duration_end_date,
            "department_name": department_name,
            "name_of_funding_agency": name_of_funding_agency,
            "name_of_principal_investigator": name_of_principal_investigator,
            "project_name": project_name,
            "duration_start_date": duration_start_date,
            "government_type": government_type,
            "year_of_award": year_of_award
            }})
        headers = {'Content-Type': 'application/json'}

        # Make API call to create the collaborative students/workshop
        response = requests.request("POST", url, headers=headers, data=payload)

        if response.status_code == 200:
            # Successfully created the collaborative students/workshop
            file_data = response.json()
            psk_id = file_data.get('psk_id')  # Get psk_id from the response

            # Handle file uploads (optional step)
            upload_url = f"{API_STUDIO_URL}crudapp/upload/media/naac01_government_grants_dc1_media"
        uploaded_files = request.FILES.getlist('file')

        if not uploaded_files:
            messages.error(request, message="No files selected for upload.")
            return render(request, 'government_templates/extension_create.html')
        fields = ['GA', 'GSL', 'UC', 'DR']
        for field in fields:
            current_year = datetime.now().year
            staff_id = username
            for field, uploaded_file in zip(fields, uploaded_files):
                validate_file_size(uploaded_file)
                validate_file_format(uploaded_file)
                file_type = uploaded_file.content_type
                custom_filename = f"{staff_id}_{field}_{current_year}_{uploaded_file.name}"
                print(f"Generated filename: {custom_filename}")  # Print the filename for each iteration
                payload = {'parent_psk_id': psk_id}
                files = {'media': (custom_filename, uploaded_file, file_type)}
                upload_headers = {'api_name': 'naac01_government_grants_dc1_media'}

                # Make API call to upload the file
                upload_response = requests.post(upload_url, headers=upload_headers, data=payload, files=files)

                if upload_response.status_code != 200:
                    # File upload failed
                    messages.error(request,
                                    message=f"File upload failed for {uploaded_file.name}. Error: {upload_response.text}")
                    return redirect('government_grant_list')

                messages.success(request, message="Documents uploaded successfully.")
           # else:
               # messages.warning(request, message="No files selected for upload.")

            # Redirect to collaborative students view after successful creation
            return redirect('government_grants_list')

        else:
            # API call failed for creating collaborative students
            messages.error(request, message="Failed to create collaborative students. Please try again.")
            return render(request, 'government_templates/government_grants_create.html')

    else:
        # If the request is GET, render the collaborative students creation form
        return render(request, 'government_templates/government_grants_create.html', {'government_data': government_data, "username": username, "staff_name": staff_name, "department": department, 'publication_year': publication_year})


# View government_grants Details
def government_grants_view(request, id):
    url = f"{API_STUDIO_URL}getapi/naac01_government_grants_dc1/{id}"
    response = requests.get(url)

    if response.status_code == 200:
        government_grants_data = response.json()
        return render(request, "government_templates/government_grants_view.html", {'government_grants': government_grants_data})

    return HttpResponse(f"Error fetching Course details: {response.text}", status=500)


# # List All government_grantss
# def government_grants_list(request):
    # # URL to get government grants data
    # url = f"{API_STUDIO_URL}getapi/all_fields/naac01_government_grants_dc1/all"
    # response = requests.get(url)

    # # Get the current username (staff_id)
    # user = get_settings(request)  # Assuming this function retrieves user settings (including staff_id)
    # username = user['username']  # Adjust if necessary to fetch staff_id or username from session
    # # username = 'AC-NT012'

    # # If no username (staff_id) is found, return the same page with an empty list
    # if not username:
        # return render(request, 'government_templates/government_grants_list.html', {'government_grants': []})

    # # If the API call was successful, filter the government grants data based on staff_id (username)
    # if response.status_code == 200:
        # government_grants = response.json()
        
        # # Filter the government grants based on the staff_id (username)
        # filtered_government_grants = [grant for grant in government_grants if grant.get('staff_id') == username]
        
    # selected_staff_id = request.GET.get('staff_id')
    
    # if selected_staff_id:
        # filtered_government_grants = [grant for grant in government_grants if grant.get('staff_id') == selected_staff_id]
        
        # # If no data is found for the username, return the same page with an empty list
        # if not filtered_government_grants:
            # return render(request, 'government_templates/government_grants_list.html', {'government_grants': government_grants})
        
        # # Return the filtered data to the template
        # return render(request, 'government_templates/government_grants_list.html', {'government_grants': filtered_government_grants})
    
    # # If the API call fails, return an empty list
    # return render(request, 'government_templates/government_grants_list.html', {'government_grants': []})
    
    
def government_grants_list(request):
    url = f"{API_STUDIO_URL}getapi/naac01_government_grants_dc1/all"
    response = requests.get(url)
    
    if response.status_code == 200:
        government_grants = response.json()
        print("government_grants:", government_grants)
    else:
        return HttpResponse("API Call Is Not Working")

    user = get_settings(request)
    username = user.get('username')
    # username = 'CS-T151'
    filtered_government_grants = [grant for grant in government_grants if grant.get('staff_id') == username]
    print("filtered_government_grants:", filtered_government_grants)
    
    selected_staff_id = request.GET.get('staff_id')
    if selected_staff_id:
        filtered_government_grants = [grant for grant in government_grants if grant.get('staff_id') == selected_staff_id]
        
    selected_department = request.GET.get('department')
    if selected_department:
        filtered_government_grants = [grant for grant in government_grants if grant.get('department_name') == selected_department]
    #from_dashboard = request.GET.get('from') == 'dashboard' or 'admin_hod_dash' or 'admin_dash' or 'department_dashboard'

    if not filtered_government_grants:
        return render(request, 'government_templates/government_grants_list.html', {"government_grants": []})
    
    return render(request, 'government_templates/government_grants_list.html', {"government_grants": filtered_government_grants, 
    #'from_dashboard': from_dashboard
    })

# Update government_grants
def government_grants_update(request, id):
    current_year = datetime.now().year
    publication_year = [f"{y}-{y+1}" for y in range(current_year, 2010, -1)]

    # Fetch government_grants details
    url = f"{API_STUDIO_URL}getapi/naac01_government_grants_dc1/{id}"
    response = requests.get(url)

    if response.status_code == 200:
        government_grants = response.json()
    else:
        return HttpResponse(f"Error fetching government_grants details: {response.text}", status=500)

    # Fetch media (child files) associated with this government_grants
    media_url = f"{API_STUDIO_URL}crudapp/get/media/naac01_government_grants_dc1_media/parent/{id}"
    media_response = requests.get(media_url)

    if media_response.status_code == 200:
        child_files = media_response.json()
    else:
        return HttpResponse(f"Failed to fetch media files: {media_response.text}", status=500)

    # Handle form submission (POST request)
    if request.method == "POST":
        # Prepare data for updating the government_grants/workshop
        update_url = f"{API_STUDIO_URL}updateapi/update/naac01_government_grants_dc1/{id}"

        payload = json.dumps({"data": {
            "amount_sanctioned": request.POST.get('amount_sanctioned', government_grants.get('amount_sanctioned')),
            "department_code": request.POST.get('department_code', government_grants.get('department_code')),
            "department_of_principal_investigator": request.POST.get('department_of_principal_investigator',
                                                                     government_grants.get(
                                                                         'department_of_principal_investigator')),
            "duration_end_date": request.POST.get('duration_end_date', government_grants.get('duration_end_date')),
            "department_name": request.POST.get('department_name', government_grants.get('department_name')),
            "name_of_funding_agency": request.POST.get('name_of_funding_agency',
                                                       government_grants.get('name_of_funding_agency')),
            "name_of_principal_investigator": request.POST.get('name_of_principal_investigator',
                                                               government_grants.get('name_of_principal_investigator')),
            "project_name": request.POST.get('project_name', government_grants.get('project_name')),
            "duration_start_date": request.POST.get('duration_start_date',
                                                    government_grants.get('duration_start_date')),
            "government_type": request.POST.get('government_type', government_grants.get('government_type')),
            "year_of_award": request.POST.get('year_of_award', government_grants.get('year_of_award'))}})
        headers = {'Content-Type': 'application/json'}
        update_response = requests.put(update_url, headers=headers, data=payload)

        if update_response.status_code != 200:
            return HttpResponse(f"Failed to update government_grants details: {update_response.text}", status=500)

        # Handle media (file) uploads
        upload_errors = []

        for child in child_files:
            upload_id = child['psk_id']
            fields = ['GA', 'GSL', 'UC', 'DR']
            
            for field in fields:
                # Dynamically get the list of uploaded files for this specific upload_id and field
                uploaded_files = request.FILES.getlist(f'file_{upload_id}_{field.upper()}')

                if not uploaded_files:
                    continue  # No files to upload for this field

                for uploaded_file in uploaded_files:
                    # Validate file size and format
                    try:
                        validate_file_size(uploaded_file)
                        validate_file_format(uploaded_file)
                    except Exception as e:
                        messages.error(request, str(e))
                        continue  # Skip file upload if validation fails

                    # Generate custom filename
                    current_year = datetime.now().year  # Get the current year
                    staff_id = request.POST.get('staff_id', government_grants.get('staff_id'))
                    custom_filename = f"{staff_id}_{field}_{current_year}_{uploaded_file.name}"
                    

                    print(f"Generated filename for field {field}: {custom_filename}")  # Log the filename for debugging

                    upload_url = f"{API_STUDIO_URL}crudapp/upload/media/naac01_government_grants_dc1_media/{upload_id}"

                    # Prepare the file payload and headers for the upload request
                    files = {'media': (custom_filename, uploaded_file, uploaded_file.content_type)}
                    payload = {'parent_psk_id': id}
                    headers = {'api_name': 'naac01_government_grants_dc1_media', 'psk_id': str(upload_id)}

                    # Upload the file
                    upload_response = requests.put(upload_url, headers=headers, data=payload, files=files)

                if upload_response.status_code != 200:
                    upload_errors.append(f"Error uploading file for {upload_id}: {upload_response.text}")

        # Provide feedback to the user
        if upload_errors:
            for error in upload_errors:
                messages.error(request, error)
        else:
            messages.success(request, "Files uploaded successfully.")

        # Redirect to the government_grants view page after successful update
        return redirect('government_grants_list')

    # If not a POST request, render the update form
    return render(request, 'government_templates/government_grants_update.html',
                  {'government_grants': government_grants, 'child_files': child_files, 'publication_year': publication_year})


# Delete government_grants
def government_grants_delete(request, id):
    # url = f"{API_STUDIO_URL}getapi/naac01_government_grants_dc1/{id}"
    # response = requests.get(url)
    # government_grants = response.json()

    # if request.method == 'POST':
    delete_url = f"{API_STUDIO_URL}deleteapi/delete/naac01_government_grants_dc1/{id}"
    delete_response = requests.delete(delete_url)

    if delete_response.status_code == 200:
        return redirect('government_grants_list')
    else:
        return HttpResponse("Failed to delete participation: " + delete_response.text)

# return render(request, 'government_templates/government_grants_delete.html', {'government_grants': government_grants})


def export_governments_grants_to_excel(parents, children=None):
    """
    Export Governments Grants data to Excel from filtered parents,
    including clickable media file links based on upload field codes.
    """
    import io
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Governments Grants"

    # Headers including media attachments
    headers = [
        "Staff ID", "Staff Name", "Department Name",
        "Project Name", "Name of Funding Agency", "Government Type",
        "Amount Sanctioned", "Year of Award", "Duration Start Date",
        "Duration End Date", "Name of Principal Investigator",
        "Department of Principal Investigator",
        "Grant Application", "Grant Sanction Letter", 
        "Utilization Certificate", "Detailed Report"
    ]
    ws.append(headers)

    # Style headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Field mapping for file categorization
    field_mapping = {
        "GA": "Grant Application",
        "GSL": "Grant Sanction Letter",
        "UC": "Utilization Certificate",
        "DR": "Detailed Report"
    }

    # Map category to column
    media_columns = {
        "Grant Application": 13,
        "Grant Sanction Letter": 14,
        "Utilization Certificate": 15,
        "Detailed Report": 16
    }

    # Populate rows
    for p in parents:
        # Initialize media categories
        media_mapping = {v: [] for v in field_mapping.values()}

        # Categorize each file based on upload code in filename
        for media in p.get("media_files", []):
            file_name = media.get("file_name", "Unknown")
            url = media.get("direct_api_url", "")
            if not url:
                continue

            # Use original filename for matching
            matched = False

            # Detect the code from filename - try multiple patterns
            for code, category in field_mapping.items():
                # Try different patterns to match the code in filename
                patterns = [
                    f"_{code}_",  # _GA_, _GSL_, _UC_, etc.
                    f"_{code}.",  # _GA.pdf, _GSL.jpg, etc.
                    f"-{code}-",  # -GA-, -GSL-, etc.
                    f"-{code}.",  # -GA.pdf, -GSL.jpg, etc.
                    f"{code}_",   # GA_, GSL_, UC_, etc.
                    f"_{code}",   # _GA, _GSL, _UC, etc.
                ]
                
                for pattern in patterns:
                    if pattern in file_name:
                        media_mapping[category].append((file_name, url))
                        matched = True
                        print(f"Matched '{file_name}' to '{category}' using pattern '{pattern}'")
                        break
                if matched:
                    break

            # If no match found with underscore patterns, try case-insensitive matching
            if not matched:
                file_name_upper = file_name.upper()
                for code, category in field_mapping.items():
                    code_upper = code.upper()
                    # Check if code appears anywhere in filename (case insensitive)
                    if code_upper in file_name_upper:
                        # Additional check to avoid partial matches
                        patterns_upper = [
                            f"_{code_upper}_",
                            f"_{code_upper}.",
                            f"-{code_upper}-",
                            f"-{code_upper}.",
                            f"{code_upper}_",
                            f"_{code_upper}"
                        ]
                        if any(pattern in file_name_upper for pattern in patterns_upper):
                            media_mapping[category].append((file_name, url))
                            matched = True
                            print(f"Matched '{file_name}' to '{category}' using case-insensitive matching")
                            break

        # Create the base info row
        base_row = [
            p.get("staff_id", ""),
            p.get("staff_name", ""),
            p.get("department_name", ""),
            p.get("project_name", ""),
            p.get("name_of_funding_agency", ""),
            p.get("government_type", ""),
            p.get("amount_sanctioned", ""),
            p.get("year_of_award", ""),
            p.get("duration_start_date", ""),
            p.get("duration_end_date", ""),
            p.get("name_of_principal_investigator", ""),
            p.get("department_of_principal_investigator", ""),
        ]

        # Append the base row first
        ws.append(base_row)
        row_idx = ws.max_row  # current row number

        # Now add media files to their respective columns
        for category, files in media_mapping.items():
            if category in media_columns:
                col_idx = media_columns[category]
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if files:
                    # Show all filenames joined by line breaks
                    filenames = [f for f, _ in files]
                    cell.value = "\n".join(filenames)
                    
                    # Add hyperlink to first file if URL exists
                    if files[0][1]:
                        cell.hyperlink = files[0][1]
                        cell.font = Font(color="0000EE", underline="single")
                    else:
                        cell.font = Font()  # Regular font if no URL
                else:
                    cell.value = "-"
                    cell.font = Font()  # Regular font
                
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-fit column widths
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            if cell.value:
                # Count the maximum line length for multiline cells
                lines = str(cell.value).split('\n')
                max_line_length = max(len(line) for line in lines)
                max_length = max(max_length, max_line_length)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

    # Adjust row heights for multiline cells
    for row_idx in range(1, ws.max_row + 1):
        max_lines = 1
        for col in range(1, len(headers) + 1):
            cell_value = str(ws.cell(row=row_idx, column=col).value or "")
            line_count = cell_value.count("\n") + 1
            max_lines = max(max_lines, line_count)
        ws.row_dimensions[row_idx].height = min(max_lines * 15, 120)

    # Freeze header and enable filters
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Return as Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="governments_grants.xlsx"'

    with io.BytesIO() as buffer:
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())

    return response


def export_governments_grants_to_pdf(parents, children=None, selected_options=None):
    """
    Export Governments Grants data to PDF (portrait), grouped by staff.
    Each staff has a mini-title "Governments Grants", then Staff info in bold labels, then table.
    Attachments are clickable links.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, PageBreak
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.lib.units import inch
    from django.http import HttpResponse

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="governments_grants.pdf"'

    # Increased margins for better spacing
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,  # Changed to portrait
        topMargin=1.0*inch,
        bottomMargin=1.0*inch,
        leftMargin=0.75*inch,
        rightMargin=0.75*inch
    )

    elements = []
    styles = getSampleStyleSheet()

    # Styles
    title_style = ParagraphStyle('Title', parent=styles['Heading2'], alignment=TA_CENTER, spaceAfter=12)
    left_style = ParagraphStyle('LeftStyle', parent=styles['Normal'], fontSize=9,
                                alignment=TA_LEFT, textColor=colors.HexColor('#2c3e50'))
    right_style = ParagraphStyle('RightStyle', parent=styles['Normal'], fontSize=9,
                                 alignment=TA_RIGHT, textColor=colors.HexColor('#2c3e50'))
    table_header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontSize=8,
                                        alignment=TA_CENTER, textColor=colors.white, fontName='Helvetica-Bold')
    table_cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=7,
                                      alignment=TA_LEFT, textColor=colors.HexColor('#2c3e50'), leading=9)
    table_cell_center_style = ParagraphStyle('TableCellCenter', parent=styles['Normal'], fontSize=7,
                                            alignment=TA_CENTER, textColor=colors.HexColor('#2c3e50'), leading=9)
    attachment_link_style = ParagraphStyle('AttachmentLink', parent=styles['Normal'], fontSize=6,
                                           alignment=TA_LEFT, textColor=colors.HexColor('#1a5276'), leading=8)
    no_data_style = ParagraphStyle('NoData', parent=styles['Normal'], fontSize=10,
                                   alignment=TA_CENTER, textColor=colors.HexColor('#7f8c8d'),
                                   fontStyle='italic', spaceBefore=12, spaceAfter=12)

    # Field mapping for file categorization
    field_mapping = {
        "GA": "Grant Application",
        "GSL": "Grant Sanction Letter",
        "UC": "Utilization Certificate",
        "DR": "Detailed Report"
    }

    # Check if there are any parents
    if not parents:
        elements.append(Paragraph("Governments Grants", title_style))
        elements.append(Spacer(1, 0.5*inch))
        elements.append(Paragraph("No data available", no_data_style))
        doc.build(elements)
        return response

    # Sort parents by staff
    parents_sorted = sorted(parents, key=lambda x: (x.get('staff_id',''), x.get('project_name','')))
    current_staff = None
    table_data = []
    
    # Adjusted column widths for portrait orientation
    col_widths = [
        # 1.5*inch,  # Project Name
        1.8*inch,  # Funding Agency
        1.0*inch,  # Government Type
        0.8*inch,  # Amount
        0.8*inch,  # Year
        # 0.7*inch,  # Start Date
        # 0.7*inch,  # End Date
        1.5*inch,  # Principal Investigator
        # 1.0*inch,  # PI Department
        1.6*inch   # Attachments
    ]

    for parent in parents_sorted:
        staff_id = parent.get('staff_id', 'N/A')
        staff_name = parent.get('staff_name', 'N/A')
        department_name = parent.get('department_name', 'N/A')

        # New staff: flush previous table + page break
        if current_staff and current_staff != staff_id:
            # Only add table if there's data (more than just headers)
            if len(table_data) > 1:
                table = Table(table_data, colWidths=col_widths)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                    ('ALIGN', (0,0), (-1,0), 'CENTER'),  # Headers centered
                    ('ALIGN', (0,1), (0,-1), 'LEFT'),    # Project Name left aligned
                    ('ALIGN', (1,1), (1,-1), 'LEFT'),    # Funding Agency left aligned
                    ('ALIGN', (2,1), (2,-1), 'LEFT'),    # Government Type left aligned
                    ('ALIGN', (3,1), (3,-1), 'RIGHT'),   # Amount right aligned
                    ('ALIGN', (4,1), (4,-1), 'CENTER'),  # Year centered
                    ('ALIGN', (5,1), (5,-1), 'CENTER'),  # Start Date centered
                    ('ALIGN', (6,1), (6,-1), 'CENTER'),  # End Date centered
                    ('ALIGN', (7,1), (7,-1), 'LEFT'),    # Principal Investigator left aligned
                    ('ALIGN', (8,1), (8,-1), 'LEFT'),    # PI Department left aligned
                    ('ALIGN', (9,1), (9,-1), 'LEFT'),    # Attachments left aligned
                    ('VALIGN', (0,0), (-1,-1), 'TOP'),   # Top alignment for all cells
                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
                    ('FONTSIZE', (0,0), (-1,-1), 7),
                    ('LEFTPADDING', (0,0), (-1,-1), 4),
                    ('RIGHTPADDING', (0,0), (-1,-1), 4),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 3),
                    ('TOPPADDING', (0,0), (-1,-1), 3),
                ]))
                elements.append(table)
            else:
                # Only headers exist, meaning no data for this staff
                elements.append(Spacer(1, 0.3*inch))
                elements.append(Paragraph("No data available for this staff", no_data_style))
            
            elements.append(PageBreak())
            table_data = []

        if current_staff != staff_id:
            # Add mini-title for this staff
            elements.append(Paragraph("Governments Grants", title_style))
            elements.append(Spacer(1, 0.3*inch))

            # Staff info with bold labels
            left_paragraph = Paragraph(
                f"<b>Staff ID:</b> {staff_id}<br/><b>Staff Name:</b> {staff_name}", left_style)
            right_paragraph = Paragraph(f"<b>Department:</b> {department_name}", right_style)

            total_width = sum(col_widths)
            info_table = Table([[left_paragraph, right_paragraph]], colWidths=[total_width*0.5]*2)
            info_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0,0), (0,0), 'LEFT'),
                ('ALIGN', (1,0), (1,0), 'RIGHT'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('TOPPADDING', (0,0), (-1,-1), 2),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 0.3*inch))

            # Table headers
            headers = [
                'Funding Agency', 'Govt Type', 'Amount', 
                'Year', 'Principal Investigator', 'Attachments'
            ]
            table_data.append([Paragraph(h, table_header_style) for h in headers])
            current_staff = staff_id

        # Process media attachments with categorization
        media_files = parent.get('media_files', [])
        categorized_files = {v: [] for v in field_mapping.values()}
        
        # Categorize media files
        for media in media_files:
            file_name = media.get('file_name', 'Unknown')
            url = media.get('direct_api_url', '#')
            
            matched = False
            for code, category in field_mapping.items():
                patterns = [
                    f"_{code}_", f"_{code}.", f"-{code}-", f"-{code}.",
                    f"{code}_", f"_{code}"
                ]
                for pattern in patterns:
                    if pattern in file_name:
                        categorized_files[category].append((file_name, url))
                        matched = True
                        break
                if matched:
                    break
                    
            if not matched:
                file_name_upper = file_name.upper()
                for code, category in field_mapping.items():
                    code_upper = code.upper()
                    patterns_upper = [
                        f"_{code_upper}_", f"_{code_upper}.",
                        f"-{code_upper}-", f"-{code_upper}.",
                        f"{code_upper}_", f"_{code_upper}"
                    ]
                    if any(pattern in file_name_upper for pattern in patterns_upper):
                        categorized_files[category].append((file_name, url))
                        break

        # Build attachments text with categorized files
        if any(categorized_files.values()):
            media_text = []
            for category, files in categorized_files.items():
                if files:
                    for file_name, url in files:
                        # Truncate long filenames for better display
                        display_name = file_name if len(file_name) <= 15 else file_name[:25] + "..."
                        media_text.append(f'<link href="{url}" color="blue">{display_name}</link>')
            
            if media_text:
                media_paragraph = Paragraph('<br/>'.join(media_text), attachment_link_style)
            else:
                media_paragraph = Paragraph("No attachments", table_cell_style)
        else:
            media_paragraph = Paragraph("No attachments", table_cell_style)

        # Create rows with appropriate alignment styles
        row = [
            # Paragraph(str(parent.get('project_name', '')), table_cell_style),
            Paragraph(str(parent.get('name_of_funding_agency', '')), table_cell_style),
            Paragraph(str(parent.get('government_type', '')), table_cell_style),
            Paragraph(str(parent.get('amount_sanctioned', '')), table_cell_style),
            Paragraph(str(parent.get('year_of_award', '')), table_cell_center_style),
            # Paragraph(str(parent.get('duration_start_date', '')), table_cell_center_style),
            # Paragraph(str(parent.get('duration_end_date', '')), table_cell_center_style),
            Paragraph(str(parent.get('name_of_principal_investigator', '')), table_cell_style),
            # Paragraph(str(parent.get('department_of_principal_investigator', '')), table_cell_style),
            media_paragraph
        ]
        table_data.append(row)

    # Add last staff table
    if table_data:
        # Check if there's actual data (more than just headers)
        if len(table_data) > 1:
            table = Table(table_data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('ALIGN', (0,0), (-1,0), 'CENTER'),  # Headers centered
                ('ALIGN', (0,1), (0,-1), 'LEFT'),    # Project Name left aligned
                ('ALIGN', (1,1), (1,-1), 'LEFT'),    # Funding Agency left aligned
                ('ALIGN', (2,1), (2,-1), 'LEFT'),    # Government Type left aligned
                ('ALIGN', (3,1), (3,-1), 'RIGHT'),   # Amount right aligned
                ('ALIGN', (4,1), (4,-1), 'CENTER'),  # Year centered
                ('ALIGN', (5,1), (5,-1), 'CENTER'),  # Start Date centered
                ('ALIGN', (6,1), (6,-1), 'CENTER'),  # End Date centered
                ('ALIGN', (7,1), (7,-1), 'LEFT'),    # Principal Investigator left aligned
                ('ALIGN', (8,1), (8,-1), 'LEFT'),    # PI Department left aligned
                ('ALIGN', (9,1), (9,-1), 'LEFT'),    # Attachments left aligned
                ('VALIGN', (0,0), (-1,-1), 'TOP'),   # Top alignment for all cells
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
                ('FONTSIZE', (0,0), (-1,-1), 7),
                ('LEFTPADDING', (0,0), (-1,-1), 4),
                ('RIGHTPADDING', (0,0), (-1,-1), 4),
                ('BOTTOMPADDING', (0,0), (-1,-1), 3),
                ('TOPPADDING', (0,0), (-1,-1), 3),
            ]))
            elements.append(table)
        else:
            # Only headers exist, meaning no data for this staff
            elements.append(Spacer(1, 0.3*inch))
            elements.append(Paragraph("No data available for this staff", no_data_style))

    doc.build(elements)
    return response