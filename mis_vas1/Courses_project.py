import random
import string
import json
import requests
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from MIS.functions import validate_file_format, validate_file_size
from datetime import datetime
from user_management.settings_views import *
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak

API_STUDIO_URL = user_bundle_settings()

def course_key():
    url = "https://api.hcaschennai.edu.in/auth/token"
    payload = json.dumps({
        "secret_key": "C4ZoXbsAnHLjk1Xyz4QPT2eoiNx6K6fo"
    })
    headers = {'Content-Type': 'application/json'}

    response = requests.post(url, headers=headers, data=payload)
    if response.status_code == 200:
        res_data = response.json()
        return res_data.get('access_token'), res_data.get('token_type')
    return None, None

def get_course_data(access_token, token_type):
    url = "https://api.hcaschennai.edu.in/sqlviews/api/v1/auth/get_response_data"
    payload = json.dumps({
        "psk_uid": "51a531b4-bd55-491c-861d-a8d7227b325b",
        "project": "hcas",
        "data": {}
    })
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'{token_type} {access_token}'
    }

    response = requests.post(url, headers=headers, data=payload)
    if response.status_code == 200:
        return response.json()
    return []

def project_create(request):
    error_message = None
    access_token, token_type = course_key()
    if not access_token or not token_type:
        error_message = 'Failed to get access token from API.'
        return render(request, 'Courses_project_templates/project_create.html', {'error': error_message})

    course_data = get_course_data(access_token, token_type)
    if not course_data:
        error_message = 'Failed to fetch course data.'
        return render(request, 'Courses_project_templates/project_create.html', {'error': error_message})
    
    user = get_settings(request)
    username = user['username']
    # username = 'AC-NT012'
    
    
    selected_faculty = None
    for faculty in course_data:
        if faculty['stf_id'] == username:
            selected_faculty = faculty
            break
        
    staff_name = selected_faculty.get('stf_name') if selected_faculty else ''
    department_name = selected_faculty.get('department') if selected_faculty else ''
    

    if request.method == 'POST':
        staff_name = request.POST.get('staff_name')
        staff_id = username
        department_name = request.POST.get('department_name')
        selected_options = request.POST.getlist('items')
        
        for faculty in course_data:
            if faculty['stf_id'] == staff_id:
                selected_faculty = faculty
                break
            
        staff_name = selected_faculty.get('stf_name') if selected_faculty else ''
        department_name = selected_faculty.get('department') if selected_faculty else ''

        
        if selected_options:
            url = f"{API_STUDIO_URL}postapi/create/naac01_project_work_dc1"
            payload = json.dumps({"data": {"staff_name": staff_name,"staff_id": staff_id,"department_name": department_name,"course_name": ', '.join(selected_options)}})
            headers = {'Content-Type': 'application/json'}
            response = requests.post(url, headers=headers, data=payload)
            if response.status_code == 200:
                psk_id = response.json().get('psk_id')
                return redirect('project_view', id=psk_id)
        else:
            return HttpResponse("Failure: " + response.text)
    return render(request, 'Courses_project_templates/project_create.html', {'course_data': course_data, "username":username, "staff_name":staff_name, "department":department_name})

def cleanup_orphan_participations():
    """Function to delete all parent records without children"""
    parent_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_project_work_dc1/all"
    parent_res = requests.get(parent_url)
    
    if parent_res.status_code != 200:
        return
    
    all_parents = parent_res.json()
    
    for parent in all_parents:
        parent_id = parent.get("psk_id")
        if not parent_id:
            continue
        
        # Check for children
        child_url = f"{API_STUDIO_URL}getapi/naac01_project_work_dc2"
        payload = json.dumps({
            "queries": [{"field": "transaction_id", "value": parent_id, "operation": "equal"}],
            "search_type": "all"
        })
        headers = {'Content-Type': 'application/json'}
        children_response = requests.get(child_url, headers=headers, data=payload)
        
        children = children_response.json() if children_response.status_code == 200 else []
        
        if not children:  # No children found, delete parent
            delete_url = f"{API_STUDIO_URL}deleteapi/delete/naac01_project_work_dc1/{parent_id}"
            requests.delete(delete_url)
            print(f"Deleted orphan parent record: {parent_id}")

def project_view(request, id):
    # URL to fetch the course data (parent)
    url = f"{API_STUDIO_URL}getapi/naac01_project_work_dc1/{id}"
    response = requests.get(url)

    if response.status_code == 200:
        course_data = response.json()
        course_type = course_data.get('course_name', '')
        selected_options = course_type.split(', ') if course_type else []
        parent_id = course_data.get('psk_id')

        # Get children data based on the parent ID
        children_url = f"{API_STUDIO_URL}getapi/naac01_project_work_dc2"
        payload = json.dumps({"queries": [{"field": "transaction_id", "value": parent_id, "operation": "equal"}], "search_type": "all"})
        children_response = requests.get(children_url, headers={'Content-Type': 'application/json'}, data=payload)

        # Parse children data
        children_data = children_response.json() if children_response.status_code == 200 else []

        # POST request handling: check if form is submitted
        if request.method == 'POST':
            # Validate if the selected parent options have corresponding children
            missing_children = []

            # Mapping of fields for validation based on the selected options
            field_map = {
                'project': ['pro_department_code', 'pro_department_name', 'pro_program_code', 'pro_name_of_the_course', 'pro_course_code', 'pro_year_of_offering', 'pro_name_of_students_studied'],
                'field': ['field_department_code', 'field_department_name', 'field_program_code', 'field_name_of_the_course', 'field_course_code', 'field_year_of_offering', 'field_name_of_students_studied'],
                'internship': ['intern_department_code', 'intern_department_name', 'intern_program_code', 'intern_name_of_the_course', 'intern_course_code', 'intern_year_of_offering', 'intern_name_of_students_studied']
            }

            # Check if each selected option has corresponding child data
            for option in selected_options:
                option_key = option.lower()
                has_children = False

                # Iterate over the children data to find if there's a child for the current option
                for child in children_data:
                    # Check for each field in the field_map based on the option
                    required_fields = field_map.get(option_key, [])
                    if all(child.get(field) for field in required_fields):
                        has_children = True
                        break

                if not has_children:
                    missing_children.append(option)

            if missing_children:
                # If there are missing children, show an error message and return to the page
                error_message = f"The following selected {', '.join(missing_children)} do not have any data"
                messages.error(request, error_message)
                return render(request, "Courses_project_templates/project_view.html", {'project': course_data,'selected_options': selected_options,'children': children_data, 'missing_children': missing_children})

            # If validation passes, redirect to the project list page
            messages.success(request, "Form submitted successfully!")  # Success message
            return redirect('project_list')  # Assuming 'project_list' is the name of the URL for the project list page

        # If the request method is not POST, render the page normally
        return render(request, "Courses_project_templates/project_view.html", {'project': course_data,'selected_options': selected_options,'children': children_data})

    return HttpResponse(f"Error fetching course details: {response.text}", status=500)

def project_list(request):

    cleanup_orphan_participations()
    
    # URL to get project data
    url = f"{API_STUDIO_URL}getapi/all_fields/naac01_project_work_dc1/all"
    response = requests.get(url)

    if response.status_code == 200:
            projects = response.json()
            print("projects:", projects)
    # Get the current username (staff_id)
    user = get_settings(request)  # Assuming this function retrieves user settings (including staff_id)
    username = user.get('username')  # Adjust if necessary to fetch staff_id or username from session
    # username = 'AC-NT012'

    filtered_projects = [project for project in projects if project.get('staff_id') == username]
    
    # If no username (staff_id) is found, return the same page with an empty list
    # if not username:
    #     return render(request, 'Courses_project_templates/project_list.html', {'projects': []})

    # If the API call was successful, filter the project data based on staff_id (username)
    
        
        # Filter the projects based on the staff_id (username)
    
        
    #from_dashboard = request.GET.get('from') == 'dashboard' or 'admin_hod_dash' or 'admin_dash' or 'department_dashboard'
    selected_staff_id = request.GET.get('staff_id')
    
    if selected_staff_id:
        filtered_projects = [project for project in projects if project.get('staff_id') == selected_staff_id]
        
    selected_department = request.GET.get('department')
    if selected_department:
        filtered_projects = [project for project in projects if project.get('department_name') == selected_department]


        
        # If no data is found for the username, return the same page with an empty list
    if not filtered_projects:
        return render(request, 'Courses_project_templates/project_list.html', {'projects': []})
        
        # Return the filtered data to the template
        # return render(request, 'Courses_project_templates/project_list.html', {'projects': filtered_projects})
    
    # If the API call fails, return an empty list
    return render(request, 'Courses_project_templates/project_list.html', {'projects': filtered_projects, 
    #'from_dashboard': from_dashboard
    })

def project_update(request, id):
    # Step 1: Fetch the project data
    url = f"{API_STUDIO_URL}getapi/naac01_project_work_dc1/{id}"
    response = requests.get(url)

    if response.status_code == 200:
        my_project = response.json()
        selected_options = my_project.get('course_name', '').split(', ') if my_project.get('course_name') else []
    else:
        return HttpResponse(f"Error fetching project details: {response.text}", status=500)

    # Step 2: Fetch the children data (courses associated with the project)
    children_url = f"{API_STUDIO_URL}getapi/naac01_project_work_dc2"
    payload = json.dumps({"queries": [{"field": "transaction_id", "value": id, "operation": "equal"}], "search_type": "all"})
    headers = {'Content-Type': 'application/json'}
    children_response = requests.get(children_url, headers=headers, data=payload)

    if children_response.status_code == 200:
        my_child = children_response.json()
    else:
        return HttpResponse(f"Error fetching children details: {children_response.text}", status=500)

    # Step 3: Handle POST request when updating the selected courses
    if request.method == "POST":
        selected_options_from_form = request.POST.getlist('items')

        # Prepare a list to track options that cannot be deselected
        options_with_children = []

        # Mapping of fields for validation based on the selected options
        field_map = {
            'project': ['pro_department_code', 'pro_department_name', 'pro_program_code', 'pro_name_of_the_course', 'pro_course_code', 'pro_year_of_offering', 'pro_name_of_students_studied'],
            'field': ['field_department_code', 'field_department_name', 'field_program_code', 'field_name_of_the_course', 'field_course_code', 'field_year_of_offering', 'field_name_of_students_studied'],
            'internship': ['intern_department_code', 'intern_department_name', 'intern_program_code', 'intern_name_of_the_course', 'intern_course_code', 'intern_year_of_offering', 'intern_name_of_students_studied']
        }

        # Check if each selected option has corresponding child data
        for option in selected_options:
            option_key = option.lower()
            has_children = False

            # Iterate over the children data to find if there's a child for the current option
            for child in my_child:
                # Check for each field in the field_map based on the option
                required_fields = field_map.get(option_key, [])
                if all(child.get(field) for field in required_fields):
                    has_children = True
                    break

            if has_children and option not in selected_options_from_form:
                options_with_children.append(option)

        # Step 5: Show an error if invalid deselections are found
        if options_with_children:
            error_message = f"Cannot deselect the following options because they have associated children: {', '.join(options_with_children)}"
            messages.error(request, error_message)
            return render(request, 'Courses_project_templates/project_update.html', {
                'selected_options': selected_options,
                'project': my_project
            })

        # Step 6: Proceed with updating the selected options
        updated_selected_options = ', '.join(selected_options_from_form)  # Final list of selected options
        update_url = f"{API_STUDIO_URL}updateapi/update/naac01_project_work_dc1/{id}"
        payload = json.dumps({"data": {"course_name": updated_selected_options}})
        update_response = requests.put(update_url, headers=headers, data=payload)

        if update_response.status_code == 200:
            messages.success(request, "Project updated successfully.")
            return redirect('project_view', id=id)
        else:
            messages.error(request, "Failed to update project")
            return render(request, 'Courses_project_templates/project_update.html', {
                'selected_options': selected_options,
                'project': my_project
            })

    # Render the project update page for GET requests
    return render(request, 'Courses_project_templates/project_update.html', {
        'selected_options': selected_options,
        'project': my_project
    })

def project_delete(request, id):
    delete_url = f"{API_STUDIO_URL}deleteapi/delete/naac01_project_work_dc1/{id}"
    delete_response = requests.delete(delete_url)

    if delete_response.status_code == 200:
        return redirect('project_list')
    else:
        return HttpResponse(f"Failed to delete project: {delete_response.text}")

def project_child_create(request, id, val: str):
    # Helper function to generate a random alphanumeric code
    def generate_random_code(length=8):
        """Generate a random alphanumeric string of a specified length."""
        characters = string.ascii_letters + string.digits
        return ''.join(random.choice(characters) for _ in range(length))

    
    current_year = datetime.now().year
    year_of_offering = [f"{y}-{y+1}" for y in range(current_year, 2010, -1)]
    
    # Fetch course details
    url = f"{API_STUDIO_URL}getapi/naac01_project_work_dc1/{id}"
    course_parent_response = requests.get(url)

    if course_parent_response.status_code == 200:
        course_parent = course_parent_response.json()
        transaction_id = course_parent.get('psk_id')
        staff_id = course_parent.get('staff_id')
        selected_options = course_parent.get('course_name', '').split(', ') if course_parent.get('course_name') else []
    else:
        return HttpResponse("Error fetching participation details: " + course_parent_response.text)

    child_data = []
    project_fields = ['Project', 'Permission', 'Compilation', 'DRP']
    field_fields = ['PL', 'Obj_outcome', 'AR', 'DR']
    intern_fields = ['AL', 'ES', 'CL', 'DR']

    if request.method == 'POST':
        # Fetch POST data and generate random program codes if fields are empty
        pro_program_code = request.POST.get('pro_program_code', '')
        field_program_code = request.POST.get('field_program_code', '')
        intern_program_code = request.POST.get('intern_program_code', '')
        pro_course_code = request.POST.get('pro_course_code', '')
        field_course_code = request.POST.get('field_course_code', '')
        intern_course_code = request.POST.get('intern_course_code', '')

        # Generate random codes for missing fields
        pro_program_code = pro_program_code or generate_random_code()
        field_program_code = field_program_code or generate_random_code()
        intern_program_code = intern_program_code or generate_random_code()
        # pro_course_code = pro_course_code or generate_random_code()
        # field_course_code = field_course_code or generate_random_code()
        # intern_course_code = intern_course_code or generate_random_code()

        payload = {'pro_department_code': request.POST.get('pro_department_code'),'pro_department_name': request.POST.get('pro_department_name'),'pro_program_code': pro_program_code,'pro_name_of_the_course': request.POST.get('pro_name_of_the_course'),'pro_course_code': pro_course_code,'pro_year_of_offering': request.POST.get('pro_year_of_offering'),'pro_name_of_students_studied': request.POST.get('pro_name_of_students_studied'),'field_department_code': request.POST.get('field_department_code'),'field_department_name': request.POST.get('field_department_name'),'field_program_code': field_program_code,'field_name_of_the_course': request.POST.get('field_name_of_the_course'),'field_course_code': field_course_code,'field_year_of_offering': request.POST.get('field_year_of_offering'),'field_name_of_students_studied': request.POST.get('field_name_of_students_studied'),'intern_department_code': request.POST.get('intern_department_code'),'intern_department_name': request.POST.get('intern_department_name'),'intern_program_code': intern_program_code,'intern_name_of_the_course': request.POST.get('intern_name_of_the_course'),'intern_course_code': intern_course_code,'intern_year_of_offering': request.POST.get('intern_year_of_offering'),'intern_name_of_students_studied': request.POST.get('intern_name_of_students_studied'),'transaction_id': transaction_id}
        headers = {'Content-Type': 'application/json'}
        child_url = f"{API_STUDIO_URL}postapi/create/naac01_project_work_dc2"
        response = requests.post(child_url, headers=headers, data=json.dumps({"data": payload}))

        if response.status_code != 200:
            return HttpResponse("Failed to create participation child: " + response.text)

        child_data = response.json()
        child_id = child_data.get('psk_id')

        # Handle file uploads
        media_url = f"{API_STUDIO_URL}crudapp/upload/media/naac01_project_work_dc2_media/"
        uploaded_files = request.FILES.getlist('file')
        upload_errors = []
        current_year = datetime.now().year

        field_map = {'project': project_fields,'field': field_fields,'internship': intern_fields}

        fields = field_map.get(val, [])

        for idx, uploaded_file in enumerate(uploaded_files):
            validate_file_size(uploaded_file)
            validate_file_format(uploaded_file)
            file_type = uploaded_file.content_type

            field_name = fields[idx] if idx < len(fields) else "Unknown_Field"
            custom_filename = f"{staff_id}_{field_name.replace(' ', '_')}_{current_year}_{uploaded_file.name}"

            payload = {'parent_psk_id': child_id}
            files = {'media': (custom_filename, uploaded_file, file_type)}

            upload_headers = {'api_name': 'naac01_project_work_dc2_media'}
            upload_response = requests.post(media_url, headers=upload_headers, data=payload, files=files)

            if upload_response.status_code != 200:
                upload_errors.append(f"File uploading failed for {uploaded_file.name}. Error: {upload_response.text}")

        if upload_errors:
            for error in upload_errors:
                messages.error(request, message=error)
        else:
            messages.success(request, message="Documents uploaded successfully.")

        return redirect('project_view', id=id)

    return render(request, 'Courses_project_templates/project_child_create.html', {'selected_options': selected_options, 'val': val, 'project': course_parent, 'children': child_data, 'year_of_offering': year_of_offering,})

def project_child_update(request, course_id, child_id, val: str):
    # Fetch course details
    course_url = f"{API_STUDIO_URL}getapi/naac01_project_work_dc1/{course_id}"
    course_response = requests.get(course_url)
    
    current_year = datetime.now().year
    year_of_offering = [f"{y}-{y+1}" for y in range(current_year, 2010, -1)]


    if course_response.status_code == 200:
        course_parent = course_response.json()
        selected_options = course_parent.get('course_name', '').split(', ') if course_parent.get('course_name') else []
    else:
        return HttpResponse("Error fetching participation details: " + course_response.text)

    # Fetch child data
    child_url = f"{API_STUDIO_URL}getapi/naac01_project_work_dc2/{child_id}"
    response = requests.get(child_url)

    if response.status_code == 200:
        child_data = response.json()
    else:
        return HttpResponse(f"Failed to fetch child data: {response.text}")

    # Fetch child files for media
    media_url = f"{API_STUDIO_URL}crudapp/get/media/naac01_project_work_dc2_media/parent/{child_id}"
    media_response = requests.get(media_url)

    if media_response.status_code == 200:
        child_files = media_response.json()
    else:
        child_files = []

    if request.method == 'POST':
        # Prepare data for updating the child
        # Exclude the invalid fields from the payload (e.g., psk_id and psk_uid)
        update_payload = {
            "data": {
                k: request.POST.get(k, child_data.get(k)) for k in child_data if k not in ['psk_id', 'psk_uid']
            }
        }

        # Update the child data without invalid fields
        update_url = f"{API_STUDIO_URL}updateapi/update/naac01_project_work_dc2/{child_id}"
        payload = json.dumps(update_payload)
        
        update_response = requests.put(update_url, data=payload)

        if update_response.status_code != 200:
            return HttpResponse("Failed to update participation child: " + update_response.text)

        # Handle file uploads (if any files are present in the request)
        upload_errors = []
        for child_file in child_files:
            upload_id = child_file['psk_id']
            staff_id = course_parent.get('staff_id')

            field_map = {
                'project': ['Project', 'Permission', 'Compilation', 'DRP'],
                'field': ['PL', 'Obj_Outcome', 'AR', 'DR'],
                'internship': ['AL', 'ES', 'CL', 'DR']
            }

            fields = field_map.get(val, [])

            for field in fields:
                uploaded_files = request.FILES.getlist(f'file_{upload_id}_{field}')
                if not uploaded_files:
                    continue

                for uploaded_file in uploaded_files:
                    validate_file_size(uploaded_file)
                    validate_file_format(uploaded_file)

                    current_year = datetime.now().year
                    custom_filename = f"{staff_id}_{field.replace(' ', '_')}_{current_year}_{uploaded_file.name}"

                    upload_url = f"{API_STUDIO_URL}crudapp/upload/media/naac01_project_work_dc2_media/{upload_id}"
                    files = {'media': (custom_filename, uploaded_file, uploaded_file.content_type)}
                    payload = {'parent_psk_id': child_id}
                    headers = {'api_name': 'naac01_project_work_dc2_media', 'psk_id': str(upload_id)}

                    upload_response = requests.put(upload_url, data=payload, files=files)
                    if upload_response.status_code != 200:
                        upload_errors.append(f"Failed to upload file {uploaded_file.name}")

        if upload_errors:
            return HttpResponse("File upload errors: " + ", ".join(upload_errors), status=500)

        return redirect('project_view', id=course_id)

    return render(request, 'Courses_project_templates/project_child_update.html', {
        'child_data': child_data, 'selected_options': selected_options, 'val': val, 
        'project': course_id, 'child': child_id, 'child_files': child_files, 'year_of_offering':year_of_offering
    })

def project_child_delete(request, child_id, course_id):
    delete_url = f"{API_STUDIO_URL}deleteapi/delete/naac01_project_work_dc2/{child_id}"
    delete_response = requests.request("DELETE", delete_url)

    if delete_response.status_code == 200:
        return redirect('project_view', id=course_id)
    else:
        return HttpResponse("Failed to delete participation: " + delete_response.text)

import pandas as pd
import io
import requests
import json
from django.http import HttpResponse
from django.shortcuts import render
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from datetime import datetime
from openpyxl.styles import Font
from django.contrib import messages
from collections import Counter
from user_management.settings_views import get_settings
from MIS.functions import validate_file_format_faculty, validate_file_size
from user_management.settings_views import *

def filter_projects(request):
    """
    Main filtering function for project work with Excel and PDF export
    with role-based access control (Staff vs HOD)
    """
    # Step 1: Get the logged-in username
    user = get_settings(request)
    username = user.get('username', '')
    
    # Step 2: Determine user role (Staff or HOD)
    role_url = f"{API_STUDIO_URL}getapi/asa0504_01_01"
    payload = json.dumps({"queries": [{"field": "username", "value": username, "operation": "equal"}], "search_type": "first"})
    role_response = requests.post(role_url, headers={'Content-Type': 'application/json'}, data=payload)
    value_user = int(role_response.json().get("user_roles", "0").strip("{}")) if role_response.status_code == 200 else 0

    role_list = roles_tbl(request)
    user_role = next((role.get("user_role") for role in role_list if role.get("psk_id") == value_user), "Staff")
    
    # Step 3: Get research data to determine department for HOD
    access_token, token_type = course_key()
    research_data = get_course_data(access_token, token_type)
    staff_info = next((s for s in research_data if s.get("stf_id") == username), {})
    department_name = staff_info.get("department", "")
    
    # Get all parent projects
    parent_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_project_work_dc1/all"
    parent_response = requests.get(parent_url)
    
    if parent_response.status_code != 200:
        return HttpResponse("Error fetching project data")
    
    all_parents = parent_response.json()
    
    # Get all child projects
    child_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_project_work_dc2/all"
    child_response = requests.get(child_url)
    
    all_children = child_response.json() if child_response.status_code == 200 else []
    
    # Get filter parameters from request
    staff_id = request.GET.get('staff_id', '').strip()
    from_year = request.GET.get('from_year', '').strip()
    to_year = request.GET.get('to_year', '').strip()
    course_type = request.GET.get('course_type', '').strip()
    project_type = request.GET.get('project_type', '').strip()
    export_format = request.GET.get('export', '')
    
    # ROLE-BASED FILTERING: Apply department filter for HOD
    if user_role == "Hod":
        # HOD can only see staff from their own department
        department_staff = [s for s in research_data if s.get("department") == department_name]
        dept_staff_ids = [s.get("stf_id") for s in department_staff]
        
        # Filter parents to only those in HOD's department
        all_parents = [parent for parent in all_parents if parent.get('staff_id') in dept_staff_ids]
    
    # Filter parents based on user selections
    filtered_parents = all_parents
    
    if user_role == "Hod":
        if staff_id:
            filtered_parents = [p for p in filtered_parents if p.get("staff_id") == staff_id]
    else:
        # Staff can only see their own data
        filtered_parents = [p for p in filtered_parents if p.get("staff_id") == username]

    # Filter by course type
    if course_type:
        filtered_parents = [parent for parent in filtered_parents 
                          if course_type.lower() in (parent.get('course_name', '') or '').lower()]

    # Convert year values to integers for comparison
    if from_year:
        try:
            from_year_int = int(from_year.split('-')[0])
            filtered_parents = [parent for parent in filtered_parents 
                              if parent.get('academic_year') and extract_year_value(parent.get('academic_year')) >= from_year_int]
        except (ValueError, AttributeError):
            pass
    
    if to_year:
        try:
            to_year_int = int(to_year.split('-')[0])
            filtered_parents = [parent for parent in filtered_parents 
                              if parent.get('academic_year') and extract_year_value(parent.get('academic_year')) <= to_year_int]
        except (ValueError, AttributeError):
            pass
    
    # Get parent IDs for child filtering
    parent_ids = [parent.get('psk_id') for parent in filtered_parents]
    
    # Filter children based on parent IDs and project type
    filtered_children = []
    for child in all_children:
        if child.get('transaction_id') in parent_ids:
            # Filter by project type if specified
            if project_type:
                # Check if any of the project type fields contain data
                has_project_data = any(child.get(field) for field in ['pro_name_of_the_course', 'pro_course_code'])
                has_field_data = any(child.get(field) for field in ['field_name_of_the_course', 'field_course_code'])
                has_intern_data = any(child.get(field) for field in ['intern_name_of_the_course', 'intern_course_code'])
                
                if project_type == 'project' and has_project_data:
                    filtered_children.append(child)
                elif project_type == 'field' and has_field_data:
                    filtered_children.append(child)
                elif project_type == 'internship' and has_intern_data:
                    filtered_children.append(child)
                elif project_type == 'all':
                    filtered_children.append(child)
            else:
                filtered_children.append(child)
    
    # Fetch media files for each child - USING MEDIA ID INSTEAD OF CHILD ID
    for child in filtered_children:
        child_id = child.get('psk_id')
        media_url = f"{API_STUDIO_URL}crudapp/get/media/naac01_project_work_dc2_media/parent/{child_id}"
        media_response = requests.get(media_url)
        
        if media_response.status_code == 200:
            media_data = media_response.json()
            # Process each media file to get individual media IDs
            processed_media_files = []
            for media in media_data:
                # Extract media ID from the media object
                media_id = media.get('id') or media.get('media_id') or media.get('psk_id') or media.get('value_id')
                
                if media_id:
                    # Create direct API URL using media ID
                    direct_api_url = f"{API_STUDIO_URL}crudapp/view/media/naac01_project_work_dc2_media/{media_id}"
                    
                    # Add the media URL and other media info
                    processed_media = {
                        'file_name': media.get('file_name', 'Unknown'),
                        'media_id': media_id,
                        'direct_api_url': direct_api_url,
                        'original_data': media
                    }
                    processed_media_files.append(processed_media)
                else:
                    # Fallback if no media ID found
                    fallback_url = f"{API_STUDIO_URL}crudapp/view/media/naac01_project_work_dc2_media/{child_id}"
                    processed_media = {
                        'file_name': media.get('file_name', 'Unknown'),
                        'media_id': None,
                        'direct_api_url': fallback_url,
                        'original_data': media
                    }
                    processed_media_files.append(processed_media)
            
            child['media_files'] = processed_media_files
        else:
            child['media_files'] = []
    
    # Create mapping of parent ID to parent data for easy access
    parent_map = {parent['psk_id']: parent for parent in filtered_parents}
    
    # Add parent information to each child
    for child in filtered_children:
        parent_id = child.get('transaction_id')
        if parent_id in parent_map:
            child['parent_data'] = parent_map[parent_id]
    
    # Handle export formats
    if export_format:
        if export_format.lower() == 'excel':
            return export_projects_to_excel(filtered_parents, filtered_children, project_type)
        elif export_format.lower() == 'pdf':
            return export_projects_to_pdf(filtered_parents, filtered_children, project_type)
    
    # Get unique staff IDs for dropdown - filtered by role
    if user_role == "Hod":
        # HOD can see all staff in their department
        department_staff = [s for s in research_data if s.get("department") == department_name]
        dept_staff_ids = [s.get("stf_id") for s in department_staff]
        staff_ids = sorted(list(set(parent.get('staff_id') for parent in all_parents if parent.get('staff_id') in dept_staff_ids)))
    else:
        # Staff can only see themselves
        staff_ids = [username]
    
    # Get unique years for dropdown
    year_values = []
    for parent in all_parents:
        if parent.get('academic_year'):
            try:
                year_val = extract_year_value(parent.get('academic_year'))
                year_values.append(year_val)
            except (ValueError, AttributeError):
                continue
    
    # Convert back to "YYYY-YYYY" format for display
    years = sorted(list(set(f"{y}-{y+1}" for y in year_values)), reverse=True)
    
    # Get unique course types
    course_types = sorted(list(set(parent.get('course_name', '') for parent in all_parents if parent.get('course_name'))))
    
    context = {
        'parents': filtered_parents,
        'children': filtered_children,
        'staff_ids': staff_ids,
        'years': years,
        'course_types': course_types,
        'project_types': ['all', 'project', 'field', 'internship'],
        'selected_staff_id': staff_id,
        'selected_from_year': from_year,
        'selected_to_year': to_year,
        'selected_course_type': course_type,
        'selected_project_type': project_type,
        'filter_applied': any([staff_id, from_year, to_year, course_type, project_type]),
        'user_role': user_role,
        'username': username
    }
    
    return render(request, 'Courses_project_templates/filter_projects.html', context)

def extract_year_value(year_str):
    """Extract first year from academic year string or return None"""
    try:
        return int(str(year_str).split('-')[0]) if year_str else None
    except:
        return None

def export_projects_to_excel(parents, children, project_type):
    """
    Export project attachments into Excel with separate worksheets for each project type.
    """
    import io
    import re
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    # Create a workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)

    # Define worksheet configurations for each project type
    worksheet_configs = {
        'Project Work': {
            'title': 'Project Work',
            'attachment_headers': [
                "Project Work Files", "Permission Letters", "Compilation Letters", "Detailed Reports"
            ],
            'categories': {
                'Project Work Files': ['project', 'work'],
                'Permission Letters': ['permission', 'permis', 'pl'],
                'Compilation Letters': ['compilation', 'complete', 'cl'],
                'Detailed Reports': ['detailed', 'drp', 'dr']
            }
        },
        'Field Project': {
            'title': 'Field Project', 
            'attachment_headers': [
                "Permission Letters", "Objective Outcomes", "Attendance Report", "Detailed Reports"
            ],
            'categories': {
                'Permission Letters': ['permission', 'permis', 'pl'],
                'Objective Outcomes': ['objective', 'outcome', 'obj'],
                'Attendance Report': ['assessment', 'attendance', 'ar'],
                'Detailed Reports': ['detailed', 'drp', 'dr']
            }
        },
        'Internship': {
            'title': 'Internship',
            'attachment_headers': [
                "Acceptance Letter", "Evaluation Sheet", "Completion Letter", "Detailed Report"
            ],
            'categories': {
                'Acceptance Letter': ['acceptance', 'activity', 'al'],
                'Evaluation Sheet': ['evaluation', 'sheet', 'es'],
                'Completion Letter': ['completion', 'cl'],
                'Detailed Report': ['detailed', 'drp', 'dr']
            }
        }
    }

    # Common headers for all worksheets
    common_headers = [
        "Staff ID", "Staff Name", "Project Type", "Department Name", 
        "Course Name", "Course Code", "Year of Offering", "Students"
    ]

    # Regex to extract from HYPERLINK formula
    hyperlink_pattern = re.compile(r'=HYPERLINK\("([^"]+)",\s*"([^"]+)"\)')

    # Process children data and organize by project type
    projects_by_type = {
        'Project Work': [],
        'Field Project': [],
        'Internship': []
    }

    for child in children:
        # Get parent data
        parent_data = child.get('parent_data', {})
        
        # Determine project type and extract data
        if child.get('pro_name_of_the_course') or child.get('pro_course_code'):
            project_type_name = 'Project Work'
            project_data = {
                'Project Type': 'Project Work',
                'Department Name': child.get('pro_department_name', ''),
                'Course Name': child.get('pro_name_of_the_course', ''),
                'Course Code': child.get('pro_course_code', ''),
                'Year of Offering': child.get('pro_year_of_offering', ''),
                'Students': child.get('pro_name_of_students_studied', '')
            }
        elif child.get('field_name_of_the_course') or child.get('field_course_code'):
            project_type_name = 'Field Project'
            project_data = {
                'Project Type': 'Field Project',
                'Department Name': child.get('field_department_name', ''),
                'Course Name': child.get('field_name_of_the_course', ''),
                'Course Code': child.get('field_course_code', ''),
                'Year of Offering': child.get('field_year_of_offering', ''),
                'Students': child.get('field_name_of_students_studied', '')
            }
        elif child.get('intern_name_of_the_course') or child.get('intern_course_code'):
            project_type_name = 'Internship'
            project_data = {
                'Project Type': 'Internship',
                'Department Name': child.get('intern_department_name', ''),
                'Course Name': child.get('intern_name_of_the_course', ''),
                'Course Code': child.get('intern_course_code', ''),
                'Year of Offering': child.get('intern_year_of_offering', ''),
                'Students': child.get('intern_name_of_students_studied', '')
            }
        else:
            project_type_name = 'Unknown'
            project_data = {
                'Project Type': 'Unknown',
                'Department Name': '',
                'Course Name': '',
                'Course Code': '',
                'Year of Offering': '',
                'Students': ''
            }

        # Skip unknown project types
        if project_type_name == 'Unknown':
            continue

        # Process media files for this child
        media_files = child.get('media_files', [])
        categorized_files = {}

        # Initialize categories based on project type
        config = worksheet_configs[project_type_name]
        for category in config['categories'].keys():
            categorized_files[category] = []

        # Categorize files
        for media in media_files:
            file_name = media.get('file_name', '').lower()
            original_file_name = media.get('file_name', 'File')
            media_url = media.get('direct_api_url', '')
            
            if not media_url:
                continue
                
            # Create hyperlink formula
            hyperlink_formula = f'=HYPERLINK("{media_url}", "{original_file_name}")'
            
            # Categorize based on project type specific patterns
            categorized = False
            for category, keywords in config['categories'].items():
                if any(keyword in file_name for keyword in keywords):
                    categorized_files[category].append(hyperlink_formula)
                    categorized = True
                    break
            
            # If not categorized, put in first available category
            if not categorized and categorized_files:
                first_category = list(categorized_files.keys())[0]
                categorized_files[first_category].append(hyperlink_formula)

        # Create record
        record = {
            'parent_data': parent_data,
            'project_data': project_data,
            'categorized_files': categorized_files
        }
        
        projects_by_type[project_type_name].append(record)

    # Create worksheets for each project type
    for project_type_name, config in worksheet_configs.items():
        # Skip if no data for this project type
        if not projects_by_type[project_type_name]:
            continue
            
        # Create worksheet
        ws = wb.create_sheet(title=config['title'])
        
        # Create full headers list
        full_headers = common_headers + config['attachment_headers']
        ws.append(full_headers)

        # Header styling
        header_font = Font(bold=True)
        for col in range(1, len(full_headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Add data rows
        row_num = 2
        for record in projects_by_type[project_type_name]:
            parent_data = record['parent_data']
            project_data = record['project_data']
            categorized_files = record['categorized_files']

            # Create row data
            row_data = [
                parent_data.get('staff_id', ''),
                parent_data.get('staff_name', ''),
                project_data['Project Type'],
                project_data['Department Name'],
                project_data['Course Name'],
                project_data['Course Code'],
                project_data['Year of Offering'],
                project_data['Students']
            ]

            # Add attachment data for this project type
            for category in config['attachment_headers']:
                files = categorized_files.get(category, [])
                if files:
                    # Show all files separated by newlines
                    display_texts = []
                    for hyperlink_formula in files:
                        match = hyperlink_pattern.match(hyperlink_formula)
                        if match:
                            _, display_name = match.groups()
                            display_texts.append(display_name)
                    
                    if display_texts:
                        row_data.append("\n".join(display_texts))
                    else:
                        row_data.append(" - ")
                else:
                    row_data.append(" - ")

            # Add row to worksheet
            ws.append(row_data)

            # Apply hyperlinks and styling
            for col_idx in range(len(common_headers) + 1, len(full_headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                
                category = config['attachment_headers'][col_idx - len(common_headers) - 1]
                files = categorized_files.get(category, [])
                
                if files:
                    # Make the first file clickable
                    first_file = files[0]
                    match = hyperlink_pattern.match(first_file)
                    if match:
                        url, display_name = match.groups()
                        cell.hyperlink = url
                        cell.font = Font(color="0000FF", underline="single")

            row_num += 1

        # Auto-fit columns
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            
            # Check header and content length
            header_length = len(str(full_headers[col_idx-1])) if col_idx <= len(full_headers) else 0
            max_length = header_length
            
            for cell in column_cells:
                try:
                    if cell.value:
                        cell_value = str(cell.value)
                        if '\n' in cell_value:
                            lines = cell_value.split('\n')
                            longest_line = max(len(line) for line in lines)
                            max_length = max(max_length, longest_line)
                        else:
                            max_length = max(max_length, len(cell_value))
                except:
                    pass
            
            adjusted_width = min(max(8, max_length + 2), 35)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Auto-fit row heights
        for row_idx in range(1, row_num + 1):
            max_lines = 1
            for col_idx in range(1, len(full_headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value and isinstance(cell.value, str):
                    lines = cell.value.count('\n') + 1
                    max_lines = max(max_lines, lines)
            
            row_height = min(max(20, max_lines * 15), 100)
            ws.row_dimensions[row_idx].height = row_height

        # Freeze panes and add filter
        ws.freeze_panes = ws['A2']
        ws.auto_filter.ref = ws.dimensions

    # If no sheets were created, create an empty one
    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet(title="No Data")
        ws.append(["No project data found for the selected filters."])

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="project_attachments.xlsx"'

    with io.BytesIO() as buffer:
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())

    return response

from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT

# def export_projects_to_pdf(parents, children, project_type):
#     """
#     Export filtered project data to PDF format with separate pages for each staff
#     """
#     from reportlab.lib import colors
#     from reportlab.lib.pagesizes import A4
#     from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
#     from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
#     from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, PageBreak
#     from reportlab.lib.units import inch
#     from django.http import HttpResponse
#     from collections import defaultdict

#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="projects_filtered.pdf"'

#     doc = SimpleDocTemplate(
#         response,
#         pagesize=A4,
#         topMargin=0.5 * inch,
#         bottomMargin=0.5 * inch,
#         leftMargin=0.5 * inch,
#         rightMargin=0.5 * inch
#     )
#     elements = []
#     styles = getSampleStyleSheet()

#     # ============================
#     # Custom professional styles
#     # ============================
#     title_style = ParagraphStyle(
#         'CustomTitle',
#         parent=styles['Heading1'],
#         fontSize=14,
#         spaceAfter=15,
#         alignment=1,
#         textColor=colors.HexColor('#2c3e50'),
#         fontName='Helvetica-Bold'
#     )

#     left_style = ParagraphStyle(
#         'LeftStyle',
#         parent=styles['Normal'],
#         fontSize=9,
#         spaceAfter=4,
#         alignment=TA_LEFT,
#         textColor=colors.HexColor('#2c3e50')
#     )

#     right_style = ParagraphStyle(
#         'RightStyle',
#         parent=styles['Normal'],
#         fontSize=9,
#         spaceAfter=4,
#         alignment=TA_RIGHT,
#         textColor=colors.HexColor('#2c3e50')
#     )

#     table_header_style = ParagraphStyle(
#         'TableHeader',
#         parent=styles['Normal'],
#         fontSize=8,
#         alignment=TA_CENTER,
#         textColor=colors.white,
#         fontName='Helvetica-Bold'
#     )

#     table_cell_style = ParagraphStyle(
#         'TableCell',
#         parent=styles['Normal'],
#         fontSize=7,
#         alignment=TA_CENTER,
#         textColor=colors.HexColor('#2c3e50'),
#         leading=9
#     )

#     attachment_link_style = ParagraphStyle(
#         'AttachmentLink',
#         parent=styles['Normal'],
#         fontSize=6,
#         alignment=TA_CENTER,
#         textColor=colors.HexColor('#1a5276'),
#         leading=8
#     )

#     # ============================
#     # ORGANIZE DATA BY STAFF
#     # ============================
#     staff_children = defaultdict(list)
    
#     # Group children by staff ID
#     for child in children:
#         parent_data = child.get('parent_data', {})
#         staff_id = parent_data.get('staff_id', 'Unknown')
#         staff_children[staff_id].append(child)
    
#     # Get parent data for each staff
#     staff_parents = {}
#     for parent in parents:
#         staff_id = parent.get('staff_id')
#         if staff_id:
#             staff_parents[staff_id] = parent

#     # ============================
#     # CREATE PAGES FOR EACH STAFF
#     # ============================
#     for staff_index, (staff_id, staff_children_list) in enumerate(staff_children.items()):
#         # Add page break for subsequent staff (except first one)
#         if staff_index > 0:
#             elements.append(PageBreak())
        
#         # Get parent data for this staff
#         parent = staff_parents.get(staff_id, {})
        
#         # ============================
#         # TITLE
#         # ============================
#         elements.append(Paragraph("PROJECT WORK REPORT", title_style))

#         # ============================
#         # STAFF INFO
#         # ============================
#         left_content = [
#             f"<b>Staff ID:</b> {parent.get('staff_id', 'N/A')}",
#             f"<b>Name:</b> {parent.get('staff_name', 'N/A')}"
#         ]
#         right_content = [
#             f"<b>Department:</b> {parent.get('department_name', 'N/A')}",
#             f"<b>Course Name:</b> {parent.get('course_name', 'N/A')}"
#         ]

#         left_paragraphs = [Paragraph(text, left_style) for text in left_content]
#         right_paragraphs = [Paragraph(text, right_style) for text in right_content]

#         side_table_data = [
#             [left_paragraphs[0], right_paragraphs[0]],
#             [left_paragraphs[1], right_paragraphs[1]]
#         ]

#         side_table = Table(side_table_data, colWidths=[3.5 * inch, 3.5 * inch])
#         side_table.setStyle(TableStyle([
#             ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
#             ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#             ('LEFTPADDING', (0, 0), (-1, -1), 0),
#             ('RIGHTPADDING', (0, 0), (-1, -1), 0),
#             ('TOPPADDING', (0, 0), (-1, -1), 2),
#             ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
#         ]))

#         elements.append(side_table)
#         elements.append(Spacer(1, 15))

#         # ============================
#         # CHILD TABLE FOR THIS STAFF
#         # ============================
#         if staff_children_list:
#             child_data = []
#             headers = ['Project Type', 'Department', 'Course', 'Year', 'Students', 'Attachments']
#             child_data.append([Paragraph(f"<b>{h}</b>", table_header_style) for h in headers])

#             for child in staff_children_list:
#                 # Detect project type
#                 if child.get('pro_name_of_the_course'):
#                     project_type_name = 'Project Work'
#                     dept_name = child.get('pro_department_name', '')
#                     course_name = child.get('pro_name_of_the_course', '')
#                     year = child.get('pro_year_of_offering', '')
#                     students = child.get('pro_name_of_students_studied', '')
#                 elif child.get('field_name_of_the_course'):
#                     project_type_name = 'Field Project'
#                     dept_name = child.get('field_department_name', '')
#                     course_name = child.get('field_name_of_the_course', '')
#                     year = child.get('field_year_of_offering', '')
#                     students = child.get('field_name_of_students_studied', '')
#                 elif child.get('intern_name_of_the_course'):
#                     project_type_name = 'Internship'
#                     dept_name = child.get('intern_department_name', '')
#                     course_name = child.get('intern_name_of_the_course', '')
#                     year = child.get('intern_year_of_offering', '')
#                     students = child.get('intern_name_of_students_studied', '')
#                 else:
#                     project_type_name = 'Unknown'
#                     dept_name = course_name = year = students = ''

#                 # Media links
#                 media_files = child.get('media_files', [])
#                 media_links = []
#                 for media in media_files:
#                     file_name = media.get('file_name', 'Unknown')
#                     media_url = media.get('direct_api_url', '')
#                     if media_url:
#                         media_links.append(f'<a href="{media_url}" color="blue">{file_name}</a>')
#                     else:
#                         media_links.append(file_name)
#                 media_info = "<br/>".join(media_links) if media_links else "No attachments"

#                 child_data.append([
#                     Paragraph(str(project_type_name), table_cell_style),
#                     Paragraph(str(dept_name), table_cell_style),
#                     Paragraph(str(course_name), table_cell_style),
#                     Paragraph(str(year), table_cell_style),
#                     Paragraph(str(students), table_cell_style),
#                     Paragraph(media_info, attachment_link_style)
#                 ])

#             col_widths = [1.0 * inch, 1.1 * inch, 1.4 * inch, 0.7 * inch, 1.0 * inch, 2.0 * inch]

#             child_table = Table(child_data, colWidths=col_widths, repeatRows=1)
#             child_table.setStyle(TableStyle([
#                 ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
#                 ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
#                 ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#                 ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
#                 ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#                 ('FONTSIZE', (0, 0), (-1, 0), 8),
#                 ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
#                 ('TOPPADDING', (0, 0), (-1, 0), 8),

#                 ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
#                 ('FONTSIZE', (0, 1), (-1, -1), 7),

#                 ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
#                 ('LEFTPADDING', (0, 0), (-1, -1), 6),
#                 ('RIGHTPADDING', (0, 0), (-1, -1), 6),
#                 ('TOPPADDING', (0, 1), (-1, -1), 4),
#                 ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
#             ]))

#             elements.append(child_table)
            
#             # Add summary for this staff
#             elements.append(Spacer(1, 10))
#             summary_text = f"<i>Total projects for {parent.get('staff_name', 'this staff')}: {len(staff_children_list)}</i>"
#             elements.append(Paragraph(summary_text, table_cell_style))

#     # ============================
#     # HANDLE CASE WITH NO CHILDREN
#     # ============================
#     if not children and parents:
#         # If no children but we have parent data, show just the staff info
#         elements.append(Paragraph("PROJECT WORK REPORT", title_style))
        
#         parent = parents[0]
#         left_content = [
#             f"<b>Staff ID:</b> {parent.get('staff_id', 'N/A')}",
#             f"<b>Name:</b> {parent.get('staff_name', 'N/A')}"
#         ]
#         right_content = [
#             f"<b>Department:</b> {parent.get('department_name', 'N/A')}",
#             f"<b>Course Name:</b> {parent.get('course_name', 'N/A')}"
#         ]

#         left_paragraphs = [Paragraph(text, left_style) for text in left_content]
#         right_paragraphs = [Paragraph(text, right_style) for text in right_content]

#         side_table_data = [
#             [left_paragraphs[0], right_paragraphs[0]],
#             [left_paragraphs[1], right_paragraphs[1]]
#         ]

#         side_table = Table(side_table_data, colWidths=[3.5 * inch, 3.5 * inch])
#         side_table.setStyle(TableStyle([
#             ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
#             ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#             ('LEFTPADDING', (0, 0), (-1, -1), 0),
#             ('RIGHTPADDING', (0, 0), (-1, -1), 0),
#             ('TOPPADDING', (0, 0), (-1, -1), 2),
#             ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
#         ]))

#         elements.append(side_table)
#         elements.append(Spacer(1, 15))
#         elements.append(Paragraph("<i>No project data found for this staff.</i>", table_cell_style))

#     doc.build(elements)
#     return response

def export_projects_to_pdf(parents, children, project_type):
    """
    Export filtered project data to PDF format with separate pages for each staff
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, PageBreak
    from reportlab.lib.units import inch
    from django.http import HttpResponse
    from collections import defaultdict

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="projects_filtered.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch
    )
    elements = []
    styles = getSampleStyleSheet()

    # ============================
    # Custom professional styles
    # ============================
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        spaceAfter=15,
        alignment=1,
        textColor=colors.HexColor('#2c3e50'),
        fontName='Helvetica-Bold'
    )

    left_style = ParagraphStyle(
        'LeftStyle',
        parent=styles['Normal'],
        fontSize=9,
        spaceAfter=4,
        alignment=TA_LEFT,
        textColor=colors.HexColor('#2c3e50')
    )

    right_style = ParagraphStyle(
        'RightStyle',
        parent=styles['Normal'],
        fontSize=9,
        spaceAfter=4,
        alignment=TA_RIGHT,
        textColor=colors.HexColor('#2c3e50')
    )

    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontSize=8,
        alignment=TA_CENTER,
        textColor=colors.white,
        fontName='Helvetica-Bold'
    )

    table_cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontSize=7,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#2c3e50'),
        leading=9
    )

    attachment_link_style = ParagraphStyle(
        'AttachmentLink',
        parent=styles['Normal'],
        fontSize=6,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1a5276'),
        leading=8
    )

    # ADD THIS STYLE FOR NO DATA MESSAGE
    no_data_style = ParagraphStyle(
        'NoData',
        parent=styles['Normal'],
        fontSize=9,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#7f8c8d'),
        fontStyle='italic',
        spaceBefore=12,
        spaceAfter=12
    )

    # ============================
    # ORGANIZE DATA BY STAFF
    # ============================
    staff_children = defaultdict(list)
    
    # Group children by staff ID
    for child in children:
        parent_data = child.get('parent_data', {})
        staff_id = parent_data.get('staff_id', 'Unknown')
        staff_children[staff_id].append(child)
    
    # Get parent data for each staff
    staff_parents = {}
    for parent in parents:
        staff_id = parent.get('staff_id')
        if staff_id:
            staff_parents[staff_id] = parent

    # ============================
    # CREATE PAGES FOR EACH STAFF
    # ============================
    # If we have staff data (either from children or parents)
    staff_to_process = list(staff_children.keys()) if staff_children else list(staff_parents.keys())
    
    if not staff_to_process:
        # No data at all
        elements.append(Paragraph("PROJECT WORK REPORT", title_style))
        elements.append(Paragraph("No data available", no_data_style))
        doc.build(elements)
        return response

    for staff_index, staff_id in enumerate(staff_to_process):
        # Add page break for subsequent staff (except first one)
        if staff_index > 0:
            elements.append(PageBreak())
        
        # Get parent data for this staff
        parent = staff_parents.get(staff_id, {})
        staff_children_list = staff_children.get(staff_id, [])
        
        # ============================
        # TITLE
        # ============================
        elements.append(Paragraph("PROJECT WORK REPORT", title_style))

        # ============================
        # STAFF INFO
        # ============================
        left_content = [
            f"<b>Staff ID:</b> {parent.get('staff_id', 'N/A')}",
            f"<b>Name:</b> {parent.get('staff_name', 'N/A')}"
        ]
        right_content = [
            f"<b>Department:</b> {parent.get('department_name', 'N/A')}",
            f"<b>Course Name:</b> {parent.get('course_name', 'N/A')}"
        ]

        left_paragraphs = [Paragraph(text, left_style) for text in left_content]
        right_paragraphs = [Paragraph(text, right_style) for text in right_content]

        side_table_data = [
            [left_paragraphs[0], right_paragraphs[0]],
            [left_paragraphs[1], right_paragraphs[1]]
        ]

        side_table = Table(side_table_data, colWidths=[3.5 * inch, 3.5 * inch])
        side_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]))

        elements.append(side_table)
        elements.append(Spacer(1, 15))

        # ============================
        # CHILD TABLE FOR THIS STAFF
        # ============================
        if staff_children_list:
            child_data = []
            headers = ['Project Type', 'Department', 'Course', 'Year', 'Students', 'Attachments']
            child_data.append([Paragraph(f"<b>{h}</b>", table_header_style) for h in headers])

            for child in staff_children_list:
                # Detect project type
                if child.get('pro_name_of_the_course'):
                    project_type_name = 'Project Work'
                    dept_name = child.get('pro_department_name', '')
                    course_name = child.get('pro_name_of_the_course', '')
                    year = child.get('pro_year_of_offering', '')
                    students = child.get('pro_name_of_students_studied', '')
                elif child.get('field_name_of_the_course'):
                    project_type_name = 'Field Project'
                    dept_name = child.get('field_department_name', '')
                    course_name = child.get('field_name_of_the_course', '')
                    year = child.get('field_year_of_offering', '')
                    students = child.get('field_name_of_students_studied', '')
                elif child.get('intern_name_of_the_course'):
                    project_type_name = 'Internship'
                    dept_name = child.get('intern_department_name', '')
                    course_name = child.get('intern_name_of_the_course', '')
                    year = child.get('intern_year_of_offering', '')
                    students = child.get('intern_name_of_students_studied', '')
                else:
                    project_type_name = 'Unknown'
                    dept_name = course_name = year = students = ''

                # Media links
                media_files = child.get('media_files', [])
                media_links = []
                for media in media_files:
                    file_name = media.get('file_name', 'Unknown')
                    media_url = media.get('direct_api_url', '')
                    if media_url:
                        media_links.append(f'<a href="{media_url}" color="blue">{file_name}</a>')
                    else:
                        media_links.append(file_name)
                media_info = "<br/>".join(media_links) if media_links else "No attachments"

                child_data.append([
                    Paragraph(str(project_type_name), table_cell_style),
                    Paragraph(str(dept_name), table_cell_style),
                    Paragraph(str(course_name), table_cell_style),
                    Paragraph(str(year), table_cell_style),
                    Paragraph(str(students), table_cell_style),
                    Paragraph(media_info, attachment_link_style)
                ])

            col_widths = [1.0 * inch, 1.1 * inch, 1.4 * inch, 0.7 * inch, 1.0 * inch, 2.0 * inch]

            child_table = Table(child_data, colWidths=col_widths, repeatRows=1)
            child_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),

                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
                ('FONTSIZE', (0, 1), (-1, -1), 7),

                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ]))

            elements.append(child_table)
            
            # Add summary for this staff
            elements.append(Spacer(1, 10))
            summary_text = f"<i>Total projects for {parent.get('staff_name', 'this staff')}: {len(staff_children_list)}</i>"
            elements.append(Paragraph(summary_text, table_cell_style))
        else:
            # SHOW NO DATA MESSAGE WHEN NO CHILDREN
            elements.append(Paragraph("No data available for the selected staff", no_data_style))

    doc.build(elements)
    return response