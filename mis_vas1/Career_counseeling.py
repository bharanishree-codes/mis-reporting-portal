import json
import requests
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from MIS.functions import validate_file_format, validate_file_size
from datetime import date, datetime
from user_management.settings_views import *

API_STUDIO_URL = user_bundle_settings()

def career_key():

    url = "https://api.hcaschennai.edu.in/auth/token"

    payload = json.dumps({
        "secret_key": "C4ZoXbsAnHLjk1Xyz4QPT2eoiNx6K6fo"
    })
    headers = {'Content-Type': 'application/json'}

    response = requests.post(url, headers=headers, data=payload)
    if response.status_code == 200:
        res_data = response.json()
        access_token = res_data.get('access_token')
        token_type = res_data.get('token_type')
        return access_token, token_type
    return None, None


# Function to fetch staff career data
def get_career_data(access_token, token_type):
    url = "https://api.hcaschennai.edu.in/sqlviews/api/v1/auth/get_response_data"

    payload = json.dumps({
        "psk_uid": "51a531b4-bd55-491c-861d-a8d7227b325b",
        "project": "hcas",
        "data": {}
    })
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'{token_type} {access_token}'  # Fix spacing
    }

    response = requests.post(url, headers=headers, data=payload)
    if response.status_code == 200:
        return response.json()
    return []


# View to create career counseling entry
def career_counseeling_create(request):
    error_message = None

    # Step 1: Get API token
    access_token, token_type = career_key()
    if not access_token or not token_type:
        error_message = 'Failed to get access token from API.'
        return render(request, 'Career_counseeling_templates/career_counseeling_create.html', {'error': error_message})

    # Step 2: Fetch career data
    career_data = get_career_data(access_token, token_type)
    if not career_data:
        error_message = 'Failed to fetch staff data.'
        return render(request, 'Career_counseeling_templates/career_counseeling_create.html', {'error': error_message})
    
    user = get_settings(request)
    username = user['username']
    # username = 'CS-T151'
    
    selected_faculty = None
    
    for faculty in career_data:
        if faculty['stf_id'] == username:
            selected_faculty = faculty
            break
        
    if career_data:
        staff_name = selected_faculty.get('stf_name') if selected_faculty else ''
        department = selected_faculty.get('department') if selected_faculty else ''
    
    # Generate academic years for dropdown
    current_year = datetime.now().year
    publication_year = [f"{y}-{y+1}" for y in range(current_year, 2010, -1)]

    # Step 3: Handle POST form submission
    if request.method == 'POST':
        # Extract form data
        staff_name = request.POST.get('staff_name')
        staff_id = username
        department_name = request.POST.get('department_name')
        year_of_activity_str = request.POST.get('year_of_activity')  # Get as string from form
        counseling_detail = request.POST.get('counseling_detail')
        number_of_students_attended = request.POST.get('number_of_students_attended')
        number_of_students_placed = request.POST.get('number_of_students_placed')
        
        # Convert string year to date format (assuming format like "2024-2025")
        try:
            # Extract the start year from string (e.g., "2024-2025" -> 2024)
            start_year = int(year_of_activity_str.split('-')[0])
            # Create a date object (using June 1st as a default day)
            year_of_activity_date = date(start_year, 6, 1)
        except (ValueError, IndexError, AttributeError):
            error_message = "Invalid year format. Please select a valid year from the dropdown."
            return render(request, 'Career_counseeling_templates/career_counseeling_create.html', {
                'career_data': career_data, 
                "username": username, 
                "staff_name": staff_name, 
                "department": department,
                'publication_year': publication_year,  # Pass years to template
                'error': error_message
            })
        
        for faculty in career_data:
            if faculty['stf_id'] == staff_id:
                selected_faculty = faculty
                break
            
        staff_name = selected_faculty.get('stf_name') if selected_faculty else ''
        department_name = selected_faculty.get('department') if selected_faculty else ''

        # POST the data to the API - use the date object
        url = f"{API_STUDIO_URL}postapi/create/naac01_students_benefited_for_career_counseeling_dc1"
        payload = json.dumps({
            "data": {
                "staff_name": staff_name,
                "staff_id": staff_id,
                "department_name": department_name,
                "year_of_activity": year_of_activity_date.isoformat(),  # Convert date to ISO string
                "counseling_detail": counseling_detail,
                "number_of_students_attended": number_of_students_attended,
                "number_of_students_placed": number_of_students_placed
            }
        })
        headers = {'Content-Type': 'application/json'}

        response = requests.post(url, headers=headers, data=payload)

        if response.status_code == 200:
            file_data = response.json()
            psk_id = file_data.get('psk_id')

            # Handle file uploads
            upload_url = f"{API_STUDIO_URL}crudapp/upload/media/naac01_students_benefited_for_career_counseeling_dc1_media"
            uploaded_files = request.FILES.getlist('file')

            fields = ['PL', 'Circular', 'Brochure', 'RPP', 'DR', 'GTP', 'SA', 'FB', 'CS']

            for field, uploaded_file in zip(fields, uploaded_files):
                try:
                    validate_file_size(uploaded_file)
                    validate_file_format(uploaded_file)

                    current_year = datetime.now().year
                    custom_filename = f"{staff_id}_{field}_{current_year}_{uploaded_file.name}"

                    payload = {'parent_psk_id': psk_id}
                    files = {'media': (custom_filename, uploaded_file, uploaded_file.content_type)}
                    upload_headers = {'api_name': 'naac01_students_benefited_for_career_counseeling_dc1_media'}

                    upload_response = requests.post(upload_url, headers=upload_headers, data=payload, files=files)

                    if upload_response.status_code != 200:
                        messages.error(request, f"File upload failed for {uploaded_file.name}. Error: {upload_response.text}")
                        return redirect('career_counseeling_list')
                except Exception as e:
                    messages.error(request, f"Error processing file {uploaded_file.name}: {str(e)}")
                    return redirect('career_counseeling_list')

            messages.success(request, "Career Counseling entry created successfully with all files uploaded.")
            return redirect('career_counseeling_list')
        else:
            messages.error(request, "Failed to create entry. Please try again.")

    return render(request, 'Career_counseeling_templates/career_counseeling_create.html', {
        'career_data': career_data, 
        "username": username, 
        "staff_name": staff_name, 
        "department": department,
        'publication_year': publication_year  # Pass years to template
    })



# View career_counseeling Details

def career_counseeling_view(request, id):
    url = f"{API_STUDIO_URL}getapi/naac01_students_benefited_for_career_counseeling_dc1/{id}"
    response = requests.get(url)

    if response.status_code == 200:
        career_counseeling_data = response.json()
        return render(request, "career_counseeling_view.html", {'career_counseeling': career_counseeling_data})

    return HttpResponse(f"Error fetching Course details: {response.text}", status=500)



def career_counseeling_list(request):
    # URL to get career counseling data
    url = f"{API_STUDIO_URL}getapi/all_fields/naac01_students_benefited_for_career_counseeling_dc1/all"
    response = requests.get(url)

    if response.status_code == 200:
        career_counseelings = response.json()
        print("career_counseelings", career_counseelings)
    else:
        career_counseelings = []

    # Get the current username (staff_id)
    user = get_settings(request)
    username = user.get('username')  # Replace 'username' with the correct key if it's different
    # username = 'CS-T151'
    #from_dashboard = request.GET.get('from') == 'dashboard' or 'admin_hod_dash' or 'admin_dash' or 'department_dashboard'
    # If no username is found, return the same page with an empty list
    # if not username:
    #     return render(request, 'Career_counseeling_templates/career_counseeling_list.html', {'career_counseelings': []})

    # Filter the career counseling data based on the username (matching the staff_id in the data)
    filtered_career_counseelings = [counseling for counseling in career_counseelings if counseling.get('staff_id') == username]
    print("filtered_career_counseelings:", filtered_career_counseelings)
    
    selected_staff_id = request.GET.get('staff_id')
    
    if selected_staff_id:
        filtered_career_counseelings = [counseling for counseling in career_counseelings if counseling.get('staff_id') == selected_staff_id]
        
    selected_department = request.GET.get('department')
    if selected_department:
        filtered_career_counseelings = [counseeling for counseeling in career_counseelings if counseeling.get('department_name') == selected_department]

    # If no data is found for the username, return the same page with an empty list
    if not filtered_career_counseelings:
        return render(request, 'Career_counseeling_templates/career_counseeling_list.html', {'career_counseelings': []})

    # Return the filtered data to the template
    return render(request, 'Career_counseeling_templates/career_counseeling_list.html', {'career_counseelings': filtered_career_counseelings, 
    #'from_dashboard': from_dashboard
    })


# Update career_counseeling
def career_counseeling_update(request, id):
    
    current_year = datetime.now().year
    publication_year = [f"{y}-{y+1}" for y in range(current_year, 2010, -1)]

    # Fetch career_counseeling details
    url = f"{API_STUDIO_URL}getapi/naac01_students_benefited_for_career_counseeling_dc1/{id}"
    response = requests.get(url)

    if response.status_code == 200:
        career_counseeling = response.json()
        
        # Convert date format "2005-06-01" to "2005-2006" for display
        year_str = career_counseeling.get('year_of_activity', '')
        if year_str and isinstance(year_str, str) and len(year_str) >= 4:
            try:
                # Extract the year part and create range format
                start_year = year_str[:4]
                career_counseeling['year_range'] = f"{start_year}-{int(start_year) + 1}"
            except (ValueError, TypeError):
                career_counseeling['year_range'] = year_str
        else:
            career_counseeling['year_range'] = year_str
            
    else:
        return HttpResponse(f"Error fetching career_counseeling details: {response.text}", status=500)

    # Fetch media (child files) associated with this career_counseeling
    media_url = f"{API_STUDIO_URL}crudapp/get/media/naac01_students_benefited_for_career_counseeling_dc1_media/parent/{id}"
    media_response = requests.get(media_url)

    if media_response.status_code == 200:
        child_files = media_response.json()
    else:
        return HttpResponse(f"Failed to fetch media files: {media_response.text}", status=500)

    # Handle form submission (POST request)
    if request.method == "POST":
        # Prepare data for updating the career_counseeling/workshop
        update_url = f"{API_STUDIO_URL}updateapi/update/naac01_students_benefited_for_career_counseeling_dc1/{id}"

        # Convert selected year range back to date format for storage
        year_range = request.POST.get('year_of_activity')
        year_of_activity_date = None
        
        if year_range and '-' in year_range:
            try:
                start_year = int(year_range.split('-')[0])
                # Create a date object (using June 1st as default)
                year_of_activity_date = date(start_year, 6, 1).isoformat()
            except (ValueError, IndexError):
                year_of_activity_date = career_counseeling.get('year_of_activity')
        else:
            year_of_activity_date = career_counseeling.get('year_of_activity')

        payload = json.dumps({"data": {
            "department_name": request.POST.get('department_name', career_counseeling.get('department_name')),
            "staff_id": request.POST.get('staff_id', career_counseeling.get('staff_id')),
            "staff_name": request.POST.get('staff_name', career_counseeling.get('staff_name')),
            "counseling_detail": request.POST.get('counseling_detail', career_counseeling.get('counseling_detail')),
            "year_of_activity": year_of_activity_date,
            "number_of_students_attended": request.POST.get('number_of_students_attended',
                                                            career_counseeling.get('number_of_students_attended')),
            "number_of_students_placed": request.POST.get('number_of_students_placed',
                                                          career_counseeling.get('number_of_students_placed'))}})

        headers = {'Content-Type': 'application/json'}
        update_response = requests.put(update_url, headers=headers, data=payload)

        if update_response.status_code != 200:
            return HttpResponse(f"Failed to update career_counseeling details: {update_response.text}", status=500)

        # Handle media (file) uploads
        upload_errors = []

        for child in child_files:
            upload_id = child['psk_id']
            fields = ['PL', 'Circular', 'Brochure', 'RPP', 'DR', 'GTP', 'SA', 'FB', 'CS']

            for field in fields:
                uploaded_files = request.FILES.getlist(f'file_{upload_id}_{field}')

                if not uploaded_files:
                    continue  # Skip if no files are uploaded for this field

                for uploaded_file in uploaded_files:
                    # Validate file size and format
                    validate_file_size(uploaded_file)
                    validate_file_format(uploaded_file)

                    # Generate custom filename
                    current_year = datetime.now().year
                    staff_id = request.POST.get('staff_id')
                    custom_filename = f"{staff_id}_{field}_{current_year}_{uploaded_file.name}"

                    print(f"Generated filename for field {field}: {custom_filename}")  # Debug log

                    # Construct the upload URL and payload
                    upload_url = f"{API_STUDIO_URL}crudapp/upload/media/naac01_students_benefited_for_career_counseeling_dc1_media/{upload_id}"
                    files = {'media': (custom_filename, uploaded_file, uploaded_file.content_type)}
                    headers = {
                        'api_name': 'naac01_students_benefited_for_career_counseeling_dc1_media',
                        'psk_id': str(upload_id)
                    }
                    payload = {'parent_psk_id': id}

                    # Upload the file
                    upload_response = requests.put(upload_url, headers=headers, data=payload, files=files)

                if upload_response.status_code != 200:
                    upload_errors.append(f"Error uploading file for {upload_id}: {upload_response.text}")

        # Provide feedback to the user
        if upload_errors:
            for error in upload_errors:
                messages.error(request, error)
        else:
            messages.success(request, "Files uploaded successfully.")

        # Redirect to the career_counseeling view page after successful update
        return redirect('career_counseeling_list')

    # If not a POST request, render the update form
    return render(request, 'Career_counseeling_templates/career_counseeling_update.html', {
        'career_counseeling': career_counseeling, 
        'child_files': child_files, 
        'publication_year': publication_year
    })


# Delete career_counseeling
def career_counseeling_delete(request, id):
    # url = f"{API_STUDIO_URL}getapi/naac01_students_benefited_for_career_counseeling_dc1/{id}"
    # response = requests.get(url)
    # career_counseeling = response.json()
    payload = ""
    headers = {}

    # if request.method == 'POST':
    delete_url = f"{API_STUDIO_URL}deleteapi/delete/naac01_students_benefited_for_career_counseeling_dc1/{id}"
    delete_response = requests.request("DELETE", delete_url, headers=headers, data=payload)

    if delete_response.status_code == 200:
        return redirect('career_counseeling_list')
    else:
        return HttpResponse("Failed to delete participation: " + delete_response.text)

    # return render(request, 'career_counseeling_delete.html', {'career_counseeling': career_counseeling})

# def export_career_counseeling_to_pdf(parents, children=None, selected_options=None):
#     """
#     Export Books & Chapters data to PDF (landscape), grouped by staff.
#     Each staff has a mini-title "Books and Chapters", then Staff info in bold labels, then table.
#     Attachments are clickable links.
#     """
#     from reportlab.lib.pagesizes import A4, landscape
#     from reportlab.lib import colors
#     from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
#     from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, PageBreak
#     from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
#     from reportlab.lib.units import inch
#     from django.http import HttpResponse

#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="career_counseling.pdf"'

#     doc = SimpleDocTemplate(
#         response,
#         pagesize=landscape(A4),
#         topMargin=0.5*inch,
#         bottomMargin=0.5*inch,
#         leftMargin=0.5*inch,
#         rightMargin=0.5*inch
#     )

#     elements = []
#     styles = getSampleStyleSheet()

#     # Styles
#     title_style = ParagraphStyle('Title', parent=styles['Heading2'], alignment=TA_CENTER, spaceAfter=12)
#     left_style = ParagraphStyle('LeftStyle', parent=styles['Normal'], fontSize=9,
#                                 alignment=TA_LEFT, textColor=colors.HexColor('#2c3e50'))
#     right_style = ParagraphStyle('RightStyle', parent=styles['Normal'], fontSize=9,
#                                  alignment=TA_RIGHT, textColor=colors.HexColor('#2c3e50'))
#     table_header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontSize=9,
#                                         alignment=TA_CENTER, textColor=colors.white, fontName='Helvetica-Bold')
#     table_cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=8,
#                                       alignment=TA_CENTER, textColor=colors.HexColor('#2c3e50'))
#     attachment_link_style = ParagraphStyle('AttachmentLink', parent=styles['Normal'], fontSize=6,
#                                            alignment=TA_CENTER, textColor=colors.HexColor('#1a5276'))

#     # Sort parents by staff
#     parents_sorted = sorted(parents, key=lambda x: (x.get('staff_id',''), x.get('counseling_detail','')))
#     current_staff = None
#     table_data = []
#     col_widths = [1.2*inch, 3*inch, 1*inch, 1*inch, 2*inch, 2*inch]

#     for parent in parents_sorted:
#         staff_id = parent.get('staff_id', 'N/A')
#         staff_name = parent.get('staff_name', 'N/A')
#         department_name = parent.get('department_name', 'N/A')

#         # New staff: flush previous table + page break
#         if current_staff and current_staff != staff_id:
#             table = Table(table_data, colWidths=col_widths)
#             table.setStyle(TableStyle([
#                 ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
#                 ('TEXTCOLOR', (0,0), (-1,0), colors.white),
#                 ('ALIGN',(0,0),(-1,-1),'CENTER'),
#                 ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
#                 ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
#                 ('FONTSIZE', (0,0), (-1,-1), 8),
#                 ('LEFTPADDING', (0,0), (-1,-1), 4),
#                 ('RIGHTPADDING', (0,0), (-1,-1), 4),
#                 ('BOTTOMPADDING', (0,0), (-1,-1), 3),
#                 ('TOPPADDING', (0,0), (-1,-1), 3),
#             ]))
#             elements.append(table)
#             elements.append(PageBreak())
#             table_data = []

#         if current_staff != staff_id:
#             # Add mini-title for this staff
#             elements.append(Paragraph("Career Counseling", title_style))
#             elements.append(Spacer(1, 6))

#             # Staff info with bold labels
#             left_paragraph = Paragraph(
#                 f"<b>Staff ID:</b> {staff_id}<br/><b>Staff Name:</b> {staff_name}", left_style)
#             right_paragraph = Paragraph(f"<b>Department:</b> {department_name}", right_style)

#             total_width = sum(col_widths)
#             info_table = Table([[left_paragraph, right_paragraph]], colWidths=[total_width*0.5]*2)
#             info_table.setStyle(TableStyle([
#                 ('VALIGN', (0, 0), (-1, -1), 'TOP'),
#                 ('ALIGN', (0,0), (0,0), 'LEFT'),
#                 ('ALIGN', (1,0), (1,0), 'RIGHT'),
#                 ('LEFTPADDING', (0,0), (-1,-1), 0),
#                 ('RIGHTPADDING', (0,0), (-1,-1), 0),
#                 ('TOPPADDING', (0,0), (-1,-1), 2),
#                 ('BOTTOMPADDING', (0,0), (-1,-1), 4),
#             ]))
#             elements.append(info_table)
#             elements.append(Spacer(1, 6))

#             # Table headers
#             headers = ['staff Name', 'Year of Activity', 'Counseling Detail', 'No. of Students Attended', 'No. of Students Placed', 'Attachments']
#             table_data.append([Paragraph(h, table_header_style) for h in headers])
#             current_staff = staff_id

#         # Media attachments
#         media_files = parent.get('media_files', [])
#         if media_files:
#             media_text = []
#             for m in media_files:
#                 filename = m.get('file_name', 'Unknown')
#                 url = m.get('direct_api_url', '#')
#                 media_text.append(f'<link href="{url}" color="blue">{filename}</link>')
#             media_paragraph = Paragraph('<br/>'.join(media_text), attachment_link_style)
#         else:
#             media_paragraph = Paragraph("No attachments", table_cell_style)

#         row = [
#             Paragraph(str(parent.get('staff_name', '')), table_cell_style),
#             Paragraph(str(parent.get('year_of_activity', '')), table_cell_style),
#             Paragraph(str(parent.get('counseling_detail', '')), table_cell_style),
#             Paragraph(str(parent.get('number_of_students_attended', '')), table_cell_style),
#             Paragraph(str(parent.get('number_of_students_placed', '')), table_cell_style),
#             media_paragraph
#         ]
#         table_data.append(row)

#     # Add last staff table
#     if table_data:
#         table = Table(table_data, colWidths=col_widths)
#         table.setStyle(TableStyle([
#             ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
#             ('TEXTCOLOR', (0,0), (-1,0), colors.white),
#             ('ALIGN',(0,0),(-1,-1),'CENTER'),
#             ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
#             ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
#             ('FONTSIZE', (0,0), (-1,-1), 8),
#             ('LEFTPADDING', (0,0), (-1,-1), 4),
#             ('RIGHTPADDING', (0,0), (-1,-1), 4),
#             ('BOTTOMPADDING', (0,0), (-1,-1), 3),
#             ('TOPPADDING', (0,0), (-1,-1), 3),
#         ]))
#         elements.append(table)

#     doc.build(elements)
#     return response

def export_career_counseeling_to_pdf(parents, children=None, selected_options=None):
    """
    Export Books & Chapters data to PDF (landscape), grouped by staff.
    Each staff has a mini-title "Books and Chapters", then Staff info in bold labels, then table.
    Attachments are clickable links.
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, PageBreak
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.lib.units import inch
    from django.http import HttpResponse

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="career_counseling.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        topMargin=0.5*inch,
        bottomMargin=0.5*inch,
        leftMargin=0.5*inch,
        rightMargin=0.5*inch
    )

    elements = []
    styles = getSampleStyleSheet()

    # Styles
    title_style = ParagraphStyle('Title', parent=styles['Heading2'], alignment=TA_CENTER, spaceAfter=12)
    left_style = ParagraphStyle('LeftStyle', parent=styles['Normal'], fontSize=9,
                                alignment=TA_LEFT, textColor=colors.HexColor('#2c3e50'))
    right_style = ParagraphStyle('RightStyle', parent=styles['Normal'], fontSize=9,
                                 alignment=TA_RIGHT, textColor=colors.HexColor('#2c3e50'))
    table_header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontSize=9,
                                        alignment=TA_CENTER, textColor=colors.white, fontName='Helvetica-Bold')
    table_cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=8,
                                      alignment=TA_CENTER, textColor=colors.HexColor('#2c3e50'))
    attachment_link_style = ParagraphStyle('AttachmentLink', parent=styles['Normal'], fontSize=6,
                                           alignment=TA_CENTER, textColor=colors.HexColor('#1a5276'))
    no_data_style = ParagraphStyle('NoData', parent=styles['Normal'], fontSize=10,
                                   alignment=TA_CENTER, textColor=colors.HexColor('#7f8c8d'),
                                   fontStyle='italic', spaceBefore=12, spaceAfter=12)

    # Check if there are any parents
    if not parents:
        elements.append(Paragraph("Career Counseling", title_style))
        elements.append(Paragraph("No data available", no_data_style))
        doc.build(elements)
        return response

    # Sort parents by staff
    parents_sorted = sorted(parents, key=lambda x: (x.get('staff_id',''), x.get('counseling_detail','')))
    current_staff = None
    table_data = []
    col_widths = [1.2*inch, 3*inch, 1*inch, 1*inch, 2*inch, 2*inch]

    for parent in parents_sorted:
        staff_id = parent.get('staff_id', 'N/A')
        staff_name = parent.get('staff_name', 'N/A')
        department_name = parent.get('department_name', 'N/A')

        # New staff: flush previous table + page break
        if current_staff and current_staff != staff_id:
            # Only add table if there's data (more than just headers)
            if len(table_data) > 1:
                table = Table(table_data, colWidths=col_widths)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                    ('ALIGN',(0,0),(-1,-1),'CENTER'),
                    ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
                    ('FONTSIZE', (0,0), (-1,-1), 8),
                    ('LEFTPADDING', (0,0), (-1,-1), 4),
                    ('RIGHTPADDING', (0,0), (-1,-1), 4),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 3),
                    ('TOPPADDING', (0,0), (-1,-1), 3),
                ]))
                elements.append(table)
            else:
                # Only headers exist, meaning no data for this staff
                elements.append(Paragraph("No data available for this staff", no_data_style))
            
            elements.append(PageBreak())
            table_data = []

        if current_staff != staff_id:
            # Add mini-title for this staff
            elements.append(Paragraph("Career Counseling", title_style))
            elements.append(Spacer(1, 6))

            # Staff info with bold labels
            left_paragraph = Paragraph(
                f"<b>Staff ID:</b> {staff_id}<br/><b>Staff Name:</b> {staff_name}", left_style)
            right_paragraph = Paragraph(f"<b>Department:</b> {department_name}", right_style)

            total_width = sum(col_widths)
            info_table = Table([[left_paragraph, right_paragraph]], colWidths=[total_width*0.5]*2)
            info_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0,0), (0,0), 'LEFT'),
                ('ALIGN', (1,0), (1,0), 'RIGHT'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('TOPPADDING', (0,0), (-1,-1), 2),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 6))

            # Table headers
            headers = ['staff Name', 'Year of Activity', 'Counseling Detail', 'No. of Students Attended', 'No. of Students Placed', 'Attachments']
            table_data.append([Paragraph(h, table_header_style) for h in headers])
            current_staff = staff_id

        # Media attachments
        media_files = parent.get('media_files', [])
        if media_files:
            media_text = []
            for m in media_files:
                filename = m.get('file_name', 'Unknown')
                url = m.get('direct_api_url', '#')
                media_text.append(f'<link href="{url}" color="blue">{filename}</link>')
            media_paragraph = Paragraph('<br/>'.join(media_text), attachment_link_style)
        else:
            media_paragraph = Paragraph("No attachments", table_cell_style)

        row = [
            Paragraph(str(parent.get('staff_name', '')), table_cell_style),
            Paragraph(str(parent.get('year_of_activity', '')), table_cell_style),
            Paragraph(str(parent.get('counseling_detail', '')), table_cell_style),
            Paragraph(str(parent.get('number_of_students_attended', '')), table_cell_style),
            Paragraph(str(parent.get('number_of_students_placed', '')), table_cell_style),
            media_paragraph
        ]
        table_data.append(row)

    # Add last staff table
    if table_data:
        # Check if there's actual data (more than just headers)
        if len(table_data) > 1:
            table = Table(table_data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#02548b')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('ALIGN',(0,0),(-1,-1),'CENTER'),
                ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d6d6d6')),
                ('FONTSIZE', (0,0), (-1,-1), 8),
                ('LEFTPADDING', (0,0), (-1,-1), 4),
                ('RIGHTPADDING', (0,0), (-1,-1), 4),
                ('BOTTOMPADDING', (0,0), (-1,-1), 3),
                ('TOPPADDING', (0,0), (-1,-1), 3),
            ]))
            elements.append(table)
        else:
            # Only headers exist, meaning no data for this staff
            elements.append(Paragraph("No data available for this staff", no_data_style))

    doc.build(elements)
    return response

# def export_career_counseeling_to_excel(parents, children=None):
#     """
#     Export Books & Chapters data to Excel from filtered parents,
#     including clickable media file links based on upload field codes (CCP, CC, IP, POI).
#     """
#     import io
#     from django.http import HttpResponse
#     from openpyxl import Workbook
#     from openpyxl.styles import Font, Alignment

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Career Counseling"

#     # Headers including media attachments
#     headers = [
#         "Staff ID", "Staff Name", "Department", "Year of Activity",
#         "Counseling Detail", "number_of_students_attended", "number_of_students_placed",
#         "Permission Letter", "Circular", "Brochure",
#         "Resource person profile", "Detailed Report-Objective & Outcome", "Geo Tagged Photos",
#         "Student Attendance", "Feedback", "Certificate sample"
#     ]
#     ws.append(headers)

#     # Style headers
#     for col in range(1, len(headers) + 1):
#         cell = ws.cell(row=1, column=col)
#         cell.font = Font(bold=True)
#         cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

#     # Map upload field codes to friendly names
#     field_mapping = {
#         "PL": "Permission Letter",
#         "Circular": "Circular",
#         "Brochure": "Brochure",
#         "RPP": "Resource person profile",
#         "DR": "Detailed Report-Objective & Outcome",
#         "GTP": "Geo Tagged Photos",
#         "SA": "Student Attendance",
#         "FB": "Feedback",
#         "CS": "Certificate sample"
#     }

#     # Populate rows
#     for p in parents:
#         # Initialize media categories
#         media_mapping = {v: [] for v in field_mapping.values()}

#         # Categorize each file based on upload code in filename
#         for media in p.get("media_files", []):
#             file_name = media.get("file_name", "Unknown")
#             url = media.get("direct_api_url", "")
#             if not url:
#                 continue

#             lname = file_name.upper()

#             # Detect the code (CCP, CC, IP, POI) from filename
#             matched = False
#             for code, category in field_mapping.items():
#                 if f"_{code}_" in lname:
#                     media_mapping[category].append((file_name, url))
#                     matched = True
#                     break

#             # Optional fallback: if none matched, put in "Copy of the Chapters"
#             if not matched:
#                 media_mapping["Copy of the Chapters"].append((file_name, url))

#         # Create the base info row (first 13 columns)
#         base_row = [
#             p.get("staff_id", ""),
#             p.get("staff_name", ""),
#             p.get("department_name", ""),
#             p.get("year_of_activity", ""),
#             p.get("counseling_detail", ""),
#             p.get("number_of_students_attended", ""),
#             p.get("number_of_students_placed", ""),
#             # p.get("name_of_conference", ""),
#             # p.get("year_of_publication", ""),
#             # p.get("national_or_international", ""),
#             # p.get("isbn_or_issn_number", ""),
#             # p.get("affiliating_institute", ""),
#             # p.get("name_of_the_publisher", "")
#         ]

#         # Append placeholders for media columns
#         ws.append(base_row + ["", "", "", ""])
#         row_idx = ws.max_row  # current row number

#         # Map category to column
#         media_columns = {
#             "Permission Letter": 8,
#             "Circular": 9,
#             "Brochure": 10,
#             "Resource person profile": 11,
#             "Detailed Report-Objective & Outcome": 12,
#             "Geo Tagged Photos": 13,
#             "Student Attendance": 14,
#             "Feedback": 15,
#             "Certificate sample": 16
#         }

#         # Add clickable hyperlinks
#         for category, files in media_mapping.items():
#             cell = ws.cell(row=row_idx, column=media_columns[category])
#             if files:
#                 # Show all filenames joined by line breaks, hyperlink first
#                 cell.value = "\n".join([f for f, _ in files])
#                 cell.hyperlink = files[0][1]  # link to first file
#                 cell.font = Font(color="0000EE", underline="single")
#             else:
#                 cell.value = "-"
#             cell.alignment = Alignment(wrap_text=True, vertical="top")

#     # Auto-fit column widths
#     for col_idx, column_cells in enumerate(ws.columns, 1):
#         max_length = max(len(str(cell.value or "")) for cell in column_cells)
#         ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

#     # Adjust row heights for multiline cells
#     for row_idx in range(1, ws.max_row + 1):
#         max_lines = max(str(ws.cell(row=row_idx, column=col).value or "").count("\n") + 1 for col in range(1, len(headers) + 1))
#         ws.row_dimensions[row_idx].height = min(max_lines * 15, 120)

#     # Freeze header and enable filters
#     ws.freeze_panes = "A2"
#     ws.auto_filter.ref = ws.dimensions

#     # Return as Excel file
#     response = HttpResponse(
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )
#     response["Content-Disposition"] = 'attachment; filename="career_counseling.xlsx"'

#     with io.BytesIO() as buffer:
#         wb.save(buffer)
#         buffer.seek(0)
#         response.write(buffer.getvalue())

#     return response

# def export_career_counseeling_to_excel(parents, children=None):
# #     """
# #     Export Books & Chapters data to Excel from filtered parents,
# #     including clickable media file links based on upload field codes (CCP, CC, IP, POI).
# #     """
# #     import io
# #     from django.http import HttpResponse
# #     from openpyxl import Workbook
# #     from openpyxl.styles import Font, Alignment

# #     wb = Workbook()
# #     ws = wb.active
# #     ws.title = "Career Counseling"

# #     # Headers including media attachments
# #     headers = [
# #         "Staff ID", "Staff Name", "Department", "Year of Activity",
# #         "Counseling Detail", "number_of_students_attended", "number_of_students_placed",
# #         "Permission Letter", "Circular", "Brochure",
# #         "Resource person profile", "Detailed Report-Objective & Outcome", "Geo Tagged Photos",
# #         "Student Attendance", "Feedback", "Certificate sample"
# #     ]
# #     ws.append(headers)

# #     # Style headers
# #     for col in range(1, len(headers) + 1):
# #         cell = ws.cell(row=1, column=col)
# #         cell.font = Font(bold=True)
# #         cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# #     # Map upload field codes to friendly names
# #     field_mapping = {
# #         "PL": "Permission Letter",
# #         "Circular": "Circular",
# #         "Brochure": "Brochure",
# #         "RPP": "Resource person profile",
# #         "DR": "Detailed Report-Objective & Outcome",
# #         "GTP": "Geo Tagged Photos",
# #         "SA": "Student Attendance",
# #         "FB": "Feedback",
# #         "CS": "Certificate sample"
# #     }

# #     # Map category to column
# #     media_columns = {
# #         "Permission Letter": 8,
# #         "Circular": 9,
# #         "Brochure": 10,
# #         "Resource person profile": 11,
# #         "Detailed Report-Objective & Outcome": 12,
# #         "Geo Tagged Photos": 13,
# #         "Student Attendance": 14,
# #         "Feedback": 15,
# #         "Certificate sample": 16
# #     }

# #     # Populate rows
# #     for p in parents:
# #         # Initialize media categories - only use categories that exist in field_mapping
# #         media_mapping = {v: [] for v in field_mapping.values()}

# #         # Categorize each file based on upload code in filename
# #         for media in p.get("media_files", []):
# #             file_name = media.get("file_name", "Unknown")
# #             url = media.get("direct_api_url", "")
# #             if not url:
# #                 continue

# #             lname = file_name.upper()
# #             matched = False

# #             # Detect the code from filename
# #             for code, category in field_mapping.items():
# #                 if f"_{code}_" in lname:
# #                     media_mapping[category].append((file_name, url))
# #                     matched = True
# #                     break

# #             # If no match found, skip or assign to a default category
# #             # Remove the fallback to "Copy of the Chapters" since it doesn't exist in our mapping
# #             if not matched:
# #                 # You can choose to skip unmatched files or assign to a specific category
# #                 # For now, let's skip unmatched files
# #                 continue

# #         # Create the base info row
# #         base_row = [
# #             p.get("staff_id", ""),
# #             p.get("staff_name", ""),
# #             p.get("department_name", ""),
# #             p.get("year_of_activity", ""),
# #             p.get("counseling_detail", ""),
# #             p.get("number_of_students_attended", ""),
# #             p.get("number_of_students_placed", ""),
# #         ]

# #         # Append the base row first
# #         ws.append(base_row)
# #         row_idx = ws.max_row  # current row number

# #         # Now add media files to their respective columns
# #         for category, files in media_mapping.items():
# #             if category in media_columns:  # Double check the category exists
# #                 col_idx = media_columns[category]
# #                 cell = ws.cell(row=row_idx, column=col_idx)
                
# #                 if files:
# #                     # Show all filenames joined by line breaks, hyperlink first
# #                     cell.value = "\n".join([f for f, _ in files])
# #                     if files[0][1]:  # Check if URL exists
# #                         cell.hyperlink = files[0][1]  # link to first file
# #                         cell.font = Font(color="0000EE", underline="single")
# #                     else:
# #                         cell.font = Font()  # Regular font if no URL
# #                 else:
# #                     cell.value = "-"
# #                     cell.font = Font()  # Regular font
                
# #                 cell.alignment = Alignment(wrap_text=True, vertical="top")

# #     # Auto-fit column widths
# #     for col_idx, column_cells in enumerate(ws.columns, 1):
# #         max_length = 0
# #         for cell in column_cells:
# #             if cell.value:
# #                 # Count the maximum line length for multiline cells
# #                 lines = str(cell.value).split('\n')
# #                 max_line_length = max(len(line) for line in lines)
# #                 max_length = max(max_length, max_line_length)
# #         ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

# #     # Adjust row heights for multiline cells
# #     for row_idx in range(1, ws.max_row + 1):
# #         max_lines = 1
# #         for col in range(1, len(headers) + 1):
# #             cell_value = str(ws.cell(row=row_idx, column=col).value or "")
# #             line_count = cell_value.count("\n") + 1
# #             max_lines = max(max_lines, line_count)
# #         ws.row_dimensions[row_idx].height = min(max_lines * 15, 120)

# #     # Freeze header and enable filters
# #     ws.freeze_panes = "A2"
# #     ws.auto_filter.ref = ws.dimensions

# #     # Return as Excel file
# #     response = HttpResponse(
# #         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
# #     )
# #     response["Content-Disposition"] = 'attachment; filename="career_counseling.xlsx"'

# #     with io.BytesIO() as buffer:
# #         wb.save(buffer)
# #         buffer.seek(0)
# #         response.write(buffer.getvalue())

# #     return response

def export_career_counseeling_to_excel(parents, children=None):
    """
    Export Books & Chapters data to Excel from filtered parents,
    including clickable media file links based on upload field codes (CCP, CC, IP, POI).
    """
    import io
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Career Counseling"

    # Headers including media attachments
    headers = [
        "Staff ID", "Staff Name", "Department", "Year of Activity",
        "Counseling Detail", "number_of_students_attended", "number_of_students_placed",
        "Permission Letter", "Circular", "Brochure",
        "Resource person profile", "Detailed Report-Objective & Outcome", "Geo Tagged Photos",
        "Student Attendance", "Feedback", "Certificate sample"
    ]
    ws.append(headers)

    # Style headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Map upload field codes to friendly names
    field_mapping = {
        "PL": "Permission Letter",
        "Circular": "Circular",
        "Brochure": "Brochure",
        "RPP": "Resource person profile",
        "DR": "Detailed Report-Objective & Outcome",
        "GTP": "Geo Tagged Photos",
        "SA": "Student Attendance",
        "FB": "Feedback",
        "CS": "Certificate sample"
    }

    # Map category to column
    media_columns = {
        "Permission Letter": 8,
        "Circular": 9,
        "Brochure": 10,
        "Resource person profile": 11,
        "Detailed Report-Objective & Outcome": 12,
        "Geo Tagged Photos": 13,
        "Student Attendance": 14,
        "Feedback": 15,
        "Certificate sample": 16
    }

    # Populate rows
    for p in parents:
        # Initialize media categories - only use categories that exist in field_mapping
        media_mapping = {v: [] for v in field_mapping.values()}

        # Categorize each file based on upload code in filename
        for media in p.get("media_files", []):
            file_name = media.get("file_name", "Unknown")
            url = media.get("direct_api_url", "")
            if not url:
                continue

            # Use original filename for matching (case sensitive)
            matched = False

            # Detect the code from filename - try multiple patterns
            for code, category in field_mapping.items():
                # Try different patterns to match the code in filename (case sensitive)
                patterns = [
                    f"_{code}_",  # _PL_, _Circular_, _Brochure_
                    f"_{code}.",  # _PL.pdf, _Circular.jpg, etc.
                    f"-{code}-",  # -PL-, -Circular-, etc.
                    f"-{code}.",  # -PL.pdf, -Circular.jpg, etc.
                    f"{code}_",   # PL_, Circular_, Brochure_ (at start)
                    f"_{code}",   # _PL, _Circular, _Brochure (at end)
                ]
                
                for pattern in patterns:
                    if pattern in file_name:
                        media_mapping[category].append((file_name, url))
                        matched = True
                        print(f"Matched '{file_name}' to '{category}' using pattern '{pattern}'")
                        break
                if matched:
                    break

            # If no match found with patterns, try case-insensitive partial matching
            if not matched:
                file_name_upper = file_name.upper()
                for code, category in field_mapping.items():
                    code_upper = code.upper()
                    # Check if code appears anywhere in filename (case insensitive)
                    if code_upper in file_name_upper:
                        media_mapping[category].append((file_name, url))
                        matched = True
                        print(f"Matched '{file_name}' to '{category}' using case-insensitive matching")
                        break

        # Create the base info row
        base_row = [
            p.get("staff_id", ""),
            p.get("staff_name", ""),
            p.get("department_name", ""),
            p.get("year_of_activity", ""),
            p.get("counseling_detail", ""),
            p.get("number_of_students_attended", ""),
            p.get("number_of_students_placed", ""),
        ]

        # Append the base row first
        ws.append(base_row)
        row_idx = ws.max_row  # current row number

        # Now add media files to their respective columns
        for category, files in media_mapping.items():
            if category in media_columns:  # Double check the category exists
                col_idx = media_columns[category]
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if files:
                    # Show all filenames joined by line breaks
                    filenames = [f for f, _ in files]
                    cell.value = "\n".join(filenames)
                    
                    # Add hyperlink to first file if URL exists
                    if files[0][1]:
                        cell.hyperlink = files[0][1]
                        cell.font = Font(color="0000EE", underline="single")
                    else:
                        cell.font = Font()  # Regular font if no URL
                else:
                    cell.value = "-"
                    cell.font = Font()  # Regular font
                
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-fit column widths
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            if cell.value:
                # Count the maximum line length for multiline cells
                lines = str(cell.value).split('\n')
                max_line_length = max(len(line) for line in lines)
                max_length = max(max_length, max_line_length)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

    # Adjust row heights for multiline cells
    for row_idx in range(1, ws.max_row + 1):
        max_lines = 1
        for col in range(1, len(headers) + 1):
            cell_value = str(ws.cell(row=row_idx, column=col).value or "")
            line_count = cell_value.count("\n") + 1
            max_lines = max(max_lines, line_count)
        ws.row_dimensions[row_idx].height = min(max_lines * 15, 120)

    # Freeze header and enable filters
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Return as Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="career_counseling.xlsx"'

    with io.BytesIO() as buffer:
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())

    return response