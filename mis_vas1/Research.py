import hashlib
import json
import re
from typing import Counter
import requests
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from MIS.functions import validate_file_format, validate_file_size
from datetime import datetime
from user_management.settings_views import *
# from Cryptodome.Cipher import AES
from .Faculty import *
from .Courses_project import *
from .books import *
from .Career_counseeling import *
from .collaborative_students import *
from .competitve_examination import *
from .Extentions import *
from .Faculty_exchange import *
from .government import *
from .program_offered import *
from .workshop import *

API_STUDIO_URL = user_bundle_settings()

def research_key():
    url = "https://api.hcaschennai.edu.in/auth/token"
    payload = json.dumps({
        "secret_key": "C4ZoXbsAnHLjk1Xyz4QPT2eoiNx6K6fo"
    })
    headers = {'Content-Type': 'application/json'}

    response = requests.post(url, headers=headers, data=payload)
    if response.status_code == 200:
        res_data = response.json()
        access_token = res_data.get('access_token')
        token_type = res_data.get('token_type')
        return access_token, token_type
    return None, None


def get_research_data(access_token, token_type):
    url = "https://api.hcaschennai.edu.in/sqlviews/api/v1/auth/get_response_data"
    payload = json.dumps({"psk_uid": "51a531b4-bd55-491c-861d-a8d7227b325b","project": "hcas","data": {}})
    headers = {'Content-Type': 'application/json','Authorization': f'{token_type} {access_token}'}

    response = requests.post(url, headers=headers, data=payload)
    if response.status_code == 200:
        return response.json()
    return []


def research_create(request):
    error_message = None

    access_token, token_type = research_key()
    if not access_token or not token_type:
        error_message = 'Failed to get access token from API.'
        return render(request, 'Research_templates/research_create.html', {'error': error_message})

    research_data = get_research_data(access_token, token_type)
    if not research_data:
        error_message = 'Failed to fetch staff data.'
        return render(request, 'Research_templates/research_create.html', {'error': error_message})
    
    current_year = datetime.now().year
    publication_years = [f"{y}-{y+1}" for y in range(current_year, 2010, -1)]
    print("publication_year: ", publication_years)
    # selected_year = 2025
    
    user = get_settings(request)
    username = user.get('username')  # This is the username (staff_id), which seems to be the key you need
    # username = 'CS-T151'
    # print("staff_name", username)

    # Pre-filling staff name and department based on staff_id
    selected_faculty = None
    for faculty in research_data:
        # print("research_data", research_data)
        if faculty['stf_id'] == username:
            selected_faculty = faculty
            break

    staff_name = selected_faculty.get('stf_name', '') if selected_faculty else ''
    department_name = selected_faculty.get('department', '') if selected_faculty else ''
    
    # Initialize employee_id to None for GET requests
    employee_id = None  # Default value for GET request

    # Define the fields here so it's available for both GET and POST requests
    fields = []

    if request.method == 'POST':
        employee_id = username  # Only set in POST requests
        department_name = request.POST.get('department_name')
        staff_name = request.POST.get('staff_name')
        paper_title = request.POST.get('paper_title')
        author_name = request.POST.get('author_name')
        journal_name = request.POST.get('journal_name')
        publication_year = request.POST.get('publication_year')
        issn_number = request.POST.get('issn_number')
        ugc_recognition_link = request.POST.get('ugc_recognition_link')
        issue_date = request.POST.get('issue_date')
        journal_nature = request.POST.get('journal_nature')
        doi_url_link = request.POST.get('doi_url_link')
        
        selected_faculty = None

        # Loop through each faculty in the research_data to find the matching stf_id
        for faculty in research_data:
            if faculty['stf_id'] == employee_id:
                selected_faculty = faculty  # Assign the matching faculty
                break  # Stop the loop once the faculty is found

        # If a matching faculty was found, extract the data
        if selected_faculty:
            depcode = selected_faculty.get('depcode', '')
            department_name = selected_faculty.get('department', '')
            staff_name = selected_faculty.get('stf_name', '')

        url = f"{API_STUDIO_URL}postapi/create/naac01_research_article_publication_dc1"
        payload = json.dumps({
            "data": {
                "employee_id": employee_id,
                "department_name": department_name,
                "staff_name": staff_name,
                "paper_title": paper_title,
                "author_name": author_name,
                "journal_name": journal_name,
                "publication_year": publication_year,
                "issn_number": issn_number,
                "ugc_recognition_link": ugc_recognition_link,
                "issue_date": issue_date,
                "journal_nature": journal_nature,
                "doi_url_link": doi_url_link
            }
        })
        headers = {'Content-Type': 'application/json'}

        response = requests.post(url, headers=headers, data=payload)

        if response.status_code == 200:
            file_data = response.json()
            print("Status Code:", response.status_code)
            print("Response Text:", response.text)
            print("file_data:", file_data)
            psk_id = file_data.get('psk_id')
        else:
            # Print the response body for debugging purposes
            print(f"API Error Response: {response.text}")
            messages.error(request, message=f"Failed to create. Error: {response.text}")
            return render(request, 'Research_templates/research_create.html')

        # # Handle file uploads
        # upload_url = f"{API_STUDIO_URL}crudapp/upload/media/naac01_research_article_publication_dc1_media"
        # uploaded_files = request.FILES.getlist('file')

        # if not uploaded_files:
        #     messages.error(request, message="No files selected for upload.")
        #     return render(request, 'Research_templates/research_create.html')

        # fields = ['JCP', 'IP', 'IF', 'CI']
        # for field in fields:
        #     current_year = datetime.now().year
        #     employee_id = username  # Now getting employee_id in the POST block
        #     print("employee_id:", employee_id)
        #     for uploaded_file in uploaded_files:
        #         validate_file_size(uploaded_file)
        #         validate_file_format(uploaded_file)
        #         file_type = uploaded_file.content_type
        #         custom_filename = f"{employee_id}_{field}_{current_year}_{uploaded_file.name}"
        #         print(f"Generated filename: {custom_filename}")  # Print the filename for each iteration
        #         payload = {'parent_psk_id': psk_id}
        #         files = {'media': (custom_filename, uploaded_file, file_type)}
        #         upload_headers = {'api_name': 'naac01_research_article_publication_dc1_media'}

        #         # Make API call to upload the file
        #         upload_response = requests.post(upload_url, headers=upload_headers, data=payload, files=files)
        # Handle file uploads
        upload_url = f"{API_STUDIO_URL}crudapp/upload/media/naac01_research_article_publication_dc1_media"
        uploaded_files = request.FILES.getlist('file')
        fields = ['JCP', 'IP', 'IF', 'CI']  # fixed fields

        if len(uploaded_files) != len(fields):
            messages.error(request, message=f"Please upload exactly {len(fields)} files corresponding to {fields}")
            return render(request, 'Research_templates/research_create.html')

        current_year = datetime.now().year

        for field, uploaded_file in zip(fields, uploaded_files):
            validate_file_size(uploaded_file)
            validate_file_format(uploaded_file)
            file_type = uploaded_file.content_type
            custom_filename = f"{employee_id}_{field}_{current_year}_{uploaded_file.name}"
            
            payload = {'parent_psk_id': psk_id}
            files = {'media': (custom_filename, uploaded_file, file_type)}
            upload_headers = {'api_name': 'naac01_research_article_publication_dc1_media'}

            upload_response = requests.post(upload_url, headers=upload_headers, data=payload, files=files)
            if upload_response.status_code != 200:
                messages.error(request, message=f"File upload failed for {uploaded_file.name}. Error: {upload_response.text}")
                return redirect('research_list')

        messages.success(request, message="Documents uploaded successfully.")
        return redirect('research_list')

    return render(request, 'Research_templates/research_create.html', {
        'research_data': research_data,
        'fields': fields,
        'username': username,  # Pass 'username' (which is the staff_id) to the template
        'department_name': department_name,
        'staff_name': staff_name,
        'publication_years':publication_years  
    })




def research_view(request, id):
    url = f"{API_STUDIO_URL}getapi/naac01_research_article_publication_dc1/{id}"
    response = requests.get(url)

    if response.status_code == 200:
        research_data = response.json()
        return render(request, "Research_templates/research_view.html", {'research': research_data})

    return HttpResponse(f"Error fetching Course details: {response.text}", status=500)


# def research_list(request):
#     # URL to get research article publication data
#     url = f"{API_STUDIO_URL}getapi/naac01_research_article_publication_dc1/all"
#     response = requests.get(url)

#     # Check if the response is successful
#     if response.status_code == 200:
#         researchs = response.json()
#         print("researches:", researchs)
#     else:
#         # If API call fails, return an empty list and an error message
#         return render(request, 'Research_templates/research_list.html', {'researchs': [], 'error': 'Failed to fetch research data.'})

#     # Get the current user (staff_id or username)
#     user = get_settings(request)
#     username = user['username']
#     # username = 'CS-T103'

#     # Filter the research articles based on the staff_id (username)
#     filtered_researchs = [research for research in researchs if research.get('employee_id') == username]
#     print("filtered_researchs:", filtered_researchs)
    
#     # If no data is found for the username, return the same page with an empty list
#     if not filtered_researchs:
#         return render(request, 'Research_templates/research_list.html', {'researchs': [], 'error': 'No research articles found for your account.'})

#     # Return the filtered data to the template
#     return render(request, 'Research_templates/research_list.html', {'researchs': filtered_researchs})

# def research_list(request):
    # url = f"{API_STUDIO_URL}getapi/naac01_research_article_publication_dc1/all"
    # response = requests.get(url)
    
    # if response.status_code == 200:
        # researchs = response.json()
        # print("researchs:", researchs)
    # else:
        # return HttpResponse("API Call Is Not Working")

    # user = get_settings(request)
    # username = user.get('username')
    # filtered_researchs = [research for research in researchs if research.get('staff_id') == username]
    
    # selected_staff_id = request.GET.get('staff_id')
    # print("selected_staff_id:", selected_staff_id)
    # if selected_staff_id:
        # filtered_researchs = [research for research in researchs if research.get('employee_id') == selected_staff_id]
        # print("filtered_researches", filtered_researchs)
    # if not filtered_researchs:
        # return render(request, 'Research_templates/research_list.html', {"researchs": researchs})
    
    # return render(request, 'Research_templates/research_list.html', {"researchs": filtered_researchs})
    
    
    
def research_list(request):
    url = f"{API_STUDIO_URL}getapi/naac01_research_article_publication_dc1/all"
    response = requests.get(url)
    
    if response.status_code == 200:
        researchs = response.json()
        print("researchs:", researchs)
    else:
        return HttpResponse("API Call Is Not Working")

    user = get_settings(request)
    username = user.get('username')
    # username = 'CS-T151'
    
    filtered_researchs = [research for research in researchs if research.get('employee_id') == username]
    
    selected_staff_id = request.GET.get('staff_id')
    #from_dashboard = request.GET.get('from') == 'dashboard' or 'admin_hod_dash' or 'admin_dash' or 'department_dashboard'
    if selected_staff_id:
        filtered_researchs = [research for research in researchs if research.get('employee_id') == selected_staff_id]
        print("filtered_researchs:", filtered_researchs)
    
    selected_department = request.GET.get('department')
    if selected_department:
        filtered_researchs = [research for research in researchs if research.get('department_name') == selected_department]

    if not filtered_researchs:
        return render(request, 'Research_templates/research_list.html', {"researchs": []})
    
    return render(request, 'Research_templates/research_list.html', {"researchs": filtered_researchs, 
    #'from_dashboard': from_dashboard
    })


def research_update(request, id):
    # Fetch research details
    url = f"{API_STUDIO_URL}getapi/naac01_research_article_publication_dc1/{id}"
    response = requests.get(url)

    if response.status_code == 200:
        research = response.json()
    else:
        return HttpResponse(f"Error fetching research details: {response.text}", status=500)

    # Fetch media (child files) associated with this research
    media_url = f"{API_STUDIO_URL}crudapp/get/media/naac01_research_article_publication_dc1_media/parent/{id}"
    media_response = requests.get(media_url)

    if media_response.status_code == 200:
        child_files = media_response.json()
    else:
        return HttpResponse(f"Failed to fetch media files: {media_response.text}", status=500)
    
    current_year = datetime.now().year
    publication_years = [f"{y}-{y+1}" for y in range(current_year, 2010, -1)]

    # Handle form submission (POST request)
    if request.method == "POST":
        # Extract employee_id from the form, use the existing value from research if not provided in POST
        update_url = f"{API_STUDIO_URL}updateapi/update/naac01_research_article_publication_dc1/{id}"
        
        employee_id = request.POST.get('employee_id', research.get('employee_id'))

        # Debugging the employee_id to ensure it's being passed correctly
        print(f"Employee ID: {employee_id}")

        # Prepare the payload to update the research details
        payload = json.dumps({
            "data": {
                "employee_id": employee_id,
                "department_name": request.POST.get('department_name', research.get('department_name')),
                "staff_name": request.POST.get('staff_name', research.get('staff_name')),
                "paper_title": request.POST.get('paper_title', research.get('paper_title')),
                "author_name": request.POST.get('author_name', research.get('author_name')),
                "journal_name": request.POST.get('journal_name', research.get('journal_name')),
                "publication_year": request.POST.get('publication_year', research.get('publication_year')),
                "issn_number": request.POST.get('issn_number', research.get('issn_number')),
                "ugc_recognition_link": request.POST.get('ugc_recognition_link', research.get('ugc_recognition_link')),
                "issue_date": request.POST.get('issue_date', research.get('issue_date')),
                "journal_nature": request.POST.get('journal_nature', research.get('journal_nature')),
                "doi_url_link": request.POST.get('doi_url_link', research.get('doi_url_link'))
            }
        })
        
        headers = {'Content-Type': 'application/json'}
        
        # Update research details
        update_response = requests.put(update_url, headers=headers, data=payload)

        if update_response.status_code != 200:
            return HttpResponse(f"Failed to update research details: {update_response.text}", status=500)

        # Handle file uploads (media)
        upload_errors = []

        for child in child_files:
            upload_id = child['psk_id']
            fields = ['JCP', 'IP', 'IF', 'CI']  # Same fields as in research_create
            
            for field in fields:
                # Dynamically get the list of uploaded files for this specific upload_id and field
                uploaded_files = request.FILES.getlist(f'file_{upload_id}_{field.upper()}')

                if not uploaded_files:
                    continue  # No files to upload for this field

                for uploaded_file in uploaded_files:
                    # Validate file size and format
                    try:
                        validate_file_size(uploaded_file)
                        validate_file_format(uploaded_file)
                    except Exception as e:
                        messages.error(request, str(e))
                        continue  # Skip file upload if validation fails

                    # Generate custom filename
                    current_year = datetime.now().year  # Get the current year
                    custom_filename = f"{employee_id}_{field}_{current_year}_{uploaded_file.name}"
                    

                    print(f"Generated filename for field {field}: {custom_filename}")  # Log the filename for debugging

                    upload_url = f"{API_STUDIO_URL}crudapp/upload/media/naac01_research_article_publication_dc1_media/{upload_id}"

                    # Prepare the file payload and headers for the upload request
                    files = {'media': (custom_filename, uploaded_file, uploaded_file.content_type)}
                    payload = {'parent_psk_id': id}
                    headers = {'api_name': 'naac01_research_article_publication_dc1_media', 'psk_id': str(upload_id)}

                    # Upload the file
                    upload_response = requests.put(upload_url, headers=headers, data=payload, files=files)

                    if upload_response.status_code != 200:
                        upload_errors.append(f"Failed to upload {uploaded_file.name}: {upload_response.text}")
                    else:
                        print(f"File uploaded successfully: {custom_filename}")  # Log successful upload

        # If there were upload errors, handle them
        if upload_errors:
            return HttpResponse(f"Some uploads failed: {', '.join(upload_errors)}", status=500)

        # Redirect after successful update
        return redirect('research_list')

    # If not a POST request, render the update form
    return render(request, 'Research_templates/research_update.html', {'research': research, 'child_files': child_files, 'publication_years':publication_years})




# def department_dashboard(request):
#     # Define API endpoints
#     endpoints = {
#         "total_research": "getapi/all_fields/naac01_research_article_publication_dc1/all",
#         "total_participation": "getapi/all_fields/naac01_faculty_participation_dc1/all",
#         "total_books": "getapi/all_fields/naac01_books_and_chapter_dc1/all",
#         "total_certificate": "getapi/all_fields/naac01_add_on_certificate_dc1/all",
#         "total_courses": "getapi/all_fields/naac01_project_work_dc1/all",
#         "total_gov_and_non_gov": "getapi/all_fields/naac01_government_grants_dc1/all",
#         "total_workshops": "getapi/all_fields/naac01_number_of_workshop_conducted_dc1/all",
#         "total_extentions": "getapi/all_fields/naac01_no_of_extension_and_outreach_programs_dc1/all",
#         "total_competitive": "getapi/all_fields/naac01_students_benefited_exam_guidance_dc1/all",
#         "total_career_counseling": "getapi/all_fields/naac01_students_benefited_for_career_counseeling_dc1/all",
#         "total_faculty_exchange": "getapi/all_fields/naac01_collaborative_activities_for_faculty_exchange_dc1/all",
#         "total_collaborative_students": "getapi/all_fields/naac01_students_benefited_dc1/all"
#     }

#     # Get the logged-in user's username (staff_id)
#     # username = 'CS-T103'  # Hardcoded for now; replace with dynamic retrieval
    
#     user = get_settings(request)
#     username = user['username']
#     # username = 'CS-T103'

#     # Get API access token and fetch research data
#     access_token, token_type = research_key()
#     research_data = get_research_data(access_token, token_type)

#     # Determine the user role (Hod or Staff)
#     user_role = 'Staff'

#     url = f"{API_STUDIO_URL}getapi/asa0504_01_01"
#     payload = json.dumps({
#         "queries": [{"field": "username", "value": username, "operation": "equal"}],
#         "search_type": "first"
#     })
#     headers = {'Content-Type': 'application/json'}
    
#     # Make the request to fetch the role
#     response = requests.post(url, headers=headers, data=payload)

#     if response.status_code == 200:
#         # Parse the JSON response to extract the user role
#         username_val = response.json()
#         value_user = int(username_val.get('user_roles').strip('{}'))
#         user_list = roles_tbl(request)  # Function to fetch roles

#         # Determine the role of the user based on the username
#         for users in user_list:
#             if users.get('psk_id') == value_user:
#                 user_role = users.get('user_role')
#                 break

#     # Find the matching staff record for the logged-in user from research_data
#     staff_info = next((item for item in research_data if item.get("stf_id") == username), None)
#     staff_id = staff_info.get("stf_id", "N/A") if staff_info else "N/A"
#     staff_name = staff_info.get("stf_name", "N/A") if staff_info else "N/A"
#     department_name = staff_info.get("department", "N/A") if staff_info else "N/A"

#     # Initialize staff_data dictionary
#     staff_data = {staff_id: {key: 0 for key in endpoints} for staff_id in [staff_id]}
#     total = {key: 0 for key in endpoints}

#     # Define role-based data blocks
#     role_blocks = {
#         "Hod": ["total_participation", "total_books", "total_workshops"],  # Hod sees these blocks
#         "Staff": ["total_participation", "total_books", "total_workshops"]  # Staff sees these blocks too
#     }

#     # Blocks to display for the logged-in role
#     blocks_to_display = role_blocks.get(user_role, list(endpoints.keys()))
    
#     # Select department staff based on the department name for HOD login
#     department_staff = []
#     if user_role == "Hod":
#         # Hod will see all staff in the department
#         department_staff = [staff for staff in research_data if staff.get('department') == department_name]

#     # Get selected staff_id or staff_name from the request
#     staff_id_param = request.GET.get('staff_id', None)
#     staff_name_param = request.GET.get('staff_name', None)

#     # Fetch data for the selected staff
#     selected_staff_data = {}
#     if staff_id_param:
#         selected_staff_data = next((staff for staff in department_staff if staff['stf_id'] == staff_id_param), None)
#     elif staff_name_param:
#         selected_staff_data = next((staff for staff in department_staff if staff['stf_name'] == staff_name_param), None)

#     # Fetch the staff data and update staff_data for the selected staff
#     if selected_staff_data:
#         staff_id = selected_staff_data['stf_id']
#         staff_name = selected_staff_data['stf_name']
#         print("staff_id:", staff_id, "staff_name:", staff_name)

#         # Fetch data for the selected staff from the API
#         for key, value in endpoints.items():
#             # Only fetch the data for the selected staff role's data blocks
#             if key in role_blocks["Hod"]:  # For Hod, fetch only these blocks
#                 url = f"{API_STUDIO_URL}{value}"
#                 response = requests.get(url)

#                 if response.status_code == 200:
#                     data = response.json()

#                     # Filter the data based on the selected staff
#                     staff_data_endpoint = [
#                         item for item in data
#                         if item.get('stf_id') == staff_id or item.get('employee_id') == staff_id or item.get('staff_id') == staff_id
#                     ]
                    
#                     print("staff_data_endpoint:", staff_data_endpoint)

#                     # Ensure the staff_data dictionary is initialized for this staff_id
#                     if staff_id not in staff_data:
#                         staff_data[staff_id] = {}

#                     # Update individual staff data
#                     staff_data[staff_id][key] = len(staff_data_endpoint)
#                     print("staff_data:", staff_data)
#                     total[key] = len(staff_data_endpoint)
#                     print(total)

#                 else:
#                     print(f"API request failed with status code {response.status_code}")
                    
#     else:
#         print("No staff selected")

#     # Render dashboard with the filtered data
#     return render(request, 'department_dashboard.html', {
#         "staff_id": staff_id,
#         "staff_name": staff_name,
#         "department_name": department_name,
#         "filtered_data": staff_data.get(staff_id, {}),
#         "blocks_to_display": blocks_to_display,
#         "role_name": user_role,
#         "username": username,
#         "department_staff": department_staff,
#         'user_role': user_role,
#         'total':total
#     })


# def department_dashboard(request):
#     # Define API endpoints
#     endpoints = {
#         "total_research": "getapi/all_fields/naac01_research_article_publication_dc1/all",
#         "total_participation": "getapi/all_fields/naac01_faculty_participation_dc1/all",
#         "total_books": "getapi/all_fields/naac01_books_and_chapter_dc1/all",
#         "total_certificate": "getapi/all_fields/naac01_add_on_certificate_dc1/all",
#         "total_courses": "getapi/all_fields/naac01_project_work_dc1/all",
#         "total_gov_and_non_gov": "getapi/all_fields/naac01_government_grants_dc1/all",
#         "total_workshops": "getapi/all_fields/naac01_number_of_workshop_conducted_dc1/all",
#         "total_extentions": "getapi/all_fields/naac01_no_of_extension_and_outreach_programs_dc1/all",
#         "total_competitive": "getapi/all_fields/naac01_students_benefited_exam_guidance_dc1/all",
#         "total_career_counseling": "getapi/all_fields/naac01_students_benefited_for_career_counseeling_dc1/all",
#         "total_faculty_exchange": "getapi/all_fields/naac01_collaborative_activities_for_faculty_exchange_dc1/all",
#         "total_collaborative_students": "getapi/all_fields/naac01_students_benefited_dc1/all"
#     }

#     # Get the logged-in user's username (staff_id)
#     user = get_settings(request)
#     # username = user['username']
#     username = 'CS-T103'
    
#     # Get API access token and fetch research data
#     access_token, token_type = research_key()
#     research_data = get_research_data(access_token, token_type)

#     # Determine the user role (Hod or Staff)
#     user_role = 'Staff'

#     url = f"{API_STUDIO_URL}getapi/asa0504_01_01"
#     payload = json.dumps({
#         "queries": [{"field": "username", "value": username, "operation": "equal"}],
#         "search_type": "first"
#     })
#     headers = {'Content-Type': 'application/json'}
    
#     # Make the request to fetch the role
#     response = requests.post(url, headers=headers, data=payload)

#     if response.status_code == 200:
#         # Parse the JSON response to extract the user role
#         username_val = response.json()
#         value_user = int(username_val.get('user_roles').strip('{}'))
#         user_list = roles_tbl(request)  # Function to fetch roles

#         # Determine the role of the user based on the username
#         for users in user_list:
#             if users.get('psk_id') == value_user:
#                 user_role = users.get('user_role')
#                 break

#     # Find the matching staff record for the logged-in user from research_data
#     staff_info = next((item for item in research_data if item.get("stf_id") == username), None)
#     staff_id = staff_info.get("stf_id", "N/A") if staff_info else "N/A"
#     staff_name = staff_info.get("stf_name", "N/A") if staff_info else "N/A"
#     department_name = staff_info.get("department", "N/A") if staff_info else "N/A"

#     # Initialize staff_data dictionary
#     staff_data = {staff_id: {key: 0 for key in endpoints} for staff_id in [staff_id]}
#     total = {key: 0 for key in endpoints}

#     # Define role-based data blocks
#     role_blocks = {
#         "Hod": ["total_participation", "total_books", "total_workshops"],  # Hod sees these blocks
#         "Staff": ["total_participation", "total_books", "total_workshops"]  # Staff sees these blocks too
#     }

#     # Blocks to display for the logged-in role
#     blocks_to_display = role_blocks.get(user_role, list(endpoints.keys()))
    
#     # Select department staff based on the department name for HOD login
#     department_staff = []
#     if user_role == "Hod":
#         # Hod will see all staff in the department
#         department_staff = [staff for staff in research_data if staff.get('department') == department_name]

#     # Get selected staff_id or staff_name from the request
#     staff_id_param = request.GET.get('staff_id', None)
#     staff_name_param = request.GET.get('staff_name', None)

#     # Fetch data for the selected staff
#     selected_staff_data = {}
#     if staff_id_param:
#         selected_staff_data = next((staff for staff in department_staff if staff['stf_id'] == staff_id_param), None)
#     elif staff_name_param:
#         selected_staff_data = next((staff for staff in department_staff if staff['stf_name'] == staff_name_param), None)

#     # Fetch the staff data and update staff_data for the selected staff
#     if selected_staff_data:
#         staff_id = selected_staff_data['stf_id']
#         staff_name = selected_staff_data['stf_name']

#         # Fetch data for the selected staff from the API
#         for key, value in endpoints.items():
#             # Only fetch the data for the selected staff role's data blocks
#             if key in role_blocks["Hod"]:  # For Hod, fetch only these blocks
#                 url = f"{API_STUDIO_URL}{value}"
#                 response = requests.get(url)

#                 if response.status_code == 200:
#                     data = response.json()

#                     # Filter the data based on the selected staff
#                     staff_data_endpoint = [
#                         item for item in data
#                         if item.get('stf_id') == staff_id or item.get('employee_id') == staff_id or item.get('staff_id') == staff_id
#                     ]

#                     # Ensure the staff_data dictionary is initialized for this staff_id
#                     if staff_id not in staff_data:
#                         staff_data[staff_id] = {}

#                     # Update individual staff data
#                     staff_data[staff_id][key] = len(staff_data_endpoint)
#                     total[key] = len(staff_data_endpoint)

#                 else:
#                     print(f"API request failed with status code {response.status_code}")
                    
#     else:
#         print("No staff selected")

#     # Render dashboard with the filtered data
#     return render(request, 'department_dashboard.html', {
#         "staff_id": staff_id,
#         "staff_name": staff_name,
#         "department_name": department_name,
#         "filtered_data": staff_data.get(staff_id, {}),
#         "blocks_to_display": blocks_to_display,
#         "role_name": user_role,
#         "username": username,
#         "department_staff": department_staff,
#         'user_role': user_role,
#         'total': total
#     })


# def department_dashboard(request):
#     # Define API endpoints
#     endpoints = {
#         "total_research": "getapi/all_fields/naac01_research_article_publication_dc1/all",
#         "total_participation": "getapi/all_fields/naac01_faculty_participation_dc1/all",
#         "total_books": "getapi/all_fields/naac01_books_and_chapter_dc1/all",
#         "total_certificate": "getapi/all_fields/naac01_add_on_certificate_dc1/all",
#         "total_courses": "getapi/all_fields/naac01_project_work_dc1/all",
#         "total_gov_and_non_gov": "getapi/all_fields/naac01_government_grants_dc1/all",
#         "total_workshops": "getapi/all_fields/naac01_number_of_workshop_conducted_dc1/all",
#         "total_extentions": "getapi/all_fields/naac01_no_of_extension_and_outreach_programs_dc1/all",
#         "total_competitive": "getapi/all_fields/naac01_students_benefited_exam_guidance_dc1/all",
#         "total_career_counseling": "getapi/all_fields/naac01_students_benefited_for_career_counseeling_dc1/all",
#         "total_faculty_exchange": "getapi/all_fields/naac01_collaborative_activities_for_faculty_exchange_dc1/all",
#         "total_collaborative_students": "getapi/all_fields/naac01_students_benefited_dc1/all"
#     }

#     # Get the logged-in user's username (staff_id)
#     user = get_settings(request)
#     # username = user.get('username')  # You can replace this line with actual dynamic username retrieval
#     username = 'CS-T155'
    
#     # Get API access token and fetch research data
#     access_token, token_type = research_key()
#     research_data = get_research_data(access_token, token_type)

#     # Determine the user role (Hod or Staff)
#     user_role = 'Staff'

#     url = f"{API_STUDIO_URL}getapi/asa0504_01_01"
#     payload = json.dumps({"queries": [{"field": "username", "value": username, "operation": "equal"}],"search_type": "first"})
#     headers = {'Content-Type': 'application/json'}
    
#     # Make the request to fetch the role
#     response = requests.post(url, headers=headers, data=payload)

#     if response.status_code == 200:
#         username_val = response.json()
#         value_user = int(username_val.get('user_roles').strip('{}'))
#         user_list = roles_tbl(request)  # Function to fetch roles

#         # Determine the role of the user based on the username
#         for users in user_list:
#             if users.get('psk_id') == value_user:
#                 user_role = users.get('user_role')
#                 break

#     # Find the matching staff record for the logged-in user from research_data
#     staff_info = next((item for item in research_data if item.get("stf_id") == username), None)
#     staff_id = staff_info.get("stf_id", "N/A") if staff_info else "N/A"
#     staff_name = staff_info.get("stf_name", "N/A") if staff_info else "N/A"
#     department_name = staff_info.get("department", "N/A") if staff_info else "N/A"

#     # Initialize staff_data dictionary
#     staff_data = {staff_id: {key: [] for key in endpoints} for staff_id in [staff_id]}
#     total = {key: 0 for key in endpoints}

#     # Define role-based data blocks
#     role_blocks = {
#         "Hod": ["total_participation", "total_books", "total_research"],  # Hod sees these blocks
#         "Staff": ["total_participation", "total_books", "total_research"]  # Staff sees these blocks too
#     }

    
    
#     # Fetch participation data
#     part_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_faculty_participation_dc1/all"
#     part_resp = requests.get(part_url)
#     participation_data = part_resp.json() if part_resp.status_code == 200 else []

#     # Get the selected staff ID from the request parameters
#     selected_staff_id = request.GET.get('staff_id', None)

#     # If a staff ID is provided, filter the participation data for that staff
#     if selected_staff_id:
#         user_participations = [
#             p for p in participation_data if p.get("stf_id") == selected_staff_id
#         ]
#     else:
#         # If no staff ID is provided, filter for the logged-in user's participation data
#         user_participations = [
#             p for p in participation_data if p.get("stf_id") == username
#         ]

#     # Count how often each type of participation is mentioned
#     option_counter = Counter(opt.strip() for p in user_participations for opt in p.get("participation", "").split(","))

#     # Known options for display
#     options = ['Board of Studies', 'Question Paper Setting', 'Evaluation', 'Design and Development', 'Certificate Courses', 'External Examiner', 'Conference', 'Seminar', 'Workshop']

#     # Normalize keys to be safe for template variable use (no spaces or dashes)
#     normalized_counts = {opt.replace(" ", "_").replace("-", "_"): option_counter.get(opt, 0)for opt in options}

    
    
#     # Blocks to display for the logged-in role
#     blocks_to_display = role_blocks.get(user_role, list(endpoints.keys()))
    
#     # Select department staff based on the department name for HOD login
#     department_staff = []
#     user_data = []
#     if user_role == "Hod":
#         # Hod will see all staff in the department
#         department_staff = [staff for staff in research_data if staff.get('department') == department_name]

#     # Get selected staff_id or staff_name from the request
#     staff_id_param = request.GET.get('staff_id', None)
#     staff_name_param = request.GET.get('staff_name', None)

#     # Fetch data for the selected staff
#     selected_staff_data = None
#     if staff_name_param:
#         selected_staff_data = next((staff for staff in department_staff if staff['stf_name'] == staff_name_param), None)
#     elif staff_id_param:
#         selected_staff_data = next((staff for staff in department_staff if staff['stf_id'] == staff_id_param), None)
#         print("selected_staff_data:", selected_staff_data)
#         # print("selected_staff_data:", selected_staff_data)
#     else:
#     # Fallback to logged-in user data
#         selected_staff_data = staff_info

#     # If a staff is selected, fetch the data for that staff
#     if selected_staff_data:
#         staff_id = selected_staff_data['stf_id']
#         staff_name = selected_staff_data['stf_name']

#         # Fetch data for the selected staff from the API
#         for key, value in endpoints.items():
#             url = f"{API_STUDIO_URL}{value}"
#             response = requests.get(url)

            
#             if response.status_code == 200:
#                 data = response.json()
#                 # print("data", data)

#                 # Filter the data based on the selected staff
#                 staff_data_endpoint = [
#                     item for item in data
#                     if item.get('stf_id') == staff_id or item.get('employee_id') == staff_id or item.get('staff_id') == staff_id
#                 ]
#                 # Ensure the staff_data dictionary is initialized for this staff_id
#                 if staff_id not in staff_data:
#                     staff_data[staff_id] = {}
                    
#                 staff_data[staff_id][key] = staff_data_endpoint
#                 if staff_data_endpoint:
#                     user_data = staff_data_endpoint
#                     # print("user_data:", user_data)
                

#                 # Update the staff_data with the detailed list (not just the count)
#                 # staff_data[staff_id][key] = staff_data_endpoint  # Store the full data for the staff
#                 total[key] = len(staff_data_endpoint)  # Update the total count based on filtered data

#             else:
#                 print(f"API request failed with status code {response.status_code}")
                    
#     else:
#         print("No staff selected")
    


#     # Render the dashboard with the filtered data
#     return render(request, 'department_dashboard.html', {
#         "staff_id": staff_id,
#         "staff_name": staff_name,
#         "department_name": department_name,
#         "filtered_data": staff_data.get(staff_id, {}),
#         "blocks_to_display": blocks_to_display,
#         "role_name": user_role,
#         "username": username,
#         "department_staff": department_staff,
#         'user_role': user_role,
#         'total': total,
#         "user_data": user_data,
#         "participation_counts": normalized_counts,"total_entries": len(user_participations)
#     })

def department_dashboard(request):
    # Define API endpoints
    endpoints = {
        "total_research": "getapi/all_fields/naac01_research_article_publication_dc1/all",
        "total_participation": "getapi/all_fields/naac01_faculty_participation_dc1/all",
        "total_books": "getapi/all_fields/naac01_books_and_chapter_dc1/all",
        "total_certificate": "getapi/all_fields/naac01_add_on_certificate_dc1/all",
        "total_courses": "getapi/all_fields/naac01_project_work_dc1/all",
        "total_gov_and_non_gov": "getapi/all_fields/naac01_government_grants_dc1/all",
        "total_workshops": "getapi/all_fields/naac01_number_of_workshop_conducted_dc1/all",
        "total_extentions": "getapi/all_fields/naac01_no_of_extension_and_outreach_programs_dc1/all",
        "total_competitive": "getapi/all_fields/naac01_students_benefited_exam_guidance_dc1/all",
        "total_career_counseling": "getapi/all_fields/naac01_students_benefited_for_career_counseeling_dc1/all",
        "total_faculty_exchange": "getapi/all_fields/naac01_collaborative_activities_for_faculty_exchange_dc1/all",
        "total_collaborative_students": "getapi/all_fields/naac01_students_benefited_dc1/all"
    }

    # Get the logged-in user's username (staff_id)
    user = get_settings(request)
    # Safe username extraction
    username = user.get('username')

    # username = user.get('username')  # You can replace this line with actual dynamic username retrieval
    # username = 'AD-NT020'
    
    # Get API access token and fetch research data
    access_token, token_type = research_key()
    research_data = get_research_data(access_token, token_type)

    # Determine the user role (Hod or Staff)
    user_role = 'Staff'

    url = f"{API_STUDIO_URL}getapi/asa0504_01_01"
    payload = json.dumps({"queries": [{"field": "username", "value": username, "operation": "equal"}],"search_type": "first"})
    headers = {'Content-Type': 'application/json'}
    
    # Make the request to fetch the role
    response = requests.post(url, headers=headers, data=payload)

    if response.status_code == 200:
        username_val = response.json()
        # Safe user_roles extraction
        value_user = int(username_val.get('user_roles', '0').strip('{}')) if username_val and username_val.get('user_roles') else 0
        # value_user = int(username_val.get('user_roles').strip('{}'))
        user_list = roles_tbl(request)  # Function to fetch roles

        # Determine the role of the user based on the username
        for users in user_list:
            if users.get('psk_id') == value_user:
                user_role = users.get('user_role')
                break

    # Find the matching staff record for the logged-in user from research_data
    staff_info = next((item for item in research_data if item.get("stf_id") == username), None)
    staff_id = staff_info.get("stf_id", "N/A") if staff_info else "N/A"
    staff_name = staff_info.get("stf_name", "N/A") if staff_info else "N/A"
    department_name = staff_info.get("department", "N/A") if staff_info else "N/A"

    # Initialize staff_data dictionary
    staff_data = {staff_id: {key: [] for key in endpoints} for staff_id in [staff_id]}
    total = {key: 0 for key in endpoints}

    # Define role-based data blocks
    role_blocks = {
        "Hod": ["total_participation", "total_books", "total_research"],  # Hod sees these blocks
        "Staff": ["total_participation", "total_books", "total_research"]  # Staff sees these blocks too
    }

    # Blocks to display for the logged-in role
    blocks_to_display = role_blocks.get(user_role, list(endpoints.keys()))
    
    # Select department staff based on the department name for HOD login
    department_staff = []
    user_data = []
    if user_role == "Hod":
        # Hod will see all staff in the department
        department_staff = [staff for staff in research_data if staff.get('department') == department_name]

    # Get selected staff_id or staff_name from the request
    staff_id_param = request.GET.get('staff_id', None)
    staff_name_param = request.GET.get('staff_name', None)

    # Fetch data for the selected staff
    selected_staff_data = None
    if staff_name_param:
        selected_staff_data = next((staff for staff in department_staff if staff['stf_name'] == staff_name_param), None)
    elif staff_id_param:
        selected_staff_data = next((staff for staff in department_staff if staff['stf_id'] == staff_id_param), None)
        print("selected_staff_data:", selected_staff_data)

        # print("selected_staff_data:", selected_staff_data)
    else:
    # Fallback to logged-in user data
        selected_staff_data = staff_info

    # If a staff is selected, fetch the data for that staff
    if selected_staff_data:
        staff_id = selected_staff_data['stf_id']
        staff_name = selected_staff_data['stf_name']

        # Fetch data for the selected staff from the API
        for key, value in endpoints.items():
            url = f"{API_STUDIO_URL}{value}"
            response = requests.get(url)

            
            if response.status_code == 200:
                data = response.json()
                # print("data", data)

                # Filter the data based on the selected staff
                staff_data_endpoint = [
                    item for item in data
                    if item.get('stf_id') == staff_id or item.get('employee_id') == staff_id or item.get('staff_id') == staff_id
                ]
                # Ensure the staff_data dictionary is initialized for this staff_id
                if staff_id not in staff_data:
                    staff_data[staff_id] = {}
                    
                staff_data[staff_id][key] = staff_data_endpoint
                if staff_data_endpoint:
                    user_data = staff_data_endpoint
                    # print("user_data:", user_data)
                

                # Update the staff_data with the detailed list (not just the count)
                # staff_data[staff_id][key] = staff_data_endpoint  # Store the full data for the staff
                total[key] = len(staff_data_endpoint)  # Update the total count based on filtered data

            else:
                print(f"API request failed with status code {response.status_code}")
                    
    else:
        print("No staff selected")
    
    # Step 9: Count specific participation types
    part_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_faculty_participation_dc1/all"
    part_resp = requests.get(part_url)
    participation_data = part_resp.json() if part_resp.status_code == 200 else []

    # Filter only for this staff member
    user_participations = [p for p in participation_data if p.get("stf_id") == staff_id]
    print("user_participations:", user_participations)

    # Count how often each type of participation is mentioned
    option_counter = Counter(opt.strip() for p in user_participations for opt in p.get("participation", "").split(","))

    # Known options for display
    options = ['Board of Studies', 'Question Paper Setting', 'Evaluation', 'Design and Development', 'Certificate Courses', 'External Examiner', 'Conference', 'Seminar', 'Workshop']

    # Normalize keys to be safe for template variable use (no spaces or dashes)
    normalized_counts = {opt.replace(" ", "_").replace("-", "_"): option_counter.get(opt, 0)for opt in options}
    print("normalized_counts:", normalized_counts)


    # Render the dashboard with the filtered data
    return render(request, 'department_dashboard.html', {
        "staff_id": staff_id,
        "staff_name": staff_name,
        "department_name": department_name,
        "filtered_data": staff_data.get(staff_id, {}),
        "blocks_to_display": blocks_to_display,
        "role_name": user_role,
        "username": username,
        "department_staff": department_staff,
        'user_role': user_role,
        'total': total,
        "user_data": user_data,
        "participation_counts": normalized_counts,"total_entries": len(user_participations)
    })


# def admin_dashboard(request):
#     # Define API endpoints
#     endpoints = {
#         "total_research": "getapi/all_fields/naac01_research_article_publication_dc1/all",
#         "total_participation": "getapi/all_fields/naac01_faculty_participation_dc1/all",
#         "total_books": "getapi/all_fields/naac01_books_and_chapter_dc1/all",
#         "total_certificate": "getapi/all_fields/naac01_add_on_certificate_dc1/all",
#         "total_courses": "getapi/all_fields/naac01_project_work_dc1/all",
#         "total_gov_and_non_gov": "getapi/all_fields/naac01_government_grants_dc1/all",
#         "total_workshops": "getapi/all_fields/naac01_number_of_workshop_conducted_dc1/all",
#         "total_extentions": "getapi/all_fields/naac01_no_of_extension_and_outreach_programs_dc1/all",
#         "total_competitive": "getapi/all_fields/naac01_students_benefited_exam_guidance_dc1/all",
#         "total_career_counseling": "getapi/all_fields/naac01_students_benefited_for_career_counseeling_dc1/all",
#         "total_faculty_exchange": "getapi/all_fields/naac01_collaborative_activities_for_faculty_exchange_dc1/all",
#         "total_collaborative_students": "getapi/all_fields/naac01_students_benefited_dc1/all"
#     }

#     # Get API access token and fetch research data
#     access_token, token_type = research_key()
#     if not access_token or not token_type:
#         return render(request, 'admin_dashboard.html', {"error": "Failed to get API access token."})

#     # Fetch research data from API using the access token
#     research_data = get_research_data(access_token, token_type)
#     if not research_data:
#         return render(request, 'admin_dashboard.html', {"error": "Failed to fetch research data."})

#     # Extract all staff IDs (without any filtering by admin status)
#     all_staff_ids = {item["stf_id"] for item in research_data if item.get("stf_id")}
    
#     # Initialize staff data and total counts for each endpoint
#     staff_data = {staff_id: {key: 0 for key in endpoints} for staff_id in all_staff_ids}
#     total = {key: 0 for key in endpoints}  # Initialize total with all expected keys

#     # Loop through all endpoints and fetch data for each
#     for key, value in endpoints.items():
#         url = f"{API_STUDIO_URL}{value}"
#         response = requests.get(url)

#         staff_data_endpoint = []  # This will store the full data corresponding to the staff IDs
#         staff_ids_endpoint = []   # This will store the staff IDs

#         if response.status_code == 200:
#             data = response.json()
#             # Extract staff IDs and corresponding data
#             for item in data:
#                 staff_id = item.get('stf_id') or item.get('employee_id') or item.get('staff_id')

#                 if staff_id:  # Only add non-empty staff_ids
#                     staff_ids_endpoint.append(staff_id)  # Add staff_id to the list
#                     staff_data_endpoint.append(item)  # Add full data for this staff_id

#             # Update the total count for the current endpoint
#             total[key] = len(staff_data_endpoint)
            
#             # Update the individual staff data for the current endpoint
#             for item in staff_data_endpoint:
#                 staff_id = item.get('staff_id') or item.get('stf_id') or item.get('employee_id')
#                 if staff_id in staff_data:
#                     staff_data[staff_id][key] += 1

#     # Department filter logic (if any department filtering is needed)
#     department_name = request.GET.get("department", "ALL")  # Default to "ALL" if no department is provided
#     selected_staff_id = request.GET.get("staff_id")

#     # If a staff ID is selected, filter data for that staff ID, otherwise show data for all staff
#     if selected_staff_id:
#         # Show data only for the selected staff ID
#         filtered_data = staff_data.get(selected_staff_id, {})
#     else:
#         # Show data for all staff by default
#         filtered_data = total

#     # Find the department of the selected staff (only when staff_id is selected)
#     staff_department = None
#     if selected_staff_id:
#         for item in research_data:
#             if item.get('stf_id') == selected_staff_id:
#                 staff_department = item.get('department', "Unknown")
#                 break

#     # Map departments to blocks of data (no need to filter by department if showing all data)
#     blocks_to_display = list(endpoints.keys())  # Default to showing all blocks if no specific department

#     # Initialize merged_graph_data
#     merged_graph_data = {
#         "Departments": {},
#         "total_research": [],
#         "total_participation": [],
#         "total_books": [],
#         "total_workshops": [],
#         "total_extentions": [],
#         "total_certificate": [],
#         "total_courses": [],
#         "total_gov_and_non_gov": [],
#         "total_faculty_exchange": [],
#         "total_competitive": [],
#         "total_career_counseling": [],
#         "total_collaborative_students": []  # Include total_collaborative_students in merged graph data
#     }

#     # Loop over each block to fill the graph data
#     for block in blocks_to_display:
#         if block in total:
#             # Add department data to the merged graph data for the selected department
#             merged_graph_data["Departments"][staff_department] = merged_graph_data["Departments"].get(staff_department, 0) + 1
#             merged_graph_data[block].append(total.get(block, 0))  # Append the corresponding data for each block

#     # Debugging output to check the structure of the merged data
#     print("Merged Graph Data: ", merged_graph_data)

#     # Render dashboard with the filtered blocks and data
#     return render(request, 'admin_dashboard.html', {
#         "staff_ids": sorted(all_staff_ids),
#         "selected_staff_id": selected_staff_id,
#         "filtered_data": filtered_data,
#         "blocks_to_display": blocks_to_display,
#         "department_name": staff_department or department_name,
#         "merged_graph_data": merged_graph_data
#     })


# def admin_hod_dash(request):
#     endpoints = {
#         "total_research": "getapi/all_fields/naac01_research_article_publication_dc1/all",
#         "total_participation": "getapi/all_fields/naac01_faculty_participation_dc1/all",
#         "total_books": "getapi/all_fields/naac01_books_and_chapter_dc1/all",
#         "total_certificate": "getapi/all_fields/naac01_add_on_certificate_dc1/all",
#         "total_courses": "getapi/all_fields/naac01_project_work_dc1/all",
#         "total_gov_and_non_gov": "getapi/all_fields/naac01_government_grants_dc1/all",
#         "total_workshops": "getapi/all_fields/naac01_number_of_workshop_conducted_dc1/all",
#         "total_extentions": "getapi/all_fields/naac01_no_of_extension_and_outreach_programs_dc1/all",
#         "total_competitive": "getapi/all_fields/naac01_students_benefited_exam_guidance_dc1/all",
#         "total_career_counseling": "getapi/all_fields/naac01_students_benefited_for_career_counseeling_dc1/all",
#         "total_faculty_exchange": "getapi/all_fields/naac01_collaborative_activities_for_faculty_exchange_dc1/all",
#         "total_collaborative_students": "getapi/all_fields/naac01_students_benefited_dc1/all"
#     }

#     access_token, token_type = research_key()
#     if not access_token or not token_type:
#         return render(request, 'dashboard.html', {"error": "Failed to get API access token."})

#     research_data = get_research_data(access_token, token_type)
#     if not research_data:
#         return render(request, 'dashboard.html', {"error": "Failed to fetch research data."})

#     staff_id_info = {}
#     staff_name_info = {}
#     all_departments = set()
#     staff_data = {}
#     for item in research_data:
#         staff_id = item.get('stf_id')
#         staff_name = item.get('stf_name')
#         department = item.get('department')
#         if staff_id and staff_name:
#             staff_id_info[staff_id] = {"name": staff_name, "department": department}
#             staff_name_info[staff_name] = {"id": staff_id, "department": department}
#             all_departments.add(department)
            
#             staff_data[staff_id] = {key:0 for key in endpoints}
#     print("all_departments:", list(all_departments))
#     total = {key: 0 for key in endpoints}
   
#     # Populate endpoint data into staff_data and total
#     for key, endpoint in endpoints.items():
#         url = f"{API_STUDIO_URL}{endpoint}"
#         response = requests.get(url)
#         if response.status_code == 200:
#             data = response.json()
#             for item in data:
#                 staff_id = item.get("stf_id") or item.get("employee_id") or item.get("staff_id")
#                 if staff_id:
#                     total[key] += 1
#                     if staff_id in staff_data:
#                         staff_data[staff_id][key] += 1

#     # GET request values
#     selected_staff_id = request.GET.get("staff_id")
#     selected_staff_name = request.GET.get("staff_name")
#     selected_department = request.GET.get("department")
#     # selected_role = request.GET.get('role_blocks', None)
    

#     # Synchronize based on selection
#     if selected_staff_id and not selected_staff_name:
#         selected_staff_name = staff_id_info.get(selected_staff_id, {}).get("name")
#         selected_department = staff_id_info.get(selected_staff_id, {}).get("department")
#     elif selected_staff_name and not selected_staff_id:
#         selected_staff_id = staff_name_info.get(selected_staff_name, {}).get("id")
#         selected_department = staff_name_info.get(selected_staff_name, {}).get("department")
#     elif selected_staff_name and selected_staff_id:
#         selected_department = staff_id_info.get(selected_staff_id, {}).get("department")

#     # Filter staff by department (if selected)
#     if selected_department:
#         filtered_staff_map = {sid: info["name"] for sid, info in staff_id_info.items() if info.get("department") == selected_department}
#     else:
#         filtered_staff_map = {sid: info["name"] for sid, info in staff_id_info.items()}

#     filtered_staff_ids = sorted(filtered_staff_map.keys())
#     filtered_staff_names = sorted(filtered_staff_map.values())

#     # Determine data to show
#     if selected_staff_id:
#         filtered_data = staff_data.get(selected_staff_id, {})
#     elif selected_department:
#         filtered_data = {key: 0 for key in endpoints}
#         for sid in filtered_staff_map:
#             for key in endpoints:
#                 filtered_data[key] += staff_data.get(sid, {}).get(key, 0)
#     else:
#         filtered_data = total
    
#     # role_blocks = {
#     #     "Hod": ["total_research", "total_gov_and_non_gov", "total_certificate", "total_extentions", "total_competitive", "total_career_counseling", "total_faculty_exchange", 'total_collaborative_students', 'total_courses'],
#     #     "Staff": ["total_participation", "total_books", "total_workshops"]
#     # }

#     # Build department-wise summary for graph
#     blocks_to_display = (list(endpoints.keys()))
#     merged_graph_data = {"Departments": {}}
#     for block in blocks_to_display:
#         merged_graph_data[block] = []

#     # department_map = {}
#     # for sid, info in staff_id_info.items():
#     #     dept = info.get("department")
#     #     if dept:
#     #         department_map.setdefault(dept, set()).add(sid)

#     # for dept, staff_ids in department_map.items():
#     #     merged_graph_data["Departments"][dept] = dept
#     #     for block in blocks_to_display:
#     #         count = sum(staff_data.get(sid, {}).get(block, 0) for sid in staff_ids)
#     #         merged_graph_data[block].append(count)

#     return render(request, 'admin_hod_dash.html', {"staff_id": staff_id,"departments": sorted(all_departments),"staff_ids": filtered_staff_ids,"staff_names": filtered_staff_names,"filtered_staff_map": filtered_staff_map,"selected_staff_id": selected_staff_id,"selected_staff_name": selected_staff_name,"selected_department": selected_department,"filtered_data": filtered_data,"blocks_to_display": blocks_to_display,"merged_graph_data": merged_graph_data,"staff_id_info": json.dumps(staff_id_info),"staff_name_info": json.dumps(staff_name_info),})



def admin_hod_dash(request):
    endpoints = {
        "total_participation": "getapi/all_fields/naac01_faculty_participation_dc1/all",
        "total_research": "getapi/all_fields/naac01_research_article_publication_dc1/all",
        "total_books": "getapi/all_fields/naac01_books_and_chapter_dc1/all",
        "total_certificate": "getapi/all_fields/naac01_add_on_certificate_dc1/all",
        "total_courses": "getapi/all_fields/naac01_project_work_dc1/all",
        "total_gov_and_non_gov": "getapi/all_fields/naac01_government_grants_dc1/all",
        "total_workshops": "getapi/all_fields/naac01_number_of_workshop_conducted_dc1/all",
        "total_extentions": "getapi/all_fields/naac01_no_of_extension_and_outreach_programs_dc1/all",
        "total_competitive": "getapi/all_fields/naac01_students_benefited_exam_guidance_dc1/all",
        "total_career_counseling": "getapi/all_fields/naac01_students_benefited_for_career_counseeling_dc1/all",
        "total_faculty_exchange": "getapi/all_fields/naac01_collaborative_activities_for_faculty_exchange_dc1/all",
        "total_collaborative_students": "getapi/all_fields/naac01_students_benefited_dc1/all"
    }

    access_token, token_type = research_key()
    if not access_token or not token_type:
        return render(request, 'dashboard.html', {"error": "Failed to get API access token."})

    research_data = get_research_data(access_token, token_type)
    if not research_data:
        return render(request, 'dashboard.html', {"error": "Failed to fetch research data."})

    all_departments = set()
    for item in research_data:
        department = item.get('department')
        if department:
            all_departments.add(department)

    # Initialize data structures
    total_counts = {key: 0 for key in endpoints}
    selected_department = request.GET.get("department")

    # Fetch and process data for all endpoints
    for key, endpoint in endpoints.items():
        url = f"{API_STUDIO_URL}{endpoint}"
        response = requests.get(url)
        if response.status_code == 200:
            data = response.json()
            for item in data:
                department = item.get("department") or item.get("department_name")
                if department:
                    total_counts[key] += 1

    # If no department selected, use total_counts as filtered_data
    if not selected_department:
        filtered_data = total_counts.copy()
    else:
        # If department selected, filter the data
        filtered_data = {key: 0 for key in endpoints}
        for key, endpoint in endpoints.items():
            url = f"{API_STUDIO_URL}{endpoint}"
            response = requests.get(url)
            if response.status_code == 200:
                data = response.json()
                for item in data:
                    department = item.get("department") or item.get("department_name")
                    if department == selected_department:
                        filtered_data[key] += 1

    # Build department-wise summary for graph
    blocks_to_display = list(endpoints.keys())
    merged_graph_data = {"Departments": {}}
    for block in blocks_to_display:
        merged_graph_data[block] = []

    return render(request, 'admin_hod_dash.html', {"departments": sorted(all_departments),"blocks_to_display": blocks_to_display,"merged_graph_data": merged_graph_data,"total": total_counts,"filtered_data": filtered_data,"selected_department": selected_department,"showing_all": not bool(selected_department)})



# def dashboard(request):
    # user = get_settings(request)
    # username = user.get('username')

    # # Define endpoints
    # endpoints = {
        # "total_participation": "naac01_faculty_participation_dc1",
        # "total_research": "naac01_research_article_publication_dc1",
        # "total_books": "naac01_books_and_chapter_dc1",
        # "total_certificate": "naac01_add_on_certificate_dc1",
        # "total_courses": "naac01_project_work_dc1",
        # "total_gov_and_non_gov": "naac01_government_grants_dc1",
        # "total_workshops": "naac01_number_of_workshop_conducted_dc1",
        # "total_extentions": "naac01_no_of_extension_and_outreach_programs_dc1",
        # "total_competitive": "naac01_students_benefited_exam_guidance_dc1",
        # "total_career_counseling": "naac01_students_benefited_for_career_counseeling_dc1",
        # "total_faculty_exchange": "naac01_collaborative_activities_for_faculty_exchange_dc1",
        # "total_collaborative_students": "naac01_students_benefited_dc1"
    # }

    # access_token, token_type = research_key()
    # research_data = get_research_data(access_token, token_type)

    # staff_info = next((s for s in research_data if s.get("stf_id") == username), {})
    # staff_id = staff_info.get("stf_id", "N/A")
    # staff_name = staff_info.get("stf_name", "N/A")
    # department_name = staff_info.get("department", "N/A")

    # # Get user role
    # role_url = "https://api.hcaschennai.edu.in/getapi/asa0504_01_01"
    # payload = json.dumps({
        # "queries": [{"field": "username", "value": username, "operation": "equal"}],
        # "search_type": "first"
    # })
    # role_response = requests.post(role_url, headers={'Content-Type': 'application/json'}, data=payload)
    # value_user = int(role_response.json().get("user_roles", "0").strip("{}")) if role_response.status_code == 200 else 0
    # role_list = roles_tbl(request)
    # user_role = next((role.get("user_role") for role in role_list if role.get("psk_id") == value_user), "Staff")

    # # Determine relevant staff list
    # if user_role == "Hod":
        # department_staff = [s for s in research_data if s.get("department") == department_name]
        # relevant_staff = [s.get("stf_id") for s in department_staff if s.get("stf_id")]
    # else:
        # relevant_staff = [username]

    # # Initialize staff_data with 0 for each staff and endpoint
    # staff_data = {
        # staff_id: {key: 0 for key in endpoints.keys()}
        # for staff_id in relevant_staff
    # }

    # total = {}

    # # Populate staff_data with actual API data counts
    # for key, value in endpoints.items():
        # url = f"{API_STUDIO_URL}getapi/all_fields/{value}/all"
        # resp = requests.get(url)
        # data = resp.json() if resp.status_code == 200 else []

        # for item in data:
            # for staff_id in relevant_staff:
                # if item.get("stf_id") == staff_id or item.get("employee_id") == staff_id or item.get("staff_id") == staff_id:
                    # staff_data[staff_id][key] += 1

        # total[key] = sum(staff_data[staff_id][key] for staff_id in relevant_staff)

    # # Determine which data to show
    # filtered_data = {}
    # if user_role == "Hod":
        # filtered_data = {
            # key: sum(staff_data[staff_id][key] for staff_id in relevant_staff)
            # for key in endpoints.keys()
        # }
    # else:
        # filtered_data = staff_data.get(username, {})

    # # Blocks by role
    # blocks = {
        # "Hod": ["total_workshops", "total_gov_and_non_gov", "total_certificate", "total_extentions", "total_competitive", "total_career_counseling", "total_faculty_exchange", "total_collaborative_students", "total_courses"],
        # "Staff": ["total_participation", "total_research", "total_books"]
    # }

    # # Count participation types
    # part_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_faculty_participation_dc1/all"
    # part_resp = requests.get(part_url)
    # participation_data = part_resp.json() if part_resp.status_code == 200 else []
    # user_participations = [p for p in participation_data if p.get("stf_id") == username]

    # option_counter = Counter(opt.strip() for p in user_participations for opt in p.get("participation", "").split(","))
    # options = ['Board of Studies', 'Question Paper Setting', 'Evaluation', 'Design and Development',
               # 'Certificate Courses', 'External Examiner', 'Conference', 'Seminar', 'Workshop']
    # normalized_counts = {
        # opt.replace(" ", "_").replace("-", "_"): option_counter.get(opt, 0)
        # for opt in options
    # }

    # return render(request, 'dashboard.html', {
        # "staff_id": staff_id,
        # "staff_name": staff_name,
        # "department_name": department_name,
        # "filtered_data": filtered_data,
        # "blocks_to_display": blocks.get(user_role, list(endpoints.keys())),
        # "role_name": user_role,
        # "username": username,
        # "participation_counts": normalized_counts,
        # "total_entries": len(user_participations),
        # "staff_data": staff_data,
        # "selected_department": department_name,
        # "selected_staff_id": username
    # })







# def dashboard(request):
#     # Step 1: Get the logged-in username (default to a test user if not available)
#     user = get_settings(request)
#     # username = user.get('username')
#     username = 'CS-T155'
    
#     # Step 2: Define model keys mapped to display labels (used for counting later)
#     endpoints = {
#         "total_participation": "naac01_faculty_participation_dc1",
#         "total_research": "naac01_research_article_publication_dc1",
#         "total_books": "naac01_books_and_chapter_dc1",
#         "total_certificate": "naac01_add_on_certificate_dc1",
#         "total_courses": "naac01_project_work_dc1",
#         "total_gov_and_non_gov": "naac01_government_grants_dc1",
#         "total_workshops": "naac01_number_of_workshop_conducted_dc1",
#         "total_extentions": "naac01_no_of_extension_and_outreach_programs_dc1",
#         "total_competitive": "naac01_students_benefited_exam_guidance_dc1",
#         "total_career_counseling": "naac01_students_benefited_for_career_counseeling_dc1",
#         "total_faculty_exchange": "naac01_collaborative_activities_for_faculty_exchange_dc1",
#         "total_collaborative_students": "naac01_students_benefited_dc1"
#     }

#     # Step 3: Get access token to pull research data
#     access_token, token_type = research_key()
#     research_data = get_research_data(access_token, token_type)

#     # Step 4: Find staff info from research data using staff_id
#     staff_info = next((s for s in research_data if s.get("stf_id") == username), {})
#     staff_id = staff_info.get("stf_id", "N/A")
#     staff_name = staff_info.get("stf_name", "N/A")
#     department_name = staff_info.get("department", "N/A")

#     # Step 5: Count data entries for each endpoint only for the logged-in staff
#     total = {}
#     staff_data = {username: {}}

#     for key, value in endpoints.items():
#         url = f"{API_STUDIO_URL}getapi/all_fields/{value}/all"
#         resp = requests.get(url)
#         data = resp.json() if resp.status_code == 200 else []
        

#         # Filter by staff_id or similar fields
#         staff_items = [emp_id for emp_id in data if emp_id.get("stf_id") == username or emp_id.get("employee_id") == username or emp_id.get("staff_id") == username]

#         # Count total and store for staff
#         total[key] = len(staff_items)
#         staff_data[username][key] = len(staff_items)
#     print("total:", total)
#     print("staff_data:", staff_data)

#     # Step 6: Determine user role from ASA user-role table via POST API call
#     role_url = "https://api.hcaschennai.edu.in/getapi/asa0504_01_01"
#     payload = json.dumps({"queries": [{"field": "username", "value": username, "operation": "equal"}],"search_type": "first"})
#     role_response = requests.post(role_url, headers={'Content-Type': 'application/json'}, data=payload)
#     value_user = int(role_response.json().get("user_roles", "0").strip("{}")) if role_response.status_code == 200 else 0

#     # Step 7: Match user_role ID with data from roles_tbl
#     role_list = roles_tbl(request)
#     user_role = next((role.get("user_role") for role in role_list if role.get("psk_id") == value_user), "Staff")

#     # Step 8: Define which metrics are visible for each role
#     blocks = {"Hod": ["total_workshops", "total_gov_and_non_gov", "total_certificate", "total_extentions", "total_competitive", "total_career_counseling", "total_faculty_exchange", "total_collaborative_students", "total_courses"],
#             "Staff": ["total_participation", "total_research", "total_books"]}

#     # Step 9: Count specific participation types
#     part_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_faculty_participation_dc1/all"
#     part_resp = requests.get(part_url)
#     participation_data = part_resp.json() if part_resp.status_code == 200 else []

#     # Filter only for this staff member
#     user_participations = [p for p in participation_data if p.get("stf_id") == username]

#     # Count how often each type of participation is mentioned
#     option_counter = Counter(opt.strip() for p in user_participations for opt in p.get("participation", "").split(","))

#     # Known options for display
#     options = ['Board of Studies', 'Question Paper Setting', 'Evaluation', 'Design and Development', 'Certificate Courses', 'External Examiner', 'Conference', 'Seminar', 'Workshop']

#     # Normalize keys to be safe for template variable use (no spaces or dashes)
#     normalized_counts = {opt.replace(" ", "_").replace("-", "_"): option_counter.get(opt, 0)for opt in options}

#     # Step 10: Render the dashboard with all required context variables
#     return render(request, 'dashboard.html', {"staff_id": staff_id,"staff_name": staff_name,"department_name": department_name,"filtered_data": staff_data.get(username, {}),"blocks_to_display": blocks.get(user_role, list(endpoints.keys())),"role_name": user_role,"username": username,"participation_counts": normalized_counts,"total_entries": len(user_participations)})


def dashboard(request):
    # Step 1: Get the logged-in username (or default for testing)
    user = get_settings(request)
    username = user.get('username')
    # username = 'CS-T155'

    # Step 2: Define endpoints
    endpoints = {
        "total_participation": "naac01_faculty_participation_dc1",
        "total_research": "naac01_research_article_publication_dc1",
        "total_books": "naac01_books_and_chapter_dc1",
        "total_certificate": "naac01_add_on_certificate_dc1",
        "total_courses": "naac01_project_work_dc1",
        "total_gov_and_non_gov": "naac01_government_grants_dc1",
        "total_workshops": "naac01_number_of_workshop_conducted_dc1",
        "total_extentions": "naac01_no_of_extension_and_outreach_programs_dc1",
        "total_competitive": "naac01_students_benefited_exam_guidance_dc1",
        "total_career_counseling": "naac01_students_benefited_for_career_counseeling_dc1",
        "total_faculty_exchange": "naac01_collaborative_activities_for_faculty_exchange_dc1",
        "total_collaborative_students": "naac01_students_benefited_dc1"
    }

    # Step 3: Research data
    access_token, token_type = research_key()
    research_data = get_research_data(access_token, token_type)

    
    # Step 4: Staff info
    staff_info = next((s for s in research_data if s.get("stf_id") == username), {})
    staff_id = staff_info.get("stf_id", "N/A")
    staff_name = staff_info.get("stf_name", "N/A")
    department_name = staff_info.get("department", "N/A")
    print("department_name:",department_name)

    # Step 5: Role detection
    role_url = "https://api.hcaschennai.edu.in/getapi/asa0504_01_01"
    payload = json.dumps({
        "queries": [{"field": "username", "value": username, "operation": "equal"}],
        "search_type": "first"
    })
    role_response = requests.post(role_url, headers={'Content-Type': 'application/json'}, data=payload)
    value_user = int(role_response.json().get("user_roles", "0").strip("{}")) if role_response.status_code == 200 else 0

    role_list = roles_tbl(request)
    user_role = next((role.get("user_role") for role in role_list if role.get("psk_id") == value_user), "Staff")

    # Step 6: Data collection
    total = {}
    staff_data = {}

    if user_role == "Hod":
        # Get all staff in same department
        department_staff = [s for s in research_data if s.get("department") == department_name]
        dept_ids = [s.get("stf_id") for s in department_staff]

        # Aggregate for all staff in dept
        for key, value in endpoints.items():
            url = f"{API_STUDIO_URL}getapi/all_fields/{value}/all"
            resp = requests.get(url)
            data = resp.json() if resp.status_code == 200 else []

            dept_items = [rec for rec in data if rec.get("stf_id") in dept_ids or rec.get("staff_id") in dept_ids or rec.get("employee_id") in dept_ids]
            total[key] = len(dept_items)

        print("HOD department totals:", total)

    else:  # Staff
        staff_data[username] = {}
        for key, value in endpoints.items():
            url = f"{API_STUDIO_URL}getapi/all_fields/{value}/all"
            resp = requests.get(url)
            data = resp.json() if resp.status_code == 200 else []

            staff_items = [rec for rec in data if rec.get("stf_id") == username or rec.get("staff_id") == username or rec.get("employee_id") == username]
            total[key] = len(staff_items)
            staff_data[username][key] = len(staff_items)

        print("Staff totals:", staff_data)

    # Step 7: Participation breakdown
    part_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_faculty_participation_dc1/all"
    part_resp = requests.get(part_url)
    participation_data = part_resp.json() if part_resp.status_code == 200 else []

    if user_role == "Hod":
        user_participations = [p for p in participation_data if p.get("stf_id") in dept_ids]
    else:
        user_participations = [p for p in participation_data if p.get("stf_id") == username]

    option_counter = Counter(opt.strip() for p in user_participations for opt in p.get("participation", "").split(","))

    options = [
        'Board of Studies', 'Question Paper Setting', 'Evaluation', 'Design and Development',
        'Certificate Courses', 'External Examiner', 'Conference', 'Seminar', 'Workshop'
    ]

    normalized_counts = {
        opt.replace(" ", "_").replace("-", "_"): option_counter.get(opt, 0)
        for opt in options
    }

    # Step 8: Role → visible blocks
    blocks = {
        "Hod": ["total_workshops", "total_gov_and_non_gov", "total_certificate", "total_extentions",
                "total_competitive", "total_career_counseling", "total_faculty_exchange",
                "total_collaborative_students", "total_courses"],
        "Staff": ["total_participation", "total_research", "total_books"]
    }

    # Step 9: Render
    return render(request, 'dashboard.html', {
        "staff_id": staff_id,
        "staff_name": staff_name,
        "department_name": department_name,
        "filtered_data": total if user_role == "Hod" else staff_data.get(username, {}),
        "blocks_to_display": blocks.get(user_role, list(endpoints.keys())),
        "role_name": user_role,
        "username": username,
        "participation_counts": normalized_counts,
        "total_entries": len(user_participations)
    })



# def dashboard(request):
#     # Define API endpoints
#     endpoints = {
#         "total_participation": "getapi/all_fields/naac01_faculty_participation_dc1/all",
#         "total_research": "getapi/all_fields/naac01_research_article_publication_dc1/all",
#         "total_books": "getapi/all_fields/naac01_books_and_chapter_dc1/all",
#         "total_certificate": "getapi/all_fields/naac01_add_on_certificate_dc1/all",
#         "total_courses": "getapi/all_fields/naac01_project_work_dc1/all",
#         "total_gov_and_non_gov": "getapi/all_fields/naac01_government_grants_dc1/all",
#         "total_workshops": "getapi/all_fields/naac01_number_of_workshop_conducted_dc1/all",
#         "total_extentions": "getapi/all_fields/naac01_no_of_extension_and_outreach_programs_dc1/all",
#         "total_competitive": "getapi/all_fields/naac01_students_benefited_exam_guidance_dc1/all",
#         "total_career_counseling": "getapi/all_fields/naac01_students_benefited_for_career_counseeling_dc1/all",
#         "total_faculty_exchange": "getapi/all_fields/naac01_collaborative_activities_for_faculty_exchange_dc1/all",
#         "total_collaborative_students": "getapi/all_fields/naac01_students_benefited_dc1/all"
#     }

#     # Get API access token and fetch research data
#     access_token, token_type = research_key()

#     # Fetch research data from API using the access token
#     research_data = get_research_data(access_token, token_type)

#     # Get the logged-in user's username (staff_id)
#     user = get_settings(request)
#     # username = user.get('username')  # Static username for testing
#     username= 'CS-T151'
    
#     # Find the matching staff record for the logged-in user
#     staff_info = next((item for item in research_data if item.get("stf_id") == username), None)
    
#     # Initialize staff details
#     staff_id = staff_info.get("stf_id", "N/A") if staff_info else "N/A"
#     staff_name = staff_info.get("stf_name", "N/A") if staff_info else "N/A"
#     department_name = staff_info.get("department", "N/A") if staff_info else "N/A"

#     # Initialize staff data and total counts for each endpoint
#     staff_data = {staff_id: {key: 0 for key in endpoints} for staff_id in [staff_id]}
#     total = {key: 0 for key in endpoints}  # Initialize total with all expected keys
    
#     # Loop through all endpoints and fetch data for each
#     for key, value in endpoints.items():
#         url = f"{API_STUDIO_URL}{value}"
#         response = requests.get(url)

#         staff_data_endpoint = []  # This will store the full data corresponding to the staff IDs

#         if response.status_code == 200:
#             data = response.json()
#             # Extract staff IDs and corresponding data
#             for item in data:
#                 staff_id_from_api = item.get('stf_id') or item.get('employee_id') or item.get('staff_id')

#                 if staff_id_from_api == username:  # Filter data for the logged-in user
#                     staff_data_endpoint.append(item)  # Add full data for this staff_id
#                     print("staff_data_endpoint:", staff_data_endpoint)

#             # Update the total count for the current endpoint
#             total[key] = len(staff_data_endpoint)
            
#             # Update the individual staff data for the current endpoint
#             for item in staff_data_endpoint:
#                 staff_data[staff_id][key] += 1
    
#     user_role = 'Staff'
                
#     # Fetch the roles from the roles_tbl
#     url = "https://api.hcaschennai.edu.in/getapi/asa0504_01_01"
#     payload = json.dumps({"queries": [{"field": "username", "value": username, "operation": "equal"}], "search_type": "first"})
#     headers = {'Content-Type': 'application/json'}

#     # Make the request to the API
#     response = requests.post(url, headers=headers, data=payload)

#     if response.status_code == 200:
#         # Parse the JSON response
#         username_val = response.json()
#         value_user = int(username_val.get('user_roles').strip('{}'))

#         # Fetch the roles from the roles_tbl (roles_tbl function should return user data)
#         user_list = roles_tbl(request)  # This should return a list of users (dictionaries)

#         # Find the user role by comparing the psk_id with the value_user
#         user_role = None
#         for users in user_list:
#             if users.get('psk_id') == value_user:  # Match psk_id with the value_user
#                 user_role = users.get('user_role')
#                 print("user_role:", user_role)
#                 break
        
#         # If no role found, default to 'Staff'
#         if user_role is None:
#             user_role = "Staff"
        
#     else:
#         print(f"Failed to fetch user role. Status code: {response.status_code}")
    
#     # Define role-based data blocks
#     role_blocks = {
#         "Hod": ["total_workshops", "total_gov_and_non_gov", "total_certificate", "total_extentions", "total_competitive", "total_career_counseling", "total_faculty_exchange", 'total_collaborative_students', 'total_courses'],
#         "Staff": ["total_participation", "total_research", "total_books"]
#     }
#     # user_role = 'Staff'
#     # Determine which blocks (metrics) to display based on the role
#     blocks_to_display = role_blocks.get(user_role, list(endpoints.keys()))
#     print("blocks_to_display:", blocks_to_display)


#     # Render dashboard with the filtered blocks and data for the logged-in user
#     return render(request, 'dashboard.html', {
#         "staff_id": staff_id,
#         "staff_name": staff_name,
#         "department_name": department_name,
#         "filtered_data": staff_data.get(staff_id, {}),
#         "blocks_to_display": blocks_to_display,
#         "role_name": user_role,
#         "username": username,
#     })

def admin_dash(request):
    # Define all API endpoints
    endpoints = {
        "total_participation": "getapi/all_fields/naac01_faculty_participation_dc1/all",
        "total_research": "getapi/all_fields/naac01_research_article_publication_dc1/all",
        "total_books": "getapi/all_fields/naac01_books_and_chapter_dc1/all",
        "total_certificate": "getapi/all_fields/naac01_add_on_certificate_dc1/all",
        "total_courses": "getapi/all_fields/naac01_project_work_dc1/all",
        "total_gov_and_non_gov": "getapi/all_fields/naac01_government_grants_dc1/all",
        "total_workshops": "getapi/all_fields/naac01_number_of_workshop_conducted_dc1/all",
        "total_extentions": "getapi/all_fields/naac01_no_of_extension_and_outreach_programs_dc1/all",
        "total_competitive": "getapi/all_fields/naac01_students_benefited_exam_guidance_dc1/all",
        "total_career_counseling": "getapi/all_fields/naac01_students_benefited_for_career_counseeling_dc1/all",
        "total_faculty_exchange": "getapi/all_fields/naac01_collaborative_activities_for_faculty_exchange_dc1/all",
        "total_collaborative_students": "getapi/all_fields/naac01_students_benefited_dc1/all"
    }

    access_token, token_type = research_key()
    research_data = get_research_data(access_token, token_type)

    # Extract department names
    departments = sorted(set(staff.get('department') for staff in research_data if staff.get('department')))

    # Read request parameters
    selected_dept = request.GET.get('department')
    selected_staff_id = request.GET.get('staff_id')
    selected_staff_name = request.GET.get('staff_name')

    # Filter staff by department (for dropdown filtering)
    selected_department = []
    if selected_dept:
        selected_department = [staff for staff in research_data if staff.get('department') == selected_dept]

    # Determine selected staff
    staff_info = None
    if selected_staff_name:
        staff_info = next((staff for staff in selected_department if staff.get('stf_name') == selected_staff_name), None)
    elif selected_staff_id:
        staff_info = next((staff for staff in selected_department if staff.get('stf_id') == selected_staff_id), None)

    # Auto-select first staff in department if no staff is selected
    if not staff_info and selected_dept and selected_department:
        staff_info = {}

    # Reset identifiers if no match is found
    if not staff_info:
        selected_staff_id = None
        selected_staff_name = None
    else:
        selected_staff_id = staff_info.get('stf_id')
        selected_staff_name = staff_info.get('stf_name')
        selected_dept = staff_info.get('department')

    # If no staff found, render only the selection form
    if not staff_info:
        return render(request, 'admin_staff_dash.html', {"departments": departments,"department_staff": selected_department,"message": "Select a department or staff to view data.","selected_dept": selected_dept,"selected_staff_id": selected_staff_id,"selected_staff_name": selected_staff_name})

    # Fetch data for selected staff
    staff_data = {selected_staff_id: {key: 0 for key in endpoints}}

    for key, endpoint in endpoints.items():
        url = f"{API_STUDIO_URL}{endpoint}"
        response = requests.get(url)
        if response.status_code == 200:
            for item in response.json():
                staff_id_from_api = item.get('stf_id') or item.get('employee_id') or item.get('staff_id')
                if staff_id_from_api == selected_staff_id:
                    staff_data[selected_staff_id][key] += 1

    # Return the updated page with filtered data based on selected staff, department, and staff ID
    return render(request, 'admin_staff_dash.html', {"departments": departments,"department_staff": selected_department,"staff_id": selected_staff_id,"staff_name": selected_staff_name,"department_name": selected_dept,"filtered_data": staff_data.get(selected_staff_id, {}),"blocks_to_display": list(endpoints.keys()),"selected_dept": selected_dept,"selected_staff_id": selected_staff_id,"selected_staff_name": selected_staff_name})


# def admin_hod_dash(request):
#     # Define all API endpoints
#     endpoints = {
#         "total_research": "getapi/all_fields/naac01_research_article_publication_dc1/all",
#         "total_participation": "getapi/all_fields/naac01_faculty_participation_dc1/all",
#         "total_books": "getapi/all_fields/naac01_books_and_chapter_dc1/all",
#         "total_certificate": "getapi/all_fields/naac01_add_on_certificate_dc1/all",
#         "total_courses": "getapi/all_fields/naac01_project_work_dc1/all",
#         "total_gov_and_non_gov": "getapi/all_fields/naac01_government_grants_dc1/all",
#         "total_workshops": "getapi/all_fields/naac01_number_of_workshop_conducted_dc1/all",
#         "total_extentions": "getapi/all_fields/naac01_no_of_extension_and_outreach_programs_dc1/all",
#         "total_competitive": "getapi/all_fields/naac01_students_benefited_exam_guidance_dc1/all",
#         "total_career_counseling": "getapi/all_fields/naac01_students_benefited_for_career_counseeling_dc1/all",
#         "total_faculty_exchange": "getapi/all_fields/naac01_collaborative_activities_for_faculty_exchange_dc1/all",
#         "total_collaborative_students": "getapi/all_fields/naac01_students_benefited_dc1/all"
#     }

#     access_token, token_type = research_key()
#     research_data = get_research_data(access_token, token_type)

#     # Extract department names
#     departments = sorted(set(staff.get('department') for staff in research_data if staff.get('department')))

#     # Read request parameters
#     selected_dept = request.GET.get('department')
#     selected_staff_id = request.GET.get('staff_id')
#     selected_staff_name = request.GET.get('staff_name')

#     # Filter staff by department (for dropdown filtering)
#     selected_department = []
#     if selected_dept:
#         selected_department = [staff for staff in research_data if staff.get('department') == selected_dept]

#     # Determine selected staff
#     staff_info = None
#     if selected_staff_id:
#         staff_info = next((staff for staff in selected_department if staff.get('stf_id') == selected_staff_id), None)
#     elif selected_staff_name:
#         staff_info = next((staff for staff in selected_department if staff.get('stf_name') == selected_staff_name), None)

#     # Auto-select first staff in department if no staff is selected
#     if not staff_info and selected_dept and selected_department:
#         staff_info = {}

#     # Reset identifiers if no match is found
#     if not staff_info:
#         selected_staff_id = None
#         selected_staff_name = None
#     else:
#         selected_staff_id = staff_info.get('stf_id')
#         selected_staff_name = staff_info.get('stf_name')
#         selected_dept = staff_info.get('department')

#     # If no staff found, render only the selection form
#     if not staff_info:
#         return render(request, 'admin_hod_dash.html', {"departments": departments,"department_staff": selected_department,"message": "Select a department or staff to view data.","selected_dept": selected_dept,"selected_staff_id": selected_staff_id,"selected_staff_name": selected_staff_name})

#     # Fetch data for selected staff
#     staff_data = {selected_staff_id: {key: 0 for key in endpoints}}

#     for key, endpoint in endpoints.items():
#         url = f"{API_STUDIO_URL}{endpoint}"
#         response = requests.get(url)
#         if response.status_code == 200:
#             for item in response.json():
#                 staff_id_from_api = item.get('stf_id') or item.get('employee_id') or item.get('staff_id')
#                 if staff_id_from_api == selected_staff_id:
#                     staff_data[selected_staff_id][key] += 1

#     # Return the updated page with filtered data based on selected staff, department, and staff ID
#     return render(request, 'admin_hod_dash.html', {"departments": departments,"department_staff": selected_department,"staff_id": selected_staff_id,"staff_name": selected_staff_name,"department_name": selected_dept,"filtered_data": staff_data.get(selected_staff_id, {}),"blocks_to_display": list(endpoints.keys()),"selected_dept": selected_dept,"selected_staff_id": selected_staff_id,"selected_staff_name": selected_staff_name})




# from Crypto.Cipher import AES
from Cryptodome.Cipher import AES
import base64


# def decrypt_aes(encrypted_text, key):
#     encrypted_text = encrypted_text.replace('-', '+').replace('_', '/')
#     encrypted_text = base64.b64decode(encrypted_text)

#     cipher = AES.new(key.encode(), AES.MODE_ECB)  # AES-128 ECB mode
#     decrypted_bytes = cipher.decrypt(encrypted_text)

#     # Remove padding
#     decrypted_text = decrypted_bytes.rstrip(b"\x00").decode()

#     return decrypted_text


def unpad_pkcs7(padded_bytes):
    padding_len = padded_bytes[-1]
    if padding_len < 1 or padding_len > 16:
        return padded_bytes  # maybe no padding, just return as is
    return padded_bytes[:-padding_len]

def decrypt_aes(encrypted_text, key):
    encrypted_text = encrypted_text.replace('-', '+').replace('_', '/')
    encrypted_bytes = base64.b64decode(encrypted_text)

    cipher = AES.new(key.encode(), AES.MODE_ECB)  # AES-128 ECB mode
    decrypted_bytes = cipher.decrypt(encrypted_bytes)

    # Remove PKCS#7 padding properly
    decrypted_bytes = unpad_pkcs7(decrypted_bytes)

    try:
        decrypted_text = decrypted_bytes.decode('utf-8')
    except UnicodeDecodeError:
        decrypted_text = decrypted_bytes.decode('utf-8', errors='replace')
        print("Warning: UnicodeDecodeError in decrypt_aes, replaced invalid chars.")

    return decrypted_text




def data_fetch(request):
    url = f"{API_STUDIO_URL}auth/token"
    payload = json.dumps({"secret_key": "DMUOJtffkZWIoNpj9eXTFSCzKJMBRjFP"})
    headers = {'Content-Type': 'application/json'}
    response = requests.request("POST", url, headers=headers, data=payload)
    if response.status_code == 200:
        res_data = response.json()
        access_token = res_data.get('access_token')
        token_type = res_data.get('token_type')
    else:
        return HttpResponse("Failed to get access token", status=500)

    url = f"{API_STUDIO_URL}sqlviews/api/v1/auth/get_response_data"
    payload = json.dumps({"psk_uid": "6e15c9c4-4de8-4bc6-9a40-9c7ecae35bb2","project": "hcas","data": {}})
    headers = {'Content-Type': 'application/json','Authorization': f'{token_type} {access_token}'}
    response = requests.request("POST", url, headers=headers, data=payload)
    if response.status_code == 200:
        pykit_data = response.json()
        print("pykit_data:", pykit_data)
    else:
        return HttpResponse("Failed to fetch pykit data", status=500)

    url = f"{API_STUDIO_URL}getapi/all_fields/asa0508_01_01/all"
    payload = {}
    headers = {}
    response = requests.request("GET", url, headers=headers, data=payload)
    if response.status_code == 200:
        total_pykit_staff = response.json()
        print("total_pykit_staff:", total_pykit_staff)
    else:
        return HttpResponse("Failed to fetch existing staff data", status=500)

    encryption_key = "MZysriwB2EpRrgsi"

 # This allows processing even if the staff table is empty
    if pykit_data:
        for data in pykit_data:
            username = data['b2eid']
            matched_user = None
            print("username:", username)

            # If table is empty, this just won't match — and that's fine
            for existing in total_pykit_staff:
                if existing.get('user_id') == username:
                    matched_user = existing
                    break

            encrypted_password = data['PWD_']
            decrypted_password = decrypt_aes(encrypted_password, encryption_key)

            active_status = int(data.get('statusflag')) == 0
            
            if active_status == 0:
                print("active")
            else:
                print("In Active")
            
            payload = json.dumps({"data": {"username": data['ID_'],"user_first_name": data['FIRST_'],"user_last_name": data['LAST_'],"user_id":data['b2eid'],"user_email_id": data['stf_email'],"password": decrypted_password,"active": active_status}})
            headers = {'Content-Type': 'application/json'}

            if matched_user:
                user_id = matched_user.get('psk_id')
                url = f"{API_STUDIO_URL}updateapi/update/asa0508_01_01/{user_id}"
                response = requests.request("PUT", url, headers=headers, data=payload)
                if response.status_code == 200:
                    print(f"Updated user: {username}")
                else:
                    print(f"Failed to update user {username}: {response.text}")
            else:
                url = f"{API_STUDIO_URL}postapi/create/asa0508_01_01"
                response = requests.request("POST", url, headers=headers, data=payload)
                if response.status_code == 200:
                    print(f"Created new user: {username}")
                else:
                    print(f"Failed to create user {username}: {response.text}")

        return redirect('pykit_staff_list')

  

def hcas_user(request):
    url = "https://api.hcaschennai.edu.in/getapi/asa0504_01_01/all"
    payload = ""
    headers = {}
    response = requests.request("GET", url, headers=headers, data=payload)
    if response.status_code == 200:
        pykit_user = response.json()
    else:
        return HttpResponse("Failed to fetch pykit_user", status=500)

    url = "https://api.hcaschennai.edu.in/getapi/all_fields/asa0508_01_01/all"
    payload = ""
    headers = {}
    response = requests.request("GET", url, headers=headers, data=payload)
    if response.status_code == 200:
        mid_user = response.json()
    else:
        return HttpResponse("Failed to fetch mid_user", status=500)

    if mid_user:
        for data in mid_user:
            username = data['username'].strip()
            print("username==:", username)
            matched_user = None
            

            for existing in pykit_user:
                if existing.get('username') == username:
                    matched_user = existing
                    break
            raw_password = data['password']
            md5_hash = hashlib.md5(raw_password.encode()).hexdigest()
       
            ERP_ID = data['user_id'] 

            payload = json.dumps({"data": {"user_roles": str({1}), "user_id": ERP_ID,"username": username,"first_name": data['user_first_name'],"user_type": 'user',"last_name": data['user_last_name'],"email": data['user_email_id'],"password": md5_hash,"active": data['active']}})
            update_payload = json.dumps({"data": {"username": username,"first_name": data['user_first_name'],"last_name": data['user_last_name'], "email": data['user_email_id'],"password": md5_hash,"active": data['active']}})
            headers = {'Content-Type': 'application/json'}
            
          

            if matched_user:
                user_id = matched_user.get('psk_id')
                url = f"{API_STUDIO_URL}updateapi/update/asa0504_01_01/{user_id}"
                response = requests.request("PUT", url, headers=headers, data=update_payload)
                if response.status_code == 200:
                    res_data = response.json()
                    
                    url = f"https://mis.hcaschennai.edu.in/pykit_user_id_base/{username}/"
                    dj_payload = json.dumps({"username": username,"last_name": res_data['psk_id'], "email": res_data['email'],  "password": md5_hash})
                    response = requests.put(url, headers=headers, data=dj_payload)
                    if response.status_code == 200:
                        print(f"✅ Updated user: {username}")
                    else:
                        
                        print(f"❌ Failed to update user {username}: {response.text}")
                else:
                    print(f"Failed to update user {username}: {response.text}")
            
                    
         
            else:
                url = f"{API_STUDIO_URL}postapi/create/asa0504_01_01"
                response = requests.request("POST", url, headers=headers, data=payload)
                if response.status_code == 200:
                    print(f"Created new user: {username}")
                    res_data = response.json()
                    url = f"https://mis.hcaschennai.edu.in/pykit_user/"
                    dj_payload = json.dumps({"username": username,"first_name": "user","last_name": res_data['psk_id'],"email": data.get('email', ''),"password": data.get('password', ''), "user_roles": str({1})})

                    response = requests.post(url, headers=headers, data=dj_payload)
                    if response.status_code == 200:
                        print(f"🆕 Created new user: {username}")
                    else:
                        print(f"❌ Failed to create user {username}: {response.text}")
                else:
                    print(f"Failed to create user {username}: {response.text}")
                
                
               

    
               

                
    return redirect("pykit_staff_list")


    
def mail_validation_filter(request):
    url = f"{API_STUDIO_URL}getapi/asa0508_01_01/all"
    response = requests.get(url)

    valid_users = []
    invalid_users = []
    all_users = []

    if response.status_code == 200:
        all_users = response.json()

        for user in all_users:
            email_id = user.get('user_email_id', '')
            if email_id and re.match(r"[^@]+@[^@]+\.[^@]+", email_id):
                valid_users.append(user)
            else:
                invalid_users.append(user)

    # Get filter from GET query param
    email_filter = request.GET.get('filter', 'all')

    if email_filter == 'valid':
        filtered_staff = valid_users
    elif email_filter == 'invalid':
        filtered_staff = invalid_users
    else:
        filtered_staff = all_users

    return render(request, 'pykit_staff_list.html', {'total_pykit_staff': filtered_staff,'selected_filter': email_filter,})


def pykit_staff_list(request):
    url = f"{API_STUDIO_URL}getapi/asa0508_01_01/all"
    payload = {}
    headers = {}
    
    response = requests.request("GET", url, headers=headers, data=payload)
    if response.status_code == 200:
        total_pykit_staff = response.json()
        print("total_pykit_staff:", total_pykit_staff)
    return render(request, 'pykit_staff_list.html', {"total_pykit_staff":total_pykit_staff})

def user_menu(request):
    return render(request, 'Research_templates/user_menus.html')

def research_delete(request, id):
    delete_url = f"{API_STUDIO_URL}deleteapi/delete/naac01_research_article_publication_dc1/{id}"
    delete_response = requests.delete(delete_url)

    if delete_response.status_code == 200:
        return redirect('research_list')
    else:
        return HttpResponse("Failed to delete participation: " + delete_response.text)
 
import pandas as pd
import io
import requests
import json
from django.http import HttpResponse
from django.shortcuts import render
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from datetime import datetime
from openpyxl.styles import Font, Alignment
from django.contrib import messages
from collections import Counter
from user_management.settings_views import get_settings
from MIS.functions import validate_file_format_faculty, validate_file_size
from user_management.settings_views import *


import pandas as pd
from django.http import HttpResponse
import io
import requests
import json
from datetime import datetime
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

# def combined_filter_view(request):
#     """
#     Unified view for Faculty Participations, Project Courses, Books & Chapters, and Research Articles.
#     Implements role-based filtering (HOD, NAAC, Staff) with export support.
#     Dynamic year filtering based on selected staff.
#     """
#     import io, json, requests
#     from django.http import HttpResponse
#     from datetime import datetime

#     user = get_settings(request)
#     username = user.get('username', '')

#     # --- User role ---
#     role_url = f"{API_STUDIO_URL}getapi/asa0504_01_01"
#     payload = json.dumps({
#         "queries": [{"field": "username", "value": username, "operation": "equal"}],
#         "search_type": "first"
#     })
#     role_resp = requests.post(role_url, headers={'Content-Type': 'application/json'}, data=payload)
#     value_user = int(role_resp.json().get("user_roles", "0").strip("{}")) if role_resp.status_code == 200 else 0

#     role_list = roles_tbl(request)
#     user_role = next(
#         (role.get("user_role") for role in role_list if role.get("psk_id") == value_user),
#         "Staff"
#     )

#     access_token, token_type = research_key()
#     research_data = get_research_data(access_token, token_type)

#     # --- Filter parameters ---
#     staff_id = request.GET.get('staff_id', '').strip()
#     staff_name = request.GET.get('staff_name', '').strip()
#     selected_department = request.GET.get('department', '').strip()
#     year = request.GET.get('year', '').strip()
#     selected_form = request.GET.get('form_type', '').strip()
#     export_format = request.GET.get('export', '').strip()
#     project_type = request.GET.get('project_type', 'all')

#     # --- Staff & Department filtering ---
#     if user_role == "Hod":
#         staff_info = next((s for s in research_data if s.get("stf_id") == username), {})
#         department_name = staff_info.get("department", "")
#         staff_list = [s for s in research_data if s.get("department") == department_name]
#         departments = [department_name]
#     elif user_role == "Naac":
#         staff_list = research_data
#         departments = sorted(list({s.get("department", "Unknown") for s in staff_list}))
#         if selected_department:
#             staff_list = [s for s in staff_list if s.get("department") == selected_department]

#         # Filter staff by department
#         if selected_department:
#             staff_list = [s for s in staff_list if s.get("department") == selected_department]

#     else:
#         staff_list = [s for s in research_data if s.get("stf_id") == username]
#         departments = []

#     staff_data = [{'id': s.get("stf_id"), 'name': s.get("stf_name"), 'department': s.get("department", "Unknown")} for s in staff_list]

#     # --- Determine API endpoints ---
#     parent_url, child_url, media_api, parent_id_field, child_parent_field = None, None, None, None, None
#     staff_field, year_field = None, None
#     export_fn_excel, export_fn_pdf = None, None
#     has_children = True

#     if selected_form == "participations":
#         parent_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_faculty_participation_dc1/all"
#         child_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_faculty_participation_dc2/all"
#         media_api = "naac01_faculty_participation_dc2_media"
#         parent_id_field = 'psk_id'
#         child_parent_field = 'transaction_id'
#         staff_field = 'stf_id'
#         year_field = 'year'
#         export_fn_excel = export_to_excel
#         export_fn_pdf = export_to_pdf

#     elif selected_form == "projects":
#         parent_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_project_work_dc1/all"
#         child_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_project_work_dc2/all"
#         media_api = "naac01_project_work_dc2_media"
#         parent_id_field = 'psk_id'
#         child_parent_field = 'transaction_id'
#         staff_field = 'staff_id'
#         year_field = 'academic_year'
#         export_fn_excel = export_projects_to_excel
#         export_fn_pdf = export_projects_to_pdf

#     elif selected_form == "books_and_chapters":
#         parent_url = f"{API_STUDIO_URL}getapi/naac01_books_and_chapter_dc1/all"
#         media_api = "naac01_books_and_chapter_dc1_media"
#         parent_id_field = 'psk_id'
#         staff_field = "staff_id"
#         year_field = "year_of_publication"
#         export_fn_pdf = export_books_and_chapters_to_pdf
#         export_fn_excel = export_books_and_chapters_to_excel
#         has_children = False

#     elif selected_form == "research_articles":
#         parent_url = f"{API_STUDIO_URL}getapi/naac01_research_article_publication_dc1/all"
#         media_api = "naac01_research_article_publication_dc1_media"
#         parent_id_field = 'psk_id'
#         staff_field = "employee_id"
#         year_field = "publication_year"
#         export_fn_pdf = export_research_articles_to_pdf   # implement like books & chapters
#         export_fn_excel = export_research_articles_to_excel
#         has_children = False

#     else:
#         current_year = datetime.now().year
#         context = {
#             'staff_data': staff_data,
#             'departments': departments,
#             'selected_department': selected_department,
#             'years': [f"{y}-{y+1}" for y in range(current_year-5, current_year+1)],
#             'selected_staff_id': staff_id,
#             'selected_staff_name': staff_name,
#             'selected_year': year,
#             'selected_form': selected_form,
#             'user_role': user_role,
#             'filter_applied': False
#         }
#         return render(request, 'combined_filter.html', context)

#     # --- Fetch parent data ---
#     parent_resp = requests.get(parent_url)
#     all_parents = parent_resp.json() if parent_resp.status_code == 200 else []

#     # --- Filter parents by staff ---
#     if staff_id:
#         filtered_parents = [p for p in all_parents if str(p.get(staff_field, "")) == staff_id]
#     elif staff_name:
#         filtered_parents = [p for p in all_parents if p.get("stf_name", "") == staff_name or p.get("staff_name", "") == staff_name]
#     else:
#         filtered_parents = [p for p in all_parents if p.get(staff_field) in [s['id'] for s in staff_data]]

#     # --- Attach staff info for books & chapters / research ---
#     if selected_form in ["books_and_chapters", "research_articles"]:
#         for p in filtered_parents:
#             staff_info = next((s for s in research_data if s.get("stf_id") == p.get(staff_field)), {})
#             p["staff_name"] = staff_info.get("stf_name", "N/A")
#             p["department_name"] = staff_info.get("department", "N/A")

#     # --- Year filter ---
#     available_years = sorted(list({str(p.get(year_field, "")) for p in filtered_parents}), reverse=True)
#     if year and year != "all":
#         filtered_parents = [p for p in filtered_parents if str(p.get(year_field, "")) == year]

#     # --- Attach media ---
#     for p in filtered_parents:
#         pid = p.get(parent_id_field)
#         media_resp = requests.get(f"{API_STUDIO_URL}crudapp/get/media/{media_api}/parent/{pid}")
#         media_files = []
#         if media_resp.status_code == 200:
#             for m in media_resp.json():
#                 media_id = m.get('id') or m.get('psk_id') or m.get('value_id') or pid
#                 media_files.append({
#                     'file_name': m.get('file_name', 'Unknown'),
#                     'media_id': media_id,
#                     'direct_api_url': f"{API_STUDIO_URL}crudapp/view/media/{media_api}/{media_id}"
#                 })
#         p['media_files'] = media_files

#     # --- Fetch children if applicable ---
#     filtered_children = []
#     if has_children:
#         child_resp = requests.get(child_url)
#         all_children = child_resp.json() if child_resp.status_code == 200 else []
#         parent_ids = [p.get(parent_id_field) for p in filtered_parents]
#         filtered_children = [c for c in all_children if c.get(child_parent_field) in parent_ids]

#         # Attach child media
#         for child in filtered_children:
#             child_id = child.get(parent_id_field)
#             media_resp = requests.get(f"{API_STUDIO_URL}crudapp/get/media/{media_api}/parent/{child_id}")
#             processed_media = []
#             if media_resp.status_code == 200:
#                 for m in media_resp.json():
#                     media_id = m.get('id') or m.get('psk_id') or m.get('value_id') or child_id
#                     processed_media.append({
#                         'file_name': m.get('file_name', 'Unknown'),
#                         'media_id': media_id,
#                         'direct_api_url': f"{API_STUDIO_URL}crudapp/view/media/{media_api}/{media_id or child_id}"
#                     })
#             child['media_files'] = processed_media

#         # Merge parent info
#         parent_map = {p[parent_id_field]: p for p in filtered_parents}
#         for child in filtered_children:
#             pid = child.get(child_parent_field)
#             if pid in parent_map:
#                 child['parent_data'] = parent_map[pid]

#     # --- Export if requested ---
#     if export_format.lower() == "pdf":
#         if selected_form == "projects":
#             return export_fn_pdf(filtered_parents, filtered_children, project_type)
#         return export_fn_pdf(filtered_parents) if not has_children else export_fn_pdf(filtered_parents, filtered_children)
#     elif export_format.lower() == "excel" and export_fn_excel:
#         if selected_form == "projects":
#             return export_fn_excel(filtered_parents, filtered_children, project_type)
#         return export_fn_excel(filtered_parents) if not has_children else export_fn_excel(filtered_parents, filtered_children)

#     # --- Render template ---
#     context = {
#         'staff_data': staff_data,
#         'departments': departments,
#         'selected_department': selected_department,
#         'years': available_years,
#         'selected_staff_id': staff_id,
#         'selected_staff_name': staff_name,
#         'selected_year': year,
#         'selected_form': selected_form,
#         'user_role': user_role,
#         'filter_applied': any([staff_id, staff_name, selected_department, year, selected_form]),
#         'parents': filtered_parents,
#         'children': filtered_children,
#         'data_type': selected_form
#     }

#     return render(request, 'combined_filter.html', context)

PARTICIPATION_OPTIONS = [
    'Board of Studies', 'Question Paper Setting', 'Evaluation', 
    'Add-On ', 'Certificate Courses',
    'External Examiner', 'Conference', 'Seminar', 'Workshop', 'FDP'
]

# def combined_filter_view(request):
#     """
#     Unified view for Faculty Participations, Project Courses, Books & Chapters, and Research Articles.
#     Implements role-based filtering (HOD, NAAC, Staff) with export support.
#     Dynamic year filtering based on selected staff.
#     """
#     import io, json, requests
#     from django.http import HttpResponse
#     from datetime import datetime

#     user = get_settings(request)
#     username = user.get('username', '')

#     # --- User role ---
#     role_url = f"{API_STUDIO_URL}getapi/asa0504_01_01"
#     payload = json.dumps({
#         "queries": [{"field": "username", "value": username, "operation": "equal"}],
#         "search_type": "first"
#     })
#     role_resp = requests.post(role_url, headers={'Content-Type': 'application/json'}, data=payload)
#     value_user = int(role_resp.json().get("user_roles", "0").strip("{}")) if role_resp.status_code == 200 else 0

#     role_list = roles_tbl(request)
#     user_role = next(
#         (role.get("user_role") for role in role_list if role.get("psk_id") == value_user),
#         "Staff"
#     )

#     access_token, token_type = research_key()
#     research_data = get_research_data(access_token, token_type)

#     # --- Filter parameters ---
#     staff_id = request.GET.get('staff_id', '').strip()
#     staff_name = request.GET.get('staff_name', '').strip()
#     selected_department = request.GET.get('department', '').strip()
#     year = request.GET.get('year', '').strip()
#     selected_form = request.GET.get('form_type', '').strip()
#     export_format = request.GET.get('export', '').strip()
#     project_type = request.GET.get('project_type', 'all')
#     # ADDED: Participation options filter
#     selected_options = request.GET.getlist('options')

#     # --- Staff & Department filtering ---
#     if user_role == "Hod":
#         staff_info = next((s for s in research_data if s.get("stf_id") == username), {})
#         department_name = staff_info.get("department", "")
#         staff_list = [s for s in research_data if s.get("department") == department_name]
#         departments = [department_name]
#     elif user_role == "Naac":
#         staff_list = research_data
#         departments = sorted(list({s.get("department", "Unknown") for s in staff_list}))
#         if selected_department:
#             staff_list = [s for s in staff_list if s.get("department") == selected_department]

#         # Filter staff by department
#         if selected_department:
#             staff_list = [s for s in staff_list if s.get("department") == selected_department]

#     else:
#         staff_list = [s for s in research_data if s.get("stf_id") == username]
#         departments = []

#     staff_data = [{'id': s.get("stf_id"), 'name': s.get("stf_name"), 'department': s.get("department", "Unknown")} for s in staff_list]

#     # --- Determine API endpoints ---
#     parent_url, child_url, media_api, parent_id_field, child_parent_field = None, None, None, None, None
#     staff_field, year_field = None, None
#     export_fn_excel, export_fn_pdf = None, None
#     has_children = True

#     if selected_form == "participations":
#         parent_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_faculty_participation_dc1/all"
#         child_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_faculty_participation_dc2/all"
#         media_api = "naac01_faculty_participation_dc2_media"
#         parent_id_field = 'psk_id'
#         child_parent_field = 'transaction_id'
#         staff_field = 'stf_id'
#         year_field = 'year'
#         export_fn_excel = export_to_excel
#         export_fn_pdf = export_to_pdf

#     elif selected_form == "projects":
#         parent_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_project_work_dc1/all"
#         child_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_project_work_dc2/all"
#         media_api = "naac01_project_work_dc2_media"
#         parent_id_field = 'psk_id'
#         child_parent_field = 'transaction_id'
#         staff_field = 'staff_id'
#         year_field = 'academic_year'
#         export_fn_excel = export_projects_to_excel
#         export_fn_pdf = export_projects_to_pdf

#     elif selected_form == "books_and_chapters":
#         parent_url = f"{API_STUDIO_URL}getapi/naac01_books_and_chapter_dc1/all"
#         media_api = "naac01_books_and_chapter_dc1_media"
#         parent_id_field = 'psk_id'
#         staff_field = "staff_id"
#         year_field = "year_of_publication"
#         export_fn_pdf = export_books_and_chapters_to_pdf
#         export_fn_excel = export_books_and_chapters_to_excel
#         has_children = False

#     elif selected_form == "research_articles":
#         parent_url = f"{API_STUDIO_URL}getapi/naac01_research_article_publication_dc1/all"
#         media_api = "naac01_research_article_publication_dc1_media"
#         parent_id_field = 'psk_id'
#         staff_field = "employee_id"
#         year_field = "publication_year"
#         export_fn_pdf = export_research_articles_to_pdf   # implement like books & chapters
#         export_fn_excel = export_research_articles_to_excel
#         has_children = False

#     else:
#         current_year = datetime.now().year
#         context = {
#             'staff_data': staff_data,
#             'departments': departments,
#             'selected_department': selected_department,
#             'years': [f"{y}-{y+1}" for y in range(current_year-5, current_year+1)],
#             'selected_staff_id': staff_id,
#             'selected_staff_name': staff_name,
#             'selected_year': year,
#             'selected_form': selected_form,
#             'user_role': user_role,
#             'filter_applied': False,
#             # ADDED: Participation options for the template
#             'options': PARTICIPATION_OPTIONS if 'PARTICIPATION_OPTIONS' in globals() else [
#                 "Board of Studies", "Question Paper Setting", "Evaluation", 
#                 "Add-On Program", "Certificate Courses", "External Examiner",
#                 "Conference", "Seminar", "Workshop", "FDP"
#             ],
#             'selected_options': selected_options
#         }
#         return render(request, 'combined_filter.html', context)

#     # --- Fetch parent data ---
#     parent_resp = requests.get(parent_url)
#     all_parents = parent_resp.json() if parent_resp.status_code == 200 else []

#     # --- Filter parents by staff ---
#     if staff_id:
#         filtered_parents = [p for p in all_parents if str(p.get(staff_field, "")) == staff_id]
#     elif staff_name:
#         filtered_parents = [p for p in all_parents if p.get("stf_name", "") == staff_name or p.get("staff_name", "") == staff_name]
#     else:
#         filtered_parents = [p for p in all_parents if p.get(staff_field) in [s['id'] for s in staff_data]]

#     # --- ADDED: Participation options filtering ---
#     if selected_form == "participations" and selected_options:
#         filtered_parents = [parent for parent in filtered_parents 
#                           if parent.get('participation') and 
#                           any(opt in parent['participation'] for opt in selected_options)]

#     # --- Attach staff info for books & chapters / research ---
#     if selected_form in ["books_and_chapters", "research_articles"]:
#         for p in filtered_parents:
#             staff_info = next((s for s in research_data if s.get("stf_id") == p.get(staff_field)), {})
#             p["staff_name"] = staff_info.get("stf_name", "N/A")
#             p["department_name"] = staff_info.get("department", "N/A")

#     # --- Year filter ---
#     available_years = sorted(list({str(p.get(year_field, "")) for p in filtered_parents}), reverse=True)
#     if year and year != "all":
#         filtered_parents = [p for p in filtered_parents if str(p.get(year_field, "")) == year]

#     # --- Attach media ---
#     for p in filtered_parents:
#         pid = p.get(parent_id_field)
#         media_resp = requests.get(f"{API_STUDIO_URL}crudapp/get/media/{media_api}/parent/{pid}")
#         media_files = []
#         if media_resp.status_code == 200:
#             for m in media_resp.json():
#                 media_id = m.get('id') or m.get('psk_id') or m.get('value_id') or pid
#                 media_files.append({'file_name': m.get('file_name', 'Unknown'),'media_id': media_id,'direct_api_url': f"{API_STUDIO_URL}crudapp/view/media/{media_api}/{media_id}"})
#         p['media_files'] = media_files

#     # --- ADDED: Pre-split participation strings for template ---
#     if selected_form == "participations":
#         for parent in filtered_parents:
#             if parent.get('participation'):
#                 # Split by comma and strip whitespace from each part
#                 parent['participation_list'] = [part.strip() for part in parent['participation'].split(',')]
#             else:
#                 parent['participation_list'] = []

#     # --- Fetch children if applicable ---
#     filtered_children = []
#     if has_children:
#         child_resp = requests.get(child_url)
#         all_children = child_resp.json() if child_resp.status_code == 200 else []
#         parent_ids = [p.get(parent_id_field) for p in filtered_parents]
#         filtered_children = [c for c in all_children if c.get(child_parent_field) in parent_ids]

#         # Attach child media
#         for child in filtered_children:
#             child_id = child.get(parent_id_field)
#             media_resp = requests.get(f"{API_STUDIO_URL}crudapp/get/media/{media_api}/parent/{child_id}")
#             processed_media = []
#             if media_resp.status_code == 200:
#                 for m in media_resp.json():
#                     media_id = m.get('id') or m.get('psk_id') or m.get('value_id') or child_id
#                     processed_media.append({
#                         'file_name': m.get('file_name', 'Unknown'),
#                         'media_id': media_id,
#                         'direct_api_url': f"{API_STUDIO_URL}crudapp/view/media/{media_api}/{media_id or child_id}"
#                     })
#             child['media_files'] = processed_media

#         # Merge parent info
#         parent_map = {p[parent_id_field]: p for p in filtered_parents}
#         for child in filtered_children:
#             pid = child.get(child_parent_field)
#             if pid in parent_map:
#                 child['parent_data'] = parent_map[pid]

#     # --- Export if requested ---
#     if export_format.lower() == "pdf":
#         if selected_form == "projects":
#             return export_fn_pdf(filtered_parents, filtered_children, project_type)
#         # ADDED: Pass selected_options for participations export
#         elif selected_form == "participations":
#             return export_fn_pdf(filtered_parents, filtered_children, selected_options)
#         return export_fn_pdf(filtered_parents) if not has_children else export_fn_pdf(filtered_parents, filtered_children)
#     elif export_format.lower() == "excel" and export_fn_excel:
#         if selected_form == "projects":
#             return export_fn_excel(filtered_parents, filtered_children, project_type)
#         # ADDED: Pass selected_options for participations export
#         elif selected_form == "participations":
#             return export_fn_excel(filtered_parents, filtered_children, selected_options)
#         return export_fn_excel(filtered_parents) if not has_children else export_fn_excel(filtered_parents, filtered_children)

#     # --- Render template ---
#     context = {
#         'staff_data': staff_data,
#         'departments': departments,
#         'selected_department': selected_department,
#         'years': available_years,
#         'selected_staff_id': staff_id,
#         'selected_staff_name': staff_name,
#         'selected_year': year,
#         'selected_form': selected_form,
#         'user_role': user_role,
#         'filter_applied': any([staff_id, staff_name, selected_department, year, selected_form, selected_options]),
#         'parents': filtered_parents,
#         'children': filtered_children,
#         'data_type': selected_form,
#         # ADDED: Participation options for the template
#         'options': PARTICIPATION_OPTIONS if 'PARTICIPATION_OPTIONS' in globals() else [
#             "Board of Studies", "Question Paper Setting", "Evaluation", 
#             "Add-On Program", "Certificate Courses", "External Examiner",
#             "Conference", "Seminar", "Workshop", "FDP"
#         ],
#         'selected_options': selected_options
#     }

#     return render(request, 'combined_filter.html', context)

# def staff_filter_view(request):
#     """
#     Unified view for Faculty Participations, Project Courses, Books & Chapters, Research Articles, and Career Counseling.
#     Implements role-based filtering (HOD, NAAC, Staff) with export support.
#     Dynamic year filtering based on selected staff.
#     """
#     import io, json, requests
#     from django.http import HttpResponse
#     from datetime import datetime

#     user = get_settings(request)
#     username = user.get('username', '')

#     # --- User role ---
#     role_url = f"{API_STUDIO_URL}getapi/asa0504_01_01"
#     payload = json.dumps({
#         "queries": [{"field": "username", "value": username, "operation": "equal"}],
#         "search_type": "first"
#     })
#     role_resp = requests.post(role_url, headers={'Content-Type': 'application/json'}, data=payload)
#     value_user = int(role_resp.json().get("user_roles", "0").strip("{}")) if role_resp.status_code == 200 else 0

#     role_list = roles_tbl(request)
#     user_role = next(
#         (role.get("user_role") for role in role_list if role.get("psk_id") == value_user),
#         "Staff"
#     )

#     access_token, token_type = research_key()
#     research_data = get_research_data(access_token, token_type)

#     # --- Filter parameters ---
#     staff_id = request.GET.get('staff_id', '').strip()
#     staff_name = request.GET.get('staff_name', '').strip()
#     selected_department = request.GET.get('department', '').strip()
#     year = request.GET.get('year', '').strip()
#     selected_form = request.GET.get('form_type', '').strip()
#     export_format = request.GET.get('export', '').strip()
#     project_type = request.GET.get('project_type', 'all')
#     # ADDED: Participation options filter
#     selected_options = request.GET.getlist('options')

#     # --- Staff & Department filtering ---
#     if user_role == "Hod":
#         staff_info = next((s for s in research_data if s.get("stf_id") == username), {})
#         department_name = staff_info.get("department", "")
#         staff_list = [s for s in research_data if s.get("department") == department_name]
#         departments = [department_name]
#     elif user_role == "Naac":
#         staff_list = research_data
#         departments = sorted(list({s.get("department", "Unknown") for s in staff_list}))
#         if selected_department:
#             staff_list = [s for s in staff_list if s.get("department") == selected_department]

#         # Filter staff by department
#         if selected_department:
#             staff_list = [s for s in staff_list if s.get("department") == selected_department]

#     else:
#         staff_list = [s for s in research_data if s.get("stf_id") == username]
#         departments = []

#     staff_data = [{'id': s.get("stf_id"), 'name': s.get("stf_name"), 'department': s.get("department", "Unknown")} for s in staff_list]

#     # --- Determine API endpoints ---
#     parent_url, child_url, media_api, parent_id_field, child_parent_field = None, None, None, None, None
#     staff_field, year_field = None, None
#     export_fn_excel, export_fn_pdf = None, None
#     has_children = True

#     if selected_form == "participations":
#         parent_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_faculty_participation_dc1/all"
#         child_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_faculty_participation_dc2/all"
#         media_api = "naac01_faculty_participation_dc2_media"
#         parent_id_field = 'psk_id'
#         child_parent_field = 'transaction_id'
#         staff_field = 'stf_id'
#         year_field = 'year'
#         export_fn_excel = export_to_excel
#         export_fn_pdf = export_to_pdf

#     elif selected_form == "projects":
#         parent_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_project_work_dc1/all"
#         child_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_project_work_dc2/all"
#         media_api = "naac01_project_work_dc2_media"
#         parent_id_field = 'psk_id'
#         child_parent_field = 'transaction_id'
#         staff_field = 'staff_id'
#         year_field = 'academic_year'
#         export_fn_excel = export_projects_to_excel
#         export_fn_pdf = export_projects_to_pdf

#     elif selected_form == "books_and_chapters":
#         parent_url = f"{API_STUDIO_URL}getapi/naac01_books_and_chapter_dc1/all"
#         media_api = "naac01_books_and_chapter_dc1_media"
#         parent_id_field = 'psk_id'
#         staff_field = "staff_id"
#         year_field = "year_of_publication"
#         export_fn_pdf = export_books_and_chapters_to_pdf
#         export_fn_excel = export_books_and_chapters_to_excel
#         has_children = False

#     elif selected_form == "research_articles":
#         parent_url = f"{API_STUDIO_URL}getapi/naac01_research_article_publication_dc1/all"
#         media_api = "naac01_research_article_publication_dc1_media"
#         parent_id_field = 'psk_id'
#         staff_field = "employee_id"
#         year_field = "publication_year"
#         export_fn_pdf = export_research_articles_to_pdf
#         export_fn_excel = export_research_articles_to_excel
#         has_children = False

#     # ADDED: Career Counseling form type
#     elif selected_form == "career_counseling":
#         parent_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_students_benefited_for_career_counseeling_dc1/all"
#         media_api = "naac01_students_benefited_for_career_counseeling_dc1_media"
#         parent_id_field = 'psk_id'
#         staff_field = "staff_id"
#         year_field = "year_of_activity"
#         export_fn_pdf = export_career_counseeling_to_pdf
#         export_fn_excel = export_career_counseeling_to_excel
#         has_children = False
    
#     elif selected_form == "collaborative_students":
#         parent_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_students_benefited_dc1/all"
#         media_api = "naac01_students_benefited_dc1_media"
#         parent_id_field = 'psk_id'
#         staff_field = "staff_id"
#         year_field = "year"
#         export_fn_pdf = export_collaborative_students_to_pdf
#         export_fn_excel = export_collaborative_students_to_excel
#         has_children = False
#     # ADDED: Competitive Examination form type
#     elif selected_form == "competitive_examination":
#         # Try different URL variations
#         parent_url = f"{API_STUDIO_URL}getapi/naac01_students_benefited_exam_guidance_dc1/all"
#         media_api = "naac01_students_benefited_exam_guidance_dc1_media"
#         parent_id_field = 'psk_id'
#         staff_field = "staff_id"
#         year_field = "year_of_activity"
#         export_fn_pdf = export_competitive_examination_to_pdf
#         export_fn_excel = export_competitive_examination_to_excel
#         has_children = False
#     elif selected_form == "Extentions":
#         # Try different URL variations
#         parent_url = f"{API_STUDIO_URL}getapi/naac01_no_of_extension_and_outreach_programs_dc1/all"
#         media_api = "naac01_no_of_extension_and_outreach_programs_dc1_media"
#         parent_id_field = 'psk_id'
#         staff_field = "staff_id"
#         year_field = "year_of_activity"
#         export_fn_pdf = export_Extentions_to_pdf
#         export_fn_excel = export_Extentions_to_excel
#         has_children = False
#     elif selected_form == "Collaborative_Faculty":
#         # Try different URL variations
#         parent_url = f"{API_STUDIO_URL}getapi/naac01_collaborative_activities_for_faculty_exchange_dc1/all"
#         media_api = "naac01_collaborative_activities_for_faculty_exchange_dc1_media"
#         parent_id_field = 'psk_id'
#         staff_field = "staff_id"
#         year_field = "collaboration_year"
#         export_fn_pdf = export_collaborative_faculty_to_pdf
#         export_fn_excel = export_collaborative_faculty_to_excel
#         has_children = False
#     elif selected_form == "governments_grants":
#         # Try different URL variations
#         parent_url = f"{API_STUDIO_URL}getapi/naac01_government_grants_dc1/all"
#         media_api = "naac01_government_grants_dc1_media"
#         parent_id_field = 'psk_id'
#         staff_field = "staff_id"
#         year_field = "year_of_award"
#         export_fn_pdf = export_governments_grants_to_pdf
#         export_fn_excel = export_governments_grants_to_excel
#         has_children = False
#     elif selected_form == "program_offered":
#         # Try different URL variations
#         parent_url = f"{API_STUDIO_URL}getapi/naac01_add_on_certificate_dc1/all"
#         media_api = "naac01_add_on_certificate_dc1_media"
#         parent_id_field = 'psk_id'
#         staff_field = "staff_id"
#         year_field = "year_of_offering"
#         export_fn_pdf = export_programs_offered_to_pdf
#         export_fn_excel = export_programs_offered_to_excel
#         has_children = False
#     elif selected_form == "event_workshop":
#         # Try different URL variations
#         parent_url = f"{API_STUDIO_URL}getapi/naac01_number_of_workshop_conducted_dc1/all"
#         media_api = "naac01_number_of_workshop_conducted_dc1_media"
#         parent_id_field = 'psk_id'
#         staff_field = "staff_id"
#         year_field = "year_of_offering"
#         export_fn_pdf = export_workshops_to_pdf
#         export_fn_excel = export_workshops_to_excel
#         has_children = False
#     else:
#         current_year = datetime.now().year
#         context = {
#             'staff_data': staff_data,
#             'departments': departments,
#             'selected_department': selected_department,
#             'years': [f"{y}-{y+1}" for y in range(current_year-5, current_year+1)],
#             'selected_staff_id': staff_id,
#             'selected_staff_name': staff_name,
#             'selected_year': year,
#             'selected_form': selected_form,
#             'user_role': user_role,
#             'filter_applied': False,
#             # ADDED: Participation options for the template
#             'options': PARTICIPATION_OPTIONS if 'PARTICIPATION_OPTIONS' in globals() else [
#                 "Board of Studies", "Question Paper Setting", "Evaluation", 
#                 "Add-On Program", "Certificate Courses", "External Examiner",
#                 "Conference", "Seminar", "Workshop", "FDP"
#             ],
#             'selected_options': selected_options
#         }
#         return render(request, 'combined_filter.html', context)

#     # --- Fetch parent data ---
#     parent_resp = requests.get(parent_url)
#     all_parents = parent_resp.json() if parent_resp.status_code == 200 else []

#     # --- Filter parents by staff ---
#     if staff_id:
#         filtered_parents = [p for p in all_parents if str(p.get(staff_field, "")) == staff_id]
#     elif staff_name:
#         filtered_parents = [p for p in all_parents if p.get("stf_name", "") == staff_name or p.get("staff_name", "") == staff_name]
#     else:
#         filtered_parents = [p for p in all_parents if p.get(staff_field) in [s['id'] for s in staff_data]]

#     # --- ADDED: Participation options filtering ---
#     if selected_form == "participations" and selected_options:
#         filtered_parents = [parent for parent in filtered_parents 
#                           if parent.get('participation') and 
#                           any(opt in parent['participation'] for opt in selected_options)]

#     # --- Attach staff info for books & chapters / research / career counseling ---
#     if selected_form in ["books_and_chapters", "research_articles", "career_counseling, collaborative_students, competitive_examination, Extentions, Collaborative_Faculty, governments_grants, event_workshop"]:
#         for p in filtered_parents:
#             staff_info = next((s for s in research_data if s.get("stf_id") == p.get(staff_field)), {})
#             p["staff_name"] = staff_info.get("stf_name", "N/A")
#             p["department_name"] = staff_info.get("department", "N/A")

#     # --- Year filter ---
#     available_years = sorted(list({str(p.get(year_field, "")) for p in filtered_parents}), reverse=True)
#     if year and year != "all":
#         filtered_parents = [p for p in filtered_parents if str(p.get(year_field, "")) == year]

#     # --- Attach media ---
#     for p in filtered_parents:
#         pid = p.get(parent_id_field)
#         media_resp = requests.get(f"{API_STUDIO_URL}crudapp/get/media/{media_api}/parent/{pid}")
#         media_files = []
#         if media_resp.status_code == 200:
#             for m in media_resp.json():
#                 media_id = m.get('id') or m.get('psk_id') or m.get('value_id') or pid
#                 media_files.append({'file_name': m.get('file_name', 'Unknown'),'media_id': media_id,'direct_api_url': f"{API_STUDIO_URL}crudapp/view/media/{media_api}/{media_id}"})
#         p['media_files'] = media_files

#     # --- ADDED: Pre-split participation strings for template ---
#     if selected_form == "participations":
#         for parent in filtered_parents:
#             if parent.get('participation'):
#                 # Split by comma and strip whitespace from each part
#                 parent['participation_list'] = [part.strip() for part in parent['participation'].split(',')]
#             else:
#                 parent['participation_list'] = []

#     # --- Fetch children if applicable ---
#     filtered_children = []
#     if has_children:
#         child_resp = requests.get(child_url)
#         all_children = child_resp.json() if child_resp.status_code == 200 else []
#         parent_ids = [p.get(parent_id_field) for p in filtered_parents]
#         filtered_children = [c for c in all_children if c.get(child_parent_field) in parent_ids]

#         # Attach child media
#         for child in filtered_children:
#             child_id = child.get(parent_id_field)
#             media_resp = requests.get(f"{API_STUDIO_URL}crudapp/get/media/{media_api}/parent/{child_id}")
#             processed_media = []
#             if media_resp.status_code == 200:
#                 for m in media_resp.json():
#                     media_id = m.get('id') or m.get('psk_id') or m.get('value_id') or child_id
#                     processed_media.append({
#                         'file_name': m.get('file_name', 'Unknown'),
#                         'media_id': media_id,
#                         'direct_api_url': f"{API_STUDIO_URL}crudapp/view/media/{media_api}/{media_id or child_id}"
#                     })
#             child['media_files'] = processed_media

#         # Merge parent info
#         parent_map = {p[parent_id_field]: p for p in filtered_parents}
#         for child in filtered_children:
#             pid = child.get(child_parent_field)
#             if pid in parent_map:
#                 child['parent_data'] = parent_map[pid]

#     # --- Export if requested ---
#     if export_format.lower() == "pdf":
#         if selected_form == "projects":
#             return export_fn_pdf(filtered_parents, filtered_children, project_type)
#         # ADDED: Pass selected_options for participations export
#         elif selected_form == "participations":
#             return export_fn_pdf(filtered_parents, filtered_children, selected_options)
#         return export_fn_pdf(filtered_parents) if not has_children else export_fn_pdf(filtered_parents, filtered_children)
#     elif export_format.lower() == "excel" and export_fn_excel:
#         if selected_form == "projects":
#             return export_fn_excel(filtered_parents, filtered_children, project_type)
#         # ADDED: Pass selected_options for participations export
#         elif selected_form == "participations":
#             return export_fn_excel(filtered_parents, filtered_children, selected_options)
#         return export_fn_excel(filtered_parents) if not has_children else export_fn_excel(filtered_parents, filtered_children)

#     # --- Render template ---
#     context = {
#         'staff_data': staff_data,
#         'departments': departments,
#         'selected_department': selected_department,
#         'years': available_years,
#         'selected_staff_id': staff_id,
#         'selected_staff_name': staff_name,
#         'selected_year': year,
#         'selected_form': selected_form,
#         'user_role': user_role,
#         'filter_applied': any([staff_id, staff_name, selected_department, year, selected_form, selected_options]),
#         'parents': filtered_parents,
#         'children': filtered_children,
#         'data_type': selected_form,
#         # ADDED: Participation options for the template
#         'options': PARTICIPATION_OPTIONS if 'PARTICIPATION_OPTIONS' in globals() else [
#             "Board of Studies", "Question Paper Setting", "Evaluation", 
#             "Add-On Program", "Certificate Courses", "External Examiner",
#             "Conference", "Seminar", "Workshop", "FDP"
#         ],
#         'selected_options': selected_options
#     }

#     return render(request, 'combined_filter.html', context)


def staff_filter_view(request):
    """
    Unified view for Faculty Participations, Project Courses, Books & Chapters, Research Articles, and Career Counseling.
    Implements role-based filtering (HOD, NAAC, Staff) with export support.
    Dynamic year filtering based on selected staff.
    """
    import io, json, requests
    from django.http import HttpResponse
    from datetime import datetime

    user = get_settings(request)
    username = user.get('username', '')

    # --- User role ---
    role_url = f"{API_STUDIO_URL}getapi/asa0504_01_01"
    payload = json.dumps({
        "queries": [{"field": "username", "value": username, "operation": "equal"}],
        "search_type": "first"
    })
    role_resp = requests.post(role_url, headers={'Content-Type': 'application/json'}, data=payload)
    value_user = int(role_resp.json().get("user_roles", "0").strip("{}")) if role_resp.status_code == 200 else 0

    role_list = roles_tbl(request)
    user_role = next(
        (role.get("user_role") for role in role_list if role.get("psk_id") == value_user),
        "Staff"
    )

    access_token, token_type = research_key()
    research_data = get_research_data(access_token, token_type)

   
    

    # --- Filter parameters ---
    staff_id = request.GET.get('staff_id', '').strip()
    staff_name = request.GET.get('staff_name', '').strip()
   
    selected_department = request.GET.get('department', '').strip()
    year = request.GET.get('year', '').strip()
    selected_form = request.GET.get('form_type', '').strip()
    export_format = request.GET.get('export', '').strip()
    project_type = request.GET.get('project_type', 'all')
    # ADDED: Participation options filter
    selected_options = request.GET.getlist('options')

    # --- Staff & Department filtering ---
    if user_role == "Hod":
        staff_info = next((s for s in research_data if s.get("stf_id") == username), {})
        department_name = staff_info.get("department", "")
        staff_list = [s for s in research_data if s.get("department") == department_name]
        departments = [department_name]
    elif user_role == "Naac":
        staff_list = research_data
        departments = sorted(list({s.get("department", "Unknown") for s in staff_list}))
        if selected_department:
            staff_list = [s for s in staff_list if s.get("department") == selected_department]
    else:
        staff_list = [s for s in research_data if s.get("stf_id") == username]


        departments = []

    staff_info = next((s for s in research_data if s.get("stf_id") == username), {})

    department_name = staff_info.get("department", "N/A")
    print("department_name:",department_name)    

    staff_info = next((s for s in research_data if s.get("stf_id") == username), {})
    user_department = staff_info.get("department", "N/A")
    print("user_department:", user_department)  
    selected_department = request.GET.get('department', '').strip()

    staff_data = [{'id': s.get("stf_id"), 'name': s.get("stf_name"), 'department': s.get("department", "Unknown")} for s in staff_list]

    # --- Determine API endpoints ---
    parent_url, child_url, media_api, parent_id_field, child_parent_field = None, None, None, None, None
    staff_field, year_field = None, None
    export_fn_excel, export_fn_pdf = None, None
    has_children = True

    if selected_form == "participations":
        parent_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_faculty_participation_dc1/all"
        child_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_faculty_participation_dc2/all"
        media_api = "naac01_faculty_participation_dc2_media"
        parent_id_field = 'psk_id'
        child_parent_field = 'transaction_id'
        staff_field = 'stf_id'
        year_field = 'year'
        export_fn_excel = export_to_excel
        export_fn_pdf = export_to_pdf

    elif selected_form == "projects":
        parent_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_project_work_dc1/all"
        child_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_project_work_dc2/all"
        media_api = "naac01_project_work_dc2_media"
        parent_id_field = 'psk_id'
        child_parent_field = 'transaction_id'
        staff_field = 'staff_id'
        year_field = 'academic_year'
        export_fn_excel = export_projects_to_excel
        export_fn_pdf = export_projects_to_pdf

    elif selected_form == "books_and_chapters":
        parent_url = f"{API_STUDIO_URL}getapi/naac01_books_and_chapter_dc1/all"
        media_api = "naac01_books_and_chapter_dc1_media"
        parent_id_field = 'psk_id'
        staff_field = "staff_id"
        year_field = "year_of_publication"
        export_fn_pdf = export_books_and_chapters_to_pdf
        export_fn_excel = export_books_and_chapters_to_excel
        has_children = False

    elif selected_form == "research_articles":
        parent_url = f"{API_STUDIO_URL}getapi/naac01_research_article_publication_dc1/all"
        media_api = "naac01_research_article_publication_dc1_media"
        parent_id_field = 'psk_id'
        staff_field = "employee_id"
        year_field = "publication_year"
        export_fn_pdf = export_research_articles_to_pdf
        export_fn_excel = export_research_articles_to_excel
        has_children = False

    # ADDED: Career Counseling form type
    elif selected_form == "career_counseling":
        parent_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_students_benefited_for_career_counseeling_dc1/all"
        media_api = "naac01_students_benefited_for_career_counseeling_dc1_media"
        parent_id_field = 'psk_id'
        staff_field = "staff_id"
        year_field = "year_of_activity"
        export_fn_pdf = export_career_counseeling_to_pdf
        export_fn_excel = export_career_counseeling_to_excel
        has_children = False
    
    elif selected_form == "collaborative_students":
        parent_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_students_benefited_dc1/all"
        media_api = "naac01_students_benefited_dc1_media"
        parent_id_field = 'psk_id'
        staff_field = "staff_id"
        year_field = "year"
        export_fn_pdf = export_collaborative_students_to_pdf
        export_fn_excel = export_collaborative_students_to_excel
        has_children = False
    # ADDED: Competitive Examination form type
    elif selected_form == "competitive_examination":
        # Try different URL variations
        parent_url = f"{API_STUDIO_URL}getapi/naac01_students_benefited_exam_guidance_dc1/all"
        media_api = "naac01_students_benefited_exam_guidance_dc1_media"
        parent_id_field = 'psk_id'
        staff_field = "staff_id"
        year_field = "year_of_activity"
        export_fn_pdf = export_competitive_examination_to_pdf
        export_fn_excel = export_competitive_examination_to_excel
        has_children = False
    elif selected_form == "Extentions":
        # Try different URL variations
        parent_url = f"{API_STUDIO_URL}getapi/naac01_no_of_extension_and_outreach_programs_dc1/all"
        media_api = "naac01_no_of_extension_and_outreach_programs_dc1_media"
        parent_id_field = 'psk_id'
        staff_field = "staff_id"
        year_field = "year_of_activity"
        export_fn_pdf = export_Extentions_to_pdf
        export_fn_excel = export_Extentions_to_excel
        has_children = False
    elif selected_form == "Collaborative_Faculty":
        # Try different URL variations
        parent_url = f"{API_STUDIO_URL}getapi/naac01_collaborative_activities_for_faculty_exchange_dc1/all"
        media_api = "naac01_collaborative_activities_for_faculty_exchange_dc1_media"
        parent_id_field = 'psk_id'
        staff_field = "staff_id"
        year_field = "collaboration_year"
        export_fn_pdf = export_collaborative_faculty_to_pdf
        export_fn_excel = export_collaborative_faculty_to_excel
        has_children = False
    elif selected_form == "governments_grants":
        # Try different URL variations
        parent_url = f"{API_STUDIO_URL}getapi/naac01_government_grants_dc1/all"
        media_api = "naac01_government_grants_dc1_media"
        parent_id_field = 'psk_id'
        staff_field = "staff_id"
        year_field = "year_of_award"
        export_fn_pdf = export_governments_grants_to_pdf
        export_fn_excel = export_governments_grants_to_excel
        has_children = False
    elif selected_form == "program_offered":
        # Try different URL variations
        parent_url = f"{API_STUDIO_URL}getapi/naac01_add_on_certificate_dc1/all"
        media_api = "naac01_add_on_certificate_dc1_media"
        parent_id_field = 'psk_id'
        staff_field = "staff_id"
        year_field = "year_of_offering"
        export_fn_pdf = export_programs_offered_to_pdf
        export_fn_excel = export_programs_offered_to_excel
        has_children = False
    elif selected_form == "event_workshop":
        # Try different URL variations
        parent_url = f"{API_STUDIO_URL}getapi/naac01_number_of_workshop_conducted_dc1/all"
        media_api = "naac01_number_of_workshop_conducted_dc1_media"
        parent_id_field = 'psk_id'
        staff_field = "staff_id"
        year_field = "year_of_offering"
        export_fn_pdf = export_workshops_to_pdf
        export_fn_excel = export_workshops_to_excel
        has_children = False
    else:
        current_year = datetime.now().year
        context = {
            'staff_data': staff_data,
            'departments': departments,
            'selected_department': selected_department,
            'years': get_available_years(),  # FIXED: Use the function
            'selected_staff_id': staff_id,
            'selected_staff_name': staff_name,
            'selected_year': year,
            'selected_form': selected_form,
            'user_role': user_role,
            'user_department': user_department,  
            'filter_applied': False,
            'department_name': department_name,
            # ADDED: Participation options for the template
            'options': PARTICIPATION_OPTIONS if 'PARTICIPATION_OPTIONS' in globals() else [
                "Board of Studies", "Question Paper Setting", "Evaluation", 
                "Add-On Program", "Certificate Courses", "External Examiner",
                "Conference", "Seminar", "Workshop", "FDP"
            ],
            'selected_options': selected_options
        }
        return render(request, 'combined_filter.html', context)

    # --- Fetch parent data ---
    parent_resp = requests.get(parent_url)
    all_parents = parent_resp.json() if parent_resp.status_code == 200 else []

    # --- Filter parents by staff ---
    if staff_id:
        filtered_parents = [p for p in all_parents if str(p.get(staff_field, "")).strip() == staff_id.strip()]
    elif staff_name:
        filtered_parents = [p for p in all_parents if p.get("stf_name", "").strip() == staff_name.strip() or p.get("staff_name", "").strip() == staff_name.strip()]
    else:
        filtered_parents = [p for p in all_parents if p.get(staff_field) in [s['id'] for s in staff_data]]

    # --- ADDED: Participation options filtering ---
    if selected_form == "participations" and selected_options:
        filtered_parents = [parent for parent in filtered_parents 
                          if parent.get('participation') and 
                          any(opt in parent['participation'] for opt in selected_options)]

    # --- FIXED: Year filter with normalization ---
    def normalize_year(year_str):
        """Normalize year format for comparison"""
        if not year_str:
            return ""
        year_str = str(year_str).strip()
        
        # Handle different year formats
        if '-' in year_str:
            # Format: "2023-2024" or "2023-24"
            parts = year_str.split('-')
            if len(parts) == 2:
                # Extract just the start year for comparison
                return parts[0]
        return year_str

    # Get available years from filtered data
    available_years_from_data = []
    for p in filtered_parents:
        raw_year = p.get(year_field, "")
        if raw_year:
            normalized = normalize_year(raw_year)
            if normalized and normalized not in available_years_from_data:
                available_years_from_data.append(normalized)
    
    available_years_from_data = sorted(available_years_from_data, reverse=True)

    # Apply year filter using the function's years
    if year and year != "all" and year != "":
        normalized_filter_year = normalize_year(year)
        filtered_parents = [p for p in filtered_parents 
                          if normalize_year(p.get(year_field, "")) == normalized_filter_year]
    
    # Use the function for available years in template
    available_years_for_template = get_available_years()

    # --- FIXED: Attach staff info for various form types ---
    form_types_needing_staff_info = [
        "books_and_chapters", "research_articles", "career_counseling", 
        "collaborative_students", "competitive_examination", "Extentions", 
        "Collaborative_Faculty", "governments_grants", "event_workshop", "program_offered"
    ]
    
    if selected_form in form_types_needing_staff_info:
        for p in filtered_parents:
            staff_info = next((s for s in research_data if s.get("stf_id") == p.get(staff_field)), {})
            p["staff_name"] = staff_info.get("stf_name", "N/A")
            p["department_name"] = staff_info.get("department", "N/A")

    # --- Attach media ---
    for p in filtered_parents:
        pid = p.get(parent_id_field)
        media_resp = requests.get(f"{API_STUDIO_URL}crudapp/get/media/{media_api}/parent/{pid}")
        media_files = []
        if media_resp.status_code == 200:
            for m in media_resp.json():
                media_id = m.get('id') or m.get('psk_id') or m.get('value_id') or pid
                media_files.append({
                    'file_name': m.get('file_name', 'Unknown'),
                    'media_id': media_id,
                    'direct_api_url': f"{API_STUDIO_URL}crudapp/view/media/{media_api}/{media_id}"
                })
        p['media_files'] = media_files

    # --- ADDED: Pre-split participation strings for template ---
    if selected_form == "participations":
        for parent in filtered_parents:
            if parent.get('participation'):
                # Split by comma and strip whitespace from each part
                parent['participation_list'] = [part.strip() for part in parent['participation'].split(',')]
            else:
                parent['participation_list'] = []

    # --- Fetch children if applicable ---
    filtered_children = []
    if has_children:
        child_resp = requests.get(child_url)
        all_children = child_resp.json() if child_resp.status_code == 200 else []
        parent_ids = [p.get(parent_id_field) for p in filtered_parents]
        filtered_children = [c for c in all_children if c.get(child_parent_field) in parent_ids]

        # Attach child media
        for child in filtered_children:
            child_id = child.get(parent_id_field)
            media_resp = requests.get(f"{API_STUDIO_URL}crudapp/get/media/{media_api}/parent/{child_id}")
            processed_media = []
            if media_resp.status_code == 200:
                for m in media_resp.json():
                    media_id = m.get('id') or m.get('psk_id') or m.get('value_id') or child_id
                    processed_media.append({
                        'file_name': m.get('file_name', 'Unknown'),
                        'media_id': media_id,
                        'direct_api_url': f"{API_STUDIO_URL}crudapp/view/media/{media_api}/{media_id or child_id}"
                    })
            child['media_files'] = processed_media

        # Merge parent info
        parent_map = {p[parent_id_field]: p for p in filtered_parents}
        for child in filtered_children:
            pid = child.get(child_parent_field)
            if pid in parent_map:
                child['parent_data'] = parent_map[pid]

    # --- Export if requested ---
    if export_format.lower() == "pdf":
        if selected_form == "projects":
            return export_fn_pdf(filtered_parents, filtered_children, project_type)
        # ADDED: Pass selected_options for participations export
        elif selected_form == "participations":
            return export_fn_pdf(filtered_parents, filtered_children, selected_options)
        return export_fn_pdf(filtered_parents) if not has_children else export_fn_pdf(filtered_parents, filtered_children)
    elif export_format.lower() == "excel" and export_fn_excel:
        if selected_form == "projects":
            return export_fn_excel(filtered_parents, filtered_children, project_type)
        # ADDED: Pass selected_options for participations export
        elif selected_form == "participations":
            return export_fn_excel(filtered_parents, filtered_children, selected_options)
        return export_fn_excel(filtered_parents) if not has_children else export_fn_excel(filtered_parents, filtered_children)

    # --- Render template ---
    context = {
        'staff_data': staff_data,
        'departments': departments,
        'selected_department': selected_department,
        'user_department': user_department,
        'years': available_years_for_template,  # FIXED: Use function result
        'selected_staff_id': staff_id,
        'selected_staff_name': staff_name,
        'selected_year': year,
        'selected_form': selected_form,
        'user_role': user_role,
        'filter_applied': any([staff_id, staff_name, selected_department, year, selected_form, selected_options]),
        'parents': filtered_parents,
        'children': filtered_children,
        'data_type': selected_form,
        # ADDED: Participation options for the template
        'options': PARTICIPATION_OPTIONS if 'PARTICIPATION_OPTIONS' in globals() else [
            "Board of Studies", "Question Paper Setting", "Evaluation", 
            "Add-On Program", "Certificate Courses", "External Examiner",
            "Conference", "Seminar", "Workshop", "FDP"
        ],
        'selected_options': selected_options
    }

    return render(request, 'combined_filter.html', context)


def get_available_years():
    """Return list of years from 2020 to current year"""
    current_year = datetime.now().year
    return [str(y) for y in range(2020, current_year + 1)]

def hod_filter_view(request):
    """
    Unified view for Faculty Participations, Project Courses, Books & Chapters, Research Articles, and Career Counseling.
    Implements role-based filtering (HOD, NAAC, Staff) with export support.
    Dynamic year filtering based on selected staff.
    """
    import io, json, requests
    from django.http import HttpResponse
    from datetime import datetime

    user = get_settings(request)
    username = user.get('username', '')

    # --- User role ---
    role_url = f"{API_STUDIO_URL}getapi/asa0504_01_01"
    payload = json.dumps({
        "queries": [{"field": "username", "value": username, "operation": "equal"}],
        "search_type": "first"
    })
    role_resp = requests.post(role_url, headers={'Content-Type': 'application/json'}, data=payload)
    value_user = int(role_resp.json().get("user_roles", "0").strip("{}")) if role_resp.status_code == 200 else 0

    role_list = roles_tbl(request)
    user_role = next(
        (role.get("user_role") for role in role_list if role.get("psk_id") == value_user),
        "Staff"
    )

    access_token, token_type = research_key()
    research_data = get_research_data(access_token, token_type)

    # --- Filter parameters ---
    staff_id = request.GET.get('staff_id', '').strip()
    staff_name = request.GET.get('staff_name', '').strip()
    selected_department = request.GET.get('department', '').strip()
    year = request.GET.get('year', '').strip()
    selected_form = request.GET.get('form_type', '').strip()
    export_format = request.GET.get('export', '').strip()
    project_type = request.GET.get('project_type', 'all')
    # ADDED: Participation options filter
    selected_options = request.GET.getlist('options')




        # --- Staff & Department filtering ---
    # --- Staff & Department filtering ---

    # if user_role == "Hod":
    #     # HOD can only see their own data
    #     staff_list = [s for s in research_data if s.get("stf_id") == username]
    #     if staff_list:
    #         department_name = staff_list[0].get("department", "")
    #         departments = [department_name] if department_name else []
    #     else:
    #         departments = []
    #     print(f"🔍 [FILTER DEBUG] HOD mode - Staff: {len(staff_list)}, Dept: {departments}")

    if user_role == "Hod":
    # HOD can only see their own data
        staff_list = [s for s in research_data if s.get("stf_id") == username]
        if staff_list:
            user_department = staff_list[0].get("department", "")  # CHANGE: department_name → user_department
            departments = [user_department] if user_department else []
        else:
            user_department = "N/A"  # ADD THIS LINE
            departments = []
        print(f"🔍 [FILTER DEBUG] HOD mode - Staff: {len(staff_list)}, Dept: {user_department}")

    # elif user_role == "Staff":
    #     # Regular staff can only see their own data
    #     staff_list = [s for s in research_data if s.get("stf_id") == username]
    #     departments = []
    #     print(f"🔍 [FILTER DEBUG] Staff mode - Staff: {len(staff_list)}")

    elif user_role == "Staff":
    # Regular staff can only see their own data
        staff_list = [s for s in research_data if s.get("stf_id") == username]
        if staff_list:
            user_department = staff_list[0].get("department", "N/A")  # ADD THIS LINE
        else:
            user_department = "N/A"  # ADD THIS LINE
        departments = []
        print(f"🔍 [FILTER DEBUG] Staff mode - Staff: {len(staff_list)}, Dept: {user_department}")

    else:
    # NAAC and other roles can see ALL HODs and ALL staff
        user_department = "All Departments"  # ADD THIS LINE
        staff_list = research_data.copy()  # All staff
        
        # Get all unique departments from ALL staff
        departments = sorted(list({s.get("department", "Unknown") for s in research_data if s.get("department")}))
        print(f"🔍 [FILTER DEBUG] NAAC/Other initial - All staff: {len(staff_list)}, All depts: {len(departments)}")
        
        # Apply department filter if selected
        if selected_department:
            staff_list = [s for s in staff_list if s.get("department") == selected_department]
            print(f"🔍 [FILTER DEBUG] After dept filter '{selected_department}': {len(staff_list)} staff")
        
        # Apply staff ID filter if selected
        if staff_id:
            staff_list = [s for s in staff_list if str(s.get("stf_id", "")) == staff_id]
            print(f"🔍 [FILTER DEBUG] After staff_id filter '{staff_id}': {len(staff_list)} staff")
        
        # Apply staff name filter if selected  
        if staff_name:
            staff_list = [s for s in staff_list if s.get("stf_name", "") == staff_name]
            print(f"🔍 [FILTER DEBUG] After staff_name filter '{staff_name}': {len(staff_list)} staff")

    # Final staff data for template
    staff_data = [{'id': s.get("stf_id"), 'name': s.get("stf_name"), 'department': s.get("department", "Unknown")} for s in staff_list]
    print(f"🔍 [FILTER DEBUG] Final staff_data: {len(staff_data)} records")

    # --- Determine API endpoints ---
    parent_url, child_url, media_api, parent_id_field, child_parent_field = None, None, None, None, None
    staff_field, year_field = None, None
    export_fn_excel, export_fn_pdf = None, None
    has_children = True

    if selected_form == "participations":
        parent_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_faculty_participation_dc1/all"
        child_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_faculty_participation_dc2/all"
        media_api = "naac01_faculty_participation_dc2_media"
        parent_id_field = 'psk_id'
        child_parent_field = 'transaction_id'
        staff_field = 'stf_id'
        year_field = 'year'
        export_fn_excel = export_to_excel
        export_fn_pdf = export_to_pdf

    elif selected_form == "projects":
        parent_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_project_work_dc1/all"
        child_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_project_work_dc2/all"
        media_api = "naac01_project_work_dc2_media"
        parent_id_field = 'psk_id'
        child_parent_field = 'transaction_id'
        staff_field = 'staff_id'
        year_field = 'academic_year'
        export_fn_excel = export_projects_to_excel
        export_fn_pdf = export_projects_to_pdf

    elif selected_form == "books_and_chapters":
        parent_url = f"{API_STUDIO_URL}getapi/naac01_books_and_chapter_dc1/all"
        media_api = "naac01_books_and_chapter_dc1_media"
        parent_id_field = 'psk_id'
        staff_field = "staff_id"
        year_field = "year_of_publication"
        export_fn_pdf = export_books_and_chapters_to_pdf
        export_fn_excel = export_books_and_chapters_to_excel
        has_children = False

    elif selected_form == "research_articles":
        parent_url = f"{API_STUDIO_URL}getapi/naac01_research_article_publication_dc1/all"
        media_api = "naac01_research_article_publication_dc1_media"
        parent_id_field = 'psk_id'
        staff_field = "employee_id"
        year_field = "publication_year"
        export_fn_pdf = export_research_articles_to_pdf
        export_fn_excel = export_research_articles_to_excel
        has_children = False

    # ADDED: Career Counseling form type
    elif selected_form == "career_counseling":
        parent_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_students_benefited_for_career_counseeling_dc1/all"
        media_api = "naac01_students_benefited_for_career_counseeling_dc1_media"
        parent_id_field = 'psk_id'
        staff_field = "staff_id"
        year_field = "year_of_activity"
        export_fn_pdf = export_career_counseeling_to_pdf
        export_fn_excel = export_career_counseeling_to_excel
        has_children = False
    
    elif selected_form == "collaborative_students":
        parent_url = f"{API_STUDIO_URL}getapi/all_fields/naac01_students_benefited_dc1/all"
        media_api = "naac01_students_benefited_dc1_media"
        parent_id_field = 'psk_id'
        staff_field = "staff_id"
        year_field = "year"
        export_fn_pdf = export_collaborative_students_to_pdf
        export_fn_excel = export_collaborative_students_to_excel
        has_children = False
    # ADDED: Competitive Examination form type
    elif selected_form == "competitive_examination":
        # Try different URL variations
        parent_url = f"{API_STUDIO_URL}getapi/naac01_students_benefited_exam_guidance_dc1/all"
        media_api = "naac01_students_benefited_exam_guidance_dc1_media"
        parent_id_field = 'psk_id'
        staff_field = "staff_id"
        year_field = "year_of_activity"
        export_fn_pdf = export_competitive_examination_to_pdf
        export_fn_excel = export_competitive_examination_to_excel
        has_children = False
    elif selected_form == "Extentions":
        # Try different URL variations
        parent_url = f"{API_STUDIO_URL}getapi/naac01_no_of_extension_and_outreach_programs_dc1/all"
        media_api = "naac01_no_of_extension_and_outreach_programs_dc1_media"
        parent_id_field = 'psk_id'
        staff_field = "staff_id"
        year_field = "year_of_activity"
        export_fn_pdf = export_Extentions_to_pdf
        export_fn_excel = export_Extentions_to_excel
        has_children = False
    elif selected_form == "Collaborative_Faculty":
        # Try different URL variations
        parent_url = f"{API_STUDIO_URL}getapi/naac01_collaborative_activities_for_faculty_exchange_dc1/all"
        media_api = "naac01_collaborative_activities_for_faculty_exchange_dc1_media"
        parent_id_field = 'psk_id'
        staff_field = "staff_id"
        year_field = "collaboration_year"
        export_fn_pdf = export_collaborative_faculty_to_pdf
        export_fn_excel = export_collaborative_faculty_to_excel
        has_children = False
    elif selected_form == "governments_grants":
        # Try different URL variations
        parent_url = f"{API_STUDIO_URL}getapi/naac01_government_grants_dc1/all"
        media_api = "naac01_government_grants_dc1_media"
        parent_id_field = 'psk_id'
        staff_field = "staff_id"
        year_field = "year_of_award"
        export_fn_pdf = export_governments_grants_to_pdf
        export_fn_excel = export_governments_grants_to_excel
        has_children = False
    elif selected_form == "program_offered":
        # Try different URL variations
        parent_url = f"{API_STUDIO_URL}getapi/naac01_add_on_certificate_dc1/all"
        media_api = "naac01_add_on_certificate_dc1_media"
        parent_id_field = 'psk_id'
        staff_field = "staff_id"
        year_field = "year_of_offering"
        export_fn_pdf = export_programs_offered_to_pdf
        export_fn_excel = export_programs_offered_to_excel
        has_children = False
    elif selected_form == "event_workshop":
        # Try different URL variations
        parent_url = f"{API_STUDIO_URL}getapi/naac01_number_of_workshop_conducted_dc1/all"
        media_api = "naac01_number_of_workshop_conducted_dc1_media"
        parent_id_field = 'psk_id'
        staff_field = "staff_id"
        year_field = "year_of_offering"
        export_fn_pdf = export_workshops_to_pdf
        export_fn_excel = export_workshops_to_excel
        has_children = False
    else:
        current_year = datetime.now().year
        context = {
            'staff_data': staff_data,
            'departments': departments,
            'selected_department': selected_department,
            'user_department': user_department,  
            'years': [f"{y}-{y+1}" for y in range(current_year-5, current_year+1)],
            'selected_staff_id': staff_id,
            'selected_staff_name': staff_name,
            'selected_year': year,
            'selected_form': selected_form,
            'user_role': user_role,
            'filter_applied': False,
            # ADDED: Participation options for the template
            'options': PARTICIPATION_OPTIONS if 'PARTICIPATION_OPTIONS' in globals() else [
                "Board of Studies", "Question Paper Setting", "Evaluation", 
                "Add-On Program", "Certificate Courses", "External Examiner",
                "Conference", "Seminar", "Workshop", "FDP"
            ],
            'selected_options': selected_options
        }
        return render(request, 'hod_filter.html', context)

    # --- Fetch parent data ---
    parent_resp = requests.get(parent_url)
    all_parents = parent_resp.json() if parent_resp.status_code == 200 else []

    # --- Filter parents by staff ---
    if staff_id:
        filtered_parents = [p for p in all_parents if str(p.get(staff_field, "")) == staff_id]
    elif staff_name:
        filtered_parents = [p for p in all_parents if p.get("stf_name", "") == staff_name or p.get("staff_name", "") == staff_name]
    else:
        filtered_parents = [p for p in all_parents if p.get(staff_field) in [s['id'] for s in staff_data]]

    # --- ADDED: Participation options filtering ---
    if selected_form == "participations" and selected_options:
        filtered_parents = [parent for parent in filtered_parents 
                          if parent.get('participation') and 
                          any(opt in parent['participation'] for opt in selected_options)]

    # --- Attach staff info for books & chapters / research / career counseling ---
    if selected_form in ["books_and_chapters", "research_articles", "career_counseling, collaborative_students, competitive_examination, Extentions, Collaborative_Faculty, governments_grants, event_workshop"]:
        for p in filtered_parents:
            staff_info = next((s for s in research_data if s.get("stf_id") == p.get(staff_field)), {})
            p["staff_name"] = staff_info.get("stf_name", "N/A")
            p["department_name"] = staff_info.get("department", "N/A")

    # --- Year filter ---
    available_years = sorted(list({str(p.get(year_field, "")) for p in filtered_parents}), reverse=True)
    if year and year != "all":
        filtered_parents = [p for p in filtered_parents if str(p.get(year_field, "")) == year]

    # --- Attach media ---
    for p in filtered_parents:
        pid = p.get(parent_id_field)
        media_resp = requests.get(f"{API_STUDIO_URL}crudapp/get/media/{media_api}/parent/{pid}")
        media_files = []
        if media_resp.status_code == 200:
            for m in media_resp.json():
                media_id = m.get('id') or m.get('psk_id') or m.get('value_id') or pid
                media_files.append({'file_name': m.get('file_name', 'Unknown'),'media_id': media_id,'direct_api_url': f"{API_STUDIO_URL}crudapp/view/media/{media_api}/{media_id}"})
        p['media_files'] = media_files

    # --- ADDED: Pre-split participation strings for template ---
    if selected_form == "participations":
        for parent in filtered_parents:
            if parent.get('participation'):
                # Split by comma and strip whitespace from each part
                parent['participation_list'] = [part.strip() for part in parent['participation'].split(',')]
            else:
                parent['participation_list'] = []

    # --- Fetch children if applicable ---
    filtered_children = []
    if has_children:
        child_resp = requests.get(child_url)
        all_children = child_resp.json() if child_resp.status_code == 200 else []
        parent_ids = [p.get(parent_id_field) for p in filtered_parents]
        filtered_children = [c for c in all_children if c.get(child_parent_field) in parent_ids]

        # Attach child media
        for child in filtered_children:
            child_id = child.get(parent_id_field)
            media_resp = requests.get(f"{API_STUDIO_URL}crudapp/get/media/{media_api}/parent/{child_id}")
            processed_media = []
            if media_resp.status_code == 200:
                for m in media_resp.json():
                    media_id = m.get('id') or m.get('psk_id') or m.get('value_id') or child_id
                    processed_media.append({
                        'file_name': m.get('file_name', 'Unknown'),
                        'media_id': media_id,
                        'direct_api_url': f"{API_STUDIO_URL}crudapp/view/media/{media_api}/{media_id or child_id}"
                    })
            child['media_files'] = processed_media

        # Merge parent info
        parent_map = {p[parent_id_field]: p for p in filtered_parents}
        for child in filtered_children:
            pid = child.get(child_parent_field)
            if pid in parent_map:
                child['parent_data'] = parent_map[pid]

    # --- Export if requested ---
    if export_format.lower() == "pdf":
        if selected_form == "projects":
            return export_fn_pdf(filtered_parents, filtered_children, project_type)
        # ADDED: Pass selected_options for participations export
        elif selected_form == "participations":
            return export_fn_pdf(filtered_parents, filtered_children, selected_options)
        return export_fn_pdf(filtered_parents) if not has_children else export_fn_pdf(filtered_parents, filtered_children)
    elif export_format.lower() == "excel" and export_fn_excel:
        if selected_form == "projects":
            return export_fn_excel(filtered_parents, filtered_children, project_type)
        # ADDED: Pass selected_options for participations export
        elif selected_form == "participations":
            return export_fn_excel(filtered_parents, filtered_children, selected_options)
        return export_fn_excel(filtered_parents) if not has_children else export_fn_excel(filtered_parents, filtered_children)

    # --- Render template ---
    context = {
        'staff_data': staff_data,
        'departments': departments,
        'selected_department': selected_department,
        'years': available_years,
        'selected_staff_id': staff_id,
        'selected_staff_name': staff_name,
        'selected_year': year,
        'selected_form': selected_form,
        'user_role': user_role,
        'filter_applied': any([staff_id, staff_name, selected_department, year, selected_form, selected_options]),
        'parents': filtered_parents,
        'children': filtered_children,
        'data_type': selected_form,
        # ADDED: Participation options for the template
        'options': PARTICIPATION_OPTIONS if 'PARTICIPATION_OPTIONS' in globals() else [
            "Board of Studies", "Question Paper Setting", "Evaluation", 
            "Add-On Program", "Certificate Courses", "External Examiner",
            "Conference", "Seminar", "Workshop", "FDP"
        ],
        'selected_options': selected_options
    }

    return render(request, 'hod_filter.html', context)



def get_available_years():
    """
    Return a sorted list of years for dropdowns.
    Pulls from parent/project data or static range.
    """
    # Example static range
    current_year = datetime.now().year
    return [str(y) for y in range(current_year - 5, current_year + 1)]



# def export_research_articles_to_pdf(parents, children=None):
#     """
#     Export Research Articles to PDF (landscape), grouped by staff.
#     Media files are listed only once per research article without category labels.
#     Fields: Paper Title, Journal Name, Year, ISSN, Attachments
#     """
#     from reportlab.lib.pagesizes import A4, landscape
#     from reportlab.lib import colors
#     from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
#     from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, PageBreak
#     from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
#     from reportlab.lib.units import inch
#     from django.http import HttpResponse

#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="research_articles.pdf"'

#     doc = SimpleDocTemplate(
#         response,
#         pagesize=landscape(A4),
#         topMargin=0.5 * inch,
#         bottomMargin=0.5 * inch,
#         leftMargin=0.5 * inch,
#         rightMargin=0.5 * inch
#     )

#     elements = []
#     styles = getSampleStyleSheet()

#     # ====== Custom Styles ======
#     title_style = ParagraphStyle('Title', parent=styles['Heading2'], alignment=TA_CENTER, spaceAfter=12)
#     table_header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontSize=9,
#                                         alignment=TA_CENTER, textColor=colors.white, fontName='Helvetica-Bold')
#     table_cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=8,
#                                       alignment=TA_LEFT, wordWrap='CJK')
#     attachment_link_style = ParagraphStyle('AttachmentLink', parent=styles['Normal'], fontSize=7,
#                                            alignment=TA_LEFT, textColor=colors.HexColor('#1a5276'), wordWrap='CJK')

#     # ====== Media Lookup ======
#     media_lookup = {}
#     if children:
#         for child in children:
#             psk_id = child.get('psk_id')
#             if psk_id:
#                 media_lookup[psk_id] = child.get('media_files', [])
#     else:
#         for p in parents:
#             media_lookup[p.get('psk_id')] = p.get('media_files', [])

#     parents_sorted = sorted(parents, key=lambda x: (x.get('employee_id', ''), x.get('paper_title', '')))
#     current_staff = None
#     table_data = []
#     col_widths = [3 * inch, 3 * inch, 0.7 * inch, 1.3 * inch, 2 * inch]

#     for parent in parents_sorted:
#         staff_id = parent.get('employee_id', 'N/A')
#         staff_name = parent.get('staff_name', 'N/A')
#         department_name = parent.get('department_name', 'N/A')
#         psk_id = parent.get('psk_id')

#         # ====== Flush Previous Staff Table ======
#         if current_staff and current_staff != staff_id and table_data:
#             table = Table(table_data, colWidths=col_widths, repeatRows=1, splitByRow=True)
#             table.setStyle(TableStyle([
#                 ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#02548b')),
#                 ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
#                 ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Center align content
#                 ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), # Vertically center content
#                 ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d6d6d6')),
#                 ('LEFTPADDING', (0, 0), (-1, -1), 4),
#                 ('RIGHTPADDING', (0, 0), (-1, -1), 4),
#                 ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
#                 ('TOPPADDING', (0, 0), (-1, -1), 4),
#             ]))
#             elements.append(table)
#             elements.append(PageBreak())
#             table_data = []

#         # ====== Add New Staff Section ======
#         if current_staff != staff_id:
#             elements.append(Paragraph("Research Articles", title_style))
#             elements.append(Spacer(1, 6))

#             # Staff info section
#             left_style = ParagraphStyle('LeftAlign', parent=table_cell_style, alignment=TA_LEFT)
#             right_style = ParagraphStyle('RightAlign', parent=table_cell_style, alignment=TA_RIGHT)
#             center_style = ParagraphStyle('CenterAlign', parent=table_cell_style, alignment=TA_CENTER)

#             # Line 1: Staff ID (left) | Department (right)
#             staff_id_para = Paragraph(f"<b>Staff ID:</b> {staff_id}", left_style)
#             dept_para = Paragraph(f"<b>Department:</b> {department_name}", right_style)
#             line1 = Table([[staff_id_para, dept_para]], colWidths=[5.5 * inch, 4.5 * inch])
#             line1.setStyle(TableStyle([
#                 ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
#                 ('ALIGN', (0, 0), (0, -1), 'LEFT'),
#                 ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
#                 ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
#                 ('TOPPADDING', (0, 0), (-1, -1), 3),
#             ]))
#             elements.append(line1)

#             # Line 2: Staff Name (left only)
#             staff_name_para = Paragraph(f"<b>Staff Name:</b> {staff_name}", left_style)
#             line2 = Table([[staff_name_para]], colWidths=[10 * inch])
#             line2.setStyle(TableStyle([
#                 ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
#                 ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
#                 ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
#                 ('TOPPADDING', (0, 0), (-1, -1), 3),
#             ]))
#             elements.append(line2)
#             elements.append(Spacer(1, 10))

#             # Table headers
#             headers = ['Paper Title', 'Journal Name', 'Year', 'ISSN', 'Attachments']
#             table_data.append([Paragraph(h, table_header_style) for h in headers])
#             current_staff = staff_id

#         # ====== Media Files ======
#         media_files = media_lookup.get(psk_id, [])
#         media_text = []
#         seen_files = set()
#         for media in media_files:
#             fname = media.get("file_name", "Unknown")
#             url = media.get("direct_api_url", "")
#             if not url or fname in seen_files:
#                 continue
#             seen_files.add(fname)
#             display_name = fname if len(fname) <= 50 else fname[:50] + "..."
#             media_text.append(f'<link href="{url}" color="blue">{display_name}</link>')

#         media_paragraph = Paragraph('<br/>'.join(media_text) if media_text else "No attachments", attachment_link_style)

#         # Data row
#         row = [
#             Paragraph(str(parent.get('paper_title', '')), table_cell_style),
#             Paragraph(str(parent.get('journal_name', '')), table_cell_style),
#             Paragraph(str(parent.get('publication_year', '')), table_cell_style),
#             Paragraph(str(parent.get('issn_number', '')), table_cell_style),
#             media_paragraph
#         ]
#         table_data.append(row)

#     # ====== Final Table Flush ======
#     if table_data:
#         table = Table(table_data, colWidths=col_widths, repeatRows=1, splitByRow=True)
#         table.setStyle(TableStyle([
#             ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#02548b')),
#             ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
#             ('ALIGN', (0, 0), (-1, -1), 'CENTER'),   # Center align for table content
#             ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Vertical middle alignment
#             ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d6d6d6')),
#             ('LEFTPADDING', (0, 0), (-1, -1), 4),
#             ('RIGHTPADDING', (0, 0), (-1, -1), 4),
#             ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
#             ('TOPPADDING', (0, 0), (-1, -1), 4),
#         ]))
#         elements.append(table)

#     doc.build(elements)
#     return response

def export_research_articles_to_pdf(parents, children=None):
    """
    Export Research Articles to PDF (landscape), grouped by staff.
    Media files are listed only once per research article without category labels.
    Fields: Paper Title, Journal Name, Year, ISSN, Attachments
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, PageBreak
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.lib.units import inch
    from django.http import HttpResponse

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="research_articles.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch
    )

    elements = []
    styles = getSampleStyleSheet()

    # ====== Custom Styles ======
    title_style = ParagraphStyle('Title', parent=styles['Heading2'], alignment=TA_CENTER, spaceAfter=12)
    table_header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontSize=9,
                                        alignment=TA_CENTER, textColor=colors.white, fontName='Helvetica-Bold')
    table_cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=8,
                                      alignment=TA_LEFT, wordWrap='CJK')
    attachment_link_style = ParagraphStyle('AttachmentLink', parent=styles['Normal'], fontSize=7,
                                           alignment=TA_LEFT, textColor=colors.HexColor('#1a5276'), wordWrap='CJK')
    no_data_style = ParagraphStyle('NoData', parent=styles['Normal'], fontSize=10,
                                   alignment=TA_CENTER, textColor=colors.HexColor('#7f8c8d'),
                                   fontStyle='italic', spaceBefore=12, spaceAfter=12)

    # ====== Check if there are any parents ======
    if not parents:
        elements.append(Paragraph("Research Articles", title_style))
        elements.append(Paragraph("No data available", no_data_style))
        doc.build(elements)
        return response

    # ====== Media Lookup ======
    media_lookup = {}
    if children:
        for child in children:
            psk_id = child.get('psk_id')
            if psk_id:
                media_lookup[psk_id] = child.get('media_files', [])
    else:
        for p in parents:
            media_lookup[p.get('psk_id')] = p.get('media_files', [])

    parents_sorted = sorted(parents, key=lambda x: (x.get('employee_id', ''), x.get('paper_title', '')))
    current_staff = None
    table_data = []
    col_widths = [3 * inch, 3 * inch, 0.7 * inch, 1.3 * inch, 2 * inch]

    for parent in parents_sorted:
        staff_id = parent.get('employee_id', 'N/A')
        staff_name = parent.get('staff_name', 'N/A')
        department_name = parent.get('department_name', 'N/A')
        psk_id = parent.get('psk_id')

        # ====== Flush Previous Staff Table ======
        if current_staff and current_staff != staff_id:
            # Only add table if there's data (more than just headers)
            if len(table_data) > 1:
                table = Table(table_data, colWidths=col_widths, repeatRows=1, splitByRow=True)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#02548b')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d6d6d6')),
                    ('LEFTPADDING', (0, 0), (-1, -1), 4),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                ]))
                elements.append(table)
            else:
                # Only headers exist, meaning no data for this staff
                elements.append(Paragraph("No data available for this staff", no_data_style))
            
            elements.append(PageBreak())
            table_data = []

        # ====== Add New Staff Section ======
        if current_staff != staff_id:
            elements.append(Paragraph("Research Articles", title_style))
            elements.append(Spacer(1, 6))

            # Staff info section
            left_style = ParagraphStyle('LeftAlign', parent=table_cell_style, alignment=TA_LEFT)
            right_style = ParagraphStyle('RightAlign', parent=table_cell_style, alignment=TA_RIGHT)
            center_style = ParagraphStyle('CenterAlign', parent=table_cell_style, alignment=TA_CENTER)

            # Line 1: Staff ID (left) | Department (right)
            staff_id_para = Paragraph(f"<b>Staff ID:</b> {staff_id}", left_style)
            dept_para = Paragraph(f"<b>Department:</b> {department_name}", right_style)
            line1 = Table([[staff_id_para, dept_para]], colWidths=[5.5 * inch, 4.5 * inch])
            line1.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
            ]))
            elements.append(line1)

            # Line 2: Staff Name (left only)
            staff_name_para = Paragraph(f"<b>Staff Name:</b> {staff_name}", left_style)
            line2 = Table([[staff_name_para]], colWidths=[10 * inch])
            line2.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
            ]))
            elements.append(line2)
            elements.append(Spacer(1, 10))

            # Table headers
            headers = ['Paper Title', 'Journal Name', 'Year', 'ISSN', 'Attachments']
            table_data.append([Paragraph(h, table_header_style) for h in headers])
            current_staff = staff_id

        # ====== Media Files ======
        media_files = media_lookup.get(psk_id, [])
        media_text = []
        seen_files = set()
        for media in media_files:
            fname = media.get("file_name", "Unknown")
            url = media.get("direct_api_url", "")
            if not url or fname in seen_files:
                continue
            seen_files.add(fname)
            display_name = fname if len(fname) <= 50 else fname[:50] + "..."
            media_text.append(f'<link href="{url}" color="blue">{display_name}</link>')

        media_paragraph = Paragraph('<br/>'.join(media_text) if media_text else "No attachments", attachment_link_style)

        pub_year = parent.get('publication_year', '')
        if isinstance(pub_year, int):
            pub_year = f"{pub_year}-{pub_year + 1}"
        elif isinstance(pub_year, str) and pub_year.isdigit():
            y = int(pub_year)
            pub_year = f"{y}-{y + 1}"
        else:
            pub_year = str(pub_year)
        # Data row
        row = [
            Paragraph(str(parent.get('paper_title', '')), table_cell_style),
            Paragraph(str(parent.get('journal_name', '')), table_cell_style),
            Paragraph(pub_year, table_cell_style),
            Paragraph(str(parent.get('issn_number', '')), table_cell_style),
            media_paragraph
        ]
        table_data.append(row)

    # ====== Final Table Flush ======
    if table_data:
        # Check if there's actual data (more than just headers)
        if len(table_data) > 1:
            table = Table(table_data, colWidths=col_widths, repeatRows=1, splitByRow=True)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#02548b')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d6d6d6')),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(table)
        else:
            # Only headers exist, meaning no data for this staff
            elements.append(Paragraph("No data available for this staff", no_data_style))

    doc.build(elements)
    return response

# def export_research_articles_to_excel(parents, children=None):
#     """
#     Export Research Articles data to Excel, including clickable media file links.
#     Media categories: Journal Cover Page, Index Page, Impact Factor, Citation Index.
#     """
#     import io
#     from django.http import HttpResponse
#     from openpyxl import Workbook
#     from openpyxl.styles import Font, Alignment

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Research Articles"

#     headers = [
#         "Employee ID", "Staff Name", "Department", "Paper Title",
#         "Author Name", "Journal Name", "Year", "ISSN Number",
#         "UGC Recognition Link", "Issue Date", "Journal Nature",
#         "DOI URL Link", "Journal Cover Page", "Index Page", "Impact Factor", "Citation Index"
#     ]
#     ws.append(headers)

#     # Style headers
#     for col in range(1, len(headers)+1):
#         cell = ws.cell(row=1, column=col)
#         cell.font = Font(bold=True)
#         cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

#     # Media categories
#     media_categories = ["Journal Cover Page", "Index Page", "Impact Factor", "Citation Index"]
#     for p in parents:
#         media_mapping = {cat: [] for cat in media_categories}
#         for media in p.get("media_files", []):
#             fname = media.get("file_name", "Unknown")
#             url = media.get("direct_api_url", "")
#             if not url:
#                 continue
#             # Map based on filename containing keywords
#             fname_upper = fname.upper()
#             if "JOURNAL COVER" in fname_upper:
#                 media_mapping["Journal Cover Page"].append((fname, url))
#             elif "INDEX" in fname_upper:
#                 media_mapping["Index Page"].append((fname, url))
#             elif "IMPACT FACTOR" in fname_upper:
#                 media_mapping["Impact Factor"].append((fname, url))
#             elif "CITATION" in fname_upper:
#                 media_mapping["Citation Index"].append((fname, url))
#             else:
#                 media_mapping["Journal Cover Page"].append((fname, url))

#         base_row = [
#             p.get("employee_id",""), p.get("staff_name",""), p.get("department_name",""),
#             p.get("paper_title",""), p.get("author_name",""), p.get("journal_name",""),
#             p.get("publication_year",""), p.get("issn_number",""), p.get("ugc_recognition_link",""),
#             p.get("issue_date",""), p.get("journal_nature",""), p.get("doi_url_link","")
#         ]
#         ws.append(base_row + ["", "", "", ""])
#         row_idx = ws.max_row

#         # Add media hyperlinks
#         for idx, cat in enumerate(media_categories, start=13):
#             files = media_mapping.get(cat, [])
#             cell = ws.cell(row=row_idx, column=idx)
#             if files:
#                 cell.value = "\n".join([f for f, _ in files])
#                 cell.hyperlink = files[0][1]
#                 cell.font = Font(color="0000EE", underline="single")
#             else:
#                 cell.value = "-"
#             cell.alignment = Alignment(wrap_text=True, vertical="top")

#     # Auto-fit columns
#     for col_idx, column_cells in enumerate(ws.columns, 1):
#         max_length = max(len(str(cell.value or "")) for cell in column_cells)
#         ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

#     for row_idx in range(1, ws.max_row+1):
#         max_lines = max(str(ws.cell(row=row_idx, column=col).value or "").count("\n")+1 for col in range(1, len(headers)+1))
#         ws.row_dimensions[row_idx].height = min(max_lines*15, 120)

#     ws.freeze_panes = "A2"
#     ws.auto_filter.ref = ws.dimensions

#     response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#     response["Content-Disposition"] = 'attachment; filename="research_articles.xlsx"'
#     with io.BytesIO() as buffer:
#         wb.save(buffer)
#         buffer.seek(0)
#         response.write(buffer.getvalue())
#     return response

def export_research_articles_to_excel(parents, children=None):
    """
    Export Research Articles data to Excel, including clickable media file links.
    Media categories: Journal Cover Page, Index Page, Impact Factor, Citation Index.
    Files are mapped based on naming convention: employeeid_field_year_index.pdf
    """
    import io
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Research Articles"

    headers = [
        "Employee ID", "Staff Name", "Department", "Paper Title",
        "Author Name", "Journal Name", "Year", "ISSN Number",
        "UGC Recognition Link", "Issue Date", "Journal Nature", "DOI URL Link",
        "Journal Cover Page", "Index Page", "Impact Factor", "Citation Index"
    ]
    ws.append(headers)

    # Style headers
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    media_categories = ["Journal Cover Page", "Index Page", "Impact Factor", "Citation Index"]
    fields = ['JCP', 'IP', 'IF', 'CI']

    for p in parents:
        # Initialize mapping
        media_mapping = {cat: [] for cat in media_categories}
        for media in p.get("media_files", []):
            fname = media.get("file_name", "Unknown")
            url = media.get("direct_api_url", "")
            if not url:
                continue

            # Determine category based on filename field
            matched = False
            for field, cat in zip(fields, media_categories):
                if f"_{field}_" in fname.upper():  # Matches "_JCP_" etc.
                    media_mapping[cat].append((fname, url))
                    matched = True
                    break
            if not matched:
                # Default to Journal Cover Page if no match
                media_mapping["Journal Cover Page"].append((fname, url))

        pub_year = p.get("publication_year", "")
        if isinstance(pub_year, int):
            pub_year = f"{pub_year}-{pub_year + 1}"
        elif isinstance(pub_year, str) and pub_year.isdigit():
            y = int(pub_year)
            pub_year = f"{y}-{y + 1}"
        else:
            pub_year = str(pub_year)        

        base_row = [
            p.get("employee_id",""), p.get("staff_name",""), p.get("department_name",""),
            p.get("paper_title",""), p.get("author_name",""), p.get("journal_name",""),
            pub_year, p.get("issn_number",""), p.get("ugc_recognition_link",""),
            p.get("issue_date",""), p.get("journal_nature",""), p.get("doi_url_link","")
        ]
        ws.append(base_row + ["", "", "", ""])
        row_idx = ws.max_row

        # Add media hyperlinks
        for idx, cat in enumerate(media_categories, start=13):
            files = media_mapping.get(cat, [])
            cell = ws.cell(row=row_idx, column=idx)
            if files:
                # Only show the first file name and hyperlink (you can extend if needed)
                cell.value = files[0][0]
                cell.hyperlink = files[0][1]
                cell.font = Font(color="0000EE", underline="single")
            else:
                cell.value = "-"
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-fit columns
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

    for row_idx in range(1, ws.max_row+1):
        max_lines = max(str(ws.cell(row=row_idx, column=col).value or "").count("\n")+1 for col in range(1, len(headers)+1))
        ws.row_dimensions[row_idx].height = min(max_lines*15, 120)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="research_articles.xlsx"'
    with io.BytesIO() as buffer:
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())
    return response
